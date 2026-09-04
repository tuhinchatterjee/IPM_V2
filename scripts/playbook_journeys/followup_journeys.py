"""Journeys K-N: the committee's own decisions, follow-up and inbound files.

Same rules as the first harness — real HTTP, real cookies, no fixtures.
"""

from __future__ import annotations

import io
import json
import sys
import urllib.error
import urllib.request
import uuid
import zipfile

sys.path.insert(0, __file__.rsplit("/", 1)[0])

from api_journeys import BASE, FAILURES, NOTES, Client, check  # noqa: E402


def upload(client: Client, path: str, filename: str, data: bytes,
           content_type: str = "application/octet-stream") -> tuple[int, object]:
    """A real multipart POST, built by hand so the parts are exactly ours."""
    boundary = "----journey" + uuid.uuid4().hex
    body = b"".join([
        f"--{boundary}\r\n".encode(),
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode(),
        f"Content-Type: {content_type}\r\n\r\n".encode(),
        data,
        f"\r\n--{boundary}--\r\n".encode(),
    ])
    req = urllib.request.Request(
        BASE + path, data=body, method="POST",
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}",
                 "Accept": "application/json"})
    try:
        with client.opener.open(req, timeout=180) as r:
            return r.status, json.loads(r.read() or b"null")
    except urllib.error.HTTPError as e:
        raw = e.read()
        try:
            return e.code, json.loads(raw or b"null")
        except Exception:
            return e.code, raw[:200]


def xlsx_with(rows: list[list[object]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Pack"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> int:
    admin = Client()
    admin.login("alex.rahman")
    _, cs, _ = admin.call("GET", "/playbook/committees")
    retail = {c["code"]: c for c in cs["committees"]}["retail-credit-risk-committee"]
    _, detail, _ = admin.call("GET", f"/playbook/committees/{retail['id']}")
    packs = detail["packs"]
    current = max(packs, key=lambda p: p["meeting_at"])
    pid = current["id"]

    # ---------------------------------------------------------------- K
    print("\nJourney K — the committee decides, and the decision is recorded")
    code, dec, _ = admin.call("POST", f"/playbook/packs/{pid}/decisions", {
        "title": "Tighten the Jeddah SME origination cut-off",
        "question": "Do we raise the cut-off score by 15 points for the Jeddah SME cohort?",
        "recommendation": "Raise it, and review after two months of bookings.",
        "alternatives": ["Leave the cut-off and tighten affordability instead",
                         "Raise it across all SME regions"],
        "impact": "Roughly 4% of Jeddah SME applications currently approved.",
    })
    check("K", "a decision can be raised", code in (200, 201), str(code)[:200])
    did = dec.get("id") if isinstance(dec, dict) else None
    check("K", "the decision is PENDING before the meeting",
          isinstance(dec, dict) and dec.get("status") in {"PENDING", "DRAFT", "OPEN"},
          str(dec.get("status") if isinstance(dec, dict) else dec)[:80])
    if did:
        code, made, _ = admin.call("POST", f"/playbook/decisions/{did}/decide", {
            "outcome": "APPROVED",
            "decision_text": "Approved. Cut-off rises 15 points for Jeddah SME from next cycle.",
            "conditions": "Reviewed after two months of bookings.",
        })
        check("K", "the chair can record the outcome", code == 200, str(made)[:160])
        check("K", "the outcome is stored against the decision",
              isinstance(made, dict) and made.get("status") == "APPROVED"
              and made.get("decided_at") and made.get("decided_by"),
              str(made.get("status") if isinstance(made, dict) else made)[:60])
        check("K", "the decision keeps its conditions",
              isinstance(made, dict) and "two months" in made.get("conditions", ""),
              str(made.get("conditions") if isinstance(made, dict) else "")[:80])
        NOTES.append(f"K: decision {did} recorded APPROVED with conditions")

    # ---------------------------------------------------------------- L
    print("\nJourney L — an action leaves the room with an owner and a date")
    code, act, _ = admin.call("POST", f"/playbook/packs/{pid}/actions", {
        "description": "Reprice the Jeddah SME cut-off in the origination policy "
                       "and confirm the booked volume impact.",
        "owner_id": 3,
        "due_date": "2026-10-15",
        "priority": "HIGH",
        "decision_id": did,
        "status": "OPEN",
    })
    check("L", "an action can be raised", code in (200, 201), str(act)[:200])
    aid = act.get("id") if isinstance(act, dict) else None
    check("L", "the action carries an owner and a due date",
          isinstance(act, dict) and act.get("owner_id") and act.get("due_date"),
          f"{act.get('owner_id') if isinstance(act, dict) else '?'} / "
          f"{act.get('due_date') if isinstance(act, dict) else '?'}")

    if aid:
        code, closed, _ = admin.call("POST", f"/playbook/actions/{aid}/close",
                                     {"evidence": "", "completed": True})
        check("L", "closing an action without evidence is refused",
              code in (400, 422), str(code))
        code, closed, _ = admin.call("POST", f"/playbook/actions/{aid}/close", {
            "evidence": "Policy PR #4471 merged; October bookings reconciled.",
            "completed": True})
        check("L", "an action closes on evidence", code == 200, str(closed)[:160])
        check("L", "the closed action keeps its evidence",
              isinstance(closed, dict)
              and "4471" in json.dumps(closed), json.dumps(closed)[:120])

        # planner follow-up: a real project, through the Planner's own service
        code, projects, _ = admin.call("GET", "/planner/projects")
        rows = (projects.get("projects") if isinstance(projects, dict) else None) or []
        check("L", "the planner has projects to hand off to", bool(rows), str(code))
        if rows:
            project_id = rows[0]["id"]
            # A planner task code is unique within its project, and the
            # action's own reference restarts at A1 each time the demo is
            # re-seeded. A run-specific code keeps the journey repeatable
            # without weakening the rule it would otherwise trip over.
            task_code = f"CJ-{uuid.uuid4().hex[:8]}"
            code, sent, _ = admin.call("POST", f"/playbook/actions/{aid}/planner",
                                       {"project_id": project_id,
                                        "task_code": task_code})
            check("L", "an action can be sent to the project planner",
                  code == 200, str(sent)[:200])
            check("L", "the action now carries its planner task",
                  isinstance(sent, dict) and sent.get("planner_task_id")
                  and sent.get("planner_project_id") == project_id,
                  str(sent.get("planner_task_id") if isinstance(sent, dict) else sent))
            # and the task really exists on the Planner side
            if isinstance(sent, dict) and sent.get("planner_task_id"):
                code, proj, _ = admin.call("GET", f"/planner/projects/{project_id}")
                found = json.dumps(proj)
                check("L", "the planner project now shows the task",
                      str(sent["planner_task_id"]) in found
                      or sent.get("reference", "zzz") in found, str(code))
                NOTES.append(f"L: action {aid} -> planner task "
                             f"{sent['planner_task_id']} in project {project_id}")
            # sending it twice is refused rather than duplicating the work
            code, again, _ = admin.call(
                "POST", f"/playbook/actions/{aid}/planner",
                {"project_id": project_id,
                 "task_code": f"CJ-{uuid.uuid4().hex[:8]}"})
            check("L", "a second handoff of the same action is refused",
                  code in (400, 409, 422), str(code))

    code, chase, _ = admin.call("GET", "/playbook/chase")
    check("L", "the chase list answers", code == 200, str(code))

    # ---------------------------------------------------------------- M
    print("\nJourney M — an inbound file is checked before it is opened")
    good = xlsx_with([
        ["Metric", "Value", "Note"],
        ["Jeddah SME approval rate", "41.2%", "From the branch pack"],
        ["<Finance> Review", "ok", "A legitimate title with angle brackets"],
    ])
    code, out = upload(admin, f"/playbook/packs/{pid}/import", "branch-pack.xlsx", good,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    check("M", "a real workbook imports", code in (200, 201), str(out)[:220])
    if code in (200, 201) and isinstance(out, dict):
        NOTES.append(f"M: import produced {json.dumps(out)[:120]}")

    hostile = [
        ("an executable renamed to xlsx", "payroll.xlsx",
         b"MZ\x90\x00" + b"\x00" * 2048,
         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("a file with no extension the product accepts", "notes.exe", good,
         "application/octet-stream"),
        ("an empty file", "empty.xlsx", b"",
         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    ]
    for label, name, data, ctype in hostile:
        code, out = upload(admin, f"/playbook/packs/{pid}/import", name, data, ctype)
        check("M", f"{label} is refused", code >= 400, str(code))

    # A traversal filename on an otherwise valid workbook is not refused —
    # the file is real. What must not survive is the NAME: no separator, no
    # `..`, nothing that could be joined onto a path later.
    code, out = upload(admin, f"/playbook/packs/{pid}/import",
                       "../../../etc/passwd.xlsx", good,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    check("M", "a valid workbook with a traversal name is accepted",
          code in (200, 201), str(code))
    if isinstance(out, dict):
        name = out.get("filename", "")
        check("M", "the traversal is stripped from the stored name",
              "/" not in name and "\\" not in name and ".." not in name, name)
        NOTES.append(f"M: '../../../etc/passwd.xlsx' stored as '{name}'")

    # a zip bomb: a small archive declaring an enormous member
    bomb = io.BytesIO()
    with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", "<Types/>")
        z.writestr("xl/worksheets/sheet1.xml", b"0" * (200 * 1024 * 1024))
    code, out = upload(admin, f"/playbook/packs/{pid}/import", "small.xlsx", bomb.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    check("M", "a zip declaring an enormous member is refused", code >= 400, str(code))

    # oversized body
    code, out = upload(admin, f"/playbook/packs/{pid}/import", "huge.xlsx",
                       b"PK\x03\x04" + b"\x00" * (60 * 1024 * 1024),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    check("M", "an oversized upload is refused", code in (400, 413, 422), str(code))
    NOTES.append("M: oversized upload refused with " + str(code))

    # ---------------------------------------------------------------- N
    print("\nJourney N — a formula in a cell leaves as text, not as a formula")
    code, blob, head = admin.call(
        "GET", f"/playbook/packs/{pid}/export?format=xlsx", raw=True)
    check("N", "the workbook downloads", code == 200 and blob[:2] == b"PK", str(code))
    if code == 200:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(blob))
        started: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v[:1] in "=+-@\t\r":
                        started.append(f"{ws.title}!{cell.coordinate}={v[:40]}")
        check("N", "no exported cell begins a formula", not started,
              "; ".join(started[:3]))
        NOTES.append(f"N: {len(wb.worksheets)} sheets scanned, "
                     f"{sum(ws.max_row for ws in wb.worksheets)} rows, "
                     f"{len(started)} formula-leading cells")

    # ---------------------------------------------------------------- O
    print("\nJourney O — a spreadsheet formula posted into a pack comes back inert")
    hostile_book = xlsx_with([
        ["Line", "Value", "Comment"],
        ["=cmd|'/c calc'!A1", "@SUM(1+1)*cmd|'/c calc'!A1", "+1+1"],
        ["-2+3+cmd|'/c calc'!A0", "=1+1", "\t=HYPERLINK(\"http://x\")"],
        ["<Finance> Review", "41.2%", "A legitimate title with angle brackets"],
    ])
    code, out = upload(admin, f"/playbook/packs/{pid}/import", "hostile.xlsx",
                       hostile_book,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    check("O", "the hostile workbook is accepted as content", code in (200, 201),
          str(out)[:200])

    code, blob, _ = admin.call(
        "GET", f"/playbook/packs/{pid}/export?format=xlsx", raw=True)
    check("O", "the pack still exports", code == 200 and blob[:2] == b"PK", str(code))
    if code == 200:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(blob))
        text = []
        started = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str):
                        text.append(v)
                        if v[:1] in "=+-@\t\r":
                            started.append(f"{ws.title}!{cell.coordinate}={v[:40]}")
        check("O", "not one cell in the exported workbook begins a formula",
              not started, "; ".join(started[:3]))
        joined = "\n".join(text)
        check("O", "the hostile text is still present, just inert",
              "calc" in joined, "the payload did not reach the export at all")
        check("O", "a legitimate angle-bracket title survives intact",
              "<Finance> Review" in joined,
              "the title was mangled or dropped by a blocklist")
        NOTES.append(f"O: {len(text)} exported strings, {len(started)} "
                     f"formula-leading; payload present={('calc' in joined)}; "
                     f"angle-bracket title present="
                     f"{('<Finance> Review' in joined)}")

    print("\n" + "=" * 68)
    for n in NOTES:
        print("  note: " + n)
    print("=" * 68)
    if FAILURES:
        print(f"\n{len(FAILURES)} FAILED:")
        for f in FAILURES:
            print("  - " + f)
        return 1
    print("\nAll journey checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
