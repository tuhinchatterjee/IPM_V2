"""§57 steps 4, 8, 9 and 10: open the downloaded files and check them."""
import json
import pathlib
import re
import sys

from openpyxl import load_workbook

DOWN = pathlib.Path("/tmp/accept")
results_path, pack_path = [
    pathlib.Path(line) for line in (DOWN / "paths.txt").read_text().split()
]
screen = json.loads((DOWN/"screen.json").read_text())
ok, bad = [], []

def check(name, condition, detail=""):
    (ok if condition else bad).append(name)
    print(("  PASS  " if condition else "  FAIL  ") + name + (f"  {detail}" if detail else ""))

# ---------------------------------------------------------- 4. results workbook
rb = load_workbook(results_path)
check("4. RESULTS is the first sheet", rb.sheetnames[0] == "RESULTS", str(rb.sheetnames))
check("4. SUMMARY is present", "SUMMARY" in rb.sheetnames)

ws = rb["RESULTS"]
grid = [[ws.cell(row=r, column=c).value for c in range(1, 6)] for r in range(1, 40)]
header = next(i for i, row in enumerate(grid)
              if all(isinstance(grid[i][c], str) and grid[i][c] for c in range(3)))
book_rows = []
for row in grid[header+1:]:
    if row[0] is None or str(row[0]) == "Total":
        break
    book_rows.append(row[:3])

# The on-screen table: header, then rows.
screen_rows = [r for r in screen[1:] if r and r[0]]
check("4. row count matches the interface",
      len(book_rows) == len(screen_rows), f"{len(book_rows)} vs {len(screen_rows)}")

#: How the interface scales a figure for reading, against the governed unit
#: the workbook carries. The screen says "usd bn" and shows 22.4; the workbook
#: says "USD mn" and carries 22,373.572. Both are labelled and both are right —
#: the screen is for reading, the workbook is the record — so the comparison
#: is made after putting them in the same unit rather than by pretending the
#: rounded figure is the exact one.
SCALE = {"bn": 1_000.0, "tn": 1_000_000.0, "mn": 1.0, "k": 0.001}


def number(text):
    m = re.search(r"-?[\d,]+\.?\d*", str(text).replace("−", "-"))
    return float(m.group().replace(",", "")) if m else None


def scale_of(heading: str) -> float:
    """How many governed units one displayed unit is worth."""
    lowered = str(heading).lower()
    for suffix, factor in SCALE.items():
        if re.search(rf"\busd\s*{suffix}\b", lowered):
            return factor
    return 1.0

factor = scale_of(screen[0][1] if screen and len(screen[0]) > 1 else "")
mismatch = []
for i, (book, seen) in enumerate(zip(book_rows, screen_rows, strict=False)):
    if str(book[0]) != str(seen[0]).strip():
        mismatch.append(f"row {i+1} label: {book[0]!r} vs {seen[0]!r}")
        continue
    on_screen = number(seen[1])
    if on_screen is None:
        continue
    expected = on_screen * factor
    # The interface rounds to one decimal for reading; the workbook carries the
    # figure the engine produced. Equal to the precision the screen showed.
    if abs(float(book[1]) - expected) > max(0.05 * factor, abs(expected) * 0.001):
        mismatch.append(f"row {i+1}: workbook {book[1]} vs screen {on_screen} "
                        f"× {factor:g}")
check("4. every value matches the interface", not mismatch, "; ".join(mismatch[:3]))

# ------------------------------------------------------ 8. the calculation pack
pb = load_workbook(pack_path)
REQUIRED = ["COVER", "SOURCE PROFILES", "RELATIONSHIPS & JOINS",
            "JOIN RECONCILIATION", "CALCULATION STEPS", "FORMULAS & QUERY",
            "VALIDATION CHECKS", "TRACE LEDGER", "INTERPRETATION EVIDENCE",
            "FINAL RESULTS"]
for sheet in REQUIRED:
    check(f"8. {sheet}", sheet in pb.sheetnames)
check("8. FINAL RESULTS is the last tab", pb.sheetnames[-1] == "FINAL RESULTS")

sql = pb["FORMULAS & QUERY"]
text = " ".join(str(sql.cell(row=r, column=1).value or "") for r in range(1, 220))
check("8. the SQL that ran is in the pack", "SELECT" in text and "read_parquet" in text)

# --------------------------------------------- 9. the rating totals reconcile
final = pb["FINAL RESULTS"]
grid = [[final.cell(row=r, column=c).value for c in range(1, 4)] for r in range(1, 60)]
# The header is the row where the dimension and the measure headings sit side
# by side. Matching on a word found the "Answer" row, which quotes the measure
# in a sentence — the shape is what distinguishes a heading from prose.
# The header is the only row with THREE headings side by side. The provenance
# block above it uses two columns, and matching on a word found the "Validation
# status" row, whose value happens to contain "invariant(s)".
head = next(i for i, row in enumerate(grid)
            if all(isinstance(cell, str) and cell for cell in row))
values = []
for row in grid[head+1:]:
    if row[0] is None:
        break
    if isinstance(row[1], (int, float)):
        values.append(float(row[1]))
total = sum(values)

profiles = pb["SOURCE PROFILES"]
source_total = None
for r in range(1, 400):
    label = str(profiles.cell(row=r, column=1).value or "")
    if label.strip().lower() in ("total ead", "total ccf-adjusted ead"):
        value = profiles.cell(row=r, column=2).value
        if isinstance(value, (int, float)):
            source_total = float(value)
            break
if source_total is None:
    # Fall back to the numeric field profile's own sum for `ead`.
    for r in range(1, 400):
        if str(profiles.cell(row=r, column=1).value or "") == "ead":
            value = profiles.cell(row=r, column=8).value
            if isinstance(value, (int, float)):
                source_total = float(value)
                break
check("9. rating-wise EAD sums to the included total",
      source_total is not None and abs(total - source_total) <= 0.05,
      f"{total:,.3f} vs {source_total if source_total is None else f'{source_total:,.3f}'}")

# ---------------------------------------------------- 10. nothing unauthorised
for path in (results_path, pack_path):
    blob = pathlib.Path(path).read_bytes()
    leaked = [t for t in (b"sk-ant-", b"ANTHROPIC_API_KEY", b"Bearer ",
                          b"Authorization:", b"password", b"/home/user")
              if t in blob]
    check(f"10. no secrets in {pathlib.Path(path).name[:34]}", not leaked, str(leaked))

print(f"\nworkbooks: {len(ok)} passed, {len(bad)} failed")
sys.exit(1 if bad else 0)
