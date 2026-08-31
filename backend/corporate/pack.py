"""The Borrower 360 pack: seventeen sheets a reviewer can work from. Phase 3.12.

An export leaves the system and stops being governed by it, so this workbook
carries with it everything that made the numbers on screen defensible: where
each field came from, which quarter's row it was read from, what the derived
graph computed and refused to compute, and which data-quality checks were
running when it did.

Two rules the workbook holds
----------------------------
**Every sheet says what it is not.** The network measures sheet carries the
ranking banner, the connectedness sheet carries B54's caveat, and the limits
sheet carries UNVERIFIED REGULATORY PARAMETER. A workbook read six months
later, away from the screen that explained them, is exactly where an
unqualified number does its damage.

**A sentinel survives the export.** NOT_AVAILABLE, NOT_APPLICABLE and
DATA_QUALITY_BLOCKED are written into the cell as text. Exporting them as
blanks - which is what a naive `to_excel` does - turns four different
statements into one empty cell, and Excel's own reading of an empty cell is
zero.
"""

from __future__ import annotations

import io
import logging
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.corporate import NOT_CLIENT_DATA, ORIGIN
from backend.corporate import graphquality as gq
from backend.corporate import graphsummary as gs
from backend.corporate import lineage as lineage_mod
from backend.corporate import network as net
from backend.corporate import service as service_mod

logger = logging.getLogger(__name__)

PACK_VERSION = "1.0.0"

HEADER_FILL = PatternFill("solid", fgColor="0B2436")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="16232F")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="6C7A8C")
BODY_FONT = Font(name="Calibri", size=10, color="16232F")

#: The seventeen sheets, in the order a reader meets them. COVER first
#: because it says what the workbook is; DATA QUALITY last but one and
#: LINEAGE last, because those are what a reviewer turns to when they want to
#: challenge something earlier in the pack.
SHEETS: tuple[str, ...] = (
    "COVER",
    "IDENTITY",
    "EXPOSURE",
    "RATING",
    "IFRS 9",
    "FINANCIALS",
    "COVENANTS",
    "COLLATERAL",
    "DELINQUENCY",
    "LIMITS",
    "GROUPS",
    "OWNERSHIP EDGES",
    "CONTROL AND UBO",
    "GUARANTEES",
    "SUPPLY CHAIN",
    "NETWORK MEASURES",
    "DATA QUALITY",
    "LINEAGE",
)

#: Lineage group -> the sheet its fields belong on.
GROUP_SHEET: dict[str, str] = {
    "IDENTITY": "IDENTITY",
    "EXPOSURE": "EXPOSURE",
    "RATING": "RATING",
    "IFRS9": "IFRS 9",
    "FINANCIALS": "FINANCIALS",
    "COVENANTS": "COVENANTS",
    "COLLATERAL": "COLLATERAL",
    "DELINQUENCY": "DELINQUENCY",
    "LIMIT": "LIMITS",
    "GRAPH SUMMARY": "NETWORK MEASURES",
    "DATA QUALITY": "DATA QUALITY",
}

#: What each sheet is NOT, written onto the sheet. A workbook read six months
#: later, away from the screen, is exactly where an unqualified number does
#: its damage.
SHEET_CAVEAT: dict[str, str] = {
    "GROUPS": (
        "These six groupings answer different questions and do not agree by "
        "design. Graph connectivity is not regulatory connectedness: the "
        "connected counterparty group is a CANDIDATE for assessment under "
        "the institution's own approved criteria, not a determination."),
    "CONTROL AND UBO": (
        "Control is binary, absorptive and transitive, and is computed over "
        "VOTING rights. It is NOT proportional ownership: 51% of 51% is 26% "
        "of the economics and 100% of the control."),
    "NETWORK MEASURES": net.NRS_LABEL + " " + net.DEBTRANK_CAVEAT,
    "LIMITS": (
        "The single-name and group limit thresholds are UNVERIFIED "
        "REGULATORY PARAMETERS carried from the framework document. They "
        "have not been confirmed as currently binding law."),
    "SUPPLY CHAIN": (
        "A supply relationship never forms a regulatory group on its own. "
        "It is one input to the interdependence test, and only a VALIDATED "
        "predicate merges a member in."),
    "OWNERSHIP EDGES": (
        "Observed assertions as filed, not a reconciled register. Where two "
        "source systems disagree, both rows are here and neither has been "
        "chosen."),
    "DATA QUALITY": (
        "A REJECT blocks the derived computation that depends on it. Fields "
        "elsewhere in this pack reading DATA_QUALITY_BLOCKED were not "
        "computed, and the reason is on this sheet."),
    "LINEAGE": (
        "No field in the Borrower 360 is AUTHORITATIVE. Every one is a copy "
        "or a derivation, and this sheet names the domain that owns it."),
}

#: The three sentinels, written as text rather than left blank. Excel reads
#: an empty cell as zero in almost every formula a reviewer will write.
SENTINELS: frozenset[str] = frozenset({
    gs.NOT_AVAILABLE, gs.NOT_APPLICABLE, gs.DATA_QUALITY_BLOCKED,
    "NOT COMPUTED", "PERMISSION_REQUIRED",
})


class PackError(RuntimeError):
    """The pack cannot be built as asked."""


def _write_block(sheet, *, title: str, note: str, columns: list[str],
                 rows: list[list[Any]], start: int = 1) -> int:
    """One titled block. Returns the next free row."""
    sheet.cell(row=start, column=1, value=title).font = TITLE_FONT
    line = start + 1
    if note:
        cell = sheet.cell(row=line, column=1, value=note)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        line += 1
    line += 1

    if columns:
        for index, name in enumerate(columns, start=1):
            cell = sheet.cell(row=line, column=index, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        line += 1

    for row in rows:
        for index, value in enumerate(row, start=1):
            cell = sheet.cell(row=line, column=index, value=_cell(value))
            cell.font = BODY_FONT
        line += 1

    widths: dict[int, int] = {}
    for row in ([columns] if columns else []) + rows:
        for index, value in enumerate(row, start=1):
            widths[index] = min(max(widths.get(index, 10),
                                    len(str(value)) + 2), 70)
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width
    return line + 1


def _cell(value: Any) -> Any:
    """One value, as Excel should hold it.

    A sentinel stays a string. A NaN becomes the string NOT COMPUTED rather
    than an empty cell, because Excel reads an empty cell as zero in almost
    every formula a reviewer will write over this pack.
    """
    if value is None:
        return "NOT COMPUTED"
    if isinstance(value, float) and pd.isna(value):
        return "NOT COMPUTED"
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _frame_rows(frame: pd.DataFrame, columns: list[str]
                ) -> list[list[Any]]:
    present = [name for name in columns if name in frame.columns]
    return [[row[name] for name in present]
            for _, row in frame[present].iterrows()]


def build(borrower_id: str, period: str, *,
          include_people: bool = True) -> bytes:
    """The whole pack for one borrower-quarter. Phase 3.12.

    `include_people` follows the caller's BORROWER_360_UBO_VIEW. When it is
    false the two people sheets are still PRESENT and say they were withheld:
    a missing sheet reads as "there is nothing here", which is a claim about
    the borrower rather than about the reader.
    """
    row = service_mod.borrower_row(borrower_id, period)
    stamp = service_mod.as_of_date(period)
    book = Workbook()
    book.remove(book.active)

    _cover(book, borrower_id, period, stamp, row, include_people)
    _field_sheets(book, row)
    _groups(book, borrower_id, period)
    _edges(book, borrower_id, period, stamp, include_people)
    _network(book, borrower_id, period)
    _quality(book, period)
    _lineage(book)

    missing = [name for name in SHEETS if name not in book.sheetnames]
    if missing:
        raise PackError(
            f"the pack is incomplete: {', '.join(missing)} were not written. "
            "A sheet the index promises and the workbook does not have is a "
            "defect a reader finds instead of a test.")
    book._sheets.sort(key=lambda sheet: SHEETS.index(sheet.title))

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _cover(book: Workbook, borrower_id: str, period: str, stamp: str,
           row: pd.Series, include_people: bool) -> None:
    sheet = book.create_sheet("COVER")
    line = _write_block(
        sheet,
        title="BORROWER 360 PACK",
        note=f"{ORIGIN} — {NOT_CLIENT_DATA}",
        columns=["Item", "Value"],
        rows=[
            ["Borrower", str(row.get("legal_name", borrower_id))],
            ["Borrower id", borrower_id],
            ["Quarter", period],
            ["As at", stamp],
            ["Sector", str(row.get("sector", ""))],
            ["Region", str(row.get("region", ""))],
            ["Internal rating", str(row.get("internal_rating", ""))],
            ["IFRS 9 stage", str(row.get("stage", ""))],
            ["Pack version", PACK_VERSION],
            ["Graph method version", net.NETWORK_VERSION],
            ["Graph policy version", net.POLICY_VERSION],
            ["Quality version", gq.QUALITY_VERSION],
            ["Natural persons included", "Yes" if include_people else
             "No — withheld by permission"],
        ])
    _write_block(
        sheet, start=line + 1,
        title="WHAT THIS PACK IS NOT",
        note="",
        columns=["Subject", "Statement"],
        rows=[
            ["Client data", NOT_CLIENT_DATA],
            ["Network Risk Score", net.NRS_LABEL],
            ["DebtRank", net.DEBTRANK_CAVEAT],
            ["Connected groups", SHEET_CAVEAT["GROUPS"]],
            ["Control", SHEET_CAVEAT["CONTROL AND UBO"]],
            ["Limits", SHEET_CAVEAT["LIMITS"]],
            ["Similarity", net.SIMILARITY_CAVEAT],
        ])
    _write_block(
        sheet, start=sheet.max_row + 2,
        title="SHEET INDEX", note="",
        columns=["#", "Sheet"],
        rows=[[index, name] for index, name in enumerate(SHEETS, start=1)])


def _field_sheets(book: Workbook, row: pd.Series) -> None:
    """One sheet per lineage group, each field with where it came from."""
    for group, sheet_name in GROUP_SHEET.items():
        fields = [entry for entry in lineage_mod.FIELDS
                  if entry.group == group and entry.name in row.index]
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            start = sheet.max_row + 2
        else:
            sheet = book.create_sheet(sheet_name)
            start = 1
        _write_block(
            sheet, start=start,
            title=f"{sheet_name} — {group}",
            note=SHEET_CAVEAT.get(sheet_name, ""),
            columns=["Field", "Value", "Unit", "Source dataset",
                     "Source field", "Source period", "Authority",
                     "Transformation"],
            rows=[[entry.name, row[entry.name], entry.unit,
                   entry.source_dataset, entry.source_field,
                   entry.source_period, entry.authority,
                   entry.transformation]
                  for entry in fields])


def _groups(book: Workbook, borrower_id: str, period: str) -> None:
    sheet = book.create_sheet("GROUPS")
    view = service_mod.group_view(borrower_id, period)
    _write_block(
        sheet,
        title="GROUP AND CONNECTEDNESS — six concepts, not reconciled",
        note=SHEET_CAVEAT["GROUPS"],
        columns=["Concept", "Value", "Answers", "Computed from", "Is NOT"],
        rows=[[concept["label"], concept.get("value", ""),
               concept["question"], concept["basis"], concept["is_not"]]
              for concept in view["concepts"]])


def _edges(book: Workbook, borrower_id: str, period: str, stamp: str,
           include_people: bool) -> None:
    """The observed relationships, each family on its own sheet."""
    plans: tuple[tuple[str, str, tuple[str, ...], bool], ...] = (
        ("OWNERSHIP EDGES", "ownership",
         ("edge_id", "edge_type", "from_node", "to_node", "ownership_pct",
          "voting_pct", "valid_from", "valid_to", "recorded_at", "source",
          "confidence"), False),
        ("CONTROL AND UBO", "control",
         ("edge_id", "edge_type", "from_node", "to_node", "voting_pct",
          "role", "source", "confidence"), True),
        ("GUARANTEES", "guarantees",
         ("edge_id", "edge_type", "from_node", "to_node", "source",
          "confidence"), False),
        ("SUPPLY CHAIN", "supply",
         ("edge_id", "edge_type", "from_node", "to_node",
          "share_of_supplier_revenue", "share_of_buyer_cogs", "source",
          "confidence"), False),
    )
    for sheet_name, view, columns, needs_people in plans:
        sheet = book.create_sheet(sheet_name)
        if needs_people and not include_people:
            _write_block(
                sheet,
                title=f"{sheet_name} — WITHHELD",
                note=("This sheet shows named natural persons and requires "
                      "BORROWER_360_UBO_VIEW. It is present and empty "
                      "because you are not permitted to see it — which is "
                      "different from this borrower having no owners."),
                columns=["Status"],
                rows=[["PERMISSION_REQUIRED"]])
            continue

        found = service_mod.ego_graph(borrower_id, period, view=view, depth=2)
        frame = pd.DataFrame(found.edges)
        _write_block(
            sheet,
            title=f"{sheet_name} — as at {stamp}, two steps from "
                  f"{borrower_id}",
            note=SHEET_CAVEAT.get(sheet_name, "") +
            (f" {found.truncation_note}" if found.truncated else ""),
            columns=[name for name in columns
                     if not frame.empty and name in frame.columns],
            rows=_frame_rows(frame, list(columns)) if not frame.empty else [])


def _network(book: Workbook, borrower_id: str, period: str) -> None:
    sheet = book["NETWORK MEASURES"] if "NETWORK MEASURES" in book.sheetnames \
        else book.create_sheet("NETWORK MEASURES")
    frame = service_mod.load(gs.GROUPS_DATASET)
    block = frame[(frame["borrower_id"] == borrower_id)
                  & (frame["period"] == period)]
    if block.empty:
        rows = [["Derived graph", gs.NOT_AVAILABLE,
                 "The derivation has not been run for this quarter."]]
    else:
        found = block.iloc[0]
        rows = []
        for column, status_column in sorted(gs.MEASURE_STATUS.items()):
            status = str(found.get(status_column, gs.NOT_AVAILABLE))
            value = (found[column] if status == gs.AVAILABLE else status)
            rows.append([column, value, status])
    _write_block(
        sheet, start=sheet.max_row + 2,
        title="NETWORK MEASURES — derived",
        note=SHEET_CAVEAT["NETWORK MEASURES"],
        columns=["Measure", "Value", "Status"],
        rows=rows)


def _quality(book: Workbook, period: str) -> None:
    sheet = book["DATA QUALITY"] if "DATA QUALITY" in book.sheetnames \
        else book.create_sheet("DATA QUALITY")
    register = service_mod.load(gs.DQ_DATASET)
    block = register[register["period"] == period]
    columns = ["check_id", "check", "status", "observed", "threshold",
               "scope", "affected_entities", "blocks"]
    _write_block(
        sheet, start=sheet.max_row + 2,
        title=f"GRAPH DATA QUALITY — {period}",
        note=SHEET_CAVEAT["DATA QUALITY"],
        columns=columns,
        rows=_frame_rows(block, columns))


def _lineage(book: Workbook) -> None:
    sheet = book.create_sheet("LINEAGE")
    _write_block(
        sheet,
        title="LINEAGE — where every Borrower 360 field comes from",
        note=SHEET_CAVEAT["LINEAGE"],
        columns=["Field", "Group", "Source domain", "Source dataset",
                 "Source field", "Source period", "Transformation",
                 "Authority", "Unit"],
        rows=[[entry.name, entry.group, entry.source_domain,
               entry.source_dataset, entry.source_field, entry.source_period,
               entry.transformation, entry.authority, entry.unit]
              for entry in lineage_mod.FIELDS])
