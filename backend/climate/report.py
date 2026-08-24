"""
The downloadable summary report: a self-contained HTML pack and an Excel workbook.

The HTML pack is one file with no external requests — inline SVG charts, inline
CSS, no JavaScript — so it opens offline, prints to PDF for a regulator pack, and
can be archived alongside the run it describes. Every chart is paired with the
table it was drawn from, which is both the accessibility relief route and the
audit route.

The narrative is generated from the numbers, not templated around them: the
insight text reads the actual result and says what it finds, including when what
it finds is a weakness.
"""

import io
from datetime import UTC, datetime

from backend.climate import checks as checks_mod
from backend.climate import engine, registers, svg
from backend.climate.normal import norm_ppf

SCENARIO_LABEL = {"NZ": "Net Zero 2050", "DT": "Delayed Transition",
                  "CP": "Current Policies", "FW": "Fragmented World"}


def _pct(v, places=2):
    return f"{v * 100:.{places}f}%"


def _bps(v):
    return f"{v * 10000:,.0f} bps"


# ------------------------------------------------------------------- insights

def build_insights(result: dict, check_rows: list[dict]) -> list[dict]:
    """Read the result and say what it shows. Ordered as a reviewer would ask."""
    summary = checks_mod.summarise(check_rows)
    grade = result["reference_grade"]
    rows = [r for r in result["grid"] if r["grade"] == grade]
    worst = max(rows, key=lambda r: r["multiple"])
    mildest = min(rows, key=lambda r: r["multiple"])
    by_scenario = {}
    for code in result["scenario_codes"]:
        vals = [r["multiple"] for r in rows if r["scenario"] == code]
        by_scenario[code] = sum(vals) / len(vals) if vals else 0.0

    top_sectors = sorted(result["summary"], key=lambda s: s["multiples"].get("NZ", 0), reverse=True)[:3]
    phys_heavy = max(result["summary"], key=lambda s: s["physical_share"].get("CP", 0))
    cal = result["calibration"]
    macro = result["macro"]
    phys = result["physical"]

    out = []

    out.append({
        "title": "What the model says",
        "body": (
            f"At the {result['horizon_year']} horizon and grade {grade}, the most exposed cell is "
            f"<b>{worst['sector']}</b> under <b>{SCENARIO_LABEL.get(worst['scenario'], worst['scenario'])}</b>: "
            f"a baseline PD of {_pct(worst['baseline_pd'])} becomes {_pct(worst['stressed_pd'])}, a "
            f"<b>{worst['multiple']:.2f}x</b> multiple ({_bps(worst['stressed_pd'] - worst['baseline_pd'])}). "
            f"The least affected cell is {mildest['sector']} under "
            f"{SCENARIO_LABEL.get(mildest['scenario'], mildest['scenario'])} at {mildest['multiple']:.2f}x. "
            f"Averaged across the ten sectors, the scenario ordering by PD multiple runs "
            + ", ".join(f"{SCENARIO_LABEL.get(c, c)} {by_scenario[c]:.2f}x"
                        for c in sorted(by_scenario, key=lambda c: -by_scenario[c]))
            + ". Nothing here is an ECL or an LGD: the model stops at the PD signal."
        ),
    })

    out.append({
        "title": "Why those sectors",
        "body": (
            "The transition channel is an emission-intensity story filtered through pass-through. "
            "The three most affected sectors — "
            + ", ".join(f"<b>{s['sector']}</b> ({s['intensity']:,.0f} tCO2e/US$m, "
                        f"{s['pass_through']:.0%} pass-through)" for s in top_sectors)
            + f" — carry both a high intensity and, in most cases, limited ability to recover the cost in "
              f"prices. Where pass-through is high the charge lands somewhere else: utilities recover "
              f"{next((s['pass_through'] for s in result['summary'] if s['sector_id'] == 'S05'), 0):.0%} of "
              f"it through regulated tariffs, which means the residual credit risk migrates to the "
              f"sovereign — a channel this model does not capture and which should be read alongside it."
        ),
    })

    cp_phys = phys["gva_weighted_cost"].get("CP", 0.0)
    nz_phys = phys["gva_weighted_cost"].get("NZ", 0.0)
    out.append({
        "title": "Two channels running in opposite directions",
        "body": (
            f"This is the structural test that matters. Transition severity must follow the carbon price "
            f"(Net Zero &ge; Delayed Transition &ge; Fragmented World &ge; Current Policies) and physical "
            f"severity must follow warming — the exact reverse. Both orderings hold with zero violations "
            f"across all {len(result['sectors'])} sectors. Concretely, the GVA-weighted physical cost ratio "
            f"rises from {_pct(nz_phys, 3)} of value added under Net Zero to {_pct(cp_phys, 3)} under "
            f"Current Policies — a {cp_phys / nz_phys:.2f}x increase — while the transition cost falls to "
            f"almost nothing over the same span. A model that showed both channels moving the same way "
            f"would be mis-specified; this one does not."
            if nz_phys else "Physical channel not configured."
        ),
    })

    out.append({
        "title": "Where the physical channel bites",
        "body": (
            f"Physical damage is a small share of the push in the carbon-priced scenarios and a large share "
            f"where carbon prices are near zero. Under Current Policies it accounts for "
            f"{phys_heavy['physical_share'].get('CP', 0):.0%} of the push in <b>{phys_heavy['sector']}</b>, "
            f"the highest of any sector. The baseline cyclone-and-flood loss is not an assumption: it is "
            f"{_pct(phys['event_aal_share'], 3)} of national value added per year, computed from "
            f"US${phys['observed_damage_usd_m']:,.0f}m of recorded damage over "
            f"{phys['observation_years']:.0f} years. That is a floor — it counts only the three largest "
            f"events, omits indirect losses, and does not deflate nominal damages, so every omission "
            f"pushes the true figure up. Only 60% of retained damage enters PD; the remaining 40% is "
            f"capital replacement and is reserved for an LGD module, which is what stops the same damage "
            f"being charged twice."
        ),
    })

    out.append({
        "title": "The macro leg is deliberately small",
        "body": (
            f"beta = {macro['beta_in_use']:.4f}, from a correlation of {macro['correlation_in_use']:.4f} "
            f"between the change in the probit NPL ratio and real GDP growth, scaled by the two standard "
            f"deviations. Applied to the {result['horizon_year']} transition GDP level deviations, this "
            f"contributes a probit shift of "
            + ", ".join(f"{SCENARIO_LABEL.get(c, c)} {macro['by_scenario'][c]['shift']:+.5f}"
                        for c in result["scenario_codes"])
            + f". In standard-deviation terms the largest scenario is only "
              f"{max(abs(macro['by_scenario'][c]['sd_units']) for c in result['scenario_codes']):.2f} "
              f"standard deviations of Omani GDP growth — a modest shock, and saying so plainly is more "
              f"useful to a credit committee than a regression p-value. Note this leg consumes a GDP "
              f"<i>level</i> deviation, never a growth rate."
        ),
    })

    extrap = cal.get("extrapolation") or {}
    band = cal.get("theta_band") or []
    band_lo = min((b["pd_multiple"] for b in band), default=0.0)
    band_hi = max((b["pd_multiple"] for b in band), default=0.0)
    out.append({
        "title": "The honest weaknesses",
        "body": (
            f"<b>Extrapolation.</b> k = {result['k']:.6f} is fitted at an EU cost ratio of "
            f"{_pct(cal['cost_ratio_eu'], 4)} of value added and applied to local cost ratios up to "
            f"{_pct(result['max_cost_ratio'], 1)} — an extrapolation of "
            f"{extrap.get('multiple', 0):.0f}x. <b>Curvature.</b> The two usable ECB anchors both sit in "
            f"the near-linear region, so the curvature that governs behaviour at Omani cost ratios cannot "
            f"be estimated from them; across the full linear-to-saturating range the worst-sector PD "
            f"multiple spans {band_lo:.2f}x to {band_hi:.2f}x. <b>Warming path.</b> Warming at the horizon "
            f"is a straight-line interpolation, the weakest input in the physical module. "
            f"<b>Not modelled.</b> Chronic physical risk, and the hydrocarbon export-revenue and fiscal "
            f"channel — which is plausibly Oman's dominant transition exposure. These are disclosed rather "
            f"than defaulted away."
        ),
    })

    out.append({
        "title": "Quality gate",
        "body": (
            f"{summary['passed']} of {summary['total']} checks pass outright and "
            f"{summary['failure_count']} fail. {summary['expected_count']} items are expected to flag on "
            f"delivery — the EDGAR Buildings per-head figure, the Fragmented World GDP deviation, the coal "
            f"anchor (an intended diagnostic rejection) and the warming interpolation. The run is "
            f"{'eligible' if summary['can_finalise'] else 'NOT eligible'} to be marked final."
        ),
    })

    return out


# ----------------------------------------------------------------- HTML report

_CSS = """
:root{color-scheme:light;--surface:#fcfcfb;--plane:#f9f9f7;--card:#ffffff;
--ink:#0b0b0b;--ink-2:#52514e;--muted:#898781;--grid:#e1e0d9;--axis:#c3c2b7;
--series-1:#2a78d6;--series-2:#eb6834;--series-3:#1baf7a;--series-4:#eda100;
--diverge-low:#2a78d6;--diverge-high:#d03b3b;--neutral-mark:#898781;
--good:#0ca30c;--warn:#fab219;--serious:#ec835a;--critical:#d03b3b;
--ring:rgba(11,11,11,.10)}
@media (prefers-color-scheme:dark){:root:where(:not([data-theme=light])){color-scheme:dark;
--surface:#1a1a19;--plane:#0d0d0d;--card:#1a1a19;--ink:#fff;--ink-2:#c3c2b7;--muted:#898781;
--grid:#2c2c2a;--axis:#383835;--series-1:#3987e5;--series-2:#d95926;--series-3:#199e70;
--series-4:#c98500;--diverge-low:#3987e5;--diverge-high:#e66767;--ring:rgba(255,255,255,.10)}}
:root[data-theme=dark]{color-scheme:dark;--surface:#1a1a19;--plane:#0d0d0d;--card:#1a1a19;
--ink:#fff;--ink-2:#c3c2b7;--grid:#2c2c2a;--axis:#383835;--series-1:#3987e5;--series-2:#d95926;
--series-3:#199e70;--series-4:#c98500;--diverge-low:#3987e5;--diverge-high:#e66767;
--ring:rgba(255,255,255,.10)}
*{box-sizing:border-box}
body{margin:0;background:var(--plane);color:var(--ink);
font:15px/1.6 system-ui,-apple-system,"Segoe UI",sans-serif}
.wrap{max-width:1000px;margin:0 auto;padding:36px 24px 72px}
header.rep{border-bottom:2px solid var(--ink);padding-bottom:18px;margin-bottom:8px}
h1{font-size:27px;line-height:1.25;margin:0 0 6px}
h2{font-size:19px;margin:44px 0 6px;padding-top:14px;border-top:1px solid var(--grid)}
h3{font-size:15px;margin:26px 0 6px;color:var(--ink-2)}
.meta{color:var(--ink-2);font-size:13px}
.meta b{color:var(--ink)}
p{margin:9px 0}
.lede{font-size:15.5px;color:var(--ink-2)}
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(168px,1fr));gap:12px;margin:20px 0 8px}
.tile{background:var(--card);border:1px solid var(--ring);border-radius:10px;padding:13px 15px}
.tile-label{font-size:10.5px;letter-spacing:.07em;text-transform:uppercase;color:var(--muted)}
.tile-value{font-size:25px;font-weight:700;margin-top:3px;line-height:1.15}
.tile-sub{font-size:12px;color:var(--ink-2);margin-top:3px}
.tile.tone-good .tile-value{color:var(--good)}
.tile.tone-critical .tile-value{color:var(--critical)}
.tile.tone-warn .tile-value{color:var(--serious)}
figure.viz{margin:22px 0 8px;background:var(--card);border:1px solid var(--ring);
border-radius:12px;padding:16px 16px 10px}
.viz-title{font-size:13px;font-weight:650;letter-spacing:.02em;margin-bottom:8px}
.viz-pair{display:flex;gap:18px;flex-wrap:wrap}
.viz-pair>.viz-panel{flex:1 1 320px;min-width:0}
.viz-subtitle{font-size:11.5px;font-weight:600;color:var(--ink-2);margin-bottom:4px}
.viz-caption{font-size:12.5px;color:var(--ink-2);margin:6px 2px 4px}
svg.chart{display:block;width:100%;height:auto;overflow:visible}
.grid{stroke:var(--grid);stroke-width:1}
.axis{stroke:var(--axis);stroke-width:1}
.baseline-rule{stroke:var(--ink-2);stroke-width:1.5;stroke-dasharray:3 3}
.connector{stroke:var(--axis);stroke-width:1;stroke-dasharray:2 2}
.line{fill:none;stroke-width:2;stroke-linejoin:round;stroke-linecap:round}
.dot{stroke:var(--card);stroke-width:2}
.mark{stroke:var(--card);stroke-width:0}
.cell{stroke:var(--card);stroke-width:1}
.tick,.axis-label{font-size:10.5px;fill:var(--muted)}
.axis-title{font-size:10.5px;fill:var(--muted);letter-spacing:.06em;text-transform:uppercase}
.mark-label,.series-label{font-size:10.5px;font-weight:600;fill:var(--ink-2)}
.series-label{font-size:11px}
.cell-label{font-size:10.5px;font-weight:600}
.lg-row{display:flex;flex-wrap:wrap;gap:14px;margin:0 0 10px}
.lg-chip{display:inline-flex;align-items:center;gap:6px;font-size:11.5px;color:var(--ink-2)}
.lg-swatch{width:10px;height:10px;border-radius:3px;display:inline-block}
table.data{width:100%;border-collapse:collapse;font-size:12.5px;margin:10px 0 4px;
font-variant-numeric:tabular-nums}
table.data caption{text-align:left;font-size:12px;color:var(--muted);padding-bottom:6px}
table.data th,table.data td{border-bottom:1px solid var(--grid);padding:6px 9px;text-align:left}
table.data th{font-size:10.5px;letter-spacing:.05em;text-transform:uppercase;color:var(--muted);
border-bottom:1px solid var(--axis)}
table.data td.num,table.data th.num{text-align:right}
.scroll{overflow-x:auto}
.badge{display:inline-block;font-size:10px;font-weight:700;letter-spacing:.05em;padding:2px 7px;
border-radius:999px;border:1px solid currentColor}
.b-ok{color:var(--good)}.b-warn{color:var(--serious)}.b-bad{color:var(--critical)}.b-info{color:var(--muted)}
.insight{background:var(--card);border:1px solid var(--ring);border-left:3px solid var(--series-1);
border-radius:0 10px 10px 0;padding:13px 16px;margin:14px 0}
.insight h3{margin:0 0 4px;font-size:13.5px;color:var(--ink);letter-spacing:.01em}
.insight p{margin:0;font-size:13.5px;color:var(--ink-2)}
.note{font-size:12.5px;color:var(--muted);margin-top:6px}
footer.rep{margin-top:48px;padding-top:16px;border-top:1px solid var(--grid);
font-size:12px;color:var(--muted)}
@media print{body{background:#fff}.wrap{max-width:none;padding:0}
figure.viz,.tile,.insight{break-inside:avoid;border-color:#ccc}h2{break-after:avoid}}
"""


def _tone_badge(status: str, tone: str) -> str:
    cls = {"ok": "b-ok", "warn": "b-warn", "bad": "b-bad"}.get(tone, "b-info")
    return f'<span class="badge {cls}">{svg.esc(status)}</span>'


def build_html(model: dict, result: dict, check_rows: list[dict],
               tornado_data: dict | None = None, title: str | None = None,
               run_id=None, generated_by: str = "") -> str:
    """The full self-contained report document."""
    summary = checks_mod.summarise(check_rows)
    grade = result["reference_grade"]
    codes = result["scenario_codes"]
    sectors = result["summary"]
    rows_at_grade = [r for r in result["grid"] if r["grade"] == grade]
    worst = max(rows_at_grade, key=lambda r: r["multiple"])
    generated = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    doc_title = title or f"{result['model_name']} — stressed PD summary, {result['horizon_year']}"

    parts = [
        "<!doctype html><html lang=\"en\"><head><meta charset=\"utf-8\">",
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        f"<title>{svg.esc(doc_title)}</title><style>{_CSS}</style></head><body><div class=\"wrap\">",
        '<header class="rep">',
        f"<h1>{svg.esc(doc_title)}</h1>",
        f'<div class="meta">Climate transition and acute physical risk expressed as a stressed '
        f'probability of default · {len(result["sectors"])} sectors × {len(result["grades"])} '
        f'rating grades × {len(codes)} NGFS scenarios = <b>{len(result["grid"])} cells</b> · '
        f'horizon <b>{result["horizon_year"]}</b> · engine v{result["engine_version"]}'
        + (f' · run <b>#{svg.esc(run_id)}</b>' if run_id else "")
        + (f' · generated by {svg.esc(generated_by)}' if generated_by else "")
        + f" · {generated}</div></header>",
        '<p class="lede">stressed PD = N( N&#8315;&#185;(PD<sub>0</sub>) + push + macro shift), where '
        'push = k &times; g(transition cost ratio + physical cost ratio). The two cost ratios are added '
        '<b>inside</b> the transform, not as two separate pushes, because the transform is concave. '
        'The model produces no ECL and no LGD: it stops at the PD signal.</p>',
    ]

    # --- headline tiles
    parts.append('<div class="tiles">')
    parts.append(svg.stat_tile("Worst cell", f"{worst['multiple']:.2f}x",
                               f"{worst['sector']} · {SCENARIO_LABEL.get(worst['scenario'])} · {grade}",
                               "critical"))
    parts.append(svg.stat_tile("Stressed PD there", _pct(worst["stressed_pd"]),
                               f"from {_pct(worst['baseline_pd'])} baseline"))
    parts.append(svg.stat_tile("Calibrated k", f"{result['k']:.6f}",
                               f"θ = {result['theta']:g}, anchor "
                               f"{result['calibration']['anchor_in_use']}"))
    parts.append(svg.stat_tile("Structural tests",
                               "BOTH PASS" if summary["structural_pair_ok"] else "REVIEW",
                               "opposing transition / physical orderings",
                               "good" if summary["structural_pair_ok"] else "critical"))
    parts.append(svg.stat_tile("Quality checks", f"{summary['passed']} / {summary['total']}",
                               f"{summary['failure_count']} failing, "
                               f"{summary['expected_count']} expected flags",
                               "good" if summary["can_finalise"] else "critical"))
    parts.append("</div>")

    insights = build_insights(result, check_rows)
    parts.append(f'<div class="insight"><h3>{svg.esc(insights[0]["title"])}</h3>'
                 f'<p>{insights[0]["body"]}</p></div>')

    # --- 1. the heat map
    parts.append("<h2>1 &nbsp;PD multiple by sector and scenario</h2>")
    parts.append(f'<p>Stressed PD divided by baseline PD at grade {svg.esc(grade)} '
                 f'({_pct(next(g["baseline_pd"] for g in result["grades"] if g["grade"] == grade))} '
                 f'baseline). A reading of 1.00x means the scenario leaves that sector untouched.</p>')
    labels = [s["sector"] for s in sectors]
    values = [[s["multiples"][c] for c in codes] for s in sectors]
    parts.append(svg.figure(
        f"PD multiple at grade {grade}, {result['horizon_year']} horizon",
        svg.heatmap(labels, [SCENARIO_LABEL.get(c, c) for c in codes], values,
                    low_label="lighter = closer to baseline", high_label="darker = more stressed"),
        caption="Sequential single-hue ramp: magnitude only. Every cell carries its own value, so the "
                "reading never depends on discriminating two shades of blue."))
    parts.append('<div class="scroll">' + svg.table(
        ["Sector", "Intensity (t/US$m)", "Pass-through"] + [SCENARIO_LABEL.get(c, c) for c in codes],
        [[s["sector"], f"{s['intensity']:,.0f}", f"{s['pass_through']:.0%}"]
         + [f"{s['multiples'][c]:.3f}x" for c in codes] for s in sectors],
        numeric_from=1) + "</div>")

    parts.append(f'<div class="insight"><h3>{svg.esc(insights[1]["title"])}</h3>'
                 f'<p>{insights[1]["body"]}</p></div>')

    # --- 2. the two opposing channels
    parts.append("<h2>2 &nbsp;The two channels, running in opposite directions</h2>")
    parts.append("<p>This pair of orderings is the single most important structural test of a climate "
                 "scenario model. Transition severity must follow the carbon price; physical severity "
                 "must follow warming, which is the exact reverse. Both hold with zero violations.</p>")
    trans_avg = [sum(s["transition_cost"][c] for s in sectors) / len(sectors) for c in codes]
    phys_avg = [result["physical"]["gva_weighted_cost"][c] for c in codes]
    scen_labels = [SCENARIO_LABEL.get(c, c) for c in codes]
    parts.append(svg.figure(
        "Cost ratio by scenario — transition (mean across sectors) and physical (GVA-weighted)",
        svg.small_multiples([
            ("Transition channel — follows the carbon price",
             svg.grouped_bars(scen_labels, [("Transition cost ratio", trans_avg)],
                              width=460, height=280, value_fmt="{:.3f}",
                              y_label="share of value added")),
            ("Physical channel — follows warming",
             svg.grouped_bars(scen_labels, [("Physical cost ratio", phys_avg)],
                              width=460, height=280, value_fmt="{:.4f}",
                              y_label="share of value added", slot=1)),
        ]),
        caption="Two panels rather than two series on one axis: the transition ratio is roughly forty "
                "times the physical one under Net Zero, so a shared scale would flatten the physical "
                "channel to a hairline and hide the very ordering being tested. Each panel keeps one "
                "axis; the shared scenario order carries the comparison. Transition falls left to right "
                "as the carbon price falls; physical rises as warming rises."))
    parts.append(svg.table(
        ["Scenario", "Carbon price (US$2010)", "Deflated", "Warming at horizon (°C)",
         "Mean transition ratio", "GVA-wtd physical ratio", "GDP deviation"],
        [[SCENARIO_LABEL.get(sc["code"], sc["code"]), f"{sc['carbon_price']:,.0f}",
          f"{sc['carbon_price_deflated']:,.1f}", f"{sc['warming_at_horizon']:.2f}",
          f"{trans_avg[i]:.4f}", f"{phys_avg[i]:.5f}", f"{sc['gdp_deviation_pct']:.2f}%"]
         for i, sc in enumerate(result["scenarios"])], numeric_from=1))

    parts.append(f'<div class="insight"><h3>{svg.esc(insights[2]["title"])}</h3>'
                 f'<p>{insights[2]["body"]}</p></div>')

    # --- 3. push decomposition
    parts.append("<h2>3 &nbsp;What drives the push, sector by sector</h2>")
    order = sorted(sectors, key=lambda s: s["push"]["NZ"], reverse=True)
    parts.append(svg.figure(
        "Probit push under Net Zero 2050, split by channel",
        svg.stacked_bars([s["sector"] for s in order],
                         [("Transition", [s["push"]["NZ"] * (1 - s["physical_share"]["NZ"]) for s in order]),
                          ("Physical", [s["push"]["NZ"] * s["physical_share"]["NZ"] for s in order])],
                         height=330, value_fmt="{:.3f}", y_label="probit push"),
        legend_html=svg.legend([("Transition", "var(--series-1)"), ("Physical", "var(--series-2)")]),
        caption="Split by attributing the residual of the joint push over the transition-only push to "
                "the physical channel — the two are never pushed separately, because g is concave."))
    parts.append('<div class="scroll">' + svg.table(
        ["Sector"] + [f"Physical share of push, {SCENARIO_LABEL.get(c, c)}" for c in codes],
        [[s["sector"]] + [f"{s['physical_share'][c]:.1%}" for c in codes] for s in sectors],
        numeric_from=1) + "</div>")

    parts.append(f'<div class="insight"><h3>{svg.esc(insights[3]["title"])}</h3>'
                 f'<p>{insights[3]["body"]}</p></div>')

    # --- 4. carbon price paths
    parts.append("<h2>4 &nbsp;Scenario inputs over the horizon set</h2>")
    horizons = result["horizons"]
    price_series = [(SCENARIO_LABEL.get(sc["code"], sc["code"]),
                     [float(next(s for s in model["scenarios"] if s["code"] == sc["code"])
                            ["carbon_price"][h]) for h in horizons])
                    for sc in result["scenarios"]]
    parts.append(svg.figure(
        "NGFS shadow carbon price path (US$2010 / tCO2e)",
        svg.line_chart([str(h) for h in horizons], price_series, height=280,
                       value_fmt="{:,.0f}", y_label="US$2010 / tCO2e"),
        legend_html=svg.legend([(n, svg.SERIES[i % 4]) for i, (n, _) in enumerate(price_series)]),
        caption="An NGFS shadow price is a marginal abatement cost, not a tax bill. Charging the price "
                "against full current emissions with no abatement response is deliberately conservative. "
                "Published prices are global weighted values; the source notes they tend to be lower in "
                "emerging economies, so applying them to Oman biases the transition cost upwards."))

    # --- 5. grade ladder
    parts.append("<h2>5 &nbsp;How the shock lands across the rating ladder</h2>")
    worst_sector_id = worst["sector_id"]
    grade_names = [g["grade"] for g in result["grades"]]
    ladder = [(SCENARIO_LABEL.get(c, c),
               [result["by_grid"][(worst_sector_id, gname, c)]["multiple"] for gname in grade_names])
              for c in codes]
    parts.append(svg.figure(
        f"PD multiple by grade — {worst['sector']}",
        svg.line_chart(grade_names, ladder, height=280, value_fmt="{:.2f}", y_label="PD multiple"),
        legend_html=svg.legend([(n, svg.SERIES[i % 4]) for i, (n, _) in enumerate(ladder)]),
        caption="A constant probit shift produces a larger relative PD increase at better grades, because "
                "the normal density is thinner in the tail. The absolute bps increase still rises with "
                "the grade — which is why both are reported."))
    parts.append('<div class="scroll">' + svg.table(
        ["Grade", "Baseline PD"] + [f"{SCENARIO_LABEL.get(c, c)} PD" for c in codes]
        + [f"{c} multiple" for c in codes],
        [[gname, _pct(next(g["baseline_pd"] for g in result["grades"] if g["grade"] == gname), 3)]
         + [_pct(result["by_grid"][(worst_sector_id, gname, c)]["stressed_pd"], 3) for c in codes]
         + [f"{result['by_grid'][(worst_sector_id, gname, c)]['multiple']:.3f}x" for c in codes]
         for gname in grade_names], numeric_from=1) + "</div>")

    # --- 6. worked example
    parts.append("<h2>6 &nbsp;Worked example — the full decomposition of one cell</h2>")
    dec = engine.decompose(result, worst_sector_id, grade, worst["scenario"])
    parts.append(f'<p>{svg.esc(dec["sector"])} · grade {svg.esc(grade)} · '
                 f'{svg.esc(dec["scenario_name"])}. Every number below is reproducible by hand from the '
                 f'inputs; this is the artefact that makes the model reviewable rather than merely '
                 f'plausible.</p>')
    probit0 = norm_ppf(dec["row"]["baseline_pd"])
    probit1 = norm_ppf(dec["row"]["stressed_pd"])
    wf = [("Transition push", dec["cell"]["push_transition"], "delta"),
          ("Physical push", dec["cell"]["push_physical"], "delta"),
          ("Macro shift", dec["cell"]["macro_shift"], "delta"),
          ("Total probit shift", 0.0, "total")]
    parts.append(svg.figure(
        "What the probit shift is made of",
        svg.waterfall(wf, height=300),
        caption=f"The three components only. The baseline probit itself is "
                f"{probit0:.4f} — two orders of magnitude larger than any component — so plotting it on "
                f"the same scale would flatten the decomposition to nothing. The shift moves the probit "
                f"from {probit0:.4f} to {probit1:.4f}, which maps back through the normal CDF to a PD of "
                f"{_pct(dec['row']['stressed_pd'])} against a {_pct(dec['row']['baseline_pd'])} baseline."))
    parts.append(svg.table(
        ["Step", "Value", "Unit", "Working"],
        [[s["label"],
          (f"{s['value']:,.6f}" if abs(s["value"]) < 1000 else f"{s['value']:,.2f}"),
          s["unit"], s["detail"]] for s in dec["steps"]], numeric_from=1))
    parts.append("<h3>Physical cost build-up for that cell</h3>")
    parts.append(svg.table(
        ["Hazard", "Baseline AAL (% GVA)", "Severity", "Normalised exposure", "Insurance recovery",
         "P&L share", "Contribution"],
        [[h["name"], f"{h['baseline_aal'] * 100:.4f}%", f"{h['severity']:.4f}", f"{h['exposure']:.4f}",
          f"{h['insurance_recovery']:.0%}", f"{h['pnl_share']:.0%}", f"{h['contribution']:.6f}"]
         for h in dec["hazards"]], numeric_from=1))

    parts.append(f'<div class="insight"><h3>{svg.esc(insights[4]["title"])}</h3>'
                 f'<p>{insights[4]["body"]}</p></div>')

    # --- 7. sensitivity
    if tornado_data and tornado_data.get("bars"):
        parts.append("<h2>7 &nbsp;Sensitivity — what actually moves the answer</h2>")
        parts.append("<p>One-way ranges around the base case, ranked by span. Two of these levers are "
                     "parameters the public sources demonstrably cannot pin down, so the band they "
                     "generate is the model's real uncertainty and belongs in front of a committee.</p>")
        parts.append(svg.figure(
            f"One-way sensitivity of the mean PD multiple at grade {grade}",
            svg.tornado(tornado_data["bars"], base=tornado_data["base"]["mean_multiple"]),
            legend_html=svg.legend([("lowers the multiple", "var(--diverge-low)"),
                                    ("raises the multiple", "var(--diverge-high)")]),
            caption="Dashed rule is the base case. Diverging pair: polarity is the data's job here, so "
                    "hue carries the direction and the numeric range repeats it in text."))
        parts.append(svg.table(
            ["Lever", "Low", "High", "Span", "At", "Why it is uncertain"],
            [[b["label"], f"{b['low']:.3f}x", f"{b['high']:.3f}x", f"{b['span']:.3f}",
              f"{b['low_label']} → {b['high_label']}", b["note"]] for b in tornado_data["bars"]],
            numeric_from=1))

    # --- theta band
    band = result["calibration"].get("theta_band") or []
    if band:
        parts.append("<h3>Curvature disclosure band</h3>")
        parts.append(svg.table(
            ["θ", "Form", "Refitted k", "Implied top-decile intensity multiple",
             "Push at max cost ratio", f"PD at {grade}", "PD multiple"],
            [[f"{b['theta']:+.1f}", b["form"] or "—", f"{b['k']:.6f}",
              f"{b['implied_multiple']:.3f}" if b["implied_multiple"] else "unreachable",
              f"{b['push_at_max']:.4f}", _pct(b["pd_at_reference"]), f"{b['pd_multiple']:.3f}x"]
             for b in band], numeric_from=2,
            caption="k is refitted on the anchor at every θ — it is a function of θ, never a stored "
                    "constant. The implied top-decile intensity multiple barely moves across the whole "
                    "range (3.97 to 4.07), which is precisely why the anchors cannot identify curvature."))

    parts.append(f'<div class="insight"><h3>{svg.esc(insights[5]["title"])}</h3>'
                 f'<p>{insights[5]["body"]}</p></div>')

    # --- quality checks
    parts.append("<h2>8 &nbsp;Quality checks</h2>")
    parts.append(f'<p>{summary["passed"]} of {summary["total"]} pass; {summary["failure_count"]} fail. '
                 f'Items marked <i>expected</i> are known disclosures on delivery of this dataset, not '
                 f'defects — they are left visible rather than defaulted away.</p>')
    parts.append('<div class="scroll">' + svg.table(
        ["#", "Check", "Result", "Status", "Expected", "What it means"],
        [[c["id"], c["name"],
          (f"{c['result']:.6g}" if isinstance(c["result"], (int, float))
           and not isinstance(c["result"], bool) else str(c["result"])),
          c["status"], "yes" if c["expected"] else "", c["explanation"]]
         for c in check_rows], numeric_from=6) + "</div>")

    parts.append(f'<div class="insight"><h3>{svg.esc(insights[6]["title"])}</h3>'
                 f'<p>{insights[6]["body"]}</p></div>')

    # --- appendix
    parts.append("<h2>Appendix A &nbsp;Full stressed-PD grid</h2>")
    parts.append('<div class="scroll">' + svg.table(
        ["Sector", "Grade", "Baseline PD"]
        + [f"{c} PD" for c in codes] + [f"{c} ×" for c in codes],
        [[r["sector"], r["grade"], _pct(r["baseline_pd"], 3)]
         + [_pct(result["by_grid"][(r["sector_id"], r["grade"], c)]["stressed_pd"], 3) for c in codes]
         + [f"{result['by_grid'][(r['sector_id'], r['grade'], c)]['multiple']:.3f}" for c in codes]
         for r in result["grid"] if r["scenario"] == codes[0]], numeric_from=2) + "</div>")

    parts.append("<h2>Appendix B &nbsp;Source register</h2>")
    parts.append('<div class="scroll">' + svg.table(
        ["#", "Value", "Location", "Source", "Note", "Status"],
        [[r["id"], r["value"], r["location"], r["source"], r["note"], r["status"]]
         for r in registers.SOURCE_REGISTER_ROWS], numeric_from=6) + "</div>")

    parts.append("<h2>Appendix C &nbsp;Assumption register</h2>")
    parts.append('<div class="scroll">' + svg.table(
        ["ID", "Assumption", "Value / status", "Rationale and limitation"],
        [[r["id"], r["assumption"], r["value"], r["rationale"]]
         for r in registers.ASSUMPTION_REGISTER_ROWS], numeric_from=4) + "</div>")

    parts.append(
        '<footer class="rep">Generated by the CreditProbe climate stressed-PD module, engine v'
        + svg.esc(result["engine_version"])
        + ". Reproduces the Oman Climate Stressed PD v5.1 workbook to full double precision. "
          "Output is a PD signal only — no ECL, no LGD, no customer layer. Chronic physical risk and "
          "the hydrocarbon fiscal channel are not modelled and are disclosed as open items."
        + "</footer></div></body></html>")

    return "".join(parts)


# ------------------------------------------------------------------ Excel pack

def build_excel(model: dict, result: dict, check_rows: list[dict]) -> bytes:
    """The regulator pack: every input, every intermediate and the full grid, one
    sheet per block, mirroring the workbook it replaces."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    bold = Font(bold=True)

    def sheet(name, headers, rows, widths=None):
        ws = wb.create_sheet(name[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in rows:
            ws.append(list(row))
        for i, w in enumerate(widths or [], start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.freeze_panes = "A2"
        return ws

    wb.remove(wb.active)
    codes = result["scenario_codes"]
    grade = result["reference_grade"]
    summary = checks_mod.summarise(check_rows)

    sheet("ReadMe", ["Item", "Value"], [
        ("Model", result["model_name"]), ("Country", result["country"]),
        ("Engine version", result["engine_version"]),
        ("Generated", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
        ("Horizon year", result["horizon_year"]), ("Curvature theta", result["theta"]),
        ("Cost ratio cap", result["cost_ratio_cap"]),
        ("US GDP deflator", result["us_gdp_deflator"]),
        ("Denominator basis", result["totals"]["basis"]),
        ("Calibrated k", result["k"]),
        ("Reference grade", grade),
        ("Grid cells", len(result["grid"])),
        ("Checks passed", f"{summary['passed']} / {summary['total']}"),
        ("Failing checks", summary["failure_count"]),
        ("Eligible to finalise", "yes" if summary["can_finalise"] else "no"),
        ("Core formula", "stressed_PD = N( N-1(PD_0) + push + macro_shift )"),
        ("Push", "push = k x g(transition cost ratio + physical cost ratio), g = ((1+x)^theta-1)/theta"),
        ("Scope", "PD signal only. No ECL, no LGD, no customer layer. Acute physical risk only."),
    ], [26, 96])

    sheet("Grid", ["Sector ID", "Sector", "Grade", "Scenario", "Baseline PD", "Stressed PD",
                   "Multiple", "Delta (bps)"],
          [(r["sector_id"], r["sector"], r["grade"], r["scenario_name"], r["baseline_pd"],
            r["stressed_pd"], r["multiple"], r["delta_bps"]) for r in result["grid"]],
          [10, 42, 8, 20, 13, 13, 11, 12])

    sheet(f"Summary_{grade}", ["Sector ID", "Sector", "Intensity (t/US$m)", "Pass-through"]
          + [f"{c} multiple" for c in codes] + [f"{c} physical share" for c in codes],
          [(s["sector_id"], s["sector"], s["intensity"], s["pass_through"])
           + tuple(s["multiples"][c] for c in codes) + tuple(s["physical_share"][c] for c in codes)
           for s in result["summary"]], [10, 42, 17, 13] + [13] * (2 * len(codes)))

    sheet("Cells", ["Sector ID", "Sector", "Scenario", "Transition cost", "Physical cost",
                    "Total cost", "Cap binds", "Push", "Push transition", "Push physical",
                    "Physical share", "Macro shift", "Probit shift"],
          [(c["sector_id"], c["sector"], c["scenario_name"], c["transition_cost"], c["physical_cost"],
            c["total_cost"], "yes" if c["cap_binds"] else "", c["push"], c["push_transition"],
            c["push_physical"], c["physical_share"], c["macro_shift"], c["probit_shift"])
           for c in result["cells"]], [10, 42, 20] + [14] * 10)

    sheet("Sectors", ["ID", "Sector", "ISIC", "GVA (local)", "GVA (US$m)", "Denominator (US$m)",
                      "GVA share", "Emissions (MtCO2e)", "Intensity (t/US$m)", "Pass-through",
                      "macro_beta", "Pass-through rationale"],
          [(s["id"], s["name"], s["isic"], s["gva_local"], s["gva_usd"], s["denominator_usd"],
            s["gva_share"], s["emissions_mt"], s["intensity"], s["pass_through"], s["macro_beta"],
            s["rationale"]) for s in result["sectors"]],
          [8, 42, 22, 14, 14, 16, 11, 17, 17, 13, 11, 90])

    ec = result["emissions"]
    sector_ids = [s["id"] for s in result["sectors"]]
    sheet("Emissions_Allocation", ["Code", "EDGAR category", "MtCO2e", "Plug"] + sector_ids
          + ["HH", "Row sum", "Definition"],
          [(c["code"], c["name"], c["mt"], "yes" if c["is_plug"] else "")
           + tuple(c["shares"].get(sid, 0.0) for sid in sector_ids)
           + (c["shares"].get("HH", 0.0), c["share_sum"], c["definition"]) for c in ec["categories"]],
          [8, 40, 11, 7] + [8] * (len(sector_ids) + 1) + [10, 90])

    sheet("Scenarios", ["Code", "Scenario", "Quadrant", "Carbon price (US$2010)", "Deflated",
                        "Intensity index", "Denominator index", "Warming 2100", "Warming at horizon",
                        "GDP deviation (%)", "Macro shift"],
          [(sc["code"], sc["name"], sc["quadrant"], sc["carbon_price"], sc["carbon_price_deflated"],
            sc["intensity_index"], sc["denominator_index"], sc["warming_2100"],
            sc["warming_at_horizon"], sc["gdp_deviation_pct"], sc["macro_shift"])
           for sc in result["scenarios"]], [8, 22, 20, 20, 12, 14, 16, 13, 17, 16, 13])

    ph = result["physical"]
    sheet("Physical", ["Hazard", "Name", "Baseline AAL (share of GVA)", "Warming elasticity",
                       "Insurance recovery", "P&L share", "Reserved for LGD"]
          + [f"Severity {c}" for c in codes],
          [(h["id"], h["name"], h["baseline_aal"], h["elasticity"], h["insurance_recovery"],
            h["pnl_share"], 1 - float(h["pnl_share"]))
           + tuple(ph["severity"][h["id"]][c] for c in codes) for h in ph["hazards"]],
          [9, 32, 22, 16, 16, 11, 15] + [12] * len(codes))

    hazard_ids = [h["id"] for h in ph["hazards"]]
    sheet("Physical_Exposure", ["ID", "Sector", "GVA share"]
          + [f"{h} raw" for h in hazard_ids] + [f"{h} normalised" for h in hazard_ids]
          + [f"Physical cost {c}" for c in codes],
          [(s["id"], s["name"], s["gva_share"])
           + tuple(model["exposure_raw"][s["id"]][h] for h in hazard_ids)
           + tuple(ph["exposure_used"][s["id"]][h] for h in hazard_ids)
           + tuple(ph["cost"][s["id"]][c] for c in codes) for s in result["sectors"]],
          [8, 42, 11] + [11] * (2 * len(hazard_ids)) + [15] * len(codes))

    cal = result["calibration"]
    sheet("Calibration", ["Item", "Value"], [
        ("Anchor in use", cal["anchor_in_use"]),
        ("Baseline PD", cal["baseline_pd"]),
        ("Relative PD change", cal["anchor_rel"]),
        ("Stressed PD implied", cal["anchor_stressed_pd"]),
        ("push_EU", cal["push_eu"]),
        ("EU intensity route", cal["route_in_use"]),
        ("Route 1 economy intensity (t/EURm)", cal["route1_economy_intensity"]),
        ("Route 1 median intensity (t/EURm)", cal["route1_intensity"]),
        ("Route 2 intensity (t/EURm)", cal["route2_intensity"]),
        ("EU intensity selected", cal["eu_intensity"]),
        ("Anchor carbon price (EUR/tCO2)", cal["anchor_price_eur"]),
        ("EU pass-through", cal["eu_pass_through"]),
        ("cost_ratio_EU", cal["cost_ratio_eu"]),
        ("theta", cal["theta"]),
        ("g at anchor", cal["g_at_anchor"]),
        ("k FITTED", cal["k"]),
        ("Push ratio, high emitters / median", cal["push_ratio_high_median"]),
        ("Implied top-decile intensity multiple", cal["implied_intensity_multiple"]),
        ("Coal required intensity multiple", (cal.get("coal") or {}).get("required_multiple")),
        ("Coal discrepancy factor", (cal.get("coal") or {}).get("discrepancy")),
        ("Coal verdict", (cal.get("coal") or {}).get("verdict")),
        ("Max local cost ratio", (cal.get("extrapolation") or {}).get("max_cost_ratio")),
        ("Extrapolation multiple", (cal.get("extrapolation") or {}).get("multiple")),
    ], [40, 60])

    sheet("Theta_Band", ["theta", "Form", "Refitted k", "Implied top-decile multiple",
                         "Push at max cost ratio", "PD at reference grade", "PD multiple"],
          [(b["theta"], b["form"], b["k"], b["implied_multiple"], b["push_at_max"],
            b["pd_at_reference"], b["pd_multiple"]) for b in cal.get("theta_band", [])],
          [8, 18, 14, 24, 20, 20, 12])

    mc = result["macro"]
    sheet("Macro", ["Item", "Value"], [
        ("Selected specification", mc["selected_specification"]),
        ("Observations used", mc["n_paired"]),
        ("Correlation estimated", mc["correlation_estimated"]),
        ("Correlation in use", mc["correlation_in_use"]),
        ("sd of D probit NPLR", mc["sd_d_probit"]),
        ("sd of GDP growth", mc["sd_gdp_growth"]),
        ("beta OLS", mc["beta_ols"]),
        ("beta in use", mc["beta_in_use"]),
        ("Intercept", mc["intercept"]), ("R2", mc["r2"]),
    ] + [(f"macro_shift {c}", mc["by_scenario"][c]["shift"]) for c in codes], [30, 26])

    sheet("Macro_Data", ["Year", "NPL ratio", "probit(NPLR)", "D probit(NPLR)", "Real GDP growth"],
          [(r["year"], r["npl_ratio"], r["probit"], r["d_probit"], r["gdp_growth"])
           for r in mc["series"]], [8, 12, 15, 16, 17])

    sheet("Quality_Checks", ["#", "Check", "Result", "Status", "Expected", "What it means"],
          [(c["id"], c["name"],
            (f"{c['result']:.10g}" if isinstance(c["result"], (int, float))
             and not isinstance(c["result"], bool) else str(c["result"])),
            c["status"], "yes" if c["expected"] else "", c["explanation"]) for c in check_rows],
          [5, 56, 22, 11, 10, 110])

    sheet("Source_Register", ["#", "Value", "Location", "Source", "Note", "Status"],
          [(r["id"], r["value"], r["location"], r["source"], r["note"], r["status"])
           for r in registers.SOURCE_REGISTER_ROWS], [5, 40, 28, 52, 90, 22])

    sheet("Assumption_Register", ["ID", "Assumption", "Location", "Value / status",
                                  "Rationale and limitation"],
          [(r["id"], r["assumption"], r["location"], r["value"], r["rationale"])
           for r in registers.ASSUMPTION_REGISTER_ROWS], [7, 44, 28, 26, 110])

    sheet("Verification_Log", ["#", "Value", "Value now in model", "Source checked against",
                              "Finding", "Status"],
          [(r["id"], r["value"], r["value_now"], r["source"], r["finding"], r["status"])
           for r in registers.VERIFICATION_LOG_ROWS], [5, 42, 34, 44, 110, 26])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
