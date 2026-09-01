"""§83's Learning Report. Twenty sheets, and three kinds of nothing.

What it is for
---------------
Somebody has to be able to take this to a model-risk committee and defend it.
That means the numbers, the sample sizes behind them, what changed, what
regressed, what is still open, and — the sheet most reports omit — the
methodology that says what each figure is allowed to claim.

Three kinds of nothing it must never contain
---------------------------------------------
§83 is explicit: no secrets, no sealed gold answers, no client rows. The
first two are checked here; the third is structural, because nothing in the
continuous-learning tables holds a client row to begin with.

Empty sheets stay
------------------
A report with fourteen sheets because six had nothing in them is a report
whose shape changes with its content, and a reader comparing this quarter to
last cannot tell a missing section from a section with nothing to say. Every
sheet is present; the empty ones say what would have been there.
"""

from __future__ import annotations

import io
import logging
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook as XlWorkbook

from backend.brain import security
from backend.exports.contract import Workbook, sheet_name, slug
from backend.exports.style import facts, note, section, table, title

logger = logging.getLogger(__name__)

REPORT_VERSION = "1.0.0"
LEARNING_REPORT = "learning_report"

#: §83's twenty sheets, in §83's order. The order is an argument: Overview
#: first, Methodology last, and Known Limitations immediately before it, so
#: a reader who stops early has still seen the caveats.
SHEETS: tuple[str, ...] = (
    "Overview", "Time Period", "Learning Captured", "Learning Approved",
    "Learning Activated", "Six Dimensions", "Development Performance",
    "Validation Performance", "Subcomponents", "Critical Failures",
    "Coverage", "Feedback Learning", "Regulatory Learning",
    "Analysis Studio Learning", "Brain Imports", "Release Timeline",
    "Model/Prompt Changes", "Regressions", "Known Limitations",
    "Methodology",
)

EXPECTED_SHEETS = 20
if len(SHEETS) != EXPECTED_SHEETS:
    raise AssertionError(
        f"§83 names {EXPECTED_SHEETS} sheets; this module has "
        f"{len(SHEETS)}.")

#: What each empty sheet says instead of being absent.
NOTHING_YET: dict[str, str] = {
    "Learning Captured": "Nothing was captured in this window.",
    "Learning Approved": "Nothing was approved in this window. That is not "
                         "the same as nothing having been captured — see "
                         "the previous sheet.",
    "Learning Activated": "Nothing reached production in this window.",
    "Six Dimensions": "No evaluation ran in this window, so no dimension "
                      "has a before and after. This is different from every "
                      "dimension being unchanged.",
    "Development Performance": "No development evaluation ran.",
    "Validation Performance": "No validation evaluation ran. Development "
                              "improvement with no out-of-sample check is "
                              "not evidence that anything generalised.",
    "Subcomponents": "No subcomponent scores were recorded.",
    "Critical Failures": "No critical failures were recorded either way. "
                         "Nothing was fixed and nothing was introduced.",
    "Coverage": "Coverage was not measured.",
    "Feedback Learning": "No feedback became a candidate in this window.",
    "Regulatory Learning": "No regulatory requirement was approved in this "
                           "window.",
    "Analysis Studio Learning": "No method changed in this window.",
    "Brain Imports": "No Brain was imported in this window.",
    "Release Timeline": "No release was activated in this window.",
    "Model/Prompt Changes": "No model role or prompt version changed.",
    "Regressions": "Nothing regressed. Worth saying explicitly rather than "
                   "leaving the reader to infer it from an absence.",
    "Known Limitations": "None were declared. That is a statement about the "
                         "review process, not a property of the system.",
}


class ReportError(Exception):
    """A report that may not be written, and why."""


def build(payload: dict[str, Any]) -> Workbook:
    """§83's workbook. Refuses to write one carrying a secret.

    The scan runs over the assembled text rather than over the source
    objects, because the leak that matters is the one that made it into a
    cell — and a check on the inputs would miss a value that arrived through
    a path nobody thought to check.
    """
    book = XlWorkbook()
    taken: set[str] = set()
    first = True

    for name in SHEETS:
        if first:
            ws = book.active
            ws.title = sheet_name(name, taken=taken)
            first = False
        else:
            ws = book.create_sheet(sheet_name(name, taken=taken))
        _sheet(ws, name, payload)

    problems = _forbidden(book)
    if problems:
        raise ReportError(
            "this Learning Report may not be written: " + "; ".join(
                problems[:8]))

    stream = io.BytesIO()
    book.save(stream)
    window = str(payload.get("window", "window"))
    return Workbook(
        filename=(f"CreditProbe_learning_report_{slug(window, limit=24)}"
                  f"_{datetime.now(UTC):%Y%m%d}.xlsx"),
        content=stream.getvalue(),
        kind=LEARNING_REPORT,
        manifest={
            "sheets": book.sheetnames,
            "window": window,
            "baseline_id": (payload.get("baseline") or {}).get("baseline_id",
                                                               ""),
            "report_version": REPORT_VERSION,
            "contains_sealed_holdout_content": False,
            "contains_client_rows": False,
        },
    )


def _forbidden(book: XlWorkbook) -> list[str]:
    """§83's three refusals, checked over what was actually written."""
    problems: list[str] = []
    for ws in book.worksheets:
        text = "\n".join(
            str(cell.value) for row in ws.iter_rows() for cell in row
            if cell.value is not None)
        for label, shown in security.scan_secrets(text):
            problems.append(f"{ws.title}: {label} present ({shown})")
        for label, count in security.scan_client_data(text):
            if label == "email address":
                continue
            problems.append(
                f"{ws.title}: {count} occurrence(s) of {label}. §83: no "
                "client rows")
    return problems


# ------------------------------------------------------------- the sheets


def _sheet(ws: Any, name: str, payload: dict[str, Any]) -> None:
    row = title(ws, name, _subtitle(name, payload))
    builder = _BUILDERS.get(name)
    written = builder(ws, row, payload) if builder else row
    if written == row and name in NOTHING_YET:
        # Empty, and saying so. A sheet that vanishes when it has nothing to
        # report changes the report's shape with its content, and a reader
        # comparing quarters cannot tell a missing section from a quiet one.
        note(ws, NOTHING_YET[name], written + 1)


def _subtitle(name: str, payload: dict[str, Any]) -> str:
    if name == "Overview":
        return str(payload.get("headline", ""))
    if name == "Methodology":
        return "What each figure on the preceding sheets is allowed to claim."
    return ""


def _overview(ws: Any, row: int, payload: dict[str, Any]) -> int:
    baseline = payload.get("baseline") or {}
    row = facts(ws, [
        ("Headline", payload.get("headline", "")),
        ("Window", payload.get("window", "")),
        ("Compared against", baseline.get("comparable_to", "no baseline")),
        ("Baseline recorded", baseline.get("created_at", "")),
        ("Brain", baseline.get("brain", "")),
    ], row + 1)
    if payload.get("these_are_not_the_same_thing"):
        row = note(ws, str(payload["these_are_not_the_same_thing"]), row + 1)
    return row


def _counts(key: str):
    def build_sheet(ws: Any, row: int, payload: dict[str, Any]) -> int:
        counts = payload.get(key) or {}
        if not counts:
            return row
        return table(ws, ["Item", "Count"],
                     [[k.replace("_", " "), v] for k, v in counts.items()],
                     row=row + 1)
    return build_sheet


def _dimensions(ws: Any, row: int, payload: dict[str, Any]) -> int:
    rows = payload.get("dimensions") or []
    if not rows:
        return row
    written = table(
        ws,
        ["Dimension", "Dev before", "Dev after", "Dev pp",
         "Validation before", "Validation after", "Validation pp",
         "Cases", "Evidence", "Verdict"],
        [[
            d.get("dimension", ""),
            (d.get("development") or {}).get("before"),
            (d.get("development") or {}).get("after"),
            (d.get("development") or {}).get("points"),
            (d.get("validation") or {}).get("before"),
            (d.get("validation") or {}).get("after"),
            (d.get("validation") or {}).get("points"),
            (d.get("validation") or {}).get("cases"),
            (d.get("validation") or {}).get("evidence"),
            d.get("verdict", ""),
        ] for d in rows],
        row=row + 1)
    return note(
        ws,
        "Development is the set that was tuned against and always looks "
        "better. Where the two disagree, the validation figure is the one "
        "the verdict rests on.",
        written + 1)


def _partition(which: str):
    def build_sheet(ws: Any, row: int, payload: dict[str, Any]) -> int:
        rows = payload.get("dimensions") or []
        if not rows:
            return row
        return table(
            ws, ["Dimension", "Before", "After", "Points", "Cases",
                 "Coverage", "Evidence", "Reads as"],
            [[
                d.get("dimension", ""),
                (d.get(which) or {}).get("before"),
                (d.get(which) or {}).get("after"),
                (d.get(which) or {}).get("points"),
                (d.get(which) or {}).get("cases"),
                (d.get(which) or {}).get("coverage"),
                (d.get(which) or {}).get("evidence"),
                (d.get(which) or {}).get("reads_as", ""),
            ] for d in rows],
            row=row + 1)
    return build_sheet


def _critical(ws: Any, row: int, payload: dict[str, Any]) -> int:
    change = payload.get("measured_change") or {}
    fixed = change.get("critical_failures_fixed")
    introduced = change.get("critical_failures_introduced")
    if fixed is None and introduced is None:
        return row
    row = facts(ws, [
        ("Critical failures fixed", fixed),
        ("Critical failures introduced", introduced),
    ], row + 1)
    return note(
        ws,
        "A critical failure introduced on the validation set blocks "
        "activation. It is not weighed against the improvements: it is a "
        "wrong answer the bank would have shown a client.",
        row + 1)


def _regressions(ws: Any, row: int, payload: dict[str, Any]) -> int:
    rows = [d for d in (payload.get("dimensions") or [])
            if d.get("verdict") in ("REGRESSED", "MIXED")]
    if not rows:
        return row
    return table(ws, ["Dimension", "Verdict", "Reads as"],
                 [[d.get("dimension", ""), d.get("verdict", ""),
                   d.get("reads_as", "")] for d in rows],
                 row=row + 1)


def _limitations(ws: Any, row: int, payload: dict[str, Any]) -> int:
    limits = list(payload.get("known_limitations") or [])
    drift = payload.get("overfitting") or {}
    if drift.get("possible_overfitting"):
        limits.append(str(drift.get("recommended_review", "")))
    hygiene = payload.get("hygiene") or {}
    limits.extend(str(f) for f in hygiene.get("findings", []))
    if not limits:
        return row
    return table(ws, ["Limitation"], [[one] for one in limits], row=row + 1)


def _methodology(ws: Any, row: int, payload: dict[str, Any]) -> int:
    from backend.continuous import measurement as ms

    row = section(ws, "What a figure on this report may claim", row + 1)
    row = table(
        ws, ["Rule", "What it means"],
        [
            ["Three forms",
             "Every change is shown as percentage points, as a relative "
             "change and as error reduction. Reporting only the relative "
             "figure is how a 2 pp move on a small base becomes 'a 40% "
             "improvement'."],
            ["Quantity is not quality",
             "Capture and improvement are separate sheets and are never "
             "added. More cases is not a better answer."],
            ["Validation outranks development",
             "Development is the set that was tuned against."],
            ["No percentage without a sample",
             f"Below {ms.MINIMUM_CASES} cases a difference is not "
             f"distinguishable from noise; below {ms.TRIVIAL_CASES} a "
             "percentage is not shown at all."],
            ["Attribution only where isolated",
             "Only changes measured on their own are attributed. The rest "
             "is reported as UNATTRIBUTED / INTERACTION rather than "
             "distributed to make a waterfall balance."],
        ],
        row=row + 1)
    return note(
        ws,
        "This report contains no secret, no sealed-holdout question or gold "
        "answer, and no client row. The sealed holdout's VERSION appears "
        "where it is relevant, which says which exam was sat without "
        "circulating it.",
        row + 1)


_BUILDERS: dict[str, Any] = {
    "Overview": _overview,
    "Time Period": lambda ws, row, payload: facts(ws, [
        ("Window", payload.get("window", "")),
        ("Windows available", ", ".join(
            payload.get("windows_available", [])[:6])),
    ], row + 1),
    "Learning Captured": _counts("learning_captured"),
    "Six Dimensions": _dimensions,
    "Development Performance": _partition("development"),
    "Validation Performance": _partition("validation"),
    "Critical Failures": _critical,
    "Regressions": _regressions,
    "Known Limitations": _limitations,
    "Methodology": _methodology,
}
