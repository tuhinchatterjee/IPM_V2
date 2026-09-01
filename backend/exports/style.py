"""
How a CreditProbe workbook looks, in one place.

A workbook that lands on a credit officer's laptop is the product, seen by
people who will never open the application. So it is styled deliberately and
identically across both exports: the same palette, the same header treatment,
the same number formats as the screen, the same status vocabulary.

Three rules, each of which exists because of a specific failure:

**Number formats come from the column contract, not from the value.** The
backend already records what each column IS — money at this scale, a percentage,
an ordinal, a count of days. Formatting from the value instead produces
73391.774000000012 in a cell a reviewer is reading, which is the right number
and looks like a bug.

**Status is never colour alone.** A PASS cell is green AND says PASS. The
workbook gets printed, photocopied and read by people who cannot distinguish
the greens, and a colour-only convention loses its meaning in all three.

**No merged-cell decoration.** Merged cells break sorting, filtering, copying
and every screen reader. The cover sheet is a plain two-column table that
happens to be styled, not a poster.
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ------------------------------------------------------------------ palette

#: Restrained institutional colours. Deep slate for structure, a single
#: accent, and status colours that survive greyscale because the text carries
#: the meaning too.
INK = "1B2430"
MUTED = "6B7684"
RULE = "D8DDE3"
ACCENT = "1F4E6B"
BAND = "F4F6F8"

PASS_FILL = "E4F1E6"
PASS_INK = "1E5B2C"
WARN_FILL = "FDF2DC"
WARN_INK = "7A5210"
FAIL_FILL = "FBE6E6"
FAIL_INK = "8B1D1D"
SKIP_FILL = "EEF0F3"
SKIP_INK = "525C68"

#: The one place a status word is mapped to a look. SKIPPED is deliberately
#: neutral rather than green: §27 says a skipped check must never read as a
#: passed one, and a grey cell saying SKIPPED cannot be mistaken for a green
#: one saying PASS.
STATUS_STYLE: dict[str, tuple[str, str]] = {
    "PASS": (PASS_FILL, PASS_INK),
    "PASSED": (PASS_FILL, PASS_INK),
    "OK": (PASS_FILL, PASS_INK),
    "WARNING": (WARN_FILL, WARN_INK),
    "WARN": (WARN_FILL, WARN_INK),
    "FAIL": (FAIL_FILL, FAIL_INK),
    "FAILED": (FAIL_FILL, FAIL_INK),
    "BLOCKED": (FAIL_FILL, FAIL_INK),
    "SKIPPED": (SKIP_FILL, SKIP_INK),
    "NOT RECORDED": (SKIP_FILL, SKIP_INK),
}

TITLE_FONT = Font(name="Calibri", size=15, bold=True, color=INK)
HEADING_FONT = Font(name="Calibri", size=11, bold=True, color=INK)
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=MUTED)
BODY_FONT = Font(name="Calibri", size=10, color=INK)
MONO_FONT = Font(name="Consolas", size=9, color=INK)
LINK_FONT = Font(name="Calibri", size=10, color=ACCENT, underline="single")

HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BAND_FILL = PatternFill("solid", fgColor=BAND)

THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
UNDERLINE = Border(bottom=Side(style="thin", color=RULE))

WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


# ------------------------------------------------------------ number formats

MONEY_0 = "#,##0"
MONEY_1 = "#,##0.0"
MONEY_2 = "#,##0.00"
PERCENT_2 = "#,##0.00"
INTEGER = "#,##0"
RATIO_4 = "#,##0.0000"
TEXT = "@"


def number_format(column: dict[str, Any]) -> str:
    """The Excel format for a column, from its semantic contract.

    Never from the value: a column of money where the first row happens to be a
    whole number is still money, and the reader who scrolls to row nine should
    not find it rendered differently.
    """
    semantic = str(column.get("semantic") or "").lower()
    decimals = column.get("decimals")
    if semantic in {"identity", "text", "period"}:
        return TEXT
    if semantic == "count":
        return INTEGER
    if semantic == "ordinal":
        return INTEGER
    if semantic == "days":
        return INTEGER
    if semantic == "percent":
        return PERCENT_2
    if semantic == "ratio":
        return RATIO_4
    if semantic == "money":
        if isinstance(decimals, int):
            return {0: MONEY_0, 1: MONEY_1}.get(decimals, MONEY_2)
        return MONEY_2
    if isinstance(decimals, int):
        return {0: INTEGER, 1: MONEY_1}.get(decimals, MONEY_2)
    return TEXT


def header_label(column: dict[str, Any]) -> str:
    """The heading a reader sees: the business label, with its unit.

    The unit belongs in the header rather than in every cell — a column that
    repeats "USD mn" twenty-five times has said the one thing that does not
    vary twenty-five times, and pushed the digits apart so the column can no
    longer be scanned.
    """
    label = str(column.get("label") or column.get("name") or "")
    unit = str(column.get("unit") or "").strip()
    if unit and unit.lower() not in label.lower():
        return f"{label} ({unit})"
    return label


# --------------------------------------------------------------- primitives


def title(ws: Worksheet, text: str, subtitle: str = "", *, row: int = 1) -> int:
    """A sheet's own heading. Returns the next free row."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    if subtitle:
        note = ws.cell(row=row + 1, column=1, value=subtitle)
        note.font = Font(name="Calibri", size=10, color=MUTED)
        return row + 3
    return row + 2


def crumb(ws: Worksheet, target: str = "COVER", *, row: int = 1) -> int:
    """A back-link on its own row, above the sheet's title.

    Above rather than beside: a link in B1 truncates the title in A1 the moment
    the title is longer than one column, and a workbook whose every heading is
    cut off at twelve characters is not one anybody navigates twice.
    """
    cell = ws.cell(row=row, column=1, value=f"\u2190 {target}")
    cell.hyperlink = f"#'{target}'!A1"
    cell.font = LINK_FONT
    return row + 1


def section(ws: Worksheet, text: str, row: int) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADING_FONT
    cell.border = UNDERLINE
    return row + 1


def facts(ws: Worksheet, pairs: list[tuple[str, Any]], row: int, *,
          label_width: int = 34, value_width: int = 78) -> int:
    """A two-column label/value block.

    Used for every metadata panel in both workbooks. Plain cells rather than
    merged ones: a merged block cannot be sorted, copied or read aloud.
    """
    for label, value in pairs:
        left = ws.cell(row=row, column=1, value=label)
        left.font = LABEL_FONT
        left.alignment = TOP
        right = ws.cell(row=row, column=2, value=_scalar(value))
        right.font = BODY_FONT
        right.alignment = WRAP
        row += 1
    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = value_width
    return row + 1


def _scalar(value: Any) -> Any:
    """Something Excel can hold in one cell.

    openpyxl raises on a list or a dict, which would turn a missing conversion
    into a failed export rather than a slightly ugly cell.
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    return str(value)


def table(ws: Worksheet, headers: list[str], rows: list[list[Any]], *,
          row: int = 1, formats: list[str] | None = None,
          widths: list[int] | None = None, band: bool = True,
          status_column: int | None = None, freeze: bool = True,
          autofilter: bool = True) -> int:
    """A styled data table. Returns the row after the last one written.

    Deliberately not an Excel "table object": a real ListObject cannot start
    anywhere but the top of a sheet without fighting the sheets that carry a
    heading first, and it brings a style that overrides the number formats.
    The frozen header, the filter and the banding are what a reader actually
    wanted from one.
    """
    top = row
    for i, head in enumerate(headers, start=1):
        cell = ws.cell(row=top, column=i, value=head)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX

    for r, values in enumerate(rows, start=top + 1):
        shade = band and (r - top) % 2 == 0
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=_scalar(value))
            cell.font = BODY_FONT
            cell.border = BOX
            cell.alignment = TOP
            if shade:
                cell.fill = BAND_FILL
            if formats and c <= len(formats) and formats[c - 1] != TEXT:
                cell.number_format = formats[c - 1]
            if status_column == c:
                _status(cell)

    for i, head in enumerate(headers, start=1):
        letter = get_column_letter(i)
        if widths and i <= len(widths):
            ws.column_dimensions[letter].width = widths[i - 1]
        else:
            longest = max(
                [len(str(head))] + [len(str(r[i - 1])) for r in rows if i <= len(r)][:200]
                or [len(str(head))]
            )
            ws.column_dimensions[letter].width = min(52, max(11, longest + 3))

    ws.row_dimensions[top].height = 28
    if autofilter and rows:
        ws.auto_filter.ref = (
            f"A{top}:{get_column_letter(len(headers))}{top + len(rows)}"
        )
    if freeze:
        ws.freeze_panes = ws.cell(row=top + 1, column=1)
    return top + len(rows) + 2


def _status(cell: Any) -> None:
    """Colour a status cell — and leave the word in it.

    §35: no colour-only meaning. The fill is an accelerant for a reader
    scanning the sheet; the text is what carries the fact.
    """
    word = str(cell.value or "").strip().upper()
    look = STATUS_STYLE.get(word)
    if not look:
        return
    fill, ink = look
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Calibri", size=10, bold=True, color=ink)


def note(ws: Worksheet, text: str, row: int, *, width: int = 110) -> int:
    """A paragraph of explanation between tables."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=10, color=MUTED)
    cell.alignment = WRAP
    if ws.column_dimensions["A"].width is None or ws.column_dimensions["A"].width < width:
        ws.column_dimensions["A"].width = min(width, 110)
    return row + 2


def code(ws: Worksheet, text: str, row: int, *, width: int = 120) -> int:
    """A block of SQL, IR or a formula, one line per row.

    Split across rows rather than crammed into one cell: Excel truncates a
    cell's display at about a thousand characters, so a long query in a single
    cell is a query nobody can read without clicking into the formula bar.
    """
    lines = str(text or "").splitlines() or [""]
    for line in lines:
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = MONO_FONT
        cell.alignment = Alignment(vertical="top")
        row += 1
    ws.column_dimensions["A"].width = max(
        ws.column_dimensions["A"].width or 0, min(width, 120)
    )
    return row + 1


def link(ws: Worksheet, row: int, column: int, text: str, target_sheet: str) -> None:
    """An internal hyperlink to another sheet in this workbook."""
    cell = ws.cell(row=row, column=column, value=text)
    cell.hyperlink = f"#'{target_sheet}'!A1"
    cell.font = LINK_FONT


def page_setup(ws: Worksheet, *, landscape: bool = True) -> None:
    """Print-friendly: landscape, fit to width, repeating header row."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False
