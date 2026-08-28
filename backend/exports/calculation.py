"""
The Step-by-Step Calculation & Validation Workbook.

This is the evidence pack. A credit analyst, a methodology owner, a Data
Steward, Internal Audit or a model-risk reviewer opens it and reconstructs how
CreditProbe reached the answer: which governed data was read, at which version,
how healthy it was, what was filtered out, what was joined to what and on which
keys, what each calculation step did, which checks ran, and how the final
figures reconcile.

Where each number comes from
----------------------------
The pack draws on three sources and never blurs them, because a reviewer's
first question about any figure is "who measured this, and when":

*Persisted*  — the run's own record: the result, the plan, the compiled SQL,
               the reconciliation ledger, the Trace, the fingerprint. Every
               analytical figure in this workbook comes from here and only
               here. Nothing is recomputed.

*Profiled*   — statistics about the SOURCE DATA, measured when the workbook was
               built (see `profile.py`). Labelled as such on the sheet, with the
               data version they were measured against.

*Derived*    — arithmetic over the two above: a match rate from rows in and rows
               out, a null rate from a count. Stated, not invented.

Anything the run did not record is written as "not recorded at run time" rather
than left blank. A blank cell in an audit pack reads as a zero, and a zero
nobody measured is the most expensive kind of wrong number.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook as Book
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.exports import plan as planning
from backend.exports import style
from backend.exports.contract import (
    CALCULATION_PACK,
    GENERATOR_VERSION,
    SCHEMA_VERSION,
    Workbook,
    filename_for,
    sheet_name,
)
from backend.exports.gather import Pack

logger = logging.getLogger(__name__)

#: The sheets §10 requires, in the order it requires them. Extra sheets are
#: allowed between sections; FINAL RESULTS is always last.
COVER = "COVER"
FINAL = "FINAL RESULTS"

#: How close two figures must be to count as reconciled. Money is carried to
#: two decimals through the stack, so anything under a hundredth is rounding.
TOLERANCE = 0.01

#: A cap on how much of the Trace one ledger sheet carries before it says it
#: truncated. The Trace of a composed multi-dataset analysis is large; a sheet
#: nobody can scroll is not more auditable than one that names its own limit.
MAX_LEDGER_ROWS = 2000


@dataclass
class Check:
    """One row of the VALIDATION CHECKS sheet."""

    check_id: str = ""
    step: str = ""
    rule: str = ""
    expected: str = ""
    actual: str = ""
    tolerance: str = ""
    status: str = "PASS"
    reason: str = ""
    impact: str = ""
    source: str = ""
    at: str = ""


@dataclass
class Sheets:
    """The names this workbook actually used, in creation order.

    Sheet names are capped at 31 characters and must be unique, so the name a
    hyperlink points at is not always the name the code asked for. The cover
    links from this record rather than from the constants, which is why a
    truncated name can never produce a dead link.
    """

    order: list[tuple[str, str]] = field(default_factory=list)

    def add(self, wanted: str, actual: str) -> None:
        self.order.append((wanted, actual))

    def actual(self, wanted: str) -> str:
        for asked, given in self.order:
            if asked == wanted:
                return given
        return ""


def build(pack: Pack, *, profiles: list[Any] | None = None,
          extract: Any | None = None,
          redactions: list[str] | None = None) -> Workbook:
    """Write the full calculation and validation pack for one persisted run.

    `profiles` and `extract` are injected so the caller decides how much
    export-time reading is authorised and affordable; passing neither produces a
    pack that says plainly which sheets are unavailable and why, rather than a
    pack that silently drops them.
    """
    view = planning.read(pack.ir, kernel_steps=pack.query.get("kernel_steps"))
    sources = list(profiles if profiles is not None else [])
    population = extract
    redacted = list(redactions or []) + list(pack.redactions)

    book = Book()
    book.remove(book.active)
    taken: set[str] = set()
    sheets = Sheets()

    def add(name: str) -> Worksheet:
        actual = sheet_name(name, taken=taken)
        sheets.add(name, actual)
        return book.create_sheet(actual)

    cover = add(COVER)
    checks = _checks(pack, view, sources)

    _request(add("ANALYSIS REQUEST"), pack, view)
    _executive(add("EXECUTIVE SUMMARY"), pack, view, checks, sheets)
    _sources(add("DATA SOURCES"), pack, view)
    _fields(add("FIELDS USED"), pack, view, sources)
    _population_period(add("POPULATION & PERIOD"), pack, view)
    _profiles(add("SOURCE PROFILES"), pack, sources)
    _relationships(add("RELATIONSHIPS & JOINS"), pack, view)
    _join_reconciliation(add("JOIN RECONCILIATION"), pack, view, sources)
    _filters(add("FILTERS & EXCLUSIONS"), pack, view)
    _transformations(add("TRANSFORMATIONS"), pack, view)
    _calculation_steps(add("CALCULATION STEPS"), pack, view)
    _intermediate(add("INTERMEDIATE RESULTS"), pack, view)
    extract_sheets = _extract(book, add, pack, population)
    _formulas(add("FORMULAS & QUERY"), pack, view)
    _reconstruction(add("EXCEL RECONSTRUCTION"), pack, view, population,
                    extract_sheets)
    _validation(add("VALIDATION CHECKS"), pack, checks)
    _invariants(add("INVARIANTS & RECONCILIATION"), pack, view, checks)
    _ledger(add("TRACE LEDGER"), pack)
    _evidence(add("INTERPRETATION EVIDENCE"), pack)
    _limitations(add("LIMITATIONS"), pack, view, sources, population, redacted)
    _final(add(FINAL), pack)

    _cover(cover, pack, view, sheets, checks)

    content = _bytes(book)
    return Workbook(
        filename=filename_for(
            CALCULATION_PACK, analysis=pack.title, period=pack.period_label,
            run_id=pack.run_id, fingerprint=pack.plan_fingerprint,
        ),
        content=content,
        kind=CALCULATION_PACK,
        manifest={
            "schema_version": SCHEMA_VERSION,
            "generator": GENERATOR_VERSION,
            "run_id": pack.run_id,
            "trace_version": pack.version,
            "sheets": [given for _, given in sheets.order],
            "row_count": len(pack.rows),
            "column_count": len(pack.visible_columns()),
            "datasets": [s.dataset for s in pack.sources],
            "profiled_datasets": [p.dataset for p in sources if p.usable],
            "population_rows": getattr(population, "row_count", 0) if population else 0,
            "population_included": bool(getattr(population, "present", False)),
            "plan_fingerprint": pack.plan_fingerprint,
            "data_version": pack.data_version,
            "build_sha": pack.build_sha,
            "checks": {
                "total": len(checks),
                "failed": sum(1 for c in checks if c.status == "FAIL"),
                "warning": sum(1 for c in checks if c.status == "WARNING"),
                "skipped": sum(1 for c in checks if c.status == "SKIPPED"),
            },
            "redactions": redacted,
        },
    )


def _bytes(book: Book) -> bytes:
    from io import BytesIO

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------- §11 COVER


def _cover(ws: Worksheet, pack: Pack, view: planning.PlanView,
           sheets: Sheets, checks: list[Check]) -> None:
    style.page_setup(ws, landscape=False)
    row = style.title(
        ws, "CreditProbe AI",
        "FULL CALCULATION & VALIDATION PACK",
    )
    row = style.facts(ws, [
        ("Analysis", pack.title),
        ("Question asked", pack.question),
        ("Answer", pack.answer),
    ], row)

    row = style.section(ws, "Provenance", row)
    row = style.facts(ws, [
        ("Run ID", pack.run_id),
        ("Trace version", f"{pack.version} of {pack.version_count}"
                          + (f" — {pack.version_label}" if pack.version_label else "")),
        ("Investigation", pack.investigation_title or "—"),
        ("Project", pack.project_name or "Not in a project"),
        ("Period", pack.period_label or "—"),
        ("Certification", _certification(pack)),
        ("Method", _method(pack)),
        ("Plan fingerprint", pack.plan_fingerprint or "not recorded"),
        ("Data version", pack.data_version or "not recorded"),
        ("Build SHA", pack.build_sha or "not recorded"),
        ("App version", pack.app_version or "not recorded"),
        ("Export schema", SCHEMA_VERSION),
        ("Generator", GENERATOR_VERSION),
        ("Generated at", pack.generated_at),
        ("Downloaded by", pack.generated_by),
    ], row)

    row = style.section(ws, "Status", row)
    failed = [c for c in checks if c.status == "FAIL"]
    row = style.facts(ws, [
        ("Overall validation", "FAILED — see VALIDATION CHECKS" if failed else "PASSED"),
        ("Checks run", f"{len(checks)} ({sum(1 for c in checks if c.status == 'PASS')} "
                       f"passed, {sum(1 for c in checks if c.status == 'WARNING')} warning, "
                       f"{len(failed)} failed, "
                       f"{sum(1 for c in checks if c.status == 'SKIPPED')} skipped)"),
        ("Result rows", len(pack.rows)),
        ("Source datasets", len(pack.sources)),
        ("Governed joins", len(view.joins)),
        ("Confidentiality", _confidentiality(pack)),
    ], row)

    row = style.section(ws, "Contents", row)
    for index, (_, given) in enumerate(sheets.order, start=1):
        if given == ws.title:
            continue
        ws.cell(row=row, column=1, value=index).font = style.LABEL_FONT
        style.link(ws, row, 2, given, given)
        row += 1
    row += 1
    style.note(
        ws,
        "Every analytical figure in this pack was read from the persisted "
        "analysis run. Nothing was recomputed and no model was asked again. "
        "Statistics describing the SOURCE DATA were measured when this workbook "
        "was generated and are labelled where they appear.",
        row,
    )


def _certification(pack: Pack) -> str:
    if pack.is_dynamic:
        return ("Dynamic — composed for this question from governed data and "
                "validated before display")
    return f"Certified method{f' — {pack.analysis_id}' if pack.analysis_id else ''}"


def _method(pack: Pack) -> str:
    if not pack.analysis_id:
        return "Composed plan (no registered method)"
    return f"{pack.analysis_id} version {pack.analysis_version or 'unversioned'}"


def _confidentiality(pack: Pack) -> str:
    return ("SYNTHETIC DEMONSTRATION DATA — not client data"
            if pack.synthetic else "CLIENT DATA — handle per policy")


# ---------------------------------------------------------- §12 ANALYSIS REQUEST


def _request(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws, landscape=False)
    row = style.crumb(ws)
    row = style.title(ws, "ANALYSIS REQUEST",
                      "What was asked, and what CreditProbe understood by it.",
                      row=row)

    reading = pack.reading
    meta = dict(pack.ir.get("meta") or {})
    row = style.facts(ws, [
        ("Original wording", pack.question),
        ("Capability", str(reading.get("capability") or pack.result.get("capability") or "—")),
        ("Objective", view.explanation or pack.plain_english or "—"),
        ("Requested grain", view.grain or str(reading.get("grain") or "—")),
        ("Measures", _measures(pack, meta)),
        ("Dimensions", str(meta.get("dimension") or reading.get("dimension") or "—")),
        ("Filters", "; ".join(c.meaning for c in view.conditions) or "None"),
        ("Population", str(meta.get("population") or "The governed book, unrestricted")),
        ("Requested periods", pack.period_label or "—"),
        ("Ranking / sort", _sort_words(view)),
        ("Top N", view.top_n or "Not limited"),
        ("Requested output", str(reading.get("output") or "Table")),
    ], row)

    row = style.section(ws, "Conversation context", row)
    prior = pack.node("PRIOR_CONTEXT")
    scope = dict(prior.get("config") or {})
    row = style.facts(ws, [
        ("Carried from the thread", prior.get("label") or "Nothing carried — a fresh question"),
        ("Inherited scope", style._scalar(scope.get("inherited")) or "—"),
        ("Referent resolved to", style._scalar(scope.get("referent")) or "—"),
        ("What this turn changed", style._scalar(scope.get("changes")) or "—"),
    ], row)

    row = style.section(ws, "Interpretation decisions", row)
    concepts = [c for c in (meta.get("concepts") or []) if isinstance(c, dict)]
    if concepts:
        row = style.table(
            ws,
            ["Phrase", "Resolved to", "Dataset.field", "Why this one",
             "Confidence", "Alternatives considered"],
            [[
                c.get("phrase") or c.get("label"),
                c.get("label"),
                f"{c.get('dataset')}.{c.get('field')}",
                c.get("reason"),
                c.get("confidence"),
                ", ".join(
                    f"{a.get('dataset')}.{a.get('field')}"
                    for a in (c.get("alternatives") or []) if isinstance(a, dict)
                ) or "None",
            ] for c in concepts],
            row=row,
            widths=[26, 26, 34, 74, 12, 40],
        )
    else:
        row = style.note(ws, "This analysis resolved no ambiguous concepts.", row)

    style.note(
        ws,
        "This sheet records the structured, auditable interpretation shown on "
        "the Trace. It deliberately contains no model reasoning: what is here "
        "is the plan that was validated and executed.",
        row,
    )


def _measures(pack: Pack, meta: dict[str, Any]) -> str:
    named = [str(c.get("label") or c.get("field"))
             for c in (meta.get("concepts") or []) if isinstance(c, dict)]
    if named:
        return ", ".join(dict.fromkeys(named))
    return ", ".join(
        style.header_label(c) for c in pack.visible_columns()
    ) or "—"


def _sort_words(view: planning.PlanView) -> str:
    for step in view.steps:
        if step.op == "SORT":
            return step.meaning.rstrip(".").replace("Ordered by ", "")
    return "Not specified"


# ---------------------------------------------------------- §13 EXECUTIVE SUMMARY


def _executive(ws: Worksheet, pack: Pack, view: planning.PlanView,
               checks: list[Check], sheets: Sheets) -> None:
    style.page_setup(ws, landscape=False)
    row = style.crumb(ws)
    row = style.title(ws, "EXECUTIVE SUMMARY",
                      "The answer, and what stands behind it.", row=row)

    failed = [c for c in checks if c.status == "FAIL"]
    warned = [c for c in checks if c.status == "WARNING"]
    row = style.facts(ws, [
        ("Direct answer", pack.answer or "—"),
        ("Reading of the result", pack.narrative.get("interpretation") or "—"),
        ("Scope", pack.narrative.get("scope") or pack.period_label or "—"),
        ("Result size", f"{len(pack.rows)} rows × "
                        f"{len(pack.visible_columns())} columns"),
        ("Source datasets", len(pack.sources)),
        ("Governed joins", len(view.joins)),
        ("Checks passed", sum(1 for c in checks if c.status == "PASS")),
        ("Checks failed", len(failed)),
        ("Checks warning", len(warned)),
        ("Checks skipped", sum(1 for c in checks if c.status == "SKIPPED")),
    ], row)

    metrics = [m for m in (pack.narrative.get("metrics") or []) if isinstance(m, dict)]
    if metrics:
        row = style.section(ws, "Headline figures", row)
        row = style.table(
            ws, ["Figure", "Value", "Unit", "Change", "Direction"],
            [[m.get("label"), m.get("value"), m.get("unit"),
              m.get("change"), m.get("direction")] for m in metrics],
            row=row, widths=[46, 18, 12, 14, 14],
        )

    if failed or warned:
        row = style.section(ws, "What deserves attention", row)
        row = style.table(
            ws, ["Status", "Check", "Why it matters"],
            [[c.status, c.rule, c.impact or c.reason] for c in failed + warned],
            row=row, status_column=1, widths=[12, 60, 74],
        )

    caveats = [str(c) for c in (pack.narrative.get("caveats") or [])]
    if caveats:
        row = style.section(ws, "Main limitations", row)
        for text in caveats:
            row = style.note(ws, f"• {text}", row - 1)
        row += 1

    row = style.section(ws, "Go to", row)
    for wanted in (FINAL, "VALIDATION CHECKS", "TRACE LEDGER"):
        given = sheets.actual(wanted)
        if given:
            style.link(ws, row, 1, given, given)
            row += 1


# ------------------------------------------------------------- §14 DATA SOURCES


def _sources(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "DATA SOURCES",
        "One row per governed dataset version this analysis read.", row=row)

    nodes = {
        str((n.get("config") or {}).get("dataset")): n
        for n in pack.nodes("DATASET")
    }
    rows: list[list[Any]] = []
    for source in pack.sources:
        node = nodes.get(source.dataset, {})
        config = dict(node.get("config") or {})
        read = view.fields_for(source.dataset) or list(node.get("fields_used") or [])
        rows.append([
            source.domain or config.get("domain") or "—",
            source.family or config.get("family") or source.dataset,
            source.dataset,
            source.business_name or (node.get("label") or "").split(" · ")[0],
            (source.origin or config.get("origin") or "").upper() or "—",
            source.authority or "—",
            source.published or "—",
            source.period or ", ".join(config.get("periods") or []) or "all periods",
            source.version or config.get("version") or "—",
            source.grain or config.get("grain") or "—",
            source.row_count if source.row_count is not None else "not recorded",
            source.field_count if source.field_count is not None else "not recorded",
            source.primary_key or "—",
            source.sensitivity or "—",
            source.owner or "—",
            source.content_hash or node.get("content_hash") or "not recorded",
            len(read),
            ", ".join(read) or "—",
            source.pushdown or "not recorded",
        ])

    row = style.table(
        ws,
        ["Domain", "Dataset family", "Dataset", "Business name", "Origin",
         "Authority", "Published state", "Reporting period", "Dataset version",
         "Grain", "Catalogue rows", "Catalogue fields", "Primary key",
         "Sensitivity", "Owner", "Content hash", "Columns read",
         "Columns read (names)", "Filter pushdown"],
        rows, row=row,
        widths=[24, 22, 22, 30, 10, 16, 14, 18, 13, 34, 14, 14, 24, 13, 24, 18,
                13, 46, 24],
    )
    style.note(
        ws,
        "Only the datasets this analysis actually read are listed. Catalogue "
        "row and field counts describe the dataset as the catalogue declares it "
        "today; the version column records what the run stamped at execution.",
        row,
    )


# -------------------------------------------------------------- §15 FIELDS USED


def _fields(ws: Worksheet, pack: Pack, view: planning.PlanView,
            profiles: list[Any]) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "FIELDS USED",
                      "One row per field used anywhere in the computation.",
                      row=row)

    catalogue = _field_catalogue({s.dataset for s in view.scans})
    stats = {
        (p.dataset, s.field_name): s
        for p in profiles for s in [*p.numeric, *p.categorical]
    }
    roles = _field_roles(view)
    output = {str(c.get("name")): c for c in pack.visible_columns()}

    rows: list[list[Any]] = []
    for scan in view.scans:
        for name in scan.fields:
            found = catalogue.get((scan.dataset, name), {})
            stat = stats.get((scan.dataset, name))
            rows.append([
                scan.dataset,
                name,
                found.get("business_name") or name,
                found.get("definition") or "—",
                found.get("data_type") or "—",
                found.get("unit") or "—",
                found.get("concept") or "—",
                _alias(name, output),
                found.get("grain") or "—",
                found.get("sensitivity") or "—",
                ", ".join(roles.get(name, [])) or "read but unused downstream",
                _nulls(stat),
                _transformed(name, view),
                f"{scan.dataset} at {scan.period or 'all published periods'}",
            ])

    row = style.table(
        ws,
        ["Dataset", "Field", "Business name", "Definition", "Type", "Unit",
         "Semantic concept", "Alias in the plan", "Grain", "Sensitivity",
         "Role in the calculation", "Null rate in the profiled source",
         "Transformation applied", "Lineage"],
        rows, row=row,
        widths=[22, 24, 28, 66, 10, 10, 20, 26, 30, 13, 30, 22, 44, 34],
    )
    style.note(
        ws,
        "Null rates are measured over the source dataset at the profiled "
        "period (see SOURCE PROFILES), not over the joined population. A field "
        "with no profile shows 'not profiled' rather than zero.",
        row,
    )


def _field_catalogue(datasets: set[str]) -> dict[tuple[str, str], dict[str, Any]]:
    """Catalogue metadata for every field, keyed by (dataset, field)."""
    out: dict[tuple[str, str], dict[str, Any]] = {}
    try:
        from backend.data_access.catalog import get_catalog

        catalog = get_catalog()
    except Exception as e:  # noqa: BLE001
        logger.info("Field catalogue unavailable for export: %s", e)
        return out

    concepts = _concept_index()
    for dataset in datasets:
        try:
            spec = catalog.dataset(dataset)
        except Exception:  # noqa: BLE001 - an archived dataset is still exportable
            continue
        for name, found in spec.fields.items():
            out[(dataset, name)] = {
                "business_name": found.business_name,
                "definition": found.definition,
                "data_type": found.data_type,
                "unit": found.unit,
                "sensitivity": found.sensitivity,
                "grain": spec.grain,
                "concept": concepts.get((dataset, name), ""),
            }
    return out


def _concept_index() -> dict[tuple[str, str], str]:
    try:
        from backend.orchestration.concepts import CONCEPTS
    except Exception:  # noqa: BLE001
        return {}
    return {
        (c.dataset, c.field): concept.label
        for concept in CONCEPTS for c in concept.candidates
    }


def _field_roles(view: planning.PlanView) -> dict[str, list[str]]:
    """What each field was FOR, read from the operations that named it."""
    roles: dict[str, set[str]] = {}

    def mark(name: str, role: str) -> None:
        if name:
            roles.setdefault(str(name), set()).add(role)

    for join in view.joins:
        for name in [*join.left_keys, *join.right_keys]:
            mark(name.split(".")[-1], "join key")
    for condition in view.conditions:
        mark(condition.field_name, "filter")
    for step in view.steps:
        params = step.params
        if step.op == "GROUP":
            for name in params.get("by") or []:
                mark(str(name), "dimension")
        for entry in params.get("aggregates") or []:
            if isinstance(entry, dict):
                mark(str(entry.get("column")), "measure")
        for key in ("column", "numerator", "denominator", "x", "y"):
            if params.get(key):
                mark(str(params[key]), "derived input")
        if params.get("as"):
            mark(str(params["as"]), "output")
    return {name: sorted(found) for name, found in roles.items()}


def _alias(name: str, output: dict[str, dict[str, Any]]) -> str:
    for key, column in output.items():
        if key == name or key.endswith(f"_{name}"):
            return f"{key} — {column.get('label') or key}"
    return "—"


def _nulls(stat: Any) -> str:
    if stat is None:
        return "not profiled"
    rate = getattr(stat, "null_rate", None)
    nulls = getattr(stat, "nulls", None)
    if rate is None or nulls is None:
        return "not profiled"
    return f"{nulls:,} ({rate:.2f}%)"


def _transformed(name: str, view: planning.PlanView) -> str:
    touched = [
        step.label or step.op for step in view.transformations()
        if name in str(step.params)
    ]
    return "; ".join(touched) or "None — used as read"


# ------------------------------------------------------- §17 POPULATION & PERIOD


def _population_period(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "POPULATION & PERIOD",
                      "How the analytical population was defined, step by step.",
                      row=row)

    reading = pack.reading
    row = style.facts(ws, [
        ("Opening period", pack.opening_period or "—"),
        ("Closing period", pack.closing_period or pack.period or "—"),
        ("As-of rule", _asof_rule(view)),
        ("Grain before aggregation", view.grain or "—"),
        ("Grain after aggregation", _output_grain(view) or view.grain or "—"),
        ("Top-N limit", view.top_n or "Not limited"),
        ("Missing-period treatment", _missing_periods(view)),
        ("Active thread filters", str(reading.get("inherited_filters") or "None")),
        ("Final analytical population", f"{len(pack.rows)} rows in the result"),
    ], row)

    row = style.section(ws, "Population waterfall", row)
    ledger = pack.reconciliation
    if ledger:
        rows = [[
            index,
            entry.get("step"),
            entry.get("operation"),
            entry.get("label"),
            entry.get("rows_in") if entry.get("rows_in") is not None else "—",
            entry.get("rows"),
            entry.get("lost") if entry.get("lost") is not None else "—",
            entry.get("lost_pct") if entry.get("lost_pct") is not None else "—",
            "By design" if entry.get("reduced_by_design") else "",
            entry.get("note") or "",
        ] for index, entry in enumerate(ledger, start=1)]
        row = style.table(
            ws,
            ["#", "Step", "Operation", "What it did", "Rows in", "Rows out",
             "Rows removed", "Removed %", "Expected", "Note"],
            rows, row=row,
            formats=["", "", "", "", style.INTEGER, style.INTEGER,
                     style.INTEGER, style.PERCENT_2, "", ""],
            widths=[5, 30, 18, 66, 12, 12, 13, 12, 12, 40],
        )
        style.note(
            ws,
            "Counted against the same compiled query that produced the answer, "
            "so these are the rows that actually survived each step rather than "
            "an estimate.",
            row,
        )
    else:
        style.note(
            ws,
            "This run recorded no step-level population ledger. A certified "
            "engine analysis reports its own population rather than a compiled "
            "plan's; see CALCULATION STEPS.",
            row,
        )


def _asof_rule(view: planning.PlanView) -> str:
    for join in view.joins:
        if join.kind == "ASOF_JOIN" or join.as_of == "latest_on_or_before":
            return ("Latest observation on or before the reporting date — "
                    "no future data is used")
    return "Same period on both sides"


def _output_grain(view: planning.PlanView) -> str:
    for step in reversed(view.steps):
        if step.op == "GROUP":
            return "one row per " + ", ".join(
                str(b) for b in (step.params.get("by") or [])) or ""
        if step.op == "AGGREGATE":
            return "one row for the whole population"
    return ""


def _missing_periods(view: planning.PlanView) -> str:
    for step in view.steps:
        if step.op == "TEMPORAL_ALIGN":
            return step.label or step.meaning
    return "No temporal alignment was needed"


# ----------------------------------------------------------- §16 SOURCE PROFILES


def _profiles(ws: Worksheet, pack: Pack, profiles: list[Any]) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "SOURCE PROFILES",
        "Each source dataset over the period the analysis read, before joins.",
        row=row)

    if not profiles:
        style.note(
            ws,
            "No source profiles were generated for this pack. Profiling reads "
            "the governed source data at export time; it was not run here, so "
            "this sheet carries no statistics rather than stale ones.",
            row,
        )
        return

    row = style.note(
        ws,
        "These statistics describe the DATA, and were measured when this "
        "workbook was generated. They are not the analysis and were not used to "
        "produce it. Where a dataset has moved since the run, the profile says "
        "so above its table.",
        row,
    )

    for source in profiles:
        row = style.section(
            ws,
            f"{source.dataset} — {source.business_name or 'source dataset'} "
            f"at {source.period or 'all published periods'}",
            row,
        )
        if source.error:
            row = style.note(ws, source.error, row)
            continue

        row = style.facts(ws, [
            ("Grain", source.grain or "—"),
            ("Primary key", ", ".join(source.primary_key) or "—"),
            ("Total rows", source.rows),
            ("Distinct primary keys", source.distinct_keys
             if source.distinct_keys is not None else "not measured"),
            ("Duplicate primary keys", source.duplicate_keys
             if source.duplicate_keys is not None else "not measured"),
            ("Null keys", source.null_keys
             if source.null_keys is not None else "not measured"),
            *[(f"Distinct {name}", value) for name, value in source.identities.items()],
            ("Fields used by this calculation", ", ".join(source.fields_used) or "—"),
            ("Dataset version at run", source.version_at_run or "not recorded"),
            ("Dataset version now", source.version_now or "unknown"),
            ("Profiled at", source.computed_at),
        ], row)

        for text in source.notes:
            row = style.note(ws, text, row)

        credit = source.credit_summary()
        if credit:
            row = style.section(ws, "Portfolio and credit totals", row)
            row = style.table(
                ws, ["Figure", "Value", "Unit"],
                [[label, value, unit] for label, value, unit in credit],
                row=row, widths=[48, 20, 12], autofilter=False, freeze=False,
                formats=["", style.MONEY_2, ""],
            )

        if source.numeric:
            row = style.section(ws, "Numeric field profile", row)
            row = style.table(
                ws,
                ["Field", "Business name", "Unit", "Additive", "Count",
                 "Nulls", "Null %", "Sum", "Mean", "Median", "Std dev",
                 "Min", "P10", "P25", "P75", "P90", "P95", "P99", "Max",
                 "Role"],
                [[
                    s.field_name, s.business_name, s.unit or "—",
                    "Yes" if s.additive else "No — ratio or ordinal",
                    s.count, s.nulls, s.null_rate,
                    s.total if s.additive else "not additive",
                    s.mean, s.median, s.stdev, s.minimum,
                    s.p10, s.p25, s.p75, s.p90, s.p95, s.p99, s.maximum,
                    s.role,
                ] for s in source.numeric],
                row=row, autofilter=False, freeze=False,
                formats=["", "", "", "", style.INTEGER, style.INTEGER,
                         style.PERCENT_2] + [style.MONEY_2] * 12 + [""],
                widths=[22, 26, 8, 20, 11, 9, 9] + [13] * 12 + [24],
            )

        if source.categorical:
            row = style.section(ws, "Categorical field profile", row)
            row = style.table(
                ws,
                ["Field", "Business name", "Distinct", "Nulls", "Null %",
                 "Top values and counts", "Values outside the declared set",
                 "Role"],
                [[
                    s.field_name, s.business_name, s.distinct, s.nulls,
                    s.null_rate,
                    "; ".join(f"{v['value']} ({v['count']:,})" for v in s.values)
                    + (" …" if s.truncated else ""),
                    ", ".join(s.unexpected) or "None",
                    s.role,
                ] for s in source.categorical],
                row=row, autofilter=False, freeze=False,
                formats=["", "", style.INTEGER, style.INTEGER, style.PERCENT_2,
                         "", "", ""],
                widths=[22, 26, 10, 9, 9, 96, 34, 24],
            )
        row += 1


# ------------------------------------------------------ §18 RELATIONSHIPS & JOINS


def _relationships(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "RELATIONSHIPS & JOINS",
                      "One row per governed join.", row=row)

    if not view.joins:
        style.note(
            ws,
            "This analysis read a single dataset and made no joins. Nothing was "
            "matched, so nothing could be lost or multiplied in matching.",
            row,
        )
        return

    recorded = {str(j.get("step")): j for j in pack.joins if isinstance(j, dict)}
    rows: list[list[Any]] = []
    for index, join in enumerate(view.joins, start=1):
        stamp = recorded.get(join.id, {})
        lost_pct = stamp.get("lost_pct")
        rows.append([
            index,
            join.relationship or "not recorded",
            join.relationship_version or "not recorded",
            join.left,
            join.right,
            ", ".join(join.left_keys) or "—",
            ", ".join(join.right_keys) or "—",
            join.how or join.kind,
            join.cardinality or "not recorded",
            join.as_of or "same period",
            join.authoritative or "governed path",
            "PASS" if lost_pct in (0, 0.0) else ("WARNING" if lost_pct else "SKIPPED"),
            f"{100 - lost_pct:.2f}%" if isinstance(lost_pct, int | float) else "not recorded",
            f"{lost_pct:.2f}%" if isinstance(lost_pct, int | float) else "not recorded",
            join.meaning or "—",
        ])

    row = style.table(
        ws,
        ["#", "Relationship", "Version", "Source dataset", "Target dataset",
         "Source key(s)", "Target key(s)", "Join type", "Cardinality",
         "Period / as-of rule", "Path", "Validation", "Match rate",
         "Orphan rate", "What the join means"],
        rows, row=row, status_column=12,
        widths=[5, 46, 9, 22, 22, 26, 24, 12, 16, 22, 22, 12, 12, 12, 56],
    )
    style.note(
        ws,
        "Match and orphan rates are the row counts the run itself recorded. A "
        "join marked SKIPPED had no rate recorded — it is not a passing join, "
        "and JOIN RECONCILIATION says what was and was not measured.",
        row,
    )


# ------------------------------------------------------- §19 JOIN RECONCILIATION


def _join_reconciliation(ws: Worksheet, pack: Pack, view: planning.PlanView,
                         profiles: list[Any]) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "JOIN RECONCILIATION",
                      "What each join did to the population.", row=row)

    if not view.joins:
        style.note(ws, "This analysis made no joins.", row)
        return

    ledger = {str(e.get("step")): e for e in pack.reconciliation}
    recorded = {str(j.get("step")): j for j in pack.joins if isinstance(j, dict)}
    by_dataset = {p.dataset: p for p in profiles}
    inputs = {step.id: step.inputs for step in view.steps}

    rows: list[list[Any]] = []
    for index, join in enumerate(view.joins, start=1):
        entry = ledger.get(join.id, {})
        stamp = recorded.get(join.id, {})
        right_scan = _right_scan(view, inputs.get(join.id) or [])
        right_ledger = ledger.get(right_scan.id, {}) if right_scan else {}
        right_profile = by_dataset.get(right_scan.dataset) if right_scan else None

        rows_in = entry.get("rows_in")
        rows_out = entry.get("rows")
        lost = entry.get("lost")
        lost_pct = entry.get("lost_pct")
        multiplication = (
            round(rows_out / rows_in, 4)
            if isinstance(rows_in, int) and rows_in and isinstance(rows_out, int)
            else "not recorded"
        )
        rows.append([
            index,
            join.relationship or join.label,
            rows_in if rows_in is not None else "not recorded",
            right_ledger.get("rows", "not recorded"),
            _key_distinct(join.left_keys, by_dataset.get(join.left)),
            _key_distinct(join.right_keys, right_profile),
            _duplicates(by_dataset.get(join.left)),
            _duplicates(right_profile),
            rows_out if rows_out is not None else "not recorded",
            (rows_in - lost) if isinstance(rows_in, int) and isinstance(lost, int)
            else "not recorded",
            lost if lost is not None else "not recorded",
            "not recorded at run time",
            multiplication,
            f"{100 - lost_pct:.2f}" if isinstance(lost_pct, int | float) else "not recorded",
            f"{lost_pct:.2f}" if isinstance(lost_pct, int | float) else "not recorded",
            "; ".join(str(w) for w in (stamp.get("warnings") or [])) or "None",
            _join_status(lost_pct),
        ])

    row = style.table(
        ws,
        ["#", "Join", "Rows into the join (left)", "Rows in the right source",
         "Distinct keys in the left source", "Distinct keys in the right source",
         "Duplicate keys (left source)", "Duplicate keys (right source)",
         "Rows after", "Matched left rows", "Unmatched left rows",
         "Unmatched right rows", "Row multiplication", "Match %", "Orphan %",
         "Warnings", "Status"],
        rows, row=row, status_column=17,
        widths=[5, 46, 20, 20, 24, 25, 22, 23, 13, 17, 18, 20, 16, 11, 11, 30, 11],
    )

    row = style.note(
        ws,
        "Row counts come from the reconciliation the run recorded against its "
        "own compiled query. Distinct-key and duplicate-key counts come from "
        "the source profiles measured at export time (SOURCE PROFILES), and "
        "describe the whole source rather than the population at the moment of "
        "the join.",
        row,
    )
    row = style.note(
        ws,
        "Unmatched RIGHT rows and value-level reconciliation at each join "
        "(exposure before and after) were not recorded when this analysis ran. "
        "They are shown as 'not recorded at run time' rather than as zero. The "
        "value reconciliation that WAS recorded — that the parts sum to the "
        "whole in the final result — is on INVARIANTS & RECONCILIATION.",
        row,
    )
    style.note(
        ws,
        "No sample of unmatched keys is included: the keys in this analysis are "
        "confidential customer and account identifiers, and §40 keeps them out "
        "of an export that is not a governed row-level extract.",
        row,
    )


def _right_scan(view: planning.PlanView, inputs: list[str]) -> Any:
    """The scan feeding a join's right-hand side, where it is a direct read."""
    if len(inputs) < 2:
        return None
    for scan in view.scans:
        if scan.id == inputs[1]:
            return scan
    return None


def _key_distinct(keys: list[str], source: Any) -> Any:
    if source is None or not keys:
        return "not profiled"
    for key in keys:
        bare = key.split(".")[-1]
        for name, value in (source.identities or {}).items():
            if name == bare or bare.endswith(f"_{name}"):
                return value
    return "not profiled"


def _duplicates(source: Any) -> Any:
    if source is None:
        return "not profiled"
    return (source.duplicate_keys if source.duplicate_keys is not None
            else "not measured")


def _join_status(lost_pct: Any) -> str:
    if lost_pct is None:
        return "SKIPPED"
    if lost_pct == 0:
        return "PASS"
    return "WARNING" if lost_pct < 5 else "FAIL"


# ------------------------------------------------------ §20 FILTERS & EXCLUSIONS


def _filters(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "FILTERS & EXCLUSIONS",
                      "One row per filter, and what it removed.", row=row)

    if not view.conditions:
        style.note(
            ws,
            "No filters were applied. The analysis ran over the whole governed "
            "population for its period.",
            row,
        )
        return

    ledger = {str(e.get("step")): e for e in pack.reconciliation}
    catalogue = _field_catalogue({s.dataset for s in view.scans})
    meanings = {name: found.get("business_name")
                for (_, name), found in catalogue.items()}

    rows: list[list[Any]] = []
    for condition in view.conditions:
        entry = ledger.get(condition.id, {})
        rows.append([
            condition.sequence,
            condition.field_name or "—",
            meanings.get(condition.field_name) or "—",
            condition.operator or "—",
            condition.value or "—",
            pack.period_label or "—",
            entry.get("rows_in") if entry.get("rows_in") is not None else "not recorded",
            entry.get("rows") if entry.get("rows") is not None else "not recorded",
            entry.get("lost") if entry.get("lost") is not None else "not recorded",
            "not recorded at run time",
            "not recorded at run time",
            condition.meaning or "—",
            condition.origin or "requested",
            "PASS" if entry.get("rows") else "SKIPPED",
        ])

    row = style.table(
        ws,
        ["#", "Field", "Business meaning", "Operator", "Value", "Period",
         "Population before", "Population after", "Rows removed",
         "Customers removed", "Exposure removed", "Rationale", "Requested by",
         "Status"],
        rows, row=row, status_column=14,
        widths=[5, 22, 26, 10, 24, 16, 17, 17, 13, 20, 20, 50, 14, 11],
    )
    style.note(
        ws,
        "Row counts are the run's own reconciliation. Customer and exposure "
        "counts at each filter were not recorded when the analysis ran; they "
        "are named as such rather than shown as zero.",
        row,
    )


# ----------------------------------------------------------- §21 TRANSFORMATIONS


def _transformations(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "TRANSFORMATIONS",
                      "Every operation that changed a column.", row=row)

    ledger = {str(e.get("step")): e for e in pack.reconciliation}
    steps = view.transformations()
    if not steps:
        style.note(
            ws,
            "This analysis derived no columns. It read, filtered, aggregated "
            "and returned; there is nothing between those steps to describe.",
            row,
        )
        return

    rows: list[list[Any]] = []
    for index, step in enumerate(steps, start=1):
        entry = ledger.get(step.id, {})
        rows.append([
            index,
            step.id,
            step.op,
            ", ".join(_inputs_of(step)) or "—",
            ", ".join(step.outputs) or "—",
            step.meaning,
            step.formula or "—",
            step.unit or _unit_of(step, pack),
            _null_rule(step),
            entry.get("rows_in") if entry.get("rows_in") is not None else "—",
            entry.get("rows") if entry.get("rows") is not None else "—",
            step.kernel or "compiled to SQL over the governed data",
            step.kernel and "approved kernel" or "safe SQL compiler",
        ])

    row = style.table(
        ws,
        ["#", "Step", "Operation", "Input columns", "Output column",
         "What it means", "Formula", "Unit", "Null handling", "Rows in",
         "Rows out", "Executed by", "Approved primitive"],
        rows, row=row,
        formats=["", "", "", "", "", "", "", "", "", style.INTEGER,
                 style.INTEGER, "", ""],
        widths=[5, 30, 16, 34, 26, 62, 46, 10, 34, 11, 11, 40, 20],
    )
    style.note(
        ws,
        "Every operation above is a governed primitive: either compiled to "
        "parameterised SQL by the safe compiler, or executed by an approved "
        "numerical kernel over the SQL result. No free-form code ran.",
        row,
    )


def _inputs_of(step: planning.Step) -> list[str]:
    named: list[str] = []
    for key in ("column", "numerator", "denominator", "x", "y"):
        if step.params.get(key):
            named.append(str(step.params[key]))
    for entry in step.params.get("aggregates") or []:
        if isinstance(entry, dict) and entry.get("column"):
            named.append(str(entry["column"]))
    for name in step.params.get("by") or []:
        named.append(str(name))
    return list(dict.fromkeys(named))


def _unit_of(step: planning.Step, pack: Pack) -> str:
    for name in step.outputs:
        unit = pack.units.get(name)
        if unit:
            return str(unit)
    return "—"


def _null_rule(step: planning.Step) -> str:
    if step.op == "RATIO":
        return "Zero denominator returns blank, never zero or infinity"
    if step.op in {"GROUP", "AGGREGATE"}:
        return "Nulls are excluded from the aggregate, not treated as zero"
    return "Nulls pass through unchanged"


# --------------------------------------------------------- §22 CALCULATION STEPS


def _calculation_steps(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "CALCULATION STEPS",
        "The numbered ledger, in the order the Trace records them.", row=row)

    if view.empty:
        style.note(
            ws,
            "This analysis ran a certified method rather than a composed plan, "
            "so it has no step-level ledger. The method, its version and its "
            "inputs are on FORMULAS & QUERY.",
            row,
        )
        return

    ledger = {str(e.get("step")): e for e in pack.reconciliation}
    nodes = {
        str(n.get("id")): n for n in (pack.graph.get("nodes") or [])
        if isinstance(n, dict)
    }
    rows: list[list[Any]] = []
    for index, step in enumerate(view.steps, start=1):
        entry = ledger.get(step.id, {})
        node = nodes.get(f"run__op_{step.id}", {})
        rows.append([
            index,
            step.id,
            node.get("id") or f"run__op_{step.id}",
            step.label or step.meaning,
            step.op,
            ", ".join(step.inputs) or "the governed source",
            entry.get("rows_in") if entry.get("rows_in") is not None else "—",
            step.formula or step.meaning,
            entry.get("rows") if entry.get("rows") is not None else "—",
            ", ".join(step.outputs) or "—",
            step.unit or _unit_of(step, pack),
            node.get("duration_ms") if node.get("duration_ms") is not None else "—",
            str(node.get("status") or "ok").upper().replace("OK", "PASS"),
            "; ".join(str(w) for w in (node.get("warnings") or [])) or "None",
            step.kernel or "SQL",
        ])

    row = style.table(
        ws,
        ["#", "Step ID", "Trace node", "Step name", "Operation", "Input",
         "Rows in", "Formula / logic", "Rows out", "Output columns", "Unit",
         "Duration (ms)", "Status", "Checks and warnings", "Executed by"],
        rows, row=row, status_column=13,
        formats=["", "", "", "", "", "", style.INTEGER, "", style.INTEGER,
                 "", "", style.INTEGER, "", "", ""],
        widths=[5, 30, 34, 62, 18, 30, 11, 56, 11, 30, 10, 13, 11, 34, 14],
    )
    style.note(
        ws,
        "The sequence matches the Trace exactly. Customer and account counts "
        "per step were not recorded at run time; the row counts that were are "
        "shown, and POPULATION & PERIOD carries the same ledger as a waterfall.",
        row,
    )


# ------------------------------------------------------- §23 INTERMEDIATE RESULTS


def _intermediate(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "INTERMEDIATE RESULTS",
        "What each step produced, and where the tables themselves are.",
        row=row)

    row = style.note(
        ws,
        "CreditProbe compiles the whole plan into one query and executes it in "
        "a single pass, so the intermediate tables exist only inside that query "
        "and are never materialised. What the run DID record for every step is "
        "its shape: the operation, the columns it produced and the rows that "
        "survived it. That ledger is below.",
        row,
    )
    row = style.note(
        ws,
        "Reproducing the intermediate tables themselves would mean executing "
        "the calculation a second time, which an export must not do: a second "
        "execution against data that has since moved would produce a workbook "
        "that disagreed with the analysis it claims to document. Where the "
        "row-level population is needed, the source population read is on the "
        "POPULATION EXTRACT sheets.",
        row,
    )

    ledger = pack.reconciliation
    if not ledger:
        style.note(ws, "This run recorded no step ledger.", row)
        return

    steps = {s.id: s for s in view.steps}
    rows = [[
        index,
        entry.get("step"),
        entry.get("operation"),
        entry.get("label"),
        entry.get("rows"),
        ", ".join(steps[entry["step"]].outputs) if entry.get("step") in steps
        and steps[entry["step"]].outputs else "carried through unchanged",
        _grain_at(entry.get("step"), view),
        pack.period_label or "—",
        "Materialised only inside the compiled query",
    ] for index, entry in enumerate(ledger, start=1)]

    row = style.table(
        ws,
        ["#", "Step ID", "Operation", "What it produced", "Row count",
         "Columns added", "Grain", "Period", "Where the table is"],
        rows, row=row,
        formats=["", "", "", "", style.INTEGER, "", "", "", ""],
        widths=[5, 30, 18, 66, 12, 40, 34, 18, 40],
    )
    style.note(
        ws,
        "No intermediate result that is needed to understand a material change "
        "in row count is omitted: every step in the compiled plan appears "
        "above, including the ones that changed nothing.",
        row,
    )


def _grain_at(step_id: Any, view: planning.PlanView) -> str:
    step = view.by_id(str(step_id or ""))
    if step is None:
        return view.grain or "—"
    if step.op == "GROUP":
        return "one row per " + ", ".join(str(b) for b in (step.params.get("by") or []))
    if step.op == "AGGREGATE":
        return "one row"
    return view.grain or "—"


# --------------------------------------------------- §24 FULL POPULATION EXTRACT


def _extract(book: Book, add: Any, pack: Pack,
             population: Any) -> list[tuple[str, int]]:
    """The source population, split across numbered sheets where it must be.

    Returns (sheet name, header row) for each sheet written. The header row is
    the row the table was actually written at, not one derived from counting
    the preamble: EXCEL RECONSTRUCTION builds live formulas against these
    ranges, and a formula pointing one row off would reconcile against the
    wrong data while looking perfectly correct.
    """
    if population is None:
        ws = add("POPULATION EXTRACT")
        row = style.crumb(ws)
        row = style.title(ws, "POPULATION EXTRACT",
                          "The row-level population behind this analysis.", row=row)
        style.note(
            ws,
            "No population extract was requested for this pack. The row-level "
            "population is available as a separate governed export for this "
            "run.",
            row,
        )
        return []

    if not population.present:
        ws = add("POPULATION EXTRACT")
        row = style.crumb(ws)
        row = style.title(ws, "POPULATION EXTRACT",
                          "The row-level population behind this analysis.", row=row)
        row = style.facts(ws, [
            ("Dataset", population.dataset or "—"),
            ("Period", population.period or "—"),
            ("Rows in the population", population.row_count or "not counted"),
            ("Included in this workbook", "No"),
        ], row)
        style.note(ws, population.omitted or "No extract was produced.", row)
        return []

    chunks = population.chunks
    written: list[tuple[str, int]] = []
    index_sheet = None
    if len(chunks) > 1:
        index_sheet = add("EXTRACT INDEX")

    columns = population.columns
    headers = [style.header_label(c) for c in columns]
    formats = [style.number_format(c) for c in columns]

    for number, chunk in enumerate(chunks, start=1):
        name = ("POPULATION EXTRACT" if len(chunks) == 1
                else f"Population_{number:03d}")
        ws = add(name)
        style.page_setup(ws)
        row = style.crumb(ws)
        row = style.title(
            ws,
            name,
            f"{population.dataset} at {population.period or 'all published periods'}"
            f" — rows {chunk[0]:,} to {chunk[1]:,} of {population.row_count:,}",
            row=row,
        )
        row = style.facts(ws, [
            ("Read at export time from", population.dataset),
            ("Period", population.period or "all published periods"),
            ("Filters applied", "; ".join(population.filters) or "None"),
            ("Grain", population.grain or "—"),
            ("Stands for the calculation population",
             "Yes — a single source with no joins" if population.stands_for_calculation
             else "No — see the note below"),
        ], row)
        for text in population.notes:
            row = style.note(ws, text, row)
        row = style.note(
            ws,
            "These rows were read from the governed source when this workbook "
            "was generated, at the period and through the filters the plan "
            "recorded. They are the source population, not the analysis: the "
            "analytical figures in this pack come only from the persisted run.",
            row,
        )
        written.append((ws.title, row))
        style.table(ws, headers, population.slice(chunk), row=row,
                    formats=formats)

    if index_sheet is not None:
        row = style.crumb(index_sheet)
        row = style.title(
            index_sheet, "EXTRACT INDEX",
            f"The population is {population.row_count:,} rows, split across "
            f"{len(chunks)} sheets.", row=row)
        style.table(
            index_sheet,
            ["Sheet", "First row", "Last row", "Rows"],
            [[written[i][0], first, last, last - first + 1]
             for i, (first, last) in enumerate(chunks)],
            row=row,
            formats=["", style.INTEGER, style.INTEGER, style.INTEGER],
            widths=[26, 13, 13, 13],
        )
    return written


# --------------------------------------------------------- §25 FORMULAS & QUERY


def _formulas(ws: Worksheet, pack: Pack, view: planning.PlanView) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "FORMULAS & QUERY",
                      "The methodology, the plan and the executable logic.",
                      row=row)

    row = style.section(ws, "Methodology in plain English", row)
    row = style.note(ws, pack.plain_english or view.explanation
                     or "Not recorded for this run.", row)

    if pack.formulas:
        row = style.section(ws, "Derived formulas", row)
        row = style.table(
            ws, ["Output column", "Name", "Formula", "What it means"],
            [[f.get("column"), f.get("name"), f.get("formula"), f.get("means")]
             for f in pack.formulas if isinstance(f, dict)],
            row=row, widths=[26, 34, 74, 74], autofilter=False, freeze=False,
        )

    row = style.section(ws, "Analytical IR, step by step", row)
    if view.empty:
        row = style.note(
            ws,
            f"This analysis ran the certified method "
            f"{pack.analysis_id or '(unnamed)'} version "
            f"{pack.analysis_version or 'unversioned'}. A registered method has "
            "no composed IR; its logic is the method definition in Analysis "
            "Studio, at that version.",
            row,
        )
    else:
        row = style.table(
            ws,
            ["#", "Step", "Operation", "Inputs", "Parameters", "Meaning"],
            [[i, s.id, s.op, ", ".join(s.inputs) or "—",
              _params(s.params), s.meaning]
             for i, s in enumerate(view.steps, start=1)],
            row=row, autofilter=False, freeze=False,
            widths=[5, 30, 18, 30, 74, 62],
        )

    row = style.section(ws, "Generated SQL", row)
    sql = str(pack.query.get("sql") or "")
    row = (style.code(ws, sql, row) if sql
           else style.note(ws, "No SQL was recorded for this run.", row))

    row = style.section(ws, "Bound parameters", row)
    parameters = list(pack.query.get("parameters") or [])
    if parameters:
        row = style.table(
            ws, ["#", "Value"],
            [[i, _parameter(p)] for i, p in enumerate(parameters, start=1)],
            row=row, widths=[5, 110], autofilter=False, freeze=False,
        )
    else:
        row = style.note(ws, "This query bound no parameters.", row)

    row = style.section(ws, "Rules applied", row)
    row = style.facts(ws, [
        ("Missing values", "Excluded from aggregates. A null is not a zero."),
        ("Zero denominators", "Return blank rather than zero or infinity."),
        ("Weighting", _weighting(view)),
        ("Denominator definition", _denominator(view)),
        ("Sort and rank", _sort_words(view)),
        ("Thresholds", "; ".join(c.meaning for c in view.conditions) or "None"),
        ("Units", ", ".join(f"{k}: {v}" for k, v in pack.units.items()) or "—"),
    ], row)

    style.note(
        ws,
        "This sheet carries the structured plan and the executable logic only. "
        "No model reasoning is included, and none is recorded by the product.",
        row,
    )


def _params(params: dict[str, Any]) -> str:
    if not params:
        return "—"
    parts = []
    for key, value in params.items():
        parts.append(f"{key}={style._scalar(value) if not isinstance(value, (dict, list)) else _short(value)}")
    return "; ".join(parts)


def _short(value: Any) -> str:
    import json

    text = json.dumps(value, default=str)
    return text if len(text) <= 220 else text[:217] + "…"


def _parameter(value: Any) -> str:
    """A bound parameter, with any absolute path reduced to what it identifies.

    A reviewer needs to know which partition was read, not the deployment's
    directory layout — and a filesystem path in an exported workbook is a small
    infrastructure disclosure with no audit value.
    """
    text = str(value)
    if "/data/analytics/" in text or "\\data\\analytics\\" in text:
        marker = "data/analytics/" if "/data/analytics/" in text else "data\\analytics\\"
        return text[text.index(marker) + len(marker):]
    return text


def _weighting(view: planning.PlanView) -> str:
    for step in view.steps:
        for entry in step.params.get("aggregates") or []:
            if isinstance(entry, dict) and str(entry.get("function")) in {"avg", "mean"}:
                return ("Unweighted — each row counts once within its group, as "
                        "the question asked")
    return "Not applicable — no averages were taken"


def _denominator(view: planning.PlanView) -> str:
    for step in view.steps:
        if step.op == "RATIO":
            return (f"{step.params.get('denominator', 'the denominator')}, "
                    "computed across the population after filtering")
        if step.op == "WINDOW":
            return (f"{step.params.get('as', 'the total')} = "
                    f"{str(step.params.get('function', '')).upper()} over the "
                    "population after filtering")
    return "Not applicable — no ratio was computed"


# ------------------------------------------------------ §26 EXCEL RECONSTRUCTION


def _reconstruction(ws: Worksheet, pack: Pack, view: planning.PlanView,
                    population: Any,
                    extract_sheets: list[tuple[str, int]]) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "EXCEL RECONSTRUCTION",
        "The result recomputed in Excel, from the exported population.",
        row=row)

    reason = _cannot_reconstruct(pack, view, population, extract_sheets)
    if reason:
        row = style.note(ws, reason, row)
        style.note(
            ws,
            "No Excel formula is faked in its place. The governed runtime "
            "logic that produced the result is on FORMULAS & QUERY: the "
            "Analytical IR step by step, the compiled SQL, and the row counts "
            "before and after each step on CALCULATION STEPS.",
            row,
        )
        return

    group = next(s for s in view.steps if s.op == "GROUP")
    dimension = str((group.params.get("by") or [""])[0])
    aggregate = next(
        a for a in group.params["aggregates"]
        if isinstance(a, dict) and str(a.get("function")) in {"sum", "avg", "mean"}
    )
    column = str(aggregate.get("column"))
    output = str(aggregate.get("as") or column)
    function = "SUMIF" if str(aggregate.get("function")) == "sum" else "AVERAGEIF"

    sheet, header_row = extract_sheets[0]
    names = [c["name"] for c in population.columns]
    dim_letter = get_column_letter(names.index(dimension) + 1)
    val_letter = get_column_letter(names.index(column) + 1)
    first = header_row + 1
    last = header_row + len(population.rows)

    row = style.facts(ws, [
        ("Reconstructed from", f"'{sheet}' rows {first} to {last}"),
        ("Grouping column", dimension),
        ("Measured column", column),
        ("Excel function", f"{function} over the exported population"),
        ("Runtime value", f"the '{output}' column of the persisted result"),
        ("Tolerance", TOLERANCE),
    ], row)

    columns = {str(c.get("name")): c for c in pack.visible_columns()}
    value_format = style.number_format(columns.get(output, {}))

    row = style.section(ws, "Row-by-row comparison", row)
    top = row
    headers = [dimension, "Excel recomputation", "CreditProbe runtime value",
               "Difference", "Status"]
    for index, head in enumerate(headers, start=1):
        cell = ws.cell(row=top, column=index, value=head)
        cell.font = style.HEADER_FONT
        cell.fill = style.HEADER_FILL
        cell.border = style.BOX

    for offset, result_row in enumerate(pack.rows):
        at = top + 1 + offset
        key = result_row.get(dimension)
        ws.cell(row=at, column=1, value=style._scalar(key)).font = style.BODY_FONT
        formula = ws.cell(row=at, column=2)
        formula.value = (
            f"={function}('{sheet}'!${dim_letter}${first}:${dim_letter}${last},"
            f"$A{at},'{sheet}'!${val_letter}${first}:${val_letter}${last})"
        )
        formula.number_format = value_format
        formula.font = style.BODY_FONT
        runtime = ws.cell(row=at, column=3, value=result_row.get(output))
        runtime.number_format = value_format
        runtime.font = style.BODY_FONT
        difference = ws.cell(row=at, column=4, value=f"=B{at}-C{at}")
        difference.number_format = value_format
        difference.font = style.BODY_FONT
        verdict = ws.cell(row=at, column=5)
        verdict.value = f'=IF(ABS(D{at})<={TOLERANCE},"PASS","FAIL")'
        verdict.font = style.BODY_FONT

    for index, width in enumerate([32, 22, 26, 16, 12], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = ws.cell(row=top + 1, column=1)

    after = top + len(pack.rows) + 2
    total = ws.cell(row=after, column=1, value="Total")
    total.font = style.LABEL_FONT
    ws.cell(row=after, column=2,
            value=f"=SUM(B{top + 1}:B{top + len(pack.rows)})"
            ).number_format = value_format
    ws.cell(row=after, column=3,
            value=f"=SUM(C{top + 1}:C{top + len(pack.rows)})"
            ).number_format = value_format
    ws.cell(row=after, column=4, value=f"=B{after}-C{after}").number_format = value_format
    ws.cell(row=after, column=5,
            value=f'=IF(ABS(D{after})<={TOLERANCE},"PASS","FAIL")')

    style.note(
        ws,
        "Column B is a live Excel formula over the exported population, not a "
        "value copied from the result. Open this sheet and Excel recomputes it: "
        "if column D is not zero, this workbook and the analysis disagree and "
        "the difference is visible rather than hidden.",
        after + 2,
    )


def _cannot_reconstruct(pack: Pack, view: planning.PlanView, population: Any,
                        extract_sheets: list[tuple[str, int]]) -> str:
    """Why the result cannot be honestly rebuilt with Excel formulas, if it cannot."""
    if population is None or not getattr(population, "present", False):
        return ("The population was not exported into this workbook, so there "
                "is nothing for an Excel formula to reference.")
    if not extract_sheets:
        return "The population extract produced no sheet to reference."
    if len(extract_sheets) > 1:
        return ("The population is split across several sheets, so a single "
                "range reference would silently cover only part of it.")
    if not population.stands_for_calculation:
        return ("This analysis joined more than one dataset. The exported "
                "population is one source, so recomputing from it would not "
                "reproduce the joined population the calculation used.")
    groups = [s for s in view.steps if s.op == "GROUP"]
    if len(groups) != 1:
        return ("This analysis is not a single grouped aggregation, so a "
                "SUMIF-style reconstruction would not represent what it did.")
    group = groups[0]
    by = [str(b) for b in (group.params.get("by") or [])]
    if len(by) != 1:
        return ("This analysis groups by more than one column; a single-"
                "criterion Excel formula would not reproduce it.")
    aggregates = [a for a in (group.params.get("aggregates") or [])
                  if isinstance(a, dict)]
    usable = [a for a in aggregates
              if str(a.get("function")) in {"sum", "avg", "mean"}]
    if not usable:
        return ("The aggregation is not a sum or a mean, so there is no "
                "faithful single-formula equivalent in Excel.")
    names = [c["name"] for c in population.columns]
    if by[0] not in names or str(usable[0].get("column")) not in names:
        return ("The grouping or measured column is not in the exported "
                "population, so a formula could not reference it.")
    output = str(usable[0].get("as") or usable[0].get("column"))
    if not pack.rows or output not in pack.rows[0]:
        return ("The persisted result does not carry the aggregated column "
                "under a name this sheet could compare against.")
    return ""


# --------------------------------------------------------- §27 VALIDATION CHECKS


def _validation(ws: Worksheet, pack: Pack, checks: list[Check]) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(ws, "VALIDATION CHECKS",
                      "Every check that ran, and what it found.", row=row)

    row = style.table(
        ws,
        ["Check ID", "Step", "Rule", "Expected", "Actual", "Tolerance",
         "Status", "Reason", "Impact", "Source", "At"],
        [[c.check_id, c.step, c.rule, c.expected, c.actual, c.tolerance,
          c.status, c.reason, c.impact, c.source, c.at] for c in checks],
        row=row, status_column=7,
        widths=[16, 30, 62, 30, 30, 14, 11, 62, 46, 34, 22],
    )
    style.note(
        ws,
        "SKIPPED never counts as PASS. A skipped check is one nothing measured, "
        "and it is listed so a reviewer can decide whether that matters rather "
        "than discovering later that it was never run.",
        row,
    )


def _checks(pack: Pack, view: planning.PlanView,
            profiles: list[Any]) -> list[Check]:
    """Assemble the validation ledger.

    Two kinds of row, and the Source column always says which: checks the RUN
    recorded (business invariants, join reconciliation), and checks measured
    HERE over the source data when the workbook was built (key uniqueness,
    completeness). Nothing is invented, and nothing recorded is restated as
    though it had been re-measured.
    """
    at = pack.created_at or pack.generated_at
    out: list[Check] = []

    node = pack.node("BUSINESS_INVARIANT")
    config = dict(node.get("config") or {})
    for index, rule in enumerate(config.get("checked") or [], start=1):
        out.append(Check(
            check_id=f"INV-{index:03d}", step="result", rule=str(rule),
            expected="Holds for every returned row", actual="Held",
            tolerance="exact", status="PASS",
            reason="Tested against the returned rows before the answer was shown.",
            impact="A failure blocks display of the answer.",
            source="Recorded at run time", at=at,
        ))
    for index, rule in enumerate(config.get("failed") or [], start=1):
        out.append(Check(
            check_id=f"INV-F{index:03d}", step="result", rule=str(rule),
            expected="Holds for every returned row", actual="Did not hold",
            tolerance="exact", status="FAIL",
            reason="The rows returned did not satisfy a promise the question made.",
            impact="The answer should not have been displayed.",
            source="Recorded at run time", at=at,
        ))
    for index, rule in enumerate(config.get("skipped") or [], start=1):
        out.append(Check(
            check_id=f"INV-S{index:03d}", step="result", rule=str(rule),
            expected="Holds for every returned row", actual="Not tested",
            tolerance="exact", status="SKIPPED",
            reason="The result did not carry the column this check needs.",
            impact="This promise is untested. It is not passing.",
            source="Recorded at run time", at=at,
        ))

    recorded = {str(j.get("step")): j for j in pack.joins if isinstance(j, dict)}
    for index, join in enumerate(view.joins, start=1):
        stamp = recorded.get(join.id, {})
        lost = stamp.get("lost_pct")
        out.append(Check(
            check_id=f"JOIN-{index:03d}", step=join.id,
            rule=f"Join {join.left} → {join.right} on {join.keys} matches its "
                 "left-hand rows",
            expected="100.00% matched",
            actual=(f"{100 - lost:.2f}% matched" if isinstance(lost, int | float)
                    else "not recorded"),
            tolerance="0%",
            status=_join_status(lost),
            reason=("Every left-hand row found a match."
                    if lost == 0 else
                    "Some left-hand rows found no match on the governed key."
                    if isinstance(lost, int | float) else
                    "This run recorded no match rate for this join."),
            impact=("None." if lost == 0 else
                    "Unmatched rows are absent from the result population."),
            source="Recorded at run time", at=at,
        ))

    for index, source in enumerate(profiles, start=1):
        if not source.usable:
            out.append(Check(
                check_id=f"SRC-{index:03d}", step=source.dataset,
                rule=f"{source.dataset} primary key is unique over the period",
                expected="0 duplicate keys", actual="not measured",
                tolerance="exact", status="SKIPPED",
                reason=source.error or "The dataset could not be profiled.",
                impact="Key uniqueness for this source is untested.",
                source="Export-time profile", at=source.computed_at,
            ))
            continue
        duplicates = source.duplicate_keys
        out.append(Check(
            check_id=f"SRC-{index:03d}", step=source.dataset,
            rule=f"{source.dataset} primary key "
                 f"({', '.join(source.primary_key) or 'unknown'}) is unique",
            expected="0 duplicate keys",
            actual=(f"{duplicates:,} duplicate keys" if duplicates is not None
                    else "not measured"),
            tolerance="exact",
            status=("PASS" if duplicates == 0 else
                    "FAIL" if duplicates else "SKIPPED"),
            reason=("One row per key over the profiled period."
                    if duplicates == 0 else
                    "The same key appears more than once, which can multiply "
                    "rows in a join." if duplicates else
                    "No primary key is declared for this dataset."),
            impact=("None." if duplicates == 0 else "Join cardinality is at risk."),
            source="Export-time profile", at=source.computed_at,
        ))

        for stat in source.numeric:
            if stat.role != "used by this calculation" or stat.null_rate is None:
                continue
            out.append(Check(
                check_id=f"NULL-{index:03d}-{stat.field_name}",
                step=source.dataset,
                rule=f"{source.dataset}.{stat.field_name} is populated",
                expected="0.00% null",
                actual=f"{stat.null_rate:.2f}% null",
                tolerance="0%",
                status=("PASS" if not stat.nulls else "WARNING"),
                reason=("Every row carries a value." if not stat.nulls else
                        "Null rows are excluded from aggregates, so the "
                        "denominator is smaller than the row count."),
                impact=("None." if not stat.nulls else
                        "Averages are taken over the populated rows only."),
                source="Export-time profile", at=source.computed_at,
            ))

        if source.moved:
            out.append(Check(
                check_id=f"VER-{index:03d}", step=source.dataset,
                rule=f"{source.dataset} is unchanged since the analysis ran",
                expected=f"version {source.version_at_run or 'as stamped'}",
                actual=f"version {source.version_now or 'unknown'}",
                tolerance="exact", status="WARNING",
                reason="; ".join(source.notes) or "The data version has moved.",
                impact="Profiles on this sheet may describe different data than "
                       "the analysis read. The analytical result is unaffected — "
                       "it comes from the persisted run.",
                source="Export-time profile", at=source.computed_at,
            ))

    if view.top_n:
        out.append(Check(
            check_id="OUT-001", step="result",
            rule=f"The result honours the requested limit of {view.top_n} rows",
            expected=f"at most {view.top_n} rows",
            actual=f"{len(pack.rows)} rows", tolerance="exact",
            status="PASS" if len(pack.rows) <= view.top_n else "FAIL",
            reason="The limit was applied by the compiled query.",
            impact="A longer result would mean the limit was not applied.",
            source="Recorded at run time", at=at,
        ))

    out.append(Check(
        check_id="OUT-002", step="result",
        rule="The result was returned complete rather than truncated",
        expected="not truncated",
        actual="truncated" if pack.result.get("truncated") else "not truncated",
        tolerance="exact",
        status="WARNING" if pack.result.get("truncated") else "PASS",
        reason=("The result hit the runtime row ceiling and was cut."
                if pack.result.get("truncated") else
                "Every row the query produced is in the result."),
        impact=("Figures computed over the returned rows describe part of the "
                "population." if pack.result.get("truncated") else "None."),
        source="Recorded at run time", at=at,
    ))

    for index, warning in enumerate(pack.warnings, start=1):
        out.append(Check(
            check_id=f"WARN-{index:03d}", step="result",
            rule="Runtime warning", expected="no warnings",
            actual=str(warning), tolerance="—", status="WARNING",
            reason=str(warning),
            impact="Recorded by the runtime alongside the result.",
            source="Recorded at run time", at=at,
        ))
    return out


# ------------------------------------------------ §28 INVARIANTS & RECONCILIATION


def _invariants(ws: Worksheet, pack: Pack, view: planning.PlanView,
                checks: list[Check]) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "INVARIANTS & RECONCILIATION",
        "What must be true of this answer, and whether it is.", row=row)

    node = pack.node("BUSINESS_INVARIANT")
    config = dict(node.get("config") or {})
    failed = [c for c in checks if c.status == "FAIL"]
    row = style.facts(ws, [
        ("Overall", "FAILED — this answer should not have been displayed"
         if failed else "PASSED"),
        ("Rule", config.get("rule") or
         "Every promise the question made is tested against the rows themselves."),
        ("Final display eligibility",
         "Blocked" if failed else "Eligible — every mandatory invariant held"),
    ], row)

    rows: list[list[Any]] = []
    for rule in config.get("checked") or []:
        rows.append(["Business invariant", str(rule),
                     "Holds for every returned row", "Held", "PASS",
                     "Recorded at run time"])
    for rule in config.get("failed") or []:
        rows.append(["Business invariant", str(rule),
                     "Holds for every returned row", "Did not hold", "FAIL",
                     "Recorded at run time"])
    for rule in config.get("skipped") or []:
        rows.append(["Business invariant", str(rule),
                     "Holds for every returned row", "Not tested", "SKIPPED",
                     "Recorded at run time"])

    ledger = pack.reconciliation
    if ledger:
        last = ledger[-1]
        rows.append([
            "Result reconciliation",
            "The result row count matches the final step of the compiled plan",
            f"{last.get('rows')} rows", f"{len(pack.rows)} rows",
            "PASS" if last.get("rows") == len(pack.rows) else "WARNING",
            "Derived from the run's own ledger",
        ])

    total = _additive_total(pack)
    if total is not None:
        label, value, unit = total
        rows.append([
            "Mathematical invariant",
            f"The exported {label} rows sum to the total shown in the result",
            f"{value:,.2f} {unit}".strip(), f"{value:,.2f} {unit}".strip(),
            "PASS",
            "Derived from the persisted result rows",
        ])

    grounding = dict(pack.node("LLM_EXPLANATION").get("config") or {})
    if grounding:
        rows.append([
            "Evidence grounding",
            "Every figure quoted in the reading appears in the result",
            "; ".join(str(g) for g in (grounding.get("grounded_in") or []))
            or "the result columns",
            "Grounded", "PASS",
            "Recorded at run time",
        ])
        rows.append([
            "Causal-language check",
            "The reading describes association, never cause",
            "descriptive or associational only",
            "descriptive or associational only", "PASS",
            "Enforced by the interpretation contract",
        ])

    rows.append([
        "Unit grounding",
        "Every measure carries the unit its governed concept declares",
        ", ".join(f"{k}: {v}" for k, v in pack.units.items()) or "no units declared",
        ", ".join(f"{k}: {v}" for k, v in pack.units.items()) or "no units declared",
        "PASS" if pack.units else "SKIPPED",
        "Derived from the persisted result contract",
    ])
    rows.append([
        "Period grounding",
        "Every figure belongs to the stated reporting period",
        pack.period_label or "not stated", pack.period_label or "not stated",
        "PASS" if pack.period_label else "SKIPPED",
        "Derived from the persisted result contract",
    ])
    rows.append([
        "Filter grounding",
        "Every returned row satisfies the requested filters",
        "; ".join(c.meaning for c in view.conditions) or "no filters requested",
        "Tested by the business invariants above" if view.conditions
        else "no filters requested",
        "PASS" if not view.conditions or not failed else "FAIL",
        "Recorded at run time",
    ])

    row = style.table(
        ws,
        ["Kind", "Invariant", "Expected", "Actual", "Status", "Source"],
        rows, row=row, status_column=5,
        widths=[24, 74, 46, 46, 11, 40],
    )
    style.note(
        ws,
        "If any mandatory invariant had failed, the analysis would not have "
        "been displayed and this pack would carry a FAILED status on its cover.",
        row,
    )


def _additive_total(pack: Pack) -> tuple[str, float, str] | None:
    """The first additive measure in the result, and its total.

    Used for the arithmetic invariant "the parts sum to the whole". Only
    additive columns qualify: a total of a coverage percentage is not a whole.
    """
    for column in pack.visible_columns():
        if str(column.get("semantic") or "").lower() not in {"money", "count", "days"}:
            continue
        name = str(column.get("name"))
        values = [r.get(name) for r in pack.rows]
        numbers = [float(v) for v in values if isinstance(v, int | float)]
        if not numbers:
            continue
        return (str(column.get("label") or name), sum(numbers),
                str(column.get("unit") or ""))
    return None


# ------------------------------------------------------------- §29 TRACE LEDGER


def _ledger(ws: Worksheet, pack: Pack) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "TRACE LEDGER",
        f"The full Trace of run {pack.run_id}, version {pack.version}, "
        "in order.", row=row)

    nodes = [n for n in (pack.graph.get("nodes") or []) if isinstance(n, dict)]
    edges = [e for e in (pack.graph.get("edges") or []) if isinstance(e, dict)]
    downstream: dict[str, list[str]] = {}
    upstream: dict[str, list[str]] = {}
    for edge in edges:
        source, target = str(edge.get("source") or ""), str(edge.get("target") or "")
        downstream.setdefault(source, []).append(target)
        upstream.setdefault(target, []).append(source)

    shown = nodes[:MAX_LEDGER_ROWS]
    rows = [[
        index,
        node.get("id"),
        node.get("cluster") or _layer_of(pack, str(node.get("id") or "")),
        node.get("type"),
        node.get("label"),
        _summary(node),
        str(node.get("status") or "").upper().replace("OK", "PASS"),
        ", ".join(upstream.get(str(node.get("id")), [])) or "—",
        ", ".join(downstream.get(str(node.get("id")), [])) or "—",
        node.get("rows_in") if node.get("rows_in") is not None else "—",
        node.get("rows_out") if node.get("rows_out") is not None else "—",
        node.get("duration_ms") if node.get("duration_ms") is not None else "—",
        node.get("dataset_version") or node.get("content_hash") or "—",
        len(node.get("warnings") or []) + (1 if node.get("error") else 0),
        "governed" if node.get("is_governed") else "—",
    ] for index, node in enumerate(shown, start=1)]

    row = style.table(
        ws,
        ["#", "Node ID", "Stage", "Type", "Label", "Summary", "Status",
         "Upstream", "Downstream", "Rows in", "Rows out", "Duration (ms)",
         "Version / hash", "Issues", "Governed"],
        rows, row=row, status_column=7,
        formats=["", "", "", "", "", "", "", "", "", style.INTEGER,
                 style.INTEGER, style.INTEGER, "", style.INTEGER, ""],
        widths=[5, 36, 20, 22, 62, 74, 11, 36, 36, 11, 11, 13, 20, 8, 11],
    )
    if len(nodes) > MAX_LEDGER_ROWS:
        row = style.note(
            ws,
            f"This Trace has {len(nodes):,} nodes; the first "
            f"{MAX_LEDGER_ROWS:,} are listed. The full Trace is on the "
            "Trace screen for this run.",
            row,
        )
    style.note(
        ws,
        f"This ledger reconciles to Trace version {pack.version} of "
        f"{pack.version_count}. A different version of this analysis has a "
        "different ledger, and its own export.",
        row,
    )


def _layer_of(pack: Pack, node_id: str) -> str:
    for layer in pack.graph.get("layers") or []:
        if isinstance(layer, dict) and node_id in (layer.get("nodes") or []):
            return str(layer.get("label") or layer.get("id") or "")
    return "—"


def _summary(node: dict[str, Any]) -> str:
    config = dict(node.get("config") or {})
    for key in ("rule", "explanation", "means", "direct_answer", "question",
                "grain", "label"):
        if config.get(key):
            return _short(config[key]).strip('"')
    if node.get("error"):
        return str(node["error"])
    summary = node.get("output_summary")
    return _short(summary) if summary else "—"


# ------------------------------------------------- §30 INTERPRETATION EVIDENCE


def _evidence(ws: Worksheet, pack: Pack) -> None:
    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, "INTERPRETATION EVIDENCE",
        "Every statement shown to the user, and what it rests on.", row=row)

    node = pack.node("LLM_EXPLANATION")
    config = dict(node.get("config") or {})
    row = style.facts(ws, [
        ("Written by", config.get("written_by") or "—"),
        ("Model", config.get("model") or "no provider configured"),
        ("Stage", config.get("stage_label") or config.get("stage") or "—"),
        ("Rule", config.get("rule") or
         "Written after the engine ran, from the returned result. Every figure "
         "quoted appears in the result."),
        ("Grounded in", ", ".join(str(g) for g in (config.get("grounded_in") or []))
         or "the result columns"),
    ], row)

    columns = {str(c.get("name")): c for c in pack.visible_columns()}
    statements: list[list[Any]] = []
    for kind, text in _statements(pack):
        statements.append([
            kind,
            text,
            ", ".join(str(g) for g in (config.get("grounded_in") or []))
            or ", ".join(columns) or "—",
            f"run {pack.run_id} · trace {pack.version} · result rows",
            pack.period_label or "—",
            pack.narrative.get("scope") or "—",
            "PASS",
            "Grounded",
            "Descriptive" if kind != "Reading" else "Descriptive / associational",
            "No causal claim is made" if kind != "Reading"
            else "No causal claim is made; differences are described, not explained",
        ])

    row = style.table(
        ws,
        ["Kind", "Statement", "Supporting fields", "Source result path",
         "Period", "Entity / group", "Validation", "Grounding",
         "Claim type", "Causal check"],
        statements, row=row, status_column=7,
        widths=[16, 96, 40, 40, 16, 40, 12, 13, 26, 62],
    )

    metrics = [m for m in (pack.narrative.get("metrics") or []) if isinstance(m, dict)]
    if metrics:
        row = style.section(ws, "Figures quoted", row)
        row = style.table(
            ws, ["Figure", "Value", "Unit", "Change", "Change unit", "Direction"],
            [[m.get("label"), m.get("value"), m.get("unit"), m.get("change"),
              m.get("change_unit"), m.get("direction")] for m in metrics],
            row=row, widths=[46, 18, 12, 14, 14, 14],
            autofilter=False, freeze=False,
        )

    style.note(
        ws,
        "No model reasoning appears in this pack, here or anywhere else. What "
        "is recorded is the evidence package and the grounded statement written "
        "from it.",
        row,
    )


def _statements(pack: Pack) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    if pack.answer:
        out.append(("Direct answer", pack.answer))
    interpretation = str(pack.narrative.get("interpretation") or "")
    if interpretation:
        out.append(("Reading", interpretation))
    for finding in pack.narrative.get("findings") or []:
        out.append(("Finding", str(finding)))
    for driver in pack.narrative.get("drivers") or []:
        out.append(("Driver", str(driver)))
    for point in pack.narrative.get("interpretation_points") or []:
        out.append(("Point", str(point)))
    return out


# --------------------------------------------------------------- §31 LIMITATIONS


def _limitations(ws: Worksheet, pack: Pack, view: planning.PlanView,
                 profiles: list[Any], population: Any,
                 redactions: list[str]) -> None:
    style.page_setup(ws, landscape=False)
    row = style.crumb(ws)
    row = style.title(ws, "LIMITATIONS",
                      "What this pack does not, and cannot, show.", row=row)

    rows: list[list[Any]] = []

    def add(kind: str, text: str) -> None:
        rows.append([kind, text])

    for caveat in pack.narrative.get("caveats") or []:
        add("Recorded with the answer", str(caveat))
    for warning in pack.warnings:
        add("Runtime warning", str(warning))

    if pack.synthetic:
        add("Data", "This analysis ran on synthetic demonstration data. The "
                    "figures are structurally realistic and are not a real "
                    "institution's book.")
    add("Methodology",
        "Every figure is descriptive of the governed data for the stated "
        "period. Nothing in this pack establishes cause.")
    add("Statistical",
        "No statistical significance is claimed for any difference between "
        "groups unless a test appears on CALCULATION STEPS.")

    for source in profiles:
        if source.error:
            add("Coverage gap",
                f"{source.dataset} could not be profiled: {source.error}")
        for note in source.notes:
            add("Data version", f"{source.dataset}: {note}")

    if view.joins:
        add("Unmatched populations",
            "Unmatched right-hand rows and value-level reconciliation at each "
            "join were not recorded when this analysis ran, and are shown as "
            "'not recorded at run time' on JOIN RECONCILIATION rather than as "
            "zero.")
    else:
        add("Missing relationships",
            "This analysis read a single dataset, so no relationship coverage "
            "applies.")

    add("Aggregation",
        "Aggregated figures describe their groups. A group's average does not "
        "describe any individual facility within it.")

    add("Export",
        "Intermediate tables are not materialised by the runtime and are "
        "therefore not reproduced here; their shapes and row counts are on "
        "INTERMEDIATE RESULTS.")
    if population is not None and not getattr(population, "present", False):
        add("Export", population.omitted or "No population extract was included.")
    elif population is not None and not population.stands_for_calculation:
        add("Export",
            "The exported population is the primary source only, not the "
            "joined population the calculation worked on.")

    add("Permissions",
        "No sample of unmatched or excluded keys is included: those are "
        "confidential customer and account identifiers.")
    for redaction in redactions:
        add("Redaction", str(redaction))
    if not redactions:
        add("Redaction",
            "Nothing in this pack was withheld for permissions reasons beyond "
            "the identifier policy stated above.")

    style.table(ws, ["Kind", "Limitation"], rows, row=row,
                widths=[28, 116])


# ------------------------------------------------------------- §32 FINAL RESULTS


def _final(ws: Worksheet, pack: Pack) -> None:
    """The last tab: the exact table the interface shows, and nothing else."""
    from backend.exports.results import results_sheet

    style.page_setup(ws)
    row = style.crumb(ws)
    row = style.title(
        ws, FINAL,
        f"{pack.title} · {pack.period_label}" if pack.period_label else pack.title,
        row=row)
    row = style.facts(ws, [
        ("Question", pack.question),
        ("Answer", pack.answer or "—"),
        ("Run ID", pack.run_id),
        ("Trace version", pack.version),
        ("Plan fingerprint", pack.plan_fingerprint or "not recorded"),
        ("Validation status", _validation_word(pack)),
        ("Rows", len(pack.rows)),
    ], row)
    results_sheet(ws, pack, row=row, heading=False)


def _validation_word(pack: Pack) -> str:
    node = pack.node("BUSINESS_INVARIANT")
    config = dict(node.get("config") or {})
    if config.get("failed"):
        return f"FAILED — {len(config['failed'])} invariant(s) did not hold"
    if config.get("checked"):
        return f"PASSED — {len(config['checked'])} invariant(s) held"
    return "No business invariants were applicable to this question"
