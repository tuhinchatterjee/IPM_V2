"""
The Results Workbook — what the analysis found.

The quick business export: what a credit officer mails to a committee, opens in
a meeting, or keeps beside a paper. Two sheets, occasionally three.

It is NOT the audit pack. It carries no lineage, no SQL, no join
reconciliation, no source profiles. That restraint is the design: a workbook
with twenty tabs is not a quick export, and a reader who has to find the
answer among nineteen technical sheets has not been given one.

The rule that matters
---------------------
**The RESULTS sheet must reconcile exactly to the screen.** Same rows, same
order, same columns in the same order, same units, same values. Not
"equivalent" — identical. A committee paper whose numbers differ from the
product they came from by one row is a paper that gets the product removed,
and the difference is always something small: a hidden column included, a sort
lost, a share re-derived at a different precision.

So this module quotes. It reads the persisted result, applies the column
contract the interface applies, and writes. It computes exactly one thing — a
total — and only where the column contract says the measure is additive.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook as XlWorkbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from backend.exports import style
from backend.exports.contract import (
    RESULTS,
    ROWS_PER_SHEET,
    Workbook,
    filename_for,
    sheet_name,
)
from backend.exports.gather import Pack

logger = logging.getLogger(__name__)

#: Semantic types whose column may honestly be totalled. A percentage, a ratio
#: and an ordinal may not: the sum of fifteen sector shares is 100 and the sum
#: of fifteen rating grades is meaningless, and a spreadsheet that prints
#: either has invented a figure.
ADDITIVE = {"money", "count", "days"}


def build(pack: Pack) -> Workbook:
    """The Results Workbook for one analysis run."""
    book = XlWorkbook()
    taken: set[str] = set()

    results = book.active
    results.title = sheet_name("RESULTS", taken=taken)
    _results_sheet(results, pack)

    summary = book.create_sheet(sheet_name("SUMMARY", taken=taken))
    _summary_sheet(summary, pack)

    chart_note = _chart_sheet(book, pack, taken)

    stream = io.BytesIO()
    book.save(stream)
    return Workbook(
        filename=filename_for(
            RESULTS, analysis=pack.title, period=pack.period_label,
            run_id=pack.run_id, fingerprint=pack.plan_fingerprint,
        ),
        content=stream.getvalue(),
        kind=RESULTS,
        manifest={
            "run_id": pack.run_id,
            "trace_version": pack.version,
            "sheets": book.sheetnames,
            "row_count": len(pack.rows),
            "column_count": len(pack.visible_columns()),
            "datasets": [s.dataset for s in pack.sources],
            "plan_fingerprint": pack.plan_fingerprint,
            "data_version": pack.data_version,
            "build_sha": pack.build_sha,
            "chart": chart_note,
            "redactions": list(pack.redactions),
        },
    )


# ---------------------------------------------------------------- RESULTS


def _results_sheet(ws: Any, pack: Pack) -> None:
    results_sheet(ws, pack)


def results_sheet(ws: Any, pack: Pack, *, row: int = 1,
                  heading: bool = True) -> None:
    """The exact final table, in the exact order the interface shows it.

    Shared with the calculation pack's FINAL RESULTS tab so the two workbooks
    cannot drift: one writer, one column order, one set of number formats. The
    pack passes `heading=False` because it has already written its own title
    and provenance block above the table.
    """
    columns = pack.visible_columns()
    headers = [style.header_label(c) for c in columns]
    formats = [style.number_format(c) for c in columns]
    rows = [[row_.get(c["name"]) for c in columns] for row_ in pack.rows]

    if heading:
        row = style.title(ws, pack.title, _scope_line(pack))
    if not rows:
        # §15: an empty result is a conclusion, not a blank sheet. The reader
        # who opens this must be told that nothing matched rather than left to
        # wonder whether the export failed.
        style.note(ws, pack.answer or "Nothing in the governed data matched "
                                      "what was asked.", row)
        return

    end = style.table(ws, headers, rows[:ROWS_PER_SHEET], row=row, formats=formats)
    _totals(ws, columns, rows, header_row=row, end_row=end - 2)
    style.page_setup(ws)


def _totals(ws: Any, columns: list[dict[str, Any]], rows: list[list[Any]],
            *, header_row: int, end_row: int) -> None:
    """A total row, for the columns where a total means something.

    A real Excel SUM over the rows above rather than a written-in number, so a
    reader who filters the table sees the total follow — and so the figure is
    visibly derived from the cells rather than asserted beside them.
    """
    additive = [
        i for i, c in enumerate(columns, start=1)
        if str(c.get("semantic") or "").lower() in ADDITIVE
        and not c.get("is_identity")
    ]
    if not additive:
        return
    line = end_row + 1
    label = ws.cell(row=line, column=1, value="Total")
    label.font = style.HEADING_FONT
    label.border = style.BOX
    for i in additive:
        letter = get_column_letter(i)
        cell = ws.cell(
            row=line, column=i,
            value=f"=SUBTOTAL(109,{letter}{header_row + 1}:{letter}{end_row})",
        )
        cell.font = style.HEADING_FONT
        cell.border = style.BOX
        cell.number_format = style.number_format(columns[i - 1])


# ---------------------------------------------------------------- SUMMARY


def _summary_sheet(ws: Any, pack: Pack) -> None:
    """What this result is, and what produced it."""
    row = style.title(ws, "Summary", "What this result covers, and how it was produced.")

    row = style.section(ws, "The question", row)
    row = style.facts(ws, [
        ("Question asked", pack.question),
        ("Answer", pack.answer),
        ("Analysis", pack.title),
    ], row)

    row = style.section(ws, "Scope", row)
    row = style.facts(ws, [
        ("Reporting period", pack.period_label or "not recorded"),
        ("Opening period", pack.opening_period or "—"),
        ("Closing period", pack.closing_period or "—"),
        ("Grain", pack.reading.get("grain") or pack.result.get("meta", {}).get("grain") or "—"),
        ("Active filters", _filters(pack) or "none"),
        ("Result rows", len(pack.rows)),
        ("Top N requested", pack.reading.get("top_n") or "—"),
    ], row)

    measures = _measures(pack)
    if measures:
        row = style.section(ws, "Measures", row)
        row = style.table(
            ws, ["Column", "Measure", "Unit", "Definition"], measures,
            row=row, freeze=False, autofilter=False,
        )

    row = style.section(ws, "Provenance", row)
    row = style.facts(ws, [
        ("Analysis run ID", pack.run_id),
        ("Trace version", f"{pack.version} of {pack.version_count}"),
        ("Status", pack.status),
        ("Certification", pack.certification or "—"),
        ("Method", _method(pack)),
        ("Datasets", "; ".join(_dataset_line(s) for s in pack.sources) or "—"),
        ("Source investigation", _origin(pack)),
        ("Plan fingerprint", pack.plan_fingerprint or "—"),
        ("Data version", pack.data_version or "—"),
        ("Build SHA", pack.build_sha or "—"),
        ("Result status", _validation_status(pack)),
        ("Generated", pack.generated_at),
        ("Downloaded by", pack.generated_by),
        ("Data classification", "Synthetic demonstration data"
                                if _is_demo(pack) else "Client data"),
    ], row)

    if pack.warnings:
        row = style.section(ws, "Warnings recorded with this result", row)
        for warning in pack.warnings:
            row = style.note(ws, f"• {warning}", row)

    style.page_setup(ws, landscape=False)


def _measures(pack: Pack) -> list[list[Any]]:
    """What each measured column IS, from the ontology's own labels."""
    out: list[list[Any]] = []
    definitions = {
        str(f.get("column")): str(f.get("formula") or f.get("means") or "")
        for f in pack.formulas if isinstance(f, dict)
    }
    for column in pack.visible_columns():
        if column.get("is_identity") or str(column.get("semantic")) in {"text", "period"}:
            continue
        out.append([
            column.get("name"),
            column.get("label") or column.get("name"),
            column.get("unit") or "—",
            definitions.get(str(column.get("name")), ""),
        ])
    return out


def _scope_line(pack: Pack) -> str:
    """The one line above the table that says what these figures cover.

    A figure computed over five carried names and read as a portfolio total is
    wrong by three orders of magnitude and looks exactly like the right answer.
    Seeing the scope before the number is the only thing that prevents it.
    """
    parts = [
        pack.period_label,
        _filters(pack),
        f"{len(pack.rows)} rows",
    ]
    return " · ".join(p for p in parts if p)


def _filters(pack: Pack) -> str:
    named = pack.reading.get("filters") or pack.filters or {}
    if isinstance(named, dict):
        return ", ".join(f"{k} = {v}" for k, v in named.items() if v not in (None, ""))
    if isinstance(named, list):
        return ", ".join(str(f) for f in named)
    return str(named or "")


def _method(pack: Pack) -> str:
    if pack.is_dynamic:
        return "Composed for this question (no registered method)"
    name = pack.analysis_id or "—"
    return f"{name} v{pack.analysis_version}" if pack.analysis_version else name


def _origin(pack: Pack) -> str:
    parts = []
    if pack.investigation_title:
        parts.append(f"{pack.investigation_title} (#{pack.investigation_id})")
    elif pack.investigation_id:
        parts.append(f"Investigation #{pack.investigation_id}")
    if pack.project_name:
        parts.append(f"Project: {pack.project_name}")
    return " · ".join(parts) or "—"


def _dataset_line(source: Any) -> str:
    bits = [source.dataset]
    if source.period:
        bits.append(source.period)
    if source.version:
        bits.append(f"v{source.version}")
    return " ".join(bits)


def _is_demo(pack: Pack) -> bool:
    return any((s.origin or "").lower() in {"demo", "synthetic"} for s in pack.sources)


def _validation_status(pack: Pack) -> str:
    """Whether the invariants the run recorded passed.

    Read from the trace's own invariant node. A result that failed a mandatory
    invariant should never have reached a screen, so a workbook reporting one
    is reporting a governance failure and says so in those words.
    """
    for node in pack.graph.get("nodes") or []:
        if isinstance(node, dict) and node.get("type") == "BUSINESS_INVARIANT":
            status = str(node.get("status") or "").lower()
            if status in {"failed", "error"}:
                return "FAILED — this result did not satisfy a mandatory invariant"
            if status == "warning":
                return "WARNING — see the calculation pack"
            return "PASS"
    return "Not recorded"


# ----------------------------------------------------------------- CHART


def _chart_sheet(book: XlWorkbook, pack: Pack, taken: set[str]) -> str:
    """A workbook-native chart, only where one would be faithful.

    §7: no decorative charts. The chart is added when the run itself chose a
    chart — the registry decided the shape supported one — and when that shape
    maps onto a form Excel draws honestly. Everything else gets no sheet, which
    is a better outcome than a bar chart of two hundred borrowers.

    The chart references the RESULTS sheet's own cells, so it cannot drift from
    the table: editing a figure moves the chart.
    """
    visual = pack.visual or {}
    chosen = str(visual.get("chart") or "")
    columns = pack.visible_columns()
    if not chosen or chosen == "table" or not pack.rows or len(columns) < 2:
        return "none — the result reads as a table"
    if len(pack.rows) > 60:
        return f"none — {len(pack.rows)} rows is past the point a chart is readable"

    x_name = str(visual.get("x") or "")
    series = [str(y) for y in (visual.get("y") or []) if y]
    names = [c["name"] for c in columns]
    if x_name not in names or not series or any(s not in names for s in series):
        return "none — the recorded chart columns are not in the exported result"

    kind = _excel_chart(chosen)
    if kind is None:
        return f"none — '{chosen}' has no faithful Excel equivalent"

    ws = book.create_sheet(sheet_name("CHART", taken=taken))
    style.title(ws, "Chart", f"{visual.get('label') or chosen} · "
                             f"{visual.get('reason') or 'chosen from the result shape'}")

    chart = kind()
    chart.title = pack.title
    chart.height, chart.width = 11, 24
    x_at = names.index(x_name) + 1
    header_row = 3  # RESULTS: title on 1, scope on 2, header on 3
    last = header_row + len(pack.rows)

    labels = Reference(book["RESULTS"], min_col=x_at, min_row=header_row + 1, max_row=last)
    for name in series:
        at = names.index(name) + 1
        data = Reference(book["RESULTS"], min_col=at, min_row=header_row, max_row=last)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    if hasattr(chart, "type") and chosen in {"bar_horizontal", "horizontal-bar"}:
        chart.type = "bar"
    ws.add_chart(chart, "A5")
    return f"{chosen} over {len(pack.rows)} rows, referencing RESULTS"


def _excel_chart(chosen: str) -> Any:
    """The Excel chart type for a recorded choice, or None.

    None where Excel has no faithful equivalent. A Sankey drawn as a bar chart
    would be a lie about the shape of the data, and no chart is not.
    """
    kinds = {
        "bar": BarChart, "bar_horizontal": BarChart, "horizontal-bar": BarChart,
        "grouped-bar": BarChart, "grouped_bar": BarChart,
        "stacked-bar": BarChart, "stacked_bar": BarChart,
        "line": LineChart, "trend": LineChart, "area": LineChart,
    }
    return kinds.get(chosen)
