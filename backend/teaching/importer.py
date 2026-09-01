"""
Importing a bank's own reviewed Q&A corpus. Part G, §5.

What a client actually has
---------------------------
Not a JSON file. A workbook. Five hundred rows a credit team wrote over two
years, in whatever columns the person who started it chose, with duplicates,
near-duplicates, contradictions, and thirty rows that are notes to
themselves. The import that assumes a clean file is the import that gets used
once.

So this reads XLSX, CSV and JSONL; matches columns by what they are called
rather than by position; validates every row against the teaching contract;
and returns a REPORT — accepted, rejected, duplicate, conflicting — before
anything is written. The report is downloadable as a workbook with the
original row numbers, because the person fixing row 287 needs to find row 287.

Nothing is approved by being imported
--------------------------------------
An imported case arrives NEEDS_REVIEW with `authoring_method = HUMAN`. It is
not retrievable, it is not in a Teaching Release, and it does not become
either by being uploaded. §6's honest count depends on this: a bank that
uploads 600 reviewed answers has 600 cases awaiting review, not 600 approved
ones, until somebody in the bank says otherwise inside CreditProbe.

Duplicates and conflicts are different problems
------------------------------------------------
A DUPLICATE is the same question with the same expected answer — harmless,
and merged rather than imported twice. A CONFLICT is the same question with a
DIFFERENT expected answer, and it is the most valuable row in the file: two
people in the bank disagree, and importing either one silently picks a winner.
Conflicts are never auto-resolved; they go to review with both sides shown.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any

from backend.teaching import families as fam
from backend.teaching import schema as sc
from backend.teaching import status as st

logger = logging.getLogger(__name__)

IMPORT_VERSION = "1.0.0"

XLSX = "XLSX"
CSV = "CSV"
JSONL = "JSONL"
IMPORT_FORMATS: tuple[str, ...] = (XLSX, CSV, JSONL)

#: The largest workbook accepted in one go. §5 asks for 500+ case batches;
#: five thousand is generous headroom and still bounded, because an unbounded
#: import is a way to run the API out of memory.
MAX_ROWS = 5000

# ---------------------------------------------------------------------------
# The template
# ---------------------------------------------------------------------------

#: (column, required, what it is). This IS the downloadable template, so the
#: file a client fills in and the contract the importer enforces cannot drift
#: apart — they are generated from the same tuple.
TEMPLATE: tuple[tuple[str, bool, str], ...] = (
    ("question", True,
     "The question, in the words somebody would actually type."),
    ("expected_answer", True,
     "What a correct answer says. Prose, not a number: the corpus teaches "
     "structure, and a stored figure goes stale with the next quarter."),
    ("family", False,
     "Which teaching family this belongs to. Left blank, CreditProbe "
     "proposes one and a reviewer confirms it."),
    ("objectives", False,
     "What a correct answer must settle, one per line or separated by ';'."),
    ("concepts", False,
     "The governed concepts involved — exposure at default, expected credit "
     "loss — separated by ';'."),
    ("datasets", False,
     "The governed datasets a correct answer reads, separated by ';'."),
    ("period", False, "The reporting period or window the question is about."),
    ("grain", False,
     "What one row of the answer should be: portfolio, segment, customer, "
     "facility."),
    ("expected_outcome", False,
     "EXECUTE, CLARIFY, UNSUPPORTED or FAIL. Blank means EXECUTE."),
    ("difficulty", False,
     "FOUNDATIONAL, INTERMEDIATE, COMPLEX, EXPERT or ADVERSARIAL."),
    ("risk_level", False, "LOW, MEDIUM, HIGH or CRITICAL."),
    ("citation", False,
     "Where the expected answer comes from: a policy, a circular reference, "
     "a committee paper. A case with a citation is worth more in review than "
     "one without."),
    ("forbidden", False,
     "What a wrong-but-plausible answer would say. The single most useful "
     "column in the file: it is what lets a case distinguish a right answer "
     "from a convincing substitute."),
    ("author", False, "Who wrote it, for the reviewer to ask."),
    ("notes", False, "Anything the reviewer should know."),
)

TEMPLATE_COLUMNS: tuple[str, ...] = tuple(c for c, _, _ in TEMPLATE)
REQUIRED_COLUMNS: tuple[str, ...] = tuple(c for c, required, _ in TEMPLATE
                                          if required)

#: Column names a bank's own workbook is likely to use. Matching on these
#: rather than demanding the template be used exactly is the difference
#: between an import that works on the file they have and one that works on
#: the file we asked for.
_ALIASES: dict[str, str] = {
    "q": "question", "prompt": "question", "user question": "question",
    "query": "question", "ask": "question",
    "a": "expected_answer", "answer": "expected_answer",
    "expected": "expected_answer", "correct answer": "expected_answer",
    "model answer": "expected_answer", "response": "expected_answer",
    "category": "family", "topic": "family", "family_id": "family",
    "objective": "objectives", "goal": "objectives",
    "concept": "concepts", "measures": "concepts", "metrics": "concepts",
    "dataset": "datasets", "sources": "datasets", "data": "datasets",
    "reporting period": "period", "as of": "period", "quarter": "period",
    "level": "grain", "granularity": "grain",
    "outcome": "expected_outcome", "action": "expected_outcome",
    "complexity": "difficulty",
    "risk": "risk_level", "severity": "risk_level",
    "source": "citation", "reference": "citation", "policy": "citation",
    "wrong answer": "forbidden", "common error": "forbidden",
    "pitfall": "forbidden", "trap": "forbidden",
    "written by": "author", "owner": "author", "sme": "author",
    "note": "notes", "comment": "notes",
}


def template_rows() -> list[dict[str, str]]:
    """The template as rows, for the downloadable workbook."""
    return [{"column": name,
             "required": "yes" if required else "no",
             "what it is": what}
            for name, required, what in TEMPLATE]


def normalise_header(name: str) -> str:
    """One of the template's column names, or "" if it is not one of them."""
    cleaned = re.sub(r"[^a-z0-9 _]+", " ", str(name or "").lower()).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    if cleaned in TEMPLATE_COLUMNS:
        return cleaned
    underscored = cleaned.replace(" ", "_")
    if underscored in TEMPLATE_COLUMNS:
        return underscored
    return _ALIASES.get(cleaned, "")


# ---------------------------------------------------------------------------
# Reading the file
# ---------------------------------------------------------------------------


class ImportError_(Exception):
    """A file that cannot be read as a corpus at all."""


def _split(value: Any) -> list[str]:
    """A multi-value cell, however the author separated it."""
    text = str(value or "").strip()
    if not text:
        return []
    parts = re.split(r"[;\n|]+", text)
    return [p.strip() for p in parts if p.strip()]


def read_rows(payload: bytes, file_format: str) -> list[dict[str, Any]]:
    """Every row, keyed by normalised column name, in file order."""
    if file_format == JSONL:
        rows = []
        for line in payload.decode("utf-8", errors="replace").splitlines():
            if not line.strip():
                continue
            try:
                found = json.loads(line)
            except json.JSONDecodeError as e:
                raise ImportError_(
                    f"line {len(rows) + 1} is not valid JSON: {e}") from e
            if not isinstance(found, dict):
                raise ImportError_(
                    f"line {len(rows) + 1} is not a JSON object")
            rows.append({normalise_header(k) or k: v
                         for k, v in found.items()})
        return rows

    if file_format == CSV:
        text = payload.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [{normalise_header(k) or (k or ""): v for k, v in row.items()}
                for row in reader]

    if file_format == XLSX:
        try:
            import openpyxl
        except Exception as e:  # noqa: BLE001
            raise ImportError_(
                "no XLSX reader is installed on this deployment; upload the "
                "corpus as CSV or JSONL") from e
        try:
            book = openpyxl.load_workbook(io.BytesIO(payload), data_only=True,
                                          read_only=True)
        except Exception as e:  # noqa: BLE001
            raise ImportError_(f"the workbook could not be opened: {e}") from e
        sheet = book[book.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        book.close()
        if not rows:
            raise ImportError_("the workbook is empty")
        header = [normalise_header(c) or str(c or "") for c in rows[0]]
        out = []
        for raw in rows[1:]:
            if not any(c is not None and str(c).strip() for c in raw):
                continue
            out.append({header[i]: raw[i] if i < len(raw) else ""
                        for i in range(len(header))})
        return out

    raise ImportError_(
        f"{file_format!r} is not an importable format: "
        + ", ".join(IMPORT_FORMATS))


# ---------------------------------------------------------------------------
# Verdicts
# ---------------------------------------------------------------------------

ACCEPTED = "ACCEPTED"
REJECTED = "REJECTED"
DUPLICATE = "DUPLICATE"
CONFLICT = "CONFLICT"

VERDICTS: tuple[str, ...] = (ACCEPTED, REJECTED, DUPLICATE, CONFLICT)

VERDICT_MEANS: dict[str, str] = {
    ACCEPTED: "Valid, and different from everything else in the file. It "
              "will be written for review — not approved.",
    REJECTED: "Something required is missing or unusable. Nothing is written.",
    DUPLICATE: "The same question with the same expected answer as an earlier "
               "row. Merged rather than written twice.",
    CONFLICT: "The same question with a DIFFERENT expected answer. Both sides "
              "go to review; CreditProbe does not pick one.",
}


@dataclass
class Row:
    """One row of the file, and what will happen to it."""

    number: int
    verdict: str
    question: str = ""
    problems: list[str] = field(default_factory=list)
    #: For a duplicate or a conflict: the row it clashes with.
    clashes_with: int = 0
    other_answer: str = ""
    case: Any = None

    def to_dict(self) -> dict[str, Any]:
        return {"row": self.number, "verdict": self.verdict,
                "means": VERDICT_MEANS.get(self.verdict, ""),
                "question": self.question[:300],
                "problems": list(self.problems),
                "clashes_with": self.clashes_with,
                "other_answer": self.other_answer[:300],
                "case_id": getattr(self.case, "case_id", "")}


@dataclass
class Report:
    """What an import would do, before it does any of it."""

    rows: list[Row] = field(default_factory=list)
    file_format: str = ""
    #: Columns in the file that matched nothing in the template. Reported
    #: rather than ignored: a column called "expected_result" that nobody
    #: mapped is 500 expected answers silently dropped.
    unmapped_columns: list[str] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)
    fatal: str = ""

    def counted(self) -> dict[str, int]:
        counts = {v: 0 for v in VERDICTS}
        for row in self.rows:
            counts[row.verdict] = counts.get(row.verdict, 0) + 1
        return counts

    @property
    def importable(self) -> list[Row]:
        return [r for r in self.rows if r.verdict == ACCEPTED]

    def sentence(self) -> str:
        if self.fatal:
            return self.fatal
        counts = self.counted()
        parts = [f"{len(self.rows)} row(s) read",
                 f"{counts[ACCEPTED]} would be written for review",
                 f"{counts[REJECTED]} rejected",
                 f"{counts[DUPLICATE]} duplicate",
                 f"{counts[CONFLICT]} in conflict"]
        line = "; ".join(parts) + "."
        if self.missing_columns:
            line += (" Required column(s) missing: "
                     + ", ".join(self.missing_columns) + ".")
        if self.unmapped_columns:
            line += (" Column(s) nothing was read from: "
                     + ", ".join(self.unmapped_columns) + ".")
        line += (" Nothing here is approved: an imported case arrives "
                 "SME_REVIEW_REQUIRED and is not retrievable until a person "
                 "in this bank approves it.")
        return line

    def to_dict(self) -> dict[str, Any]:
        return {"rows": [r.to_dict() for r in self.rows],
                "counts": self.counted(),
                "format": self.file_format,
                "unmapped_columns": list(self.unmapped_columns),
                "missing_columns": list(self.missing_columns),
                "fatal": self.fatal,
                "explanation": self.sentence(),
                "version": IMPORT_VERSION}


# ---------------------------------------------------------------------------
# Building a case
# ---------------------------------------------------------------------------

def _key(question: str) -> str:
    """What makes two questions the same question.

    Case, punctuation and whitespace removed. Deliberately not a similarity
    score: a threshold nobody can explain silently merges two questions a
    credit officer considers different.
    """
    return re.sub(r"[^a-z0-9 ]+", " ",
                  str(question or "").lower()).strip().replace("  ", " ")


def _answer_key(answer: str) -> str:
    return re.sub(r"\s+", " ", str(answer or "").strip().lower())


def _family_for(row: dict[str, Any], question: str) -> str:
    """The family a row declares, or the closest governed one to it."""
    declared = str(row.get("family") or "").strip().upper().replace(" ", "_")
    if declared in fam.BY_ID:
        return declared
    legacy = fam.from_legacy(declared)
    if legacy:
        return legacy
    lowered = question.lower()
    for family, words in (
        ("AMBIGUITY", ("which", "do you mean", "either")),
        ("ECL_MOVEMENT", ("ecl", "expected credit loss", "impairment")),
        ("RATING_MIGRATION", ("downgrade", "upgrade", "rating")),
        ("DPD_MIGRATION", ("past due", "dpd", "arrears")),
        ("CONCENTRATION", ("concentration", "largest", "top ")),
        ("RISK_APPETITE", ("appetite", "limit", "breach")),
        ("PORTFOLIO_MIX", ("mix", "by sector", "distribution")),
        ("COVENANT_AND_COLLATERAL", ("covenant", "collateral", "headroom")),
        ("EARLY_WARNING", ("early warning", "watchlist", "deteriorat")),
    ):
        if any(w in lowered for w in words) and family in fam.BY_ID:
            return family
    return "SINGLE_DOMAIN_AGGREGATION"


def build_case(row: dict[str, Any], *, number: int,
               batch: str) -> tuple[Any, list[str]]:
    """One imported row as a teaching case, plus everything wrong with it."""
    problems: list[str] = []
    question = str(row.get("question") or "").strip()
    answer = str(row.get("expected_answer") or "").strip()

    if not question:
        problems.append("no question")
    elif len(question) < 8:
        problems.append(f"the question is {len(question)} characters long, "
                        "which is not a question")
    if not answer:
        problems.append("no expected answer")
    elif len(answer) < 12:
        problems.append("the expected answer is too short to teach anything")

    outcome = str(row.get("expected_outcome") or "").strip().upper() \
        or fam.EXECUTE
    if outcome not in fam.OUTCOMES:
        problems.append(f"{outcome!r} is not an outcome: "
                        + ", ".join(fam.OUTCOMES))
        outcome = fam.EXECUTE

    difficulty = str(row.get("difficulty") or "").strip().upper() \
        or sc.INTERMEDIATE
    if difficulty not in sc.DIFFICULTIES:
        problems.append(f"{difficulty!r} is not a difficulty")
        difficulty = sc.INTERMEDIATE

    risk = str(row.get("risk_level") or "").strip().upper() or "MEDIUM"
    if risk not in sc.RISK_LEVELS:
        problems.append(f"{risk!r} is not a risk level")
        risk = "MEDIUM"

    if problems and ("no question" in problems or "no expected answer"
                     in problems):
        return None, problems

    objectives = _split(row.get("objectives")) or [answer[:200]]
    case = sc.TeachingCase(
        case_id=f"imp-{batch}-{number:04d}",
        title=question[:200],
        family_id=_family_for(row, question),
        description=(f"Imported from a client corpus, row {number}. "
                     "Awaiting review."),
        question=question,
        objectives=[sc.Objective(id=f"o{i}", text=text)
                    for i, text in enumerate(objectives)],
        concepts=_split(row.get("concepts")),
        required_datasets=_split(row.get("datasets")),
        grain=str(row.get("grain") or "").strip().lower(),
        expected_outcome=outcome,
        difficulty=difficulty,
        risk_level=risk,
        # HUMAN, and the count depends on it being true. A bank that uploads
        # 600 answers its own people wrote has 600 human-authored cases
        # awaiting review — which is exactly what §6 wants reported, and is
        # not the same thing as 600 approved ones.
        authoring_method=st.HUMAN,
        # SME_REVIEW_REQUIRED, not AUTO_VALIDATED. A case a person wrote and
        # a machine has never checked is exactly the case a person has to
        # sign for, and starting it anywhere else would let 600 uploaded
        # answers drift into a count of approved ones.
        review_status=st.SME_REVIEW_REQUIRED,
        # A client's own Q&A is client material until somebody says
        # otherwise. §47 forbids CLIENT-sensitive content as a teaching case
        # that leaves the tenant, and this is what makes that enforceable
        # rather than hoped for.
        data_sensitivity=st.CLIENT,
        source_provenance=f"import:{batch}:row{number}",
        tags=["imported", batch],
        notes=str(row.get("notes") or "").strip(),
    )
    period = str(row.get("period") or "").strip()
    if period:
        case.period_contract = {"stated": period}
    citation = str(row.get("citation") or "").strip()
    if citation:
        case.scope_contract = {**dict(case.scope_contract or {}),
                               "citation": citation}
    forbidden = _split(row.get("forbidden"))
    if forbidden:
        case.scope_contract = {**dict(case.scope_contract or {}),
                               "forbidden_behaviours": forbidden}
    author = str(row.get("author") or "").strip()
    if author:
        case.created_by = author

    case.result_contract = {**dict(case.result_contract or {}),
                            "expected_answer": answer}

    # A case expected to EXECUTE has to say what a correct answer computes.
    # A client's workbook carries the ANSWER, not the plan — the SME who
    # wrote row 287 knows what the right answer is and has never seen an
    # Analytical IR. So the contract is assembled from what the row does
    # state, and it says where it came from: a reviewer confirming this case
    # is confirming a thin contract, and should be able to see that it is
    # thin rather than discovering it later.
    if case.expected_outcome == fam.EXECUTE and not case.analytical_plan_contract:
        case.analytical_plan_contract = {
            "datasets": list(case.required_datasets),
            "concepts": list(case.concepts),
            "grain": case.grain,
            "expected_answer": answer[:500],
            "source": ("stated by the client corpus, not derived from a "
                       "plan; to be confirmed in review"),
        }

    for problem in sc.validate(case):
        problems.append(str(getattr(problem, "message", problem)))
    return case, problems


# ---------------------------------------------------------------------------
# The whole job
# ---------------------------------------------------------------------------


def preview(payload: bytes, file_format: str, *,
            batch: str) -> Report:
    """What this file would do, without writing anything.

    Always run before an import, and returned to the user as the thing they
    approve. An import whose first visible output is "487 cases created" gives
    nobody the chance to notice that column F was never read.
    """
    report = Report(file_format=file_format)
    try:
        rows = read_rows(payload, file_format)
    except ImportError_ as e:
        report.fatal = str(e)
        return report

    if not rows:
        report.fatal = "the file has no rows"
        return report
    if len(rows) > MAX_ROWS:
        report.fatal = (f"{len(rows)} rows is more than the {MAX_ROWS} this "
                        "import accepts in one batch; split the file")
        return report

    header = set()
    for row in rows:
        header.update(k for k in row if k)
    report.unmapped_columns = sorted(k for k in header
                                     if k not in TEMPLATE_COLUMNS)
    report.missing_columns = [c for c in REQUIRED_COLUMNS if c not in header]

    by_question: dict[str, tuple[int, str]] = {}
    for index, raw in enumerate(rows, start=1):
        question = str(raw.get("question") or "").strip()
        answer = str(raw.get("expected_answer") or "").strip()
        key = _key(question)

        if key and key in by_question:
            first, first_answer = by_question[key]
            same = _answer_key(answer) == _answer_key(first_answer)
            report.rows.append(Row(
                number=index,
                verdict=DUPLICATE if same else CONFLICT,
                question=question, clashes_with=first,
                other_answer=first_answer,
                problems=([] if same else [
                    "the same question already appears in this file with a "
                    "different expected answer; both go to review and neither "
                    "is chosen"])))
            continue

        case, problems = build_case(raw, number=index, batch=batch)
        if case is None or problems:
            report.rows.append(Row(number=index, verdict=REJECTED,
                                   question=question, problems=problems,
                                   case=case))
            continue
        if key:
            by_question[key] = (index, answer)
        report.rows.append(Row(number=index, verdict=ACCEPTED,
                               question=question, case=case))
    return report


def error_workbook(report: Report) -> list[dict[str, Any]]:
    """The rows a person has to fix, with their original row numbers.

    Returned as rows rather than as a file so the caller decides the format;
    the API renders it as a workbook. The row NUMBER is the point: a report
    that lists problems without saying where they are is a report nobody can
    act on.
    """
    return [{"row": r.number, "verdict": r.verdict,
             "question": r.question[:300],
             "problem": "; ".join(r.problems) or VERDICT_MEANS.get(r.verdict,
                                                                   ""),
             "clashes_with_row": r.clashes_with or "",
             "other_expected_answer": r.other_answer[:300]}
            for r in report.rows if r.verdict != ACCEPTED]


__all__ = ["ACCEPTED", "CONFLICT", "CSV", "DUPLICATE", "IMPORT_FORMATS",
           "IMPORT_VERSION", "ImportError_", "JSONL", "MAX_ROWS",
           "REJECTED", "REQUIRED_COLUMNS", "Report", "Row", "TEMPLATE",
           "TEMPLATE_COLUMNS", "VERDICTS", "VERDICT_MEANS", "XLSX",
           "build_case", "error_workbook", "normalise_header", "preview",
           "read_rows", "template_rows"]
