"""
The validation evidence workbook. §83, §55.

Ten sheets, and the reason there are ten rather than one is that this
workbook is what somebody opens when they do not believe a number in the
report. "Where does 0.7104 come from?" has to be answerable by opening a
named sheet and finding the row, not by re-running an analysis.

So the sheets mirror the report's sections and the evidence index names the
sheet for every figure it lists. The index is first, because the workbook is
entered from a question rather than read front to back.

Like the Word writer, this computes nothing. The report content model
decided every value; openpyxl only decides what it looks like in a cell.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scorecard import dashboard as dash
from backend.scorecard import report as report_mod

HEADER_FILL = PatternFill("solid", fgColor="0B2436")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="16232F")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="6C7A8C")
BODY_FONT = Font(name="Calibri", size=10, color="16232F")

#: §83's sheets, in the order a reader meets them.
SHEETS: tuple[str, ...] = (
    "EVIDENCE INDEX", "METRICS", "MONTHLY HISTORY", "VARIABLES", "PSI",
    "CSI", "EQUATION", "WOE BINS", "IMPLEMENTATION", "FINDINGS",
    "REGULATORY MAPPING",
)

#: §13's appendix mapping, and the reason it is a table of section names
#: rather than of rule numbers: naming a specific CBUAE paragraph as the
#: authority for a threshold would be claiming a sourced requirement this
#: workspace does not have. §26 and §89 both forbid that.
REGULATORY_MAPPING: tuple[tuple[str, str], ...] = (
    ("Model identification and classification", "1 Cover and document control"),
    ("Model purpose, scope and usage", "3 Model purpose, scope and usage"),
    ("Governance and independent validation", "4 Governance and independence"),
    ("Data quality and representativeness",
     "5 Development and validation data"),
    ("Default definition", "5 Development and validation data"),
    ("Conceptual soundness and design",
     "6 Model design and conceptual soundness"),
    ("Implementation verification", "7 Implementation verification"),
    ("Discriminatory power", "8.2 Discriminatory power"),
    ("Calibration and accuracy", "8.3 Calibration and accuracy"),
    ("Stability and robustness", "8.4 Stability and robustness"),
    ("Sensitivity analysis", "8.5 Sensitivity and variable diagnostics"),
    ("Ongoing monitoring", "9 Monitoring review"),
    ("Findings, severity and remediation", "10 Findings and severity"),
    ("Model risk assessment", "11 Model risk assessment"),
    ("Validation conclusion", "12 Overall validation conclusion"),
)


def _write(sheet, rows: list[list[Any]], *, title: str, note: str = "",
           columns: list[str] | None = None) -> None:
    """One block: a title, an optional note, a header row and the body."""
    sheet.cell(row=1, column=1, value=title).font = TITLE_FONT
    line = 2
    if note:
        cell = sheet.cell(row=line, column=1, value=note)
        cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        line += 1
    line += 1

    if columns:
        for index, name in enumerate(columns, start=1):
            cell = sheet.cell(row=line, column=index, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        sheet.freeze_panes = sheet.cell(row=line + 1, column=1)
        line += 1

    for row in rows:
        for index, value in enumerate(row, start=1):
            cell = sheet.cell(row=line, column=index, value=value)
            cell.font = BODY_FONT
        line += 1

    widths: dict[int, int] = {}
    for row in ([columns] if columns else []) + rows:
        for index, value in enumerate(row, start=1):
            widths[index] = min(max(widths.get(index, 10),
                                    len(str(value)) + 2), 60)
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


def _tables_from(report: report_mod.Report,
                 numbers: tuple[str, ...]) -> list[tuple[str, Any]]:
    out: list[tuple[str, Any]] = []
    for number in numbers:
        section = report.section(number)
        if section is None:
            continue
        for table in section.tables:
            out.append((f"{number} {section.title}", table))
    return out


def _stack(sheet, blocks: list[tuple[str, Any]], *, title: str,
           empty_note: str) -> None:
    """Several report tables on one sheet, one under another."""
    if not blocks:
        _write(sheet, [], title=title, note=empty_note)
        return
    sheet.cell(row=1, column=1, value=title).font = TITLE_FONT
    line = 3
    for label, table in blocks:
        cell = sheet.cell(row=line, column=1,
                          value=f"{label} — {table.caption}")
        cell.font = Font(name="Calibri", size=10, bold=True, color="16232F")
        line += 1
        for index, name in enumerate(table.columns, start=1):
            cell = sheet.cell(row=line, column=index, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        line += 1
        for row in table.rows:
            for index, value in enumerate(row, start=1):
                sheet.cell(row=line, column=index,
                           value=value).font = BODY_FONT
            line += 1
        if table.note:
            cell = sheet.cell(row=line, column=1, value=table.note)
            cell.font = NOTE_FONT
            line += 1
        line += 2
    for index in range(1, 10):
        sheet.column_dimensions[get_column_letter(index)].width = 26


def monthly_history(scorecard_type: str, model_kind: str,
                    months: list[str]) -> list[list[Any]]:
    """§83's monthly history: one row a month, and its maturity.

    The maturity column is not decoration. Without it a month with a blank
    default rate is indistinguishable from a month with no defaults, and
    those are opposite facts.
    """
    rows: list[list[Any]] = []
    for month in months:
        board = dash.build_dashboard(scorecard_type, model_kind=model_kind,
                                     month=month, curves=False).to_dict()
        summary = board.get("summary") or {}
        discrimination = board.get("discrimination") or {}
        calibration = board.get("calibration") or {}
        stability = board.get("stability") or {}
        matured = board["context"]["outcome_maturity_status"]
        rows.append([
            month, matured, summary.get("population"),
            report_mod.stat(discrimination.get("gini")),
            report_mod.stat(discrimination.get("auc")),
            report_mod.stat(discrimination.get("ks")),
            report_mod.percent(calibration.get("observed_default_rate")),
            report_mod.percent(calibration.get("average_predicted_pd")),
            report_mod.stat((stability.get("score_psi") or {}).get("index")),
        ])
    return rows


def write(report: report_mod.Report, *,
          history_months: list[str] | None = None) -> bytes:
    """The evidence workbook, as .xlsx bytes."""
    book = Workbook()
    book.remove(book.active)

    index = book.create_sheet("EVIDENCE INDEX")
    _write(
        index,
        [[e.section, e.label, e.metric, e.value_text, e.method, e.period,
          e.model_version, e.validation_state, e.workbook_sheet]
         for e in report.evidence],
        title=f"Validation evidence index — {report.report_id}",
        note=(f"{report.title}. {report_mod.SYNTHETIC_NOTICE} "
              f"Content hash {report.content_hash}."),
        columns=["Section", "Figure", "Metric", "Value", "Method", "Period",
                 "Model version", "Validation state", "Sheet"])

    _stack(book.create_sheet("METRICS"),
           _tables_from(report, ("2", "8.1", "8.2", "8.3", "8.9")),
           title="Metrics as reported",
           empty_note="No metric sections were reported for this period.")

    history = book.create_sheet("MONTHLY HISTORY")
    _write(history,
           monthly_history(report.scorecard_type, report.model_kind,
                           history_months or [report.period]),
           title="Monthly history",
           note=("A month marked NOT MATURED has no realised outcome. Its "
                 "default rate is blank because the window has not closed, "
                 "which is not the same as a month with no defaults."),
           columns=["Month", "Outcome maturity", "Population", "Gini", "AUC",
                    "KS", "Observed default rate", "Average predicted PD",
                    "Score PSI"])

    _stack(book.create_sheet("VARIABLES"), _tables_from(report, ("8.5",)),
           title="Variable diagnostics",
           empty_note="No variable diagnostics were reported.")

    psi = book.create_sheet("PSI")
    stability = report.section("8.4")
    psi_table = stability.tables[0] if stability and stability.tables else None
    _write(psi, [list(r) for r in (psi_table.rows if psi_table else [])],
           title="Population stability index",
           note=(psi_table.note if psi_table else ""),
           columns=list(psi_table.columns) if psi_table else ["Metric",
                                                              "Value"])

    csi = book.create_sheet("CSI")
    csi_table = (stability.tables[1]
                 if stability and len(stability.tables) > 1 else None)
    _write(csi, [list(r) for r in (csi_table.rows if csi_table else [])],
           title="Characteristic stability index",
           note=("CSI cut-offs are a scorecard convention, not a regulatory "
                 "threshold."),
           columns=list(csi_table.columns) if csi_table
           else ["Variable", "CSI", "Observations"])

    _stack(book.create_sheet("EQUATION"),
           _tables_from(report, ("6", "13.1")),
           title="Model equation and score mapping",
           empty_note="No equation was reported.")

    _stack(book.create_sheet("WOE BINS"), _tables_from(report, ("13.2",)),
           title="Weight of Evidence bin tables",
           empty_note="No binning specification was reported.")

    _stack(book.create_sheet("IMPLEMENTATION"), _tables_from(report, ("7",)),
           title="Implementation replication",
           empty_note="No implementation check was reported.")

    _stack(book.create_sheet("FINDINGS"), _tables_from(report, ("9", "10")),
           title="Findings and monitoring limits",
           empty_note="No findings were raised for this period.")

    _write(book.create_sheet("REGULATORY MAPPING"),
           [[topic, section] for topic, section in REGULATORY_MAPPING],
           title="CBUAE MMS/MMG-aligned section mapping",
           note=("This maps the topics the standard's section list covers to "
                 "the sections of this report that address them. It is a "
                 "mapping of structure, not an assertion that any threshold "
                 "in this report is a regulatory requirement, and CreditProbe "
                 "does not provide regulatory certification."),
           columns=["Topic", "Addressed in"])

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def default_history(scorecard_type: str, limit: int = 12) -> list[str]:
    """The last `limit` months, matured or not.

    Deliberately includes the immature tail: the point of a monthly history
    is to show the trend up to now, and cutting it at the maturity boundary
    would hide the months a validator most wants to look at.
    """
    months = dash.available_months(scorecard_type)
    return months[-limit:] if months else []


__all__ = ["SHEETS", "REGULATORY_MAPPING", "write", "default_history",
           "monthly_history"]
