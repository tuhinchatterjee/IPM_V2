"""Early Warning, as a workbook somebody can take into a meeting. Section 11L.

Two of them, because two different questions get asked:

*   the **borrower scorecard** — one name, four layers, every governed
    condition with its current value, its previous value, the line it is being
    held to and whether it is over it. What an officer takes into a review.
*   the **watchlist** — the ranked book at one reporting date, one row per
    borrower, with the risk level, the priority and the reason behind each.
    What a committee reads before deciding whose name to discuss.

Everything in both is read from the same engine the screen reads. A workbook
that recomputes anything is a workbook that will eventually disagree with the
screen it was downloaded from, and the reader will believe whichever one they
opened last.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook as Book
from openpyxl.utils import get_column_letter

from backend.early_warning import assessment as ea
from backend.early_warning import scorecard as sc
from backend.early_warning import signals as sg
from backend.early_warning import taxonomy as tx
from backend.exports import style as st

WORKBOOK_VERSION = "1.0.0"

XLSX = ("application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet")

#: Every row carries it, on every sheet, because a workbook leaves the product
#: and the disclosure has to leave with it.
ORIGIN = "SYNTHETIC_DEMO"


def _title(sheet: Any, text: str, subtitle: str = "") -> int:
    """The sheet's own heading. Returns the next free row."""
    sheet["A1"] = text
    sheet["A1"].font = st.TITLE_FONT
    row = 2
    if subtitle:
        sheet["A2"] = subtitle
        sheet["A2"].font = st.BODY_FONT
        row = 3
    return row + 1


def _header(sheet: Any, row: int, columns: list[str]) -> int:
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=name)
        cell.font = st.HEADER_FONT
        cell.fill = st.HEADER_FILL
        cell.border = st.BOX
    return row + 1


def _write(sheet: Any, row: int, values: list[Any], *, band: bool = False,
           ) -> int:
    for index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=index, value=value)
        cell.font = st.BODY_FONT
        cell.border = st.BOX
        if band:
            cell.fill = st.BAND_FILL
    return row + 1


def _widths(sheet: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _save(book: Book) -> bytes:
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _value(raw: Any) -> Any:
    """A cell Excel can hold. A NaN in a spreadsheet reads as a broken cell."""
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "Yes" if raw else "No"
    if isinstance(raw, (int, float)):
        return "" if raw != raw else round(float(raw), 4)
    return str(raw)


# ------------------------------------------------------- the borrower pack


def borrower(standing: sg.Standing) -> bytes:
    """One borrower's four-layer scorecard. Sections 11C, 11D and 11L."""
    card = sc.build(standing)
    found = card["assessment"]
    book = Book()

    # ---- ASSESSMENT. The answer before the workings, on the first sheet a
    # reader opens, because a workbook that opens on a component table asks
    # them to derive the conclusion themselves.
    sheet = book.active
    sheet.title = "ASSESSMENT"
    row = _title(sheet, f"Early Warning — {card['borrower_id']}",
                 f"{card['period']} · {ORIGIN} · thresholds owned by "
                 f"{card['owner']}, version {card['taxonomy_version']}")
    row = _write(sheet, row, ["Risk level", found["level"]], band=True)
    row = _write(sheet, row, ["What that means", found["means"]])
    row = _write(sheet, row, ["Primary concern", found["primary_concern"]])
    row = _write(sheet, row, ["What changed", found["why_now"]])
    row = _write(sheet, row, ["Corroborating families",
                              ", ".join(found["corroborating"]) or "None"])
    row += 1
    row = _write(sheet, row, ["Why", ""], band=True)
    for reason in found["reasons"]:
        row = _write(sheet, row, [reason["rule"], reason["says"]])
    if found["mitigating"]:
        row += 1
        row = _write(sheet, row, ["Evidence the other way", ""], band=True)
        for reason in found["mitigating"]:
            row = _write(sheet, row, [reason["rule"], reason["says"]])
    row += 1
    row = _write(sheet, row, ["Detection (TAC)", ""], band=True)
    for kind, count in found["tac"].items():
        row = _write(sheet, row,
                     [f"{tx.TAC_LETTER.get(kind, '')} — {kind.title()}", count])
    _widths(sheet, [28, 110])

    # ---- one sheet per layer, every condition, over the line or inside it
    for layer in card["layers"]:
        page = book.create_sheet(f"LAYER {layer['number']}")
        row = _title(page, f"Layer {layer['number']} — {layer['name']}",
                     layer["sentence"])
        row = _header(page, row, card["columns"] + ["Condition"])
        for component in layer["components"]:
            row = _write(page, row, [
                _value(component["current"]),
                _value(component["previous"]),
                _value(component["movement"]),
                _value(component["threshold"]),
                component["status"],
                component["severity"],
                component["persistence"],
                f"{component['detection_letter']} — "
                f"{component['detection'].title()}",
                component["means"],
                component["label"],
            ], band=component["status"] == sc.OVER)
        row += 1
        row = _write(page, row, ["", "", "", "", "", "", "", "",
                                 layer["matters"], ""])
        if layer["gap"]:
            _write(page, row, ["", "", "", "", "", "", "", "",
                               layer["gap"], ""])
        _widths(page, [14, 14, 12, 14, 18, 12, 14, 22, 90, 34])

    # ---- where the numbers came from
    page = book.create_sheet("SOURCE")
    row = _title(page, "Where every figure came from",
                 "Nothing in this workbook is recomputed. Every value is the "
                 "one the screen read.")
    row = _header(page, row, ["Condition", "Dataset", "Field", "Test",
                             "Threshold", "Unit", "Owner", "Version"])
    for observation in standing.observations:
        row = _write(page, row, [
            observation.label, observation.dataset, observation.field_name,
            observation.test, _value(observation.threshold),
            observation.unit, observation.threshold_owner,
            observation.threshold_version])
    _widths(page, [38, 30, 30, 14, 14, 12, 24, 10])
    return _save(book)


# ------------------------------------------------------------ the watchlist


def watchlist(standings: list[sg.Standing], *, period: str = "",
              limit: int = 500) -> bytes:
    """The ranked book at one reporting date. Sections 11B and 11L."""
    book = Book()
    sheet = book.active
    sheet.title = "WATCHLIST"
    row = _title(sheet, f"Early Warning watchlist — {period}",
                 f"{len(standings)} borrowers assessed · {ORIGIN} · "
                 f"thresholds owned by {tx.THRESHOLD_OWNER}")
    row = _header(sheet, row, [
        "Borrower", "Risk level", "Priority", "Exposure "
        f"({tx.CURRENCY} mn)", "Primary concern", "What changed",
        "Signals firing", "Families", "Worsening", "Resolved", "Why"])
    for standing in standings[:limit]:
        found = ea.assess(standing, standing.record)
        verdict = standing.verdict
        row = _write(sheet, row, [
            standing.borrower_id,
            found.level,
            verdict.label,
            _value(verdict.exposure),
            found.primary_concern,
            found.why_now,
            len(standing.fired),
            len(found.families),
            len(found.worsening),
            len(found.resolved),
            " ".join(found.because()),
        ], band=found.level == ea.HIGH)
    _widths(sheet, [18, 12, 26, 18, 34, 46, 14, 10, 11, 10, 120])

    # ---- how the level was decided, so the column is not a mystery
    page = book.create_sheet("METHODOLOGY")
    described = ea.describe()
    row = _title(page, "How the risk level is decided",
                 described["rule"]["high"])
    row = _header(page, row, ["Part", "Rule"])
    for key in ("gravity", "corroboration", "medium", "low"):
        row = _write(page, row, [key.title(), described["rule"][key]])
    row += 1
    row = _write(page, row, ["Weighed", ""], band=True)
    for entry in described["inputs"]:
        row = _write(page, row, [entry["input"], entry["rule"]])
    row += 1
    row = _write(page, row, ["Deliberately not used", ""], band=True)
    for entry in described["not_used"]:
        row = _write(page, row, [entry["input"], entry["why"]])
    _widths(page, [34, 120])
    return _save(book)


__all__ = ["ORIGIN", "WORKBOOK_VERSION", "XLSX", "borrower", "watchlist"]
