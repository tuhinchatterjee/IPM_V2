"""The project plan as a spreadsheet, in both directions.

A delivery plan arrives as a workbook far more often than it arrives typed
into a screen. So this module has to be good at two things that pull against
each other: producing a workbook a project manager will actually work in, and
reading back one they have worked in — including one they have reordered,
half-filled, or pasted into from somewhere else.

The design follows from that:

**One column contract, used by both directions.** `SHEETS` below is the whole
specification: it drives the template, the export and the parser. A round trip
— export, edit two cells, re-import — therefore cannot drift, because there is
no second list to forget to update.

**Nothing is applied from a file that has not been shown to somebody.**
Upload parses and validates and writes a staged copy; commit applies the
staged copy. A commit therefore applies exactly what the preview displayed,
not a re-read of a file that may have changed underneath.

**Omission is not deletion.** A workbook with three tasks in it, uploaded
against a project with forty, adds or updates three tasks. It does not delete
thirty-seven. Somebody exporting one workstream to work on over a weekend
must not destroy the plan by uploading it back.

**Every write goes through `service`.** Import is not a back door: the same
permission check, the same validation, the same history row and the same
audit record, with SOURCE = EXCEL_IMPORT so that a change nobody remembers
making can be traced to the file it came from.

**A cell is data, never a formula.** openpyxl writes a string beginning `=`
as a formula, so an exported task titled `=cmd|'/c calc'!A1` is a live
payload in the reader's Excel. Everything written goes through `_safe`.
"""

from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select

from backend.models.planner import (
    ACCESS_LEVELS,
    CADENCES,
    DEPENDENCY_TYPES,
    ENTITY_TYPES,
    MILESTONE_STATUSES,
    PRIORITIES,
    PROJECT_ROLES,
    PROJECT_STATUSES,
    RAID_STATUSES,
    RAID_TYPES,
    SEVERITIES,
    SOURCE_EXCEL,
    TASK_STATUSES,
    PlannerDependency,
    PlannerImport,
    PlannerMilestone,
    PlannerParticipant,
    PlannerProject,
    PlannerRaid,
    PlannerTask,
    PlannerUpdate,
    PlannerWorkstream,
)
from backend.planner import access as acl
from backend.planner import service as svc

#: An upload larger than this is refused before openpyxl sees it. A zip bomb
#: is a small file that becomes a large one, so the parsed row counts below
#: matter as much as this does.
MAX_UPLOAD_BYTES = 8 * 1024 * 1024
MAX_ROWS_PER_SHEET = 5_000
MAX_TOTAL_ROWS = 20_000

XLSX_MIME = ("application/vnd.openxmlformats-officedocument."
             "spreadsheetml.sheet")

#: A leading one of these makes Excel treat the cell as a formula or a command.
#: Checked on the way OUT as well as on the way in: a plan can be exported from
#: CreditProbe to a person who never signed in, and the export is where the
#: payload would land.
_DANGEROUS = ("=", "+", "-", "@", "\t", "\r")


def _safe(value: Any) -> Any:
    """A value Excel will show rather than run."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "YES" if value else "NO"
    if isinstance(value, (int, float, datetime, date)):
        return value
    if isinstance(value, (list, tuple, set)):
        value = ", ".join(str(v) for v in value)
    text = str(value)
    if text[:1] in _DANGEROUS:
        # A leading apostrophe is Excel's own "this is text" marker. It is
        # not shown in the cell and it survives a round trip, so a task
        # genuinely called "-5% variance" reads back as itself.
        return "'" + text
    return text


def _unsafe(value: Any) -> Any:
    """Undo `_safe` when reading, so a round trip is lossless."""
    if isinstance(value, str) and value[:2] in ("'=", "'+", "'-", "'@"):
        return value[1:]
    return value


# ===================================================== the column contract


@dataclass(frozen=True)
class Column:
    """One column, in both directions.

    `key` is the field name the service takes. `header` is what a person
    reads. `note` is what the IMPORT GUIDE says about it, and is not
    decoration: most import failures are somebody guessing at a vocabulary.
    """

    key: str
    header: str
    note: str = ""
    width: int = 18
    required: bool = False
    choices: tuple[str, ...] = ()


@dataclass(frozen=True)
class Sheet:
    name: str
    entity: str
    purpose: str
    columns: tuple[Column, ...]
    #: The column whose value identifies an existing row. Blank on a row means
    #: "this is new"; a value that matches means "update that one".
    identity: str = "code"
    #: UPDATES is history. It can be added to and never rewritten.
    append_only: bool = False

    @property
    def headers(self) -> list[str]:
        return [c.header for c in self.columns]

    def by_header(self) -> dict[str, Column]:
        return {_norm(c.header): c for c in self.columns}


def _norm(text: Any) -> str:
    """Match headers the way a person would: case and spacing do not count."""
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


PROJECT_SHEET = Sheet(
    name="PROJECT", entity="project", identity="code",
    purpose="The project itself. One row.",
    columns=(
        Column("code", "Project Code", "Must match the project you are "
               "importing into.", 18, required=True),
        Column("name", "Project Name", "", 40, required=True),
        Column("status", "Status", "", 14, choices=PROJECT_STATUSES),
        Column("priority", "Priority", "", 12, choices=PRIORITIES),
        Column("objective", "Objective", "What done looks like.", 50),
        Column("business_context", "Business Context",
               "Why the bank is doing it.", 50),
        Column("description", "Description", "", 50),
        Column("start_date", "Start Date", "YYYY-MM-DD.", 14),
        Column("target_end_date", "Target End Date", "YYYY-MM-DD.", 16),
        Column("reporting_cadence", "Reporting Cadence", "", 18,
               choices=CADENCES),
        Column("stale_after_days", "Stale After (days)",
               "How long a task may go unmentioned before it is chased.", 18),
    ))

PARTICIPANTS_SHEET = Sheet(
    name="PARTICIPANTS", entity="participant", identity="username",
    purpose="Who is on the project, and what they may do.",
    columns=(
        Column("username", "Username", "A CreditProbe username. A person who "
               "is not a user cannot be added by a spreadsheet.", 24,
               required=True),
        Column("project_role", "Project Role", "", 20, choices=PROJECT_ROLES),
        Column("access", "Access", "", 14, choices=ACCESS_LEVELS),
        Column("workstream_code", "Workstream", "Optional.", 16),
        Column("notes", "Notes", "", 40),
    ))

WORKSTREAMS_SHEET = Sheet(
    name="WORKSTREAMS", entity="workstream", identity="code",
    purpose="The major parts of the work.",
    columns=(
        Column("code", "Workstream Code", "", 18, required=True),
        Column("name", "Workstream Name", "", 36, required=True),
        Column("lead_username", "Lead", "A CreditProbe username.", 22),
        Column("description", "Description", "", 44),
        Column("start_date", "Start Date", "YYYY-MM-DD.", 14),
        Column("target_end_date", "Target End Date", "YYYY-MM-DD.", 16),
        Column("sequence", "Order", "Left to right on the timeline.", 10),
    ))

TASKS_SHEET = Sheet(
    name="TASKS", entity="task", identity="code",
    purpose="The work itself. A blank Task Code creates a new task.",
    columns=(
        Column("code", "Task Code", "Leave blank to create a new task.", 14,
               required=False),
        Column("title", "Task", "", 44, required=True),
        Column("workstream_code", "Workstream", "Must exist on the "
               "WORKSTREAMS sheet or already in the project.", 16),
        Column("parent_code", "Parent Task", "For a subtask.", 14),
        Column("owner_username", "Owner", "Who is doing it.", 22),
        Column("reviewer_username", "Reviewer", "", 22),
        Column("status", "Status", "", 15, choices=TASK_STATUSES),
        Column("percent_complete", "% Complete", "0 to 100.", 12),
        Column("priority", "Priority", "", 12, choices=PRIORITIES),
        Column("start_date", "Start Date", "YYYY-MM-DD.", 14),
        Column("due_date", "Due Date", "YYYY-MM-DD.", 14),
        Column("effort_days", "Effort (days)", "", 13),
        Column("weight", "Weight", "How much of the project this task is "
               "worth. Default 1.", 10),
        Column("critical", "Critical", "YES or NO.", 10),
        Column("blocked", "Blocked", "YES or NO. A blocked task needs a "
               "reason.", 10),
        Column("blocker_reason", "Blocked By", "", 36),
        Column("next_step", "Next Step", "", 36),
        Column("description", "Description", "", 44),
    ))

MILESTONES_SHEET = Sheet(
    name="MILESTONES", entity="milestone", identity="code",
    purpose="The dates the project is judged on.",
    columns=(
        Column("code", "Milestone Code", "Leave blank to create.", 16),
        Column("name", "Milestone", "", 40, required=True),
        Column("workstream_code", "Workstream", "", 16),
        Column("owner_username", "Owner", "", 22),
        Column("target_date", "Target Date", "YYYY-MM-DD.", 14),
        Column("status", "Status", "", 14, choices=MILESTONE_STATUSES),
        Column("critical", "Critical", "YES or NO.", 10),
        Column("description", "Description", "", 44),
    ))

DEPENDENCIES_SHEET = Sheet(
    name="DEPENDENCIES", entity="dependency", identity="",
    purpose="What has to happen before what.",
    columns=(
        Column("from_type", "Predecessor Type", "TASK or MILESTONE.", 18,
               choices=("TASK", "MILESTONE"), required=True),
        Column("from_code", "Predecessor", "Its code.", 16, required=True),
        Column("to_type", "Successor Type", "TASK or MILESTONE.", 16,
               choices=("TASK", "MILESTONE"), required=True),
        Column("to_code", "Successor", "Its code.", 16, required=True),
        Column("dependency_type", "Link", "FS finish-to-start, SS "
               "start-to-start, FF finish-to-finish, SF start-to-finish.", 10,
               choices=DEPENDENCY_TYPES),
        Column("lag_days", "Lag (days)", "", 12),
        Column("notes", "Notes", "", 36),
    ))

RAID_SHEET = Sheet(
    name="RAID", entity="raid", identity="code",
    purpose="Risks, assumptions, issues and decisions.",
    columns=(
        Column("code", "Ref", "Leave blank to create.", 12),
        Column("raid_type", "Type", "", 14, choices=RAID_TYPES,
               required=True),
        Column("title", "Title", "", 44, required=True),
        Column("description", "Description", "", 50),
        Column("severity", "Severity", "", 12, choices=SEVERITIES),
        Column("status", "Status", "", 14, choices=RAID_STATUSES),
        Column("owner_username", "Owner", "", 22),
        Column("target_date", "Target Resolution", "YYYY-MM-DD.", 18),
        Column("mitigation", "Mitigation / Action", "", 44),
        Column("resolution", "Resolution", "", 44),
    ))

UPDATES_SHEET = Sheet(
    name="UPDATES", entity="update", identity="", append_only=True,
    purpose="What has been said about the project. Existing rows are "
            "history and are never changed by an import; a row you add "
            "here is posted as a new update.",
    columns=(
        Column("entity_type", "About", "PROJECT, TASK, MILESTONE or RAID.",
               12, choices=ENTITY_TYPES),
        Column("entity_code", "Reference", "", 14),
        Column("narrative", "Update", "", 60, required=True),
        Column("blocker", "Blocked By", "", 30),
        Column("next_step", "Next Step", "", 30),
        Column("author", "Author", "Filled in on export. Ignored on import: "
               "an update is recorded as written by whoever uploaded the "
               "file.", 20),
        Column("posted_at", "When", "Filled in on export. Ignored on "
               "import.", 20),
    ))

SHEETS: tuple[Sheet, ...] = (
    PROJECT_SHEET, PARTICIPANTS_SHEET, WORKSTREAMS_SHEET, TASKS_SHEET,
    MILESTONES_SHEET, DEPENDENCIES_SHEET, RAID_SHEET, UPDATES_SHEET,
)
BY_NAME = {s.name: s for s in SHEETS}

GUIDE = "IMPORT GUIDE"


# ================================================================ styling

_INK = "1B2430"
_ACCENT = "1F4E6B"
_MUTED = "6B7684"
_BAND = "F4F6F8"

_TITLE = Font(name="Calibri", size=15, bold=True, color=_INK)
_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_BODY = Font(name="Calibri", size=10, color=_INK)
_LABEL = Font(name="Calibri", size=10, bold=True, color=_MUTED)
_HEAD_FILL = PatternFill("solid", fgColor=_ACCENT)
_BAND_FILL = PatternFill("solid", fgColor=_BAND)
_WRAP = Alignment(vertical="top", wrap_text=True)


def _write_sheet(ws: Worksheet, sheet: Sheet, rows: list[dict[str, Any]]
                 ) -> None:
    """Header row, data rows, frozen and filtered. Identical for export and
    template — the template is simply the export of an empty project."""
    for i, column in enumerate(sheet.columns, start=1):
        cell = ws.cell(row=1, column=i, value=column.header)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = column.width

    for r, values in enumerate(rows, start=2):
        for c, column in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=r, column=c, value=_safe(values.get(column.key)))
            cell.font = _BODY
            cell.alignment = _WRAP
            if r % 2 == 0:
                cell.fill = _BAND_FILL

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(sheet.columns))}{len(rows) + 1}")


def _write_guide(ws: Worksheet) -> None:
    """The sheet that stops most import failures before they happen."""
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    row = 1
    cell = ws.cell(row=row, column=1, value="Project plan workbook")
    cell.font = _TITLE
    row += 2

    for label, text in (
        ("What this is", "One project's plan. Edit it here and upload it "
                         "back into CreditProbe."),
        ("How it is applied", "Uploading shows you a preview of every change "
                              "before anything is saved. Nothing is written "
                              "until you confirm it."),
        ("Adding something", "Add a row and leave its code blank. A code is "
                             "generated for you."),
        ("Changing something", "Edit the row. Keep the code as it is — that "
                               "is how the row is matched."),
        ("Removing something", "You cannot. Deleting a row from this file "
                               "does nothing: a workbook is often a partial "
                               "extract, and an import that deleted whatever "
                               "was missing from it would destroy a plan the "
                               "first time somebody worked on one workstream "
                               "over a weekend. Cancel the task in "
                               "CreditProbe instead, or set its status to "
                               "CANCELLED here."),
        ("Dates", "Write dates as 2026-09-30. A real Excel date cell also "
                  "works."),
        ("People", "Use CreditProbe usernames. Somebody who does not have a "
                   "CreditProbe account cannot be given work by a "
                   "spreadsheet."),
        ("Who gets the credit", "Every change is recorded as made by you, "
                                "from this file, however the file says it "
                                "was authored."),
    ):
        left = ws.cell(row=row, column=1, value=label)
        left.font = _LABEL
        left.alignment = _WRAP
        right = ws.cell(row=row, column=2, value=text)
        right.font = _BODY
        right.alignment = _WRAP
        row += 1

    row += 1
    for sheet in SHEETS:
        head = ws.cell(row=row, column=1, value=sheet.name)
        head.font = Font(name="Calibri", size=11, bold=True, color=_INK)
        purpose = ws.cell(row=row, column=2, value=sheet.purpose)
        purpose.font = _BODY
        purpose.alignment = _WRAP
        row += 1
        for column in sheet.columns:
            note = column.note
            if column.choices:
                note = (note + " " if note else "") + \
                    "One of: " + ", ".join(column.choices) + "."
            if column.required:
                note = "Required. " + note
            label = ws.cell(row=row, column=1, value="    " + column.header)
            label.font = _BODY
            text = ws.cell(row=row, column=2, value=note.strip() or "—")
            text.font = Font(name="Calibri", size=10, color=_MUTED)
            text.alignment = _WRAP
            row += 1
        row += 1


# ================================================================= export


def _people_by_id(session: Any, ids: Any) -> dict[int, str]:
    from backend.db.models import User

    wanted = {int(i) for i in ids if i}
    if not wanted:
        return {}
    rows = session.execute(
        select(User.id, User.username).where(User.id.in_(wanted))).all()
    return {int(i): u for i, u in rows}


def export(session: Any, principal: Any, project_id: int) -> bytes:
    """This project, as the workbook it can be re-imported from.

    The same columns the template declares, so export → edit → import is a
    round trip rather than two formats that happen to look alike.
    """
    acl.readable(session, project_id, principal)
    project = session.get(PlannerProject, int(project_id))
    tasks = session.execute(
        select(PlannerTask).where(PlannerTask.project_id == project.id)
        .order_by(PlannerTask.code)).scalars().all()
    workstreams = session.execute(
        select(PlannerWorkstream)
        .where(PlannerWorkstream.project_id == project.id)
        .order_by(PlannerWorkstream.sequence)).scalars().all()
    milestones = session.execute(
        select(PlannerMilestone)
        .where(PlannerMilestone.project_id == project.id)
        .order_by(PlannerMilestone.target_date)).scalars().all()
    raid = session.execute(
        select(PlannerRaid).where(PlannerRaid.project_id == project.id)
        .order_by(PlannerRaid.code)).scalars().all()
    participants = session.execute(
        select(PlannerParticipant)
        .where(PlannerParticipant.project_id == project.id)).scalars().all()
    deps = session.execute(
        select(PlannerDependency)
        .where(PlannerDependency.project_id == project.id)).scalars().all()
    updates = session.execute(
        select(PlannerUpdate).where(PlannerUpdate.project_id == project.id)
        .order_by(PlannerUpdate.created_at.desc()).limit(500)).scalars().all()

    directory = _people_by_id(session, (
        [t.owner_id for t in tasks] + [t.reviewer_id for t in tasks]
        + [m.owner_id for m in milestones] + [r.owner_id for r in raid]
        + [p.user_id for p in participants] + [w.lead_id for w in workstreams]
        + [u.author_id for u in updates]))
    ws_code = {int(w.id): w.code for w in workstreams}
    task_code = {int(t.id): t.code for t in tasks}
    ms_code = {int(m.id): m.code for m in milestones}

    def who(user_id: Any) -> str:
        return directory.get(int(user_id), "") if user_id else ""

    def entity_code(kind: str, entity_id: Any) -> str:
        if entity_id is None:
            return ""
        if kind == "TASK":
            return task_code.get(int(entity_id), "")
        if kind == "MILESTONE":
            return ms_code.get(int(entity_id), "")
        return ""

    data: dict[str, list[dict[str, Any]]] = {
        "PROJECT": [{
            "code": project.code, "name": project.name,
            "status": project.status, "priority": project.priority,
            "objective": project.objective,
            "business_context": project.business_context,
            "description": project.description,
            "start_date": _iso(project.start_date),
            "target_end_date": _iso(project.target_end_date),
            "reporting_cadence": project.reporting_cadence,
            "stale_after_days": project.stale_after_days,
        }],
        "PARTICIPANTS": [{
            "username": who(p.user_id), "project_role": p.project_role,
            "access": p.access,
            "workstream_code": ws_code.get(int(p.workstream_id), "")
            if p.workstream_id else "",
            "notes": p.notes,
        } for p in participants],
        "WORKSTREAMS": [{
            "code": w.code, "name": w.name, "lead_username": who(w.lead_id),
            "description": w.description, "start_date": _iso(w.start_date),
            "target_end_date": _iso(w.target_end_date),
            "sequence": w.sequence,
        } for w in workstreams],
        "TASKS": [{
            "code": t.code, "title": t.title,
            "workstream_code": ws_code.get(int(t.workstream_id), "")
            if t.workstream_id else "",
            "parent_code": task_code.get(int(t.parent_id), "")
            if t.parent_id else "",
            "owner_username": who(t.owner_id),
            "reviewer_username": who(t.reviewer_id),
            "status": t.status, "percent_complete": t.percent_complete,
            "priority": t.priority, "start_date": _iso(t.start_date),
            "due_date": _iso(t.due_date), "effort_days": t.effort_days,
            "weight": t.weight, "critical": bool(t.critical),
            "blocked": bool(t.blocked), "blocker_reason": t.blocker_reason,
            "next_step": t.next_step, "description": t.description,
        } for t in tasks],
        "MILESTONES": [{
            "code": m.code, "name": m.name,
            "workstream_code": ws_code.get(int(m.workstream_id), "")
            if m.workstream_id else "",
            "owner_username": who(m.owner_id),
            "target_date": _iso(m.target_date), "status": m.status,
            "critical": bool(m.critical), "description": m.description,
        } for m in milestones],
        "DEPENDENCIES": [{
            "from_type": d.from_type,
            "from_code": entity_code(d.from_type, d.from_id),
            "to_type": d.to_type,
            "to_code": entity_code(d.to_type, d.to_id),
            "dependency_type": d.dependency_type, "lag_days": d.lag_days,
            "notes": d.notes,
        } for d in deps],
        "RAID": [{
            "code": r.code, "raid_type": r.raid_type, "title": r.title,
            "description": r.description, "severity": r.severity,
            "status": r.status, "owner_username": who(r.owner_id),
            "target_date": _iso(r.target_date),
            "mitigation": r.mitigation,
            "resolution": r.resolution,
        } for r in raid],
        "UPDATES": [{
            "entity_type": u.entity_type, "entity_code": u.entity_code,
            "narrative": u.narrative, "blocker": u.blocker,
            "next_step": u.next_step, "author": who(u.author_id),
            "posted_at": _iso(u.created_at),
        } for u in updates if u.narrative],
    }

    book = Workbook()
    book.remove(book.active)
    _write_guide(book.create_sheet(GUIDE))
    for sheet in SHEETS:
        _write_sheet(book.create_sheet(sheet.name), sheet,
                     data.get(sheet.name, []))
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def template() -> bytes:
    """An empty workbook with the guide and the headers. The same shape an
    export produces, so a person who fills one in is filling in the format
    the importer reads."""
    book = Workbook()
    book.remove(book.active)
    _write_guide(book.create_sheet(GUIDE))
    for sheet in SHEETS:
        _write_sheet(book.create_sheet(sheet.name), sheet, [])
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _iso(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


# ================================================================= parsing


class ImportRefused(svc.PlannerError):
    """The file cannot be read at all. Distinct from rows that fail checks:
    a workbook with forty bad rows is still a workbook, and the person gets
    forty messages. This is "that is not a workbook"."""


@dataclass
class RowIssue:
    sheet: str
    row: int
    column: str
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {"sheet": self.sheet, "row": self.row, "column": self.column,
                "message": self.message}


@dataclass
class Change:
    """One intended write, as the preview shows it and the commit applies it."""

    sheet: str
    row: int
    entity: str
    action: str          # CREATE | UPDATE | UNCHANGED
    identity: str
    label: str
    values: dict[str, Any]
    before: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {"sheet": self.sheet, "row": self.row, "entity": self.entity,
                "action": self.action, "identity": self.identity,
                "label": self.label, "values": self.values,
                "changed": sorted(k for k, v in self.values.items()
                                  if k in self.before
                                  and _differs(self.before[k], v))}


def _differs(before: Any, after: Any) -> bool:
    """Whether a workbook cell actually changes what is stored.

    Numbers are compared as numbers. A weight stored as Decimal("1.00") and
    exported as 1.0 comes back as the string "1.0", and comparing the two as
    text calls every unchanged row an update — which turns a re-imported
    export into forty history entries recording that nothing happened.
    """
    if before is None and after in (None, ""):
        return False
    if isinstance(before, bool) or isinstance(after, bool):
        return bool(before) != bool(after)
    try:
        if before is not None and after is not None:
            return float(before) != float(after)
    except (TypeError, ValueError):
        pass
    return str(before) != str(after)


def parse(content: bytes, filename: str = "") -> dict[str, list[dict]]:
    """Bytes to rows. Refuses anything that is not a plausible plan workbook.

    Every refusal here is a real upload somebody will make: the CSV they
    renamed, the 200MB export, the workbook with a million blank styled rows,
    the file whose TASKS sheet is called "Tasks (final v3)".
    """
    if not content:
        raise ImportRefused("The file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ImportRefused(
            f"That file is {len(content) // (1024 * 1024)}MB. The limit is "
            f"{MAX_UPLOAD_BYTES // (1024 * 1024)}MB — a project plan is text, "
            "so a file this large is usually embedded images or a different "
            "kind of workbook.")
    if content[:2] != b"PK":
        raise ImportRefused(
            "That is not an .xlsx workbook. A .xls or a .csv renamed to "
            ".xlsx is still not one — open it in Excel and use Save As.")

    try:
        # read_only keeps a 40MB sheet from being materialised as objects;
        # data_only takes the cached VALUE of a formula rather than its text,
        # so a due date computed with =TODAY()+7 imports as a date.
        book = load_workbook(io.BytesIO(content), read_only=True,
                             data_only=True)
    except Exception as exc:
        raise ImportRefused(
            "That workbook could not be opened. It may be password "
            f"protected or damaged. ({type(exc).__name__})") from exc

    present = {_norm(name): name for name in book.sheetnames}
    found: dict[str, list[dict]] = {}
    total = 0
    try:
        for sheet in SHEETS:
            actual = present.get(_norm(sheet.name))
            if actual is None:
                continue
            rows, count = _read_sheet(book[actual], sheet)
            total += count
            if total > MAX_TOTAL_ROWS:
                raise ImportRefused(
                    f"That workbook has more than {MAX_TOTAL_ROWS:,} rows "
                    "across its sheets. Split it by workstream.")
            found[sheet.name] = rows
    finally:
        book.close()

    if not found:
        raise ImportRefused(
            "None of the expected sheets are in that workbook. It needs at "
            "least one of: " + ", ".join(s.name for s in SHEETS)
            + ". Download the template to see the format.")
    return found


def _read_sheet(ws: Any, sheet: Sheet) -> tuple[list[dict], int]:
    """Header-matched rows. Column ORDER is not assumed anywhere.

    People reorder columns, hide them, and add their own. Matching on the
    header text means all three survive; matching on position means a hidden
    column silently shifts every value one to the left, which is the kind of
    corruption nobody notices until a quarter later.
    """
    lookup = sheet.by_header()
    header_row: list[Any] | None = None
    rows: list[dict] = []
    seen = 0

    for excel_row, values in enumerate(ws.iter_rows(values_only=True),
                                       start=1):
        if header_row is None:
            if values and any(_norm(v) in lookup for v in values):
                header_row = list(values)
            continue
        seen += 1
        if seen > MAX_ROWS_PER_SHEET:
            raise ImportRefused(
                f"The {sheet.name} sheet has more than "
                f"{MAX_ROWS_PER_SHEET:,} rows.")
        if not values or all(v in (None, "") for v in values):
            continue
        record: dict[str, Any] = {"_row": excel_row}
        for i, head in enumerate(header_row):
            column = lookup.get(_norm(head))
            if column is None or i >= len(values):
                continue
            raw = _unsafe(values[i])
            record[column.key] = (raw.strip() if isinstance(raw, str)
                                  else raw)
        rows.append(record)

    if header_row is None:
        raise ImportRefused(
            f"The {sheet.name} sheet has no header row that matches the "
            "template. Its first row should be the column names.")
    return rows, seen


# ============================================================== validation


@dataclass
class Preview:
    """What an upload would do, before it does any of it."""

    import_id: int | None
    project_id: int
    project_code: str
    filename: str
    changes: list[Change]
    issues: list[RowIssue]

    @property
    def ok(self) -> bool:
        return not self.issues

    def summary(self) -> dict[str, Any]:
        counts: dict[str, dict[str, int]] = {}
        for change in self.changes:
            bucket = counts.setdefault(
                change.entity, {"CREATE": 0, "UPDATE": 0, "UNCHANGED": 0})
            bucket[change.action] += 1
        return {
            "by_entity": counts,
            "creates": sum(b["CREATE"] for b in counts.values()),
            "updates": sum(b["UPDATE"] for b in counts.values()),
            "unchanged": sum(b["UNCHANGED"] for b in counts.values()),
            "issues": len(self.issues),
            "ok": self.ok,
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "import_id": self.import_id, "project_id": self.project_id,
            "project_code": self.project_code, "filename": self.filename,
            "summary": self.summary(),
            "changes": [c.to_dict() for c in self.changes],
            "issues": [i.to_dict() for i in self.issues],
        }


def _usernames(session: Any, names: set[str]) -> dict[str, int]:
    from backend.db.models import User

    wanted = {n.strip().lower() for n in names if n and str(n).strip()}
    if not wanted:
        return {}
    rows = session.execute(
        select(User.id, User.username, User.is_active)).all()
    return {u.lower(): int(i) for i, u, active in rows
            if u.lower() in wanted and active}


def _text_of(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def validate(session: Any, principal: Any, project_id: int,
             parsed: dict[str, list[dict]], *, filename: str = "",
             content: bytes = b"") -> Preview:
    """Decide what every row means, and say so without writing anything.

    Row-level: one bad date does not reject the workbook, it rejects that
    row and names the sheet, the row number and the column. A person fixing
    forty rows one upload at a time gives up at about the fourth.
    """
    granted = acl.require(session, project_id, principal, "EDITOR",
                          "import a plan")
    project = session.get(PlannerProject, int(project_id))
    issues: list[RowIssue] = []
    changes: list[Change] = []

    existing = _existing(session, project)
    names: set[str] = set()
    for sheet in SHEETS:
        for row in parsed.get(sheet.name, []):
            for column in sheet.columns:
                if column.key.endswith("_username") or column.key == "username":
                    value = _text_of(row.get(column.key))
                    if value:
                        names.add(value)
    directory = _usernames(session, names)

    # Codes the workbook itself introduces count as existing for the rows
    # that follow: a TASKS row may point at a workstream that is created two
    # sheets earlier in the same file, and refusing that would mean nobody
    # could ever set up a project from one workbook.
    pending: dict[str, set[str]] = {
        "workstream": set(existing["workstream"]),
        "task": set(existing["task"]),
        "milestone": set(existing["milestone"]),
    }
    for row in parsed.get("WORKSTREAMS", []):
        code = _text_of(row.get("code"))
        if code:
            pending["workstream"].add(code)
    for row in parsed.get("TASKS", []):
        code = _text_of(row.get("code"))
        if code:
            pending["task"].add(code)
    for row in parsed.get("MILESTONES", []):
        code = _text_of(row.get("code"))
        if code:
            pending["milestone"].add(code)

    context = _Context(project=project, existing=existing,
                       directory=directory, pending=pending,
                       granted=granted, issues=issues)

    for sheet in SHEETS:
        rows = parsed.get(sheet.name)
        if rows is None:
            continue
        handler = _HANDLERS[sheet.name]
        for row in rows:
            change = handler(context, sheet, row)
            if change is not None:
                changes.append(change)

    staged = PlannerImport(
        project_id=int(project.id), project_code=project.code,
        filename=(filename or "")[:300],
        file_sha256=hashlib.sha256(content).hexdigest() if content else "",
        state="VALIDATED" if not issues else "FAILED",
        staged={"changes": [
            {"sheet": c.sheet, "row": c.row, "entity": c.entity,
             "action": c.action, "identity": c.identity, "label": c.label,
             "values": _jsonable(c.values)} for c in changes]},
        findings={"issues": [i.to_dict() for i in issues]},
        uploaded_by=getattr(principal, "user_id", None))
    session.add(staged)
    session.flush()

    preview = Preview(import_id=int(staged.id), project_id=int(project.id),
                      project_code=project.code, filename=filename,
                      changes=changes, issues=issues)
    staged.summary = preview.summary()
    return preview


def _jsonable(values: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in values.items():
        if isinstance(value, datetime):
            out[key] = value.isoformat()
        elif isinstance(value, date):
            out[key] = value.isoformat()
        else:
            out[key] = value
    return out


@dataclass
class _Context:
    project: Any
    existing: dict[str, dict[str, Any]]
    directory: dict[str, int]
    pending: dict[str, set[str]]
    granted: Any
    issues: list[RowIssue]

    def fail(self, sheet: str, row: int, column: str, message: str) -> None:
        self.issues.append(RowIssue(sheet, row, column, message))

    @property
    def ws_code(self) -> dict[int, str]:
        return {int(w.id): code
                for code, w in self.existing["workstream"].items()}

    @property
    def task_code(self) -> dict[int, str]:
        return {int(t.id): code
                for code, t in self.existing["task"].items()}

    def id_of(self, kind: str, code: str) -> int | None:
        """The id of something already in the project, or None if this
        workbook is about to create it."""
        row = self.existing.get(kind.lower(), {}).get(code)
        return int(row.id) if row is not None else None


def _existing(session: Any, project: Any) -> dict[str, dict[str, Any]]:
    """Everything already in the project, by code. Four queries, not four
    hundred: a 500-row workbook that looked each row up individually would
    take a minute a person spends watching a spinner."""
    pid = int(project.id)
    return {
        "workstream": {w.code: w for w in session.execute(
            select(PlannerWorkstream).where(
                PlannerWorkstream.project_id == pid)).scalars()},
        "task": {t.code: t for t in session.execute(
            select(PlannerTask).where(
                PlannerTask.project_id == pid)).scalars()},
        "milestone": {m.code: m for m in session.execute(
            select(PlannerMilestone).where(
                PlannerMilestone.project_id == pid)).scalars()},
        "raid": {r.code: r for r in session.execute(
            select(PlannerRaid).where(
                PlannerRaid.project_id == pid)).scalars()},
        # Keyed by user id and by the pair of ends, because that is what the
        # workbook names. Without these two an unchanged export re-imported
        # reports every participant and every dependency as a CREATE, and a
        # round trip that says "37 creates" is one nobody dares confirm.
        "participant": {int(p.user_id): p for p in session.execute(
            select(PlannerParticipant).where(
                PlannerParticipant.project_id == pid)).scalars()},
        "dependency": {
            (d.from_type, int(d.from_id), d.to_type, int(d.to_id)): d
            for d in session.execute(
                select(PlannerDependency).where(
                    PlannerDependency.project_id == pid)).scalars()},
    }


# ========================================================== row handlers


def _cell(context: _Context, sheet: Sheet, row: dict, column: Column,
          *, required: bool = False) -> Any:
    """One cell, validated against its column's own rules."""
    raw = row.get(column.key)
    text = _text_of(raw)
    if not text:
        if required or column.required:
            context.fail(sheet.name, row["_row"], column.header,
                         f"{column.header} is required.")
        return None
    if column.choices:
        upper = text.upper()
        if upper not in column.choices:
            context.fail(sheet.name, row["_row"], column.header,
                         f"{text!r} is not a {column.header.lower()}. "
                         f"One of: {', '.join(column.choices)}.")
            return None
        return upper
    return raw


def _date_cell(context: _Context, sheet: Sheet, row: dict,
               column: Column) -> Any:
    raw = row.get(column.key)
    if raw in (None, "", "-"):
        return None
    try:
        return svc._as_date(raw, column.header)
    except svc.PlannerError as exc:
        context.fail(sheet.name, row["_row"], column.header, str(exc))
        return None


def _person_cell(context: _Context, sheet: Sheet, row: dict,
                 column: Column) -> Any:
    name = _text_of(row.get(column.key))
    if not name:
        return None
    user_id = context.directory.get(name.lower())
    if user_id is None:
        context.fail(
            sheet.name, row["_row"], column.header,
            f"{name!r} is not an active CreditProbe user. Work cannot be "
            "assigned by spreadsheet to somebody without an account.")
        return None
    return user_id


def _bool_cell(context: _Context, sheet: Sheet, row: dict,
               column: Column) -> bool | None:
    text = _text_of(row.get(column.key)).upper()
    if not text:
        return None
    if text in ("YES", "Y", "TRUE", "1", "X"):
        return True
    if text in ("NO", "N", "FALSE", "0"):
        return False
    context.fail(sheet.name, row["_row"], column.header,
                 f"{text!r} is not YES or NO.")
    return None


def _number_cell(context: _Context, sheet: Sheet, row: dict, column: Column,
                 *, integer: bool = False) -> Any:
    raw = row.get(column.key)
    if raw in (None, ""):
        return None
    try:
        number = float(str(raw).replace("%", "").strip())
    except (TypeError, ValueError):
        context.fail(sheet.name, row["_row"], column.header,
                     f"{raw!r} is not a number.")
        return None
    return int(round(number)) if integer else number


def _by_key(sheet: Sheet) -> dict[str, Column]:
    return {c.key: c for c in sheet.columns}


def _ref(context: _Context, sheet: Sheet, row: dict, key: str, kind: str,
         header: str) -> str | None:
    """A code pointing at something else. Checked against what exists AND
    what this workbook is about to create."""
    code = _text_of(row.get(key))
    if not code:
        return None
    if code not in context.pending[kind]:
        context.fail(sheet.name, row["_row"], header,
                     f"There is no {kind} {code!r} in this project or in "
                     "this workbook.")
        return None
    return code


def _change(sheet: Sheet, row: dict, entity: str, identity: str, label: str,
            values: dict[str, Any], before: Any) -> Change:
    """Decide whether this row creates, changes or says nothing new.

    `before` is a mapping of the same keys as `values`, or None for something
    that does not exist yet. A mapping rather than the ORM row because half
    the workbook's columns are codes — `workstream_code`, `parent_code` —
    which have no attribute of that name on the row they describe, and
    getattr would return None for every one of them and call an unchanged
    round trip an update.
    """
    live = {k: v for k, v in values.items() if v is not None}
    if before is None:
        action = "CREATE"
    else:
        action = "UPDATE" if any(
            _differs(before.get(k), v) for k, v in live.items()
        ) else "UNCHANGED"
    snapshot = ({k: _text_of(before.get(k)) for k in live}
                if before is not None else {})
    return Change(sheet=sheet.name, row=row["_row"], entity=entity,
                  action=action, identity=identity, label=label,
                  values=live, before=snapshot)


def _snapshot(row: Any, keys: Any, extra: dict[str, Any] | None = None
              ) -> dict[str, Any]:
    """The current value of every key a row's handler might set."""
    out = {k: getattr(row, k, None) for k in keys}
    out.update(extra or {})
    return out


def _project_row(context: _Context, sheet: Sheet, row: dict) -> Change | None:
    columns = _by_key(sheet)
    code = _text_of(row.get("code"))
    if code and code != context.project.code:
        context.fail(
            sheet.name, row["_row"], "Project Code",
            f"This workbook says it is project {code!r}, but it is being "
            f"imported into {context.project.code!r}. Import it into the "
            "right project, or correct the code.")
        return None
    values = {
        "name": _text_of(row.get("name")) or None,
        "status": _cell(context, sheet, row, columns["status"]),
        "priority": _cell(context, sheet, row, columns["priority"]),
        "objective": _text_of(row.get("objective")) or None,
        "business_context": _text_of(row.get("business_context")) or None,
        "description": _text_of(row.get("description")) or None,
        "start_date": _date_cell(context, sheet, row, columns["start_date"]),
        "target_end_date": _date_cell(context, sheet, row,
                                      columns["target_end_date"]),
        "reporting_cadence": _cell(context, sheet, row,
                                   columns["reporting_cadence"]),
        "stale_after_days": _number_cell(context, sheet, row,
                                         columns["stale_after_days"],
                                         integer=True),
    }
    return _change(sheet, row, "project", context.project.code,
                   context.project.code, values,
                   _snapshot(context.project, values))


def _participant_row(context: _Context, sheet: Sheet, row: dict
                     ) -> Change | None:
    columns = _by_key(sheet)
    user_id = _person_cell(context, sheet, row, columns["username"])
    if user_id is None:
        return None
    if not context.granted.at_least("OWNER"):
        access = _text_of(row.get("access")).upper()
        if access == "OWNER":
            context.fail(
                sheet.name, row["_row"], "Access",
                "Only an owner can make somebody else an owner, and a "
                "spreadsheet is not a promotion.")
            return None
    values = {
        "user_id": user_id,
        "project_role": _cell(context, sheet, row, columns["project_role"]),
        "access": _cell(context, sheet, row, columns["access"]),
        "workstream_code": _ref(context, sheet, row, "workstream_code",
                                "workstream", "Workstream"),
        "notes": _text_of(row.get("notes")) or None,
    }
    held = context.existing["participant"].get(int(user_id))
    before = None
    if held is not None:
        before = _snapshot(held, values, {
            "workstream_code": context.ws_code.get(int(held.workstream_id))
            if held.workstream_id else None})
    return _change(sheet, row, "participant",
                   _text_of(row.get("username")),
                   _text_of(row.get("username")), values, before)


def _workstream_row(context: _Context, sheet: Sheet, row: dict
                    ) -> Change | None:
    columns = _by_key(sheet)
    code = _text_of(row.get("code"))
    name = _text_of(row.get("name"))
    if not name:
        context.fail(sheet.name, row["_row"], "Workstream Name",
                     "Workstream Name is required.")
        return None
    if code:
        try:
            svc.check_code(code, "Workstream code")
        except svc.PlannerError as exc:
            context.fail(sheet.name, row["_row"], "Workstream Code", str(exc))
            return None
    before = context.existing["workstream"].get(code) if code else None
    values = {
        "code": code or None, "name": name,
        "lead_id": _person_cell(context, sheet, row,
                                columns["lead_username"]),
        "description": _text_of(row.get("description")) or None,
        "start_date": _date_cell(context, sheet, row, columns["start_date"]),
        "target_end_date": _date_cell(context, sheet, row,
                                      columns["target_end_date"]),
        "sequence": _number_cell(context, sheet, row, columns["sequence"],
                                 integer=True),
    }
    return _change(sheet, row, "workstream", code, name, values,
                   _snapshot(before, values) if before is not None else None)


def _task_row(context: _Context, sheet: Sheet, row: dict) -> Change | None:
    columns = _by_key(sheet)
    code = _text_of(row.get("code"))
    title = _text_of(row.get("title"))
    if not title:
        context.fail(sheet.name, row["_row"], "Task", "Task is required.")
        return None
    if code:
        try:
            svc.check_code(code, "Task code")
        except svc.PlannerError as exc:
            context.fail(sheet.name, row["_row"], "Task Code", str(exc))
            return None
    before = context.existing["task"].get(code) if code else None
    if code and before is None:
        # A code that names nothing is a new task, not an error: somebody
        # numbering their own rows T-101..T-140 in a fresh workbook is doing
        # a normal thing, and refusing it would make the template unusable.
        pass

    parent = _ref(context, sheet, row, "parent_code", "task", "Parent Task")
    if parent and code and parent == code:
        context.fail(sheet.name, row["_row"], "Parent Task",
                     "A task cannot be its own parent.")
        return None

    percent = _number_cell(context, sheet, row, columns["percent_complete"],
                           integer=True)
    if percent is not None and not 0 <= percent <= 100:
        context.fail(sheet.name, row["_row"], "% Complete",
                     f"{percent} is not between 0 and 100.")
        percent = None

    blocked = _bool_cell(context, sheet, row, columns["blocked"])
    reason = _text_of(row.get("blocker_reason"))
    if blocked and not reason:
        context.fail(
            sheet.name, row["_row"], "Blocked By",
            "A task marked blocked needs a reason. 'Blocked' with nothing "
            "after it tells the project manager nothing they can act on.")
        return None

    start = _date_cell(context, sheet, row, columns["start_date"])
    due = _date_cell(context, sheet, row, columns["due_date"])
    if start and due and due < start:
        context.fail(sheet.name, row["_row"], "Due Date",
                     f"The due date ({due}) is before the start date "
                     f"({start}).")
        return None

    values = {
        "code": code or None, "title": title,
        "workstream_code": _ref(context, sheet, row, "workstream_code",
                                "workstream", "Workstream"),
        "parent_code": parent,
        "owner_id": _person_cell(context, sheet, row,
                                 columns["owner_username"]),
        "reviewer_id": _person_cell(context, sheet, row,
                                    columns["reviewer_username"]),
        "status": _cell(context, sheet, row, columns["status"]),
        "percent_complete": percent,
        "priority": _cell(context, sheet, row, columns["priority"]),
        "start_date": start, "due_date": due,
        "effort_days": _number_cell(context, sheet, row,
                                    columns["effort_days"]),
        "weight": _number_cell(context, sheet, row, columns["weight"]),
        "critical": _bool_cell(context, sheet, row, columns["critical"]),
        "blocked": blocked,
        "blocker_reason": reason or None,
        "next_step": _text_of(row.get("next_step")) or None,
        "description": _text_of(row.get("description")) or None,
    }
    snapshot = None
    if before is not None:
        snapshot = _snapshot(before, values, {
            "workstream_code": context.ws_code.get(int(before.workstream_id))
            if before.workstream_id else None,
            "parent_code": context.task_code.get(int(before.parent_id))
            if before.parent_id else None})
    return _change(sheet, row, "task", code, title, values, snapshot)


def _milestone_row(context: _Context, sheet: Sheet, row: dict
                   ) -> Change | None:
    columns = _by_key(sheet)
    code = _text_of(row.get("code"))
    name = _text_of(row.get("name"))
    if not name:
        context.fail(sheet.name, row["_row"], "Milestone",
                     "Milestone is required.")
        return None
    before = context.existing["milestone"].get(code) if code else None
    values = {
        "code": code or None, "name": name,
        "workstream_code": _ref(context, sheet, row, "workstream_code",
                                "workstream", "Workstream"),
        "owner_id": _person_cell(context, sheet, row,
                                 columns["owner_username"]),
        "target_date": _date_cell(context, sheet, row,
                                  columns["target_date"]),
        "status": _cell(context, sheet, row, columns["status"]),
        "critical": _bool_cell(context, sheet, row, columns["critical"]),
        "description": _text_of(row.get("description")) or None,
    }
    snapshot = None
    if before is not None:
        snapshot = _snapshot(before, values, {
            "workstream_code": context.ws_code.get(int(before.workstream_id))
            if before.workstream_id else None})
    return _change(sheet, row, "milestone", code, name, values, snapshot)


def _dependency_row(context: _Context, sheet: Sheet, row: dict
                    ) -> Change | None:
    columns = _by_key(sheet)
    from_type = _cell(context, sheet, row, columns["from_type"],
                      required=True)
    to_type = _cell(context, sheet, row, columns["to_type"], required=True)
    if not from_type or not to_type:
        return None
    from_code = _ref(context, sheet, row, "from_code", from_type.lower(),
                     "Predecessor")
    to_code = _ref(context, sheet, row, "to_code", to_type.lower(),
                   "Successor")
    if not from_code or not to_code:
        return None
    if from_type == to_type and from_code == to_code:
        context.fail(sheet.name, row["_row"], "Successor",
                     f"{from_code} cannot depend on itself.")
        return None
    values = {
        "from_type": from_type, "from_code": from_code,
        "to_type": to_type, "to_code": to_code,
        "dependency_type": _cell(context, sheet, row,
                                 columns["dependency_type"]) or "FS",
        "lag_days": _number_cell(context, sheet, row, columns["lag_days"],
                                 integer=True),
        "notes": _text_of(row.get("notes")) or None,
    }
    label = f"{from_code} \u2192 {to_code}"
    ends = (from_type, context.id_of(from_type, from_code),
            to_type, context.id_of(to_type, to_code))
    held = (context.existing["dependency"].get(ends)
            if None not in ends else None)
    before = _snapshot(held, values, {
        "from_code": from_code, "to_code": to_code,
    }) if held is not None else None
    return _change(sheet, row, "dependency", label, label, values, before)


def _raid_row(context: _Context, sheet: Sheet, row: dict) -> Change | None:
    columns = _by_key(sheet)
    kind = _cell(context, sheet, row, columns["raid_type"], required=True)
    title = _text_of(row.get("title"))
    if not kind or not title:
        if not title:
            context.fail(sheet.name, row["_row"], "Title",
                         "Title is required.")
        return None
    code = _text_of(row.get("code"))
    before = context.existing["raid"].get(code) if code else None
    values = {
        "code": code or None, "raid_type": kind, "title": title,
        "description": _text_of(row.get("description")) or None,
        "severity": _cell(context, sheet, row, columns["severity"]),
        "status": _cell(context, sheet, row, columns["status"]),
        "owner_id": _person_cell(context, sheet, row,
                                 columns["owner_username"]),
        "target_date": _date_cell(context, sheet, row,
                                  columns["target_date"]),
        "mitigation": _text_of(row.get("mitigation")) or None,
        "resolution": _text_of(row.get("resolution")) or None,
    }
    return _change(sheet, row, "raid", code, title, values,
                   _snapshot(before, values) if before is not None else None)


def _update_row(context: _Context, sheet: Sheet, row: dict) -> Change | None:
    """A new note. Never an edit to an old one.

    The export fills in Author and When; both are ignored here. An import
    that honoured them would let anybody with a text editor write history in
    a colleague's name, which is the whole reason the history exists.
    """
    columns = _by_key(sheet)
    narrative = _text_of(row.get("narrative"))
    if not narrative:
        return None
    posted = _text_of(row.get("posted_at"))
    if posted:
        # An exported row, coming back unchanged. Re-posting it would
        # duplicate the project's history on every round trip.
        return None
    kind = _cell(context, sheet, row, columns["entity_type"]) or "PROJECT"
    entity_code = _text_of(row.get("entity_code"))
    if kind in ("TASK", "MILESTONE") and entity_code:
        if entity_code not in context.pending[kind.lower()]:
            context.fail(sheet.name, row["_row"], "Reference",
                         f"There is no {kind.lower()} {entity_code!r} in "
                         "this project.")
            return None
    values = {"entity_type": kind, "entity_code": entity_code or None,
              "narrative": narrative,
              "blocker": _text_of(row.get("blocker")) or None,
              "next_step": _text_of(row.get("next_step")) or None}
    return _change(sheet, row, "update", "", narrative[:60], values, None)


_HANDLERS = {
    "PROJECT": _project_row,
    "PARTICIPANTS": _participant_row,
    "WORKSTREAMS": _workstream_row,
    "TASKS": _task_row,
    "MILESTONES": _milestone_row,
    "DEPENDENCIES": _dependency_row,
    "RAID": _raid_row,
    "UPDATES": _update_row,
}


# ================================================================= commit


def commit(session: Any, principal: Any, import_id: int) -> dict[str, Any]:
    """Apply a staged import. All of it, or none of it.

    Atomic because the caller owns the transaction and nothing here commits:
    a service call that raises leaves the request handler to roll back, and
    a half-applied plan — forty tasks in, three workstreams missing — is
    worse than a rejected one, because nobody can tell which half is real.

    Every write goes through `service`, so the import is subject to exactly
    the checks the UI is, and lands in the history with SOURCE = EXCEL_IMPORT.
    """
    staged = session.get(PlannerImport, int(import_id))
    if staged is None:
        raise svc.PlannerError(f"There is no import {import_id}.")
    if staged.state == "COMMITTED":
        raise svc.PlannerError(
            "That import has already been applied. Upload the file again if "
            "you meant to apply it twice.")
    if staged.state != "VALIDATED":
        raise svc.PlannerError(
            f"That import is {staged.state.lower()} and cannot be applied. "
            "Only a workbook that passed its checks can be.")

    project_id = int(staged.project_id or 0)
    acl.require(session, project_id, principal, "EDITOR", "import a plan")
    if staged.uploaded_by is not None and getattr(
            principal, "user_id", None) != staged.uploaded_by:
        raise svc.PlannerError(
            "That upload belongs to somebody else. Upload the file yourself "
            "to see what it would do before applying it.")

    project = session.get(PlannerProject, project_id)
    changes = list(staged.staged.get("changes", []))
    counts = {"project": 0, "participant": 0, "workstream": 0, "task": 0,
              "milestone": 0, "dependency": 0, "raid": 0, "update": 0}

    codes = _CodeIndex(session, project)
    order = ("project", "participant", "workstream", "task", "milestone",
             "dependency", "raid", "update")
    for entity in order:
        for change in changes:
            if change["entity"] != entity or change["action"] == "UNCHANGED":
                continue
            _apply(session, principal, project, change, codes)
            counts[entity] += 1

    staged.state = "COMMITTED"
    staged.committed_at = datetime.now(UTC)
    svc.audit(session, "PLANNER_IMPORT_COMMITTED",
              actor_id=getattr(principal, "user_id", None),
              project_id=project_id, source=SOURCE_EXCEL,
              filename=staged.filename, sha256=staged.file_sha256,
              applied=counts)
    return {"import_id": int(staged.id), "project_id": project_id,
            "applied": counts,
            "total": sum(counts.values())}


class _CodeIndex:
    """Codes to ids, kept current as the import creates things.

    A TASKS row can name a workstream the same workbook created four rows
    earlier, so the map has to be written to as well as read from.
    """

    def __init__(self, session: Any, project: Any) -> None:
        pid = int(project.id)
        self.workstream = {w.code: int(w.id) for w in session.execute(
            select(PlannerWorkstream).where(
                PlannerWorkstream.project_id == pid)).scalars()}
        self.task = {t.code: int(t.id) for t in session.execute(
            select(PlannerTask).where(
                PlannerTask.project_id == pid)).scalars()}
        self.milestone = {m.code: int(m.id) for m in session.execute(
            select(PlannerMilestone).where(
                PlannerMilestone.project_id == pid)).scalars()}
        self.raid = {r.code: int(r.id) for r in session.execute(
            select(PlannerRaid).where(
                PlannerRaid.project_id == pid)).scalars()}


def _apply(session: Any, principal: Any, project: Any, change: dict,
           codes: _CodeIndex) -> None:
    entity = change["entity"]
    values = dict(change["values"])
    pid = int(project.id)
    where = f"{change['sheet']} row {change['row']}"

    if entity == "project":
        svc.update_project(session, principal, pid, source=SOURCE_EXCEL,
                           **values)
        return

    if entity == "participant":
        code = values.pop("workstream_code", None)
        svc.add_participant(
            session, principal, pid, source=SOURCE_EXCEL,
            workstream_id=codes.workstream.get(code) if code else None,
            **values)
        return

    if entity == "workstream":
        code = values.get("code")
        if code and code in codes.workstream:
            row = session.get(PlannerWorkstream, codes.workstream[code])
            for key, value in values.items():
                if key != "code" and value is not None:
                    setattr(row, key, value)
            svc.record(session, pid, entity_type="PROJECT",
                       entity_id=pid, entity_code=code, action="updated",
                       author_id=getattr(principal, "user_id", None),
                       source=SOURCE_EXCEL,
                       narrative=f"Workstream {code} updated from {where}.")
            return
        row = svc.create_workstream(session, principal, pid,
                                    source=SOURCE_EXCEL, **values)
        session.flush()
        codes.workstream[row.code] = int(row.id)
        return

    if entity == "task":
        ws_code = values.pop("workstream_code", None)
        parent_code = values.pop("parent_code", None)
        values["workstream_id"] = (codes.workstream.get(ws_code)
                                   if ws_code else None)
        values["parent_id"] = (codes.task.get(parent_code)
                               if parent_code else None)
        code = values.get("code")
        if code and code in codes.task:
            task_id = codes.task[code]
            values.pop("code", None)
            title = values.pop("title", None)
            svc.update_task(
                session, principal, task_id, source=SOURCE_EXCEL,
                narrative="", title=title, **values)
            return
        row = svc.create_task(session, principal, pid, source=SOURCE_EXCEL,
                              **values)
        session.flush()
        codes.task[row.code] = int(row.id)
        return

    if entity == "milestone":
        ws_code = values.pop("workstream_code", None)
        values["workstream_id"] = (codes.workstream.get(ws_code)
                                   if ws_code else None)
        code = values.get("code")
        if code and code in codes.milestone:
            values.pop("code", None)
            svc.update_milestone(session, principal, codes.milestone[code],
                                 source=SOURCE_EXCEL, **values)
            return
        row = svc.create_milestone(session, principal, pid,
                                   source=SOURCE_EXCEL, **values)
        session.flush()
        codes.milestone[row.code] = int(row.id)
        return

    if entity == "dependency":
        from_id = _resolve(codes, values["from_type"], values["from_code"])
        to_id = _resolve(codes, values["to_type"], values["to_code"])
        if from_id is None or to_id is None:
            raise svc.PlannerError(
                f"{where}: {values['from_code']} or {values['to_code']} no "
                "longer exists.")
        if _dependency_exists(session, pid, values["from_type"], from_id,
                              values["to_type"], to_id):
            return
        svc.create_dependency(
            session, principal, pid, source=SOURCE_EXCEL,
            from_type=values["from_type"], from_id=from_id,
            to_type=values["to_type"], to_id=to_id,
            dependency_type=values.get("dependency_type") or "FS",
            lag_days=values.get("lag_days") or 0,
            notes=values.get("notes") or "")
        return

    if entity == "raid":
        code = values.get("code")
        if code and code in codes.raid:
            values.pop("code", None)
            values.pop("raid_type", None)
            svc.update_raid(session, principal, codes.raid[code],
                            source=SOURCE_EXCEL, **values)
            return
        values.pop("code", None)
        row = svc.create_raid(session, principal, pid, source=SOURCE_EXCEL,
                              **values)
        session.flush()
        codes.raid[row.code] = int(row.id)
        return

    if entity == "update":
        kind = values.get("entity_type") or "PROJECT"
        entity_id = _resolve(codes, kind, values.get("entity_code"))
        svc.post_update(
            session, principal, pid, source=SOURCE_EXCEL,
            narrative=values["narrative"], entity_type=kind,
            entity_id=entity_id, blocker=values.get("blocker") or "",
            next_step=values.get("next_step") or "")
        return

    raise svc.PlannerError(f"{where}: unknown row type {entity!r}.")


def _resolve(codes: _CodeIndex, kind: str, code: Any) -> int | None:
    if not code:
        return None
    if kind == "TASK":
        return codes.task.get(str(code))
    if kind == "MILESTONE":
        return codes.milestone.get(str(code))
    if kind == "RAID":
        return codes.raid.get(str(code))
    return None


def _dependency_exists(session: Any, project_id: int, from_type: str,
                       from_id: int, to_type: str, to_id: int) -> bool:
    """Re-importing an unchanged workbook must not double every link."""
    found = session.execute(
        select(PlannerDependency.id).where(
            PlannerDependency.project_id == project_id,
            PlannerDependency.from_type == from_type,
            PlannerDependency.from_id == from_id,
            PlannerDependency.to_type == to_type,
            PlannerDependency.to_id == to_id)).first()
    return found is not None


__all__ = [
    "SHEETS", "BY_NAME", "GUIDE", "Column", "Sheet",
    "MAX_UPLOAD_BYTES", "MAX_ROWS_PER_SHEET", "MAX_TOTAL_ROWS", "XLSX_MIME",
    "template", "export", "parse", "validate", "commit",
    "Preview", "Change", "RowIssue", "ImportRefused",
]
