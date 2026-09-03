"""Trying to break it on purpose.

Every test here is somebody misusing the product rather than using it. They
are separated from the permission suite because the questions are different:
that file asks "does the right person get through?", this one asks "what
happens when somebody sends something nobody would send by accident".

The attacks are the ones the specification names, plus the ones the code's own
shape suggests. Where an attack fails to be an attack — because the capability
does not exist at all — the test says so, and asserts the absence.
"""

from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import Workbook

from tests.planner.conftest import PREFIX, headers

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def theirs(client, cast):
    """A second project, with only Mallory on it.

    The point of a SECOND project is that Mallory is a legitimate user with
    legitimate projects. An attacker who is refused everything proves nothing
    about a product where most people are refused only some things.
    """
    tag = uuid.uuid4().hex[:6].upper()
    made = client.post(
        f"{PREFIX}/projects", headers=headers(cast["mallory"]),
        json={"code": f"MAL-{tag}", "name": "Mallory's own project",
              "status": "ACTIVE", "manager_id": cast["mallory"]})
    assert made.status_code == 201, made.text
    pid = made.json()["project"]["id"]
    task = client.post(f"{PREFIX}/projects/{pid}/tasks",
                       headers=headers(cast["mallory"]),
                       json={"code": "T-MAL", "title": "Mallory's task",
                             "owner_id": cast["mallory"]})
    assert task.status_code == 201, task.text
    return {"id": pid, "task": task.json()["id"]}


class TestCrossProjectReferences:
    """The interesting attacks are not "read somebody else's project". They
    are "make MY project point at a row in theirs"."""

    def test_a_dependency_cannot_reach_into_another_project(
            self, client, cast, project, theirs):
        """A link from my task to theirs would leak its existence, its code
        and its dates into my project's detail response."""
        refused = client.post(
            f"{PREFIX}/projects/{project['id']}/dependencies",
            headers=headers(cast["alice"]),
            json={"predecessor_type": "TASK",
                  "predecessor_id": project["bob_task"],
                  "successor_type": "TASK",
                  "successor_id": theirs["task"],
                  "dependency_type": "FS"})
        assert refused.status_code == 422, refused.text
        assert "not part of this project" in refused.json()["detail"]["message"]

    def test_a_task_cannot_be_parented_into_another_project(
            self, client, cast, project, theirs):
        refused = client.post(
            f"{PREFIX}/projects/{project['id']}/tasks",
            headers=headers(cast["alice"]),
            json={"code": "T-SNEAK", "title": "Reaching across",
                  "parent_id": theirs["task"]})
        assert refused.status_code in (403, 404, 422), refused.text

    def test_a_raid_item_cannot_link_to_another_projects_task(
            self, client, cast, project, theirs):
        refused = client.post(
            f"{PREFIX}/projects/{project['id']}/raid",
            headers=headers(cast["alice"]),
            json={"raid_type": "RISK", "title": "Linked across",
                  "linked_entity_type": "TASK",
                  "linked_entity_id": theirs["task"]})
        assert refused.status_code == 422, refused.text
        assert "not part of this project" in refused.json()["detail"]["message"]

    def test_a_raid_item_can_link_to_a_task_in_its_own_project(
            self, client, cast, project):
        """The mirror, so the check above is a boundary and not a ban."""
        made = client.post(
            f"{PREFIX}/projects/{project['id']}/raid",
            headers=headers(cast["alice"]),
            json={"raid_type": "RISK", "title": "Linked properly",
                  "linked_entity_type": "TASK",
                  "linked_entity_id": project["bob_task"]})
        assert made.status_code == 201, made.text

    def test_a_task_in_another_project_cannot_be_updated_or_deleted(
            self, client, cast, theirs):
        assert client.patch(f"{PREFIX}/tasks/{theirs['task']}",
                            headers=headers(cast["alice"]),
                            json={"percent_complete": 100}
                            ).status_code == 404
        assert client.delete(f"{PREFIX}/tasks/{theirs['task']}",
                             headers=headers(cast["alice"])
                             ).status_code == 404


class TestMaliciousIdentifiers:
    """Ids somebody typed rather than clicked."""

    @pytest.mark.parametrize("bad", ["0", "-1", "999999999999999999",
                                     "1 OR 1=1", "../1", "null",
                                     "1;DROP TABLE planner_tasks"])
    def test_a_hostile_project_id_is_refused_cleanly(self, client, cast, bad):
        """Cleanly means: not a 500, and not a stack trace.

        A 422 from the router's own validation and a 404 from the access
        layer are both correct answers. A traceback is never one — it names
        internal paths to somebody probing.
        """
        got = client.get(f"{PREFIX}/projects/{bad}",
                         headers=headers(cast["alice"]))
        assert got.status_code in (404, 422), f"{bad}: {got.status_code}"
        assert "Traceback" not in got.text
        assert "planner_tasks" not in got.text or got.status_code == 422

    def test_the_tables_are_still_there(self, client, cast, project):
        got = client.get(f"{PREFIX}/projects/{project['id']}",
                         headers=headers(cast["alice"]))
        assert got.status_code == 200
        assert got.json()["tasks"]


class TestCodesAndText:
    def test_a_code_cannot_be_a_path(self, client, cast):
        refused = client.post(
            f"{PREFIX}/projects", headers=headers(cast["alice"]),
            json={"code": "../../etc/passwd", "name": "Traversal"})
        assert refused.status_code == 422, refused.text
        assert "usable code" in refused.json()["detail"]["message"]

    def test_a_code_cannot_be_a_formula(self, client, cast):
        """Codes end up in cells. A leading = makes one a formula."""
        refused = client.post(
            f"{PREFIX}/projects", headers=headers(cast["alice"]),
            json={"code": "=1+1", "name": "Formula"})
        assert refused.status_code == 422, refused.text

    def test_a_title_may_contain_anything_and_is_stored_as_written(
            self, client, cast, project):
        """Titles are NOT sanitised, and that is deliberate.

        A risk genuinely called "<Finance> won't sign off" must survive. The
        defence belongs where the danger is — escaping on the way into a
        spreadsheet, and React escaping on the way into the page — not in a
        blocklist that mangles honest text.
        """
        payload = "<script>alert(1)</script> & \"quotes\" & 'apostrophes'"
        made = client.post(
            f"{PREFIX}/projects/{project['id']}/tasks",
            headers=headers(cast["alice"]),
            json={"code": "T-XSS", "title": payload})
        assert made.status_code == 201, made.text
        detail = client.get(f"{PREFIX}/projects/{project['id']}",
                            headers=headers(cast["alice"])).json()
        stored = [t for t in detail["tasks"] if t["code"] == "T-XSS"][0]
        assert stored["title"] == payload


class TestValueBounds:
    @pytest.mark.parametrize("percent", [-1, 101, 1e9])
    def test_progress_outside_zero_to_one_hundred_is_refused(
            self, client, cast, project, percent):
        refused = client.patch(
            f"{PREFIX}/tasks/{project['bob_task']}",
            headers=headers(cast["alice"]),
            json={"percent_complete": percent})
        assert refused.status_code == 422, refused.text

    def test_a_negative_weight_is_refused(self, client, cast, project):
        refused = client.post(
            f"{PREFIX}/projects/{project['id']}/tasks",
            headers=headers(cast["alice"]),
            json={"code": "T-NEG", "title": "Negative weight", "weight": -5})
        assert refused.status_code == 422, refused.text

    def test_a_due_date_before_the_start_is_refused(self, client, cast,
                                                    project):
        refused = client.post(
            f"{PREFIX}/projects/{project['id']}/tasks",
            headers=headers(cast["alice"]),
            json={"code": "T-BACK", "title": "Backwards",
                  "start_date": "2026-06-01", "due_date": "2026-05-01"})
        assert refused.status_code == 422, refused.text


class TestTheGraphCannotBeMadeIllegal:
    def test_a_cycle_is_refused_and_the_loop_is_named(self, client, cast,
                                                      project):
        """Naming the loop matters. "Cycle detected" makes somebody hunt
        through forty links; naming them makes it a ten-second fix."""
        a, b = project["bob_task"], project["alice_task"]
        first = client.post(
            f"{PREFIX}/projects/{project['id']}/dependencies",
            headers=headers(cast["alice"]),
            json={"predecessor_type": "TASK", "predecessor_id": a,
                  "successor_type": "TASK", "successor_id": b})
        assert first.status_code == 201, first.text

        loop = client.post(
            f"{PREFIX}/projects/{project['id']}/dependencies",
            headers=headers(cast["alice"]),
            json={"predecessor_type": "TASK", "predecessor_id": b,
                  "successor_type": "TASK", "successor_id": a})
        assert loop.status_code == 422, loop.text
        message = loop.json()["detail"]["message"]
        assert "T-BOB" in message and "T-ALICE" in message, message

    def test_a_task_cannot_depend_on_itself(self, client, cast, project):
        refused = client.post(
            f"{PREFIX}/projects/{project['id']}/dependencies",
            headers=headers(cast["alice"]),
            json={"predecessor_type": "TASK",
                  "predecessor_id": project["bob_task"],
                  "successor_type": "TASK",
                  "successor_id": project["bob_task"]})
        assert refused.status_code == 422, refused.text


class TestConcurrency:
    def test_the_second_writer_is_told_rather_than_ignored(self, client, cast,
                                                            project):
        detail = client.get(f"{PREFIX}/projects/{project['id']}",
                            headers=headers(cast["alice"])).json()
        task = [t for t in detail["tasks"]
                if t["id"] == project["bob_task"]][0]
        version = task["version"]

        first = client.patch(
            f"{PREFIX}/tasks/{task['id']}", headers=headers(cast["alice"]),
            json={"percent_complete": 61, "expected_version": version})
        assert first.status_code == 200, first.text

        second = client.patch(
            f"{PREFIX}/tasks/{task['id']}", headers=headers(cast["bob"]),
            json={"percent_complete": 62, "expected_version": version})
        assert second.status_code == 409, second.text
        assert "somebody else" in second.json()["detail"]["message"]

    def test_the_first_write_survived(self, client, cast, project):
        detail = client.get(f"{PREFIX}/projects/{project['id']}",
                            headers=headers(cast["alice"])).json()
        task = [t for t in detail["tasks"]
                if t["id"] == project["bob_task"]][0]
        assert task["percent_complete"] == 61


class TestImportsAsAnAttackSurface:
    def test_an_import_cannot_be_committed_into_a_different_project(
            self, client, cast, project, theirs):
        """The commit route takes an import id and no project id.

        If it trusted the id alone, an upload validated against a project you
        own could be applied to one you do not.
        """
        book = Workbook()
        book.remove(book.active)
        sheet = book.create_sheet("TASKS")
        sheet.append(["Task Code", "Task"])
        sheet.append(["T-CROSS", "Applied to the wrong project"])
        buffer = io.BytesIO()
        book.save(buffer)

        mine = client.post(
            f"{PREFIX}/projects/{project['id']}/import",
            headers=headers(cast["alice"]),
            files={"file": ("p.xlsx", buffer.getvalue(), XLSX)})
        assert mine.status_code == 200, mine.text
        import_id = mine.json()["import_id"]

        stolen = client.post(f"{PREFIX}/imports/{import_id}/commit",
                             headers=headers(cast["mallory"]))
        assert stolen.status_code in (403, 404, 422), stolen.text

        theirs_detail = client.get(f"{PREFIX}/projects/{theirs['id']}",
                                   headers=headers(cast["mallory"])).json()
        assert not any(t["code"] == "T-CROSS"
                       for t in theirs_detail["tasks"])

    def test_a_workbook_of_many_rows_is_refused_before_it_is_applied(
            self, client, cast, project):
        """Not a size attack — a small file that becomes a large one."""
        book = Workbook()
        book.remove(book.active)
        sheet = book.create_sheet("TASKS")
        sheet.append(["Task Code", "Task"])
        for i in range(6000):
            sheet.append(["", f"Task {i}"])
        buffer = io.BytesIO()
        book.save(buffer)

        got = client.post(
            f"{PREFIX}/projects/{project['id']}/import",
            headers=headers(cast["alice"]),
            files={"file": ("many.xlsx", buffer.getvalue(), XLSX)})
        assert got.status_code == 422, got.status_code
        assert "rows" in got.json()["detail"]["message"]

    def test_a_wrong_mime_type_is_judged_on_content_not_on_the_label(
            self, client, cast, project):
        """The label is the caller's claim. The magic bytes are the fact."""
        got = client.post(
            f"{PREFIX}/projects/{project['id']}/import",
            headers=headers(cast["alice"]),
            files={"file": ("plan.xlsx", b"not a workbook at all", XLSX)})
        assert got.status_code == 422
        assert "not an .xlsx" in got.json()["detail"]["message"]

    def test_a_real_workbook_labelled_as_a_pdf_still_works(self, client, cast,
                                                           project):
        """The mirror of the test above, and the reason content wins.

        Browsers and email clients relabel attachments constantly. Refusing a
        genuine workbook because its Content-Type says application/pdf would
        be a support ticket a week.
        """
        book = Workbook()
        book.remove(book.active)
        sheet = book.create_sheet("TASKS")
        sheet.append(["Task Code", "Task"])
        sheet.append(["", "Mislabelled but real"])
        buffer = io.BytesIO()
        book.save(buffer)
        got = client.post(
            f"{PREFIX}/projects/{project['id']}/import",
            headers=headers(cast["alice"]),
            files={"file": ("plan.pdf", buffer.getvalue(), "application/pdf")})
        assert got.status_code == 200, got.text


class TestCapabilitiesThatDoNotExist:
    """Some attacks fail because there is nothing to attack."""

    def test_there_is_no_route_that_deletes_a_project(self, client, cast,
                                                      project):
        """Deliberate. A delivery project carries the record of what a team
        committed to and when; deleting it destroys evidence. Archiving is
        the reversible answer, and it keeps the history."""
        got = client.delete(f"{PREFIX}/projects/{project['id']}",
                            headers=headers(cast["alice"]))
        assert got.status_code == 405, got.status_code

    def test_there_is_no_route_that_edits_history(self, client, cast,
                                                  project):
        activity = client.get(f"{PREFIX}/projects/{project['id']}/activity",
                              headers=headers(cast["alice"])).json()
        assert activity["items"], "no history to try to edit"
        update_id = activity["items"][0]["id"]
        for method in ("patch", "delete"):
            got = getattr(client, method)(
                f"{PREFIX}/projects/{project['id']}/updates/{update_id}",
                headers=headers(cast["alice"]),
                **({"json": {"narrative": "rewritten"}}
                   if method == "patch" else {}))
            assert got.status_code in (404, 405), (method, got.status_code)

    def test_no_registered_tool_can_change_a_commitment(self):
        """Restated here as well as in the AI suite.

        It belongs in both: this file is what a reviewer reads when asking
        "what did you try?", and the answer to "could the assistant close a
        risk" is that there is no such tool to call.
        """
        from backend.agentic import tools as reg

        ids = {t.tool_id for t in reg.TOOLS}
        for forbidden in ("complete_task", "change_task_owner",
                          "move_due_date", "cancel_task", "close_risk",
                          "set_project_health"):
            assert forbidden not in ids
            assert forbidden in reg.NO_TOOL_EXISTS
