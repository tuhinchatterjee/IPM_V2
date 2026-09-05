"""Creating a project from a workbook.

The update path already existed: upload into a project, preview, apply. This
is the other half — the workbook that brings the project into existence — and
the thing worth proving is that nothing is written until commit, including the
project itself.

Everything goes through the real routes, because "who may create a project
from a spreadsheet" is a question about the route and not about the parser.
"""

from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import load_workbook

from backend.planner import workbook as wb
from tests.planner.conftest import PREFIX, headers

XLSX = ("application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet")


def filled(code: str, name: str = "Imported programme", *,
           tasks: int = 3, owner: str = "") -> bytes:
    """The shipped template, populated the way a person would populate it."""
    book = load_workbook(io.BytesIO(wb.template()))

    project = book["PROJECT"]
    project.cell(row=2, column=1, value=code)
    project.cell(row=2, column=2, value=name)
    project.cell(row=2, column=3, value="ACTIVE")
    project.cell(row=2, column=4, value="HIGH")
    project.cell(row=2, column=5, value="Ship the thing.")
    project.cell(row=2, column=8, value="2026-01-05")
    project.cell(row=2, column=9, value="2026-12-18")

    streams = book["WORKSTREAMS"]
    streams.cell(row=2, column=1, value="WS-DATA")
    streams.cell(row=2, column=2, value="Data and controls")

    sheet = book["TASKS"]
    for i in range(tasks):
        row = 2 + i
        sheet.cell(row=row, column=1, value=f"T-{100 + i}")
        sheet.cell(row=row, column=2, value=f"Task {i + 1}")
        sheet.cell(row=row, column=3, value="WS-DATA")
        if owner:
            sheet.cell(row=row, column=5, value=owner)

    stones = book["MILESTONES"]
    stones.cell(row=2, column=1, value="M-1")
    stones.cell(row=2, column=2, value="Data freeze")
    stones.cell(row=2, column=5, value="2026-03-31")

    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def upload(client, user_id: int, content: bytes, filename="plan.xlsx"):
    return client.post(f"{PREFIX}/imports", headers=headers(user_id),
                       files={"file": (filename, content, XLSX)})


@pytest.fixture()
def code() -> str:
    return f"NEW-{uuid.uuid4().hex[:6].upper()}"


# ================================================================= preview


def test_a_workbook_previews_a_project_it_would_create(client, cast, code):
    reply = upload(client, cast["alice"], filled(code))
    assert reply.status_code == 200, reply.text
    body = reply.json()
    assert body["mode"] == "CREATE"
    assert body["project_code"] == code
    assert body["summary"]["ok"] is True
    assert body["summary"]["updates"] == 0, \
        "a project that does not exist yet cannot have updates"
    entities = {c["entity"] for c in body["changes"]}
    assert {"project", "workstream", "task", "milestone"} <= entities


def test_the_preview_writes_nothing(client, cast, code):
    upload(client, cast["alice"], filled(code))
    listed = client.get(f"{PREFIX}/projects", headers=headers(cast["alice"]))
    assert code not in [p["code"] for p in listed.json()["projects"]], \
        "the project existed before anybody confirmed the import"


def test_a_workbook_with_no_project_row_is_refused(client, cast):
    book = load_workbook(io.BytesIO(wb.template()))
    out = io.BytesIO()
    book.save(out)
    reply = upload(client, cast["alice"], out.getvalue())
    assert reply.status_code == 422
    assert "PROJECT sheet" in reply.text


def test_a_code_that_is_taken_says_so_and_says_what_to_do(client, cast,
                                                          project):
    reply = upload(client, cast["alice"], filled(project["code"]))
    assert reply.status_code == 422
    assert "already exists" in reply.text
    assert "import the workbook there" in reply.text


def test_a_row_error_still_names_the_sheet_and_the_row(client, cast, code):
    book = load_workbook(io.BytesIO(wb.template()))
    book["PROJECT"].cell(row=2, column=1, value=code)
    book["PROJECT"].cell(row=2, column=2, value="Bad dates")
    book["TASKS"].cell(row=2, column=1, value="T-1")
    book["TASKS"].cell(row=2, column=2, value="Impossible")
    book["TASKS"].cell(row=2, column=8, value="not-a-date")
    out = io.BytesIO()
    book.save(out)
    reply = upload(client, cast["alice"], out.getvalue())
    assert reply.status_code == 200
    body = reply.json()
    assert body["summary"]["ok"] is False
    issue = body["issues"][0]
    assert issue["sheet"] == "TASKS"
    assert issue["row"] == 2


# ================================================================== commit


def test_committing_creates_the_project_and_everything_in_it(client, cast,
                                                             code):
    preview = upload(client, cast["alice"], filled(code, tasks=4)).json()
    applied = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                          headers=headers(cast["alice"]))
    assert applied.status_code == 200, applied.text
    body = applied.json()
    project_id = body["project_id"]

    detail = client.get(f"{PREFIX}/projects/{project_id}",
                        headers=headers(cast["alice"])).json()
    assert detail["project"]["code"] == code
    assert detail["project"]["status"] == "ACTIVE"
    assert len(detail["tasks"]) == 4
    assert [m["code"] for m in detail["milestones"]] == ["M-1"]


def test_the_person_who_imported_it_can_administer_it(client, cast, code):
    preview = upload(client, cast["alice"], filled(code)).json()
    applied = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                          headers=headers(cast["alice"])).json()
    detail = client.get(f"{PREFIX}/projects/{applied['project_id']}",
                        headers=headers(cast["alice"])).json()
    assert detail["access"]["access"] == "OWNER", \
        "a project nobody can administer is a project nobody can use"


def test_a_failed_workbook_cannot_be_committed(client, cast, code):
    book = load_workbook(io.BytesIO(wb.template()))
    book["PROJECT"].cell(row=2, column=1, value=code)
    book["PROJECT"].cell(row=2, column=2, value="Has a bad row")
    book["TASKS"].cell(row=2, column=1, value="T-1")
    book["TASKS"].cell(row=2, column=2, value="Impossible")
    book["TASKS"].cell(row=2, column=8, value="not-a-date")
    out = io.BytesIO()
    book.save(out)
    preview = upload(client, cast["alice"], out.getvalue()).json()

    applied = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                          headers=headers(cast["alice"]))
    assert applied.status_code == 422
    listed = client.get(f"{PREFIX}/projects", headers=headers(cast["alice"]))
    assert code not in [p["code"] for p in listed.json()["projects"]], \
        "a rejected workbook created its project anyway"


def test_somebody_else_cannot_apply_your_upload(client, cast, code):
    preview = upload(client, cast["alice"], filled(code)).json()
    applied = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                          headers=headers(cast["bob"]))
    assert applied.status_code == 422
    assert "belongs to somebody else" in applied.text


def test_applying_twice_is_refused(client, cast, code):
    preview = upload(client, cast["alice"], filled(code)).json()
    first = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                        headers=headers(cast["alice"]))
    assert first.status_code == 200
    second = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                         headers=headers(cast["alice"]))
    assert second.status_code == 422
    assert "already been applied" in second.text


def test_the_round_trip_survives_a_created_project(client, cast, code):
    """Create from a workbook, export it, re-import: nothing new, nothing
    changed. The claim the whole column contract exists to make."""
    preview = upload(client, cast["alice"], filled(code, tasks=3)).json()
    applied = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                          headers=headers(cast["alice"])).json()
    project_id = applied["project_id"]

    exported = client.get(f"{PREFIX}/projects/{project_id}/export",
                          headers=headers(cast["alice"]))
    assert exported.status_code == 200

    back = client.post(f"{PREFIX}/projects/{project_id}/import",
                       headers=headers(cast["alice"]),
                       files={"file": ("again.xlsx", exported.content, XLSX)})
    assert back.status_code == 200, back.text
    summary = back.json()["summary"]
    assert summary["ok"] is True
    assert summary["creates"] == 0, back.json()["changes"]
    assert summary["updates"] == 0, back.json()["changes"]
