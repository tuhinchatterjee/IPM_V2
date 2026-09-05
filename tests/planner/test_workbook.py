"""The plan as a spreadsheet, tested with real .xlsx bytes.

Not a mocked parser. Every test here builds an actual workbook with openpyxl,
uploads it through the real HTTP route, and reads the result back through the
API — because the failures worth catching are the ones between the layers: a
column matched by position instead of by name, a preview that says one thing
and a commit that does another, an import that deletes what it was not
told about.
"""

from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import Workbook, load_workbook

from backend.planner import workbook as wb
from tests.planner.conftest import PREFIX, headers

XLSX = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def book_of(sheets: dict[str, list[list]]) -> bytes:
    """A workbook with exactly the sheets and rows given."""
    book = Workbook()
    book.remove(book.active)
    for name, rows in sheets.items():
        ws = book.create_sheet(name)
        for row in rows:
            ws.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def upload(client, cast, project_id: int, content: bytes, who: str = "alice",
           filename: str = "plan.xlsx"):
    return client.post(
        f"{PREFIX}/projects/{project_id}/import",
        headers=headers(cast[who]),
        files={"file": (filename, content, XLSX)})


@pytest.fixture(scope="module")
def own(client, cast):
    """A project of this file's own, so imports here cannot disturb the
    permission fixture's project."""
    tag = uuid.uuid4().hex[:6].upper()
    created = client.post(
        f"{PREFIX}/projects", headers=headers(cast["alice"]),
        json={"code": f"XL-{tag}", "name": "Workbook project",
              "status": "ACTIVE", "manager_id": cast["alice"]})
    assert created.status_code == 201, created.text
    pid = created.json()["project"]["id"]
    client.post(f"{PREFIX}/projects/{pid}/participants",
                headers=headers(cast["alice"]),
                json={"user_id": cast["bob"], "project_role": "CONTRIBUTOR",
                      "access": "CONTRIBUTOR"})
    client.post(f"{PREFIX}/projects/{pid}/workstreams",
                headers=headers(cast["alice"]),
                json={"code": "WS-1", "name": "Build"})
    return {"id": pid, "code": created.json()["project"]["code"]}


class TestTemplate:
    def test_it_has_every_sheet_and_a_guide(self, client, cast):
        got = client.get(f"{PREFIX}/template", headers=headers(cast["alice"]))
        assert got.status_code == 200
        assert got.headers["content-type"].startswith(
            "application/vnd.openxmlformats")
        book = load_workbook(io.BytesIO(got.content))
        assert book.sheetnames[0] == wb.GUIDE
        assert [s.name for s in wb.SHEETS] == book.sheetnames[1:]

    def test_the_guide_says_omission_is_not_deletion(self, client, cast):
        got = client.get(f"{PREFIX}/template", headers=headers(cast["alice"]))
        guide = load_workbook(io.BytesIO(got.content))[wb.GUIDE]
        text = " ".join(str(c.value or "") for row in guide.iter_rows()
                        for c in row).lower()
        assert "deleting a row from this file does nothing" in text

    def test_every_column_is_documented(self, client, cast):
        got = client.get(f"{PREFIX}/template", headers=headers(cast["alice"]))
        guide = load_workbook(io.BytesIO(got.content))[wb.GUIDE]
        documented = {str(c.value or "").strip()
                      for row in guide.iter_rows() for c in row}
        missing = [f"{s.name}.{c.header}" for s in wb.SHEETS
                   for c in s.columns if c.header not in documented]
        assert not missing, missing


class TestParsingRefusals:
    """Files people really upload. Each refusal names what to do instead."""

    def test_a_csv_renamed_to_xlsx(self, client, cast, own):
        got = upload(client, cast, own["id"], b"code,name\nT-1,Thing\n")
        assert got.status_code == 422
        assert "not an .xlsx" in got.json()["detail"]["message"]

    def test_an_empty_file(self, client, cast, own):
        got = upload(client, cast, own["id"], b"")
        assert got.status_code == 422

    def test_a_workbook_with_none_of_our_sheets(self, client, cast, own):
        got = upload(client, cast, own["id"],
                     book_of({"Sheet1": [["a", "b"], [1, 2]]}))
        assert got.status_code == 422
        assert "IMPORT GUIDE" not in got.json()["detail"]["message"]
        assert "TASKS" in got.json()["detail"]["message"]

    def test_a_tasks_sheet_with_no_header(self, client, cast, own):
        got = upload(client, cast, own["id"],
                     book_of({"TASKS": [["just", "some", "values"]]}))
        assert got.status_code == 422
        assert "header row" in got.json()["detail"]["message"]

    def test_an_oversized_file(self, client, cast, own):
        payload = b"PK" + b"\x00" * (wb.MAX_UPLOAD_BYTES + 1)
        got = upload(client, cast, own["id"], payload)
        assert got.status_code == 422
        assert "limit is" in got.json()["detail"]["message"]


class TestRowLevelErrors:
    def test_one_bad_row_does_not_reject_the_others(self, client, cast, own):
        content = book_of({"TASKS": [
            ["Task Code", "Task", "Due Date", "Owner"],
            ["", "A good task", "2026-05-01", ""],
            ["", "A bad one", "next Thursday", ""],
            ["", "Another good task", "2026-06-01", ""],
        ]})
        got = upload(client, cast, own["id"], content)
        assert got.status_code == 200, got.text
        body = got.json()
        assert len(body["issues"]) == 1
        issue = body["issues"][0]
        assert issue["sheet"] == "TASKS"
        assert issue["row"] == 3, "the row number must be the Excel one"
        assert issue["column"] == "Due Date"
        assert not body["summary"]["ok"]

    def test_an_unknown_person_is_named(self, client, cast, own):
        content = book_of({"TASKS": [
            ["Task Code", "Task", "Owner"],
            ["", "Assigned to a ghost", "nobody-at-all"],
        ]})
        body = upload(client, cast, own["id"], content).json()
        assert body["issues"][0]["column"] == "Owner"
        assert "nobody-at-all" in body["issues"][0]["message"]

    def test_a_blocked_task_with_no_reason_is_refused(self, client, cast,
                                                      own):
        content = book_of({"TASKS": [
            ["Task Code", "Task", "Blocked", "Blocked By"],
            ["", "Stuck", "YES", ""],
        ]})
        body = upload(client, cast, own["id"], content).json()
        assert body["issues"][0]["column"] == "Blocked By"

    def test_a_failed_import_cannot_be_committed(self, client, cast, own):
        content = book_of({"TASKS": [
            ["Task Code", "Task", "Due Date"],
            ["", "Bad", "not a date"],
        ]})
        body = upload(client, cast, own["id"], content).json()
        refused = client.post(f"{PREFIX}/imports/{body['import_id']}/commit",
                              headers=headers(cast["alice"]))
        assert refused.status_code == 422
        assert "passed its checks" in refused.json()["detail"]["message"]


class TestColumnsAreMatchedByName:
    def test_reordered_and_extra_columns_still_import(self, client, cast,
                                                      own):
        """Somebody adds their own column and drags another one left.

        Matching by position would put the due date into the title. This is
        the corruption nobody notices for a quarter, so it gets a test.
        """
        content = book_of({"TASKS": [
            ["Their own note", "Due Date", "Task", "% Complete"],
            ["ignore me", "2026-07-15", "Reordered task", 25],
        ]})
        got = upload(client, cast, own["id"], content)
        assert got.status_code == 200, got.text
        body = got.json()
        assert body["summary"]["ok"], body["issues"]
        values = body["changes"][0]["values"]
        assert values["title"] == "Reordered task"
        assert values["due_date"] == "2026-07-15"
        assert values["percent_complete"] == 25


class TestPreviewThenCommit:
    def test_nothing_is_written_until_the_commit(self, client, cast, own):
        content = book_of({"TASKS": [
            ["Task Code", "Task", "Workstream", "Owner", "Due Date"],
            ["T-IMP-1", "Imported task one", "WS-1", "", "2026-08-01"],
            ["T-IMP-2", "Imported task two", "WS-1", "", "2026-09-01"],
        ]})
        preview = upload(client, cast, own["id"], content).json()
        assert preview["summary"]["creates"] == 2
        assert preview["summary"]["ok"]

        before = client.get(f"{PREFIX}/projects/{own['id']}",
                            headers=headers(cast["alice"])).json()
        assert not any(t["code"].startswith("T-IMP")
                       for t in before["tasks"]), \
            "the preview wrote to the database"

        applied = client.post(
            f"{PREFIX}/imports/{preview['import_id']}/commit",
            headers=headers(cast["alice"]))
        assert applied.status_code == 200, applied.text
        assert applied.json()["applied"]["task"] == 2

        after = client.get(f"{PREFIX}/projects/{own['id']}",
                           headers=headers(cast["alice"])).json()
        codes = {t["code"] for t in after["tasks"]}
        assert {"T-IMP-1", "T-IMP-2"} <= codes

    def test_an_import_cannot_be_applied_twice(self, client, cast, own):
        content = book_of({"TASKS": [
            ["Task Code", "Task"],
            ["T-ONCE", "Only once"],
        ]})
        preview = upload(client, cast, own["id"], content).json()
        first = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                            headers=headers(cast["alice"]))
        assert first.status_code == 200
        again = client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                            headers=headers(cast["alice"]))
        assert again.status_code == 422
        assert "already been applied" in again.json()["detail"]["message"]

    def test_somebody_elses_upload_cannot_be_committed(self, client, cast,
                                                       own):
        content = book_of({"TASKS": [["Task Code", "Task"],
                                     ["T-MINE", "Alice's upload"]]})
        preview = upload(client, cast, own["id"], content).json()
        refused = client.post(
            f"{PREFIX}/imports/{preview['import_id']}/commit",
            headers=headers(cast["bob"]))
        assert refused.status_code in (403, 422), refused.text


class TestOmissionIsNotDeletion:
    def test_a_partial_workbook_leaves_everything_else_alone(self, client,
                                                             cast, own):
        """The single most destructive thing an importer can get wrong."""
        before = client.get(f"{PREFIX}/projects/{own['id']}",
                            headers=headers(cast["alice"])).json()
        existing = {t["code"] for t in before["tasks"]}
        assert len(existing) >= 2, "fixture needs tasks to lose"

        content = book_of({"TASKS": [
            ["Task Code", "Task", "% Complete"],
            ["T-IMP-1", "Imported task one", 60],
        ]})
        preview = upload(client, cast, own["id"], content).json()
        client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                    headers=headers(cast["alice"]))

        after = client.get(f"{PREFIX}/projects/{own['id']}",
                           headers=headers(cast["alice"])).json()
        still = {t["code"] for t in after["tasks"]}
        assert existing <= still, sorted(existing - still)
        moved = [t for t in after["tasks"] if t["code"] == "T-IMP-1"][0]
        assert moved["percent_complete"] == 60


class TestRoundTrip:
    def test_export_edit_reimport(self, client, cast, own):
        """The property the whole format exists for."""
        exported = client.get(f"{PREFIX}/projects/{own['id']}/export",
                              headers=headers(cast["alice"]))
        assert exported.status_code == 200
        book = load_workbook(io.BytesIO(exported.content))
        assert [s.name for s in wb.SHEETS] == book.sheetnames[1:]

        tasks = book["TASKS"]
        headers_row = [c.value for c in tasks[1]]
        title_at = headers_row.index("Task") + 1
        percent_at = headers_row.index("% Complete") + 1
        code_at = headers_row.index("Task Code") + 1
        edited_code = tasks.cell(row=2, column=code_at).value
        tasks.cell(row=2, column=percent_at, value=35)
        tasks.cell(row=2, column=title_at,
                   value=str(tasks.cell(row=2, column=title_at).value)
                   + " (revised)")

        buffer = io.BytesIO()
        book.save(buffer)
        preview = upload(client, cast, own["id"], buffer.getvalue()).json()
        assert preview["summary"]["ok"], preview["issues"]
        assert preview["summary"]["creates"] == 0, \
            "a round trip must not create duplicates of what it exported"

        client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                    headers=headers(cast["alice"]))
        after = client.get(f"{PREFIX}/projects/{own['id']}",
                           headers=headers(cast["alice"])).json()
        task = [t for t in after["tasks"] if t["code"] == edited_code][0]
        assert task["percent_complete"] == 35
        assert task["title"].endswith("(revised)")

    def test_a_second_round_trip_changes_nothing(self, client, cast, own):
        """Re-importing an untouched export must be a no-op.

        If it is not, then every export/import cycle drifts the plan — and
        the first symptom is duplicated dependencies and a history full of
        edits nobody made.
        """
        exported = client.get(f"{PREFIX}/projects/{own['id']}/export",
                              headers=headers(cast["alice"]))
        preview = upload(client, cast, own["id"], exported.content).json()
        assert preview["summary"]["ok"], preview["issues"]
        assert preview["summary"]["creates"] == 0
        assert preview["summary"]["updates"] == 0, [
            (c["sheet"], c["row"], c["label"], c["changed"])
            for c in preview["changes"] if c["action"] == "UPDATE"]


class TestDangerousContent:
    def test_a_formula_in_a_title_is_not_a_formula_on_export(self, client,
                                                             cast, own):
        """openpyxl writes a leading `=` as a live formula.

        So a task titled `=cmd|'/c calc'!A1` becomes a payload in the Excel
        of whoever opens the export — who may be a board member who has
        never signed in.
        """
        payload = "=cmd|'/c calc'!A1"
        made = client.post(
            f"{PREFIX}/projects/{own['id']}/tasks",
            headers=headers(cast["alice"]),
            json={"code": "T-FORMULA", "title": payload})
        assert made.status_code == 201, made.text

        exported = client.get(f"{PREFIX}/projects/{own['id']}/export",
                              headers=headers(cast["alice"]))
        book = load_workbook(io.BytesIO(exported.content))
        tasks = book["TASKS"]
        title_at = [c.value for c in tasks[1]].index("Task") + 1
        cells = [tasks.cell(row=r, column=title_at)
                 for r in range(2, tasks.max_row + 1)]
        ours = [c for c in cells if payload in str(c.value or "")]
        assert ours, "the task did not make it into the export"
        for cell in ours:
            assert cell.data_type != "f", "exported as a live formula"
            assert str(cell.value).startswith("'"), \
                "not escaped as text for Excel"

    def test_the_escape_survives_a_round_trip(self, client, cast, own):
        exported = client.get(f"{PREFIX}/projects/{own['id']}/export",
                              headers=headers(cast["alice"]))
        preview = upload(client, cast, own["id"], exported.content).json()
        titles = [c["values"].get("title") for c in preview["changes"]
                  if c["entity"] == "task"]
        assert "=cmd|'/c calc'!A1" in titles, \
            "the leading apostrophe was not stripped on the way back in"

    def test_an_import_cannot_forge_an_author(self, client, cast, own):
        """The UPDATES sheet has an Author column, filled in on export.

        Honouring it would let anybody with a text editor write project
        history in a colleague's name.
        """
        content = book_of({"UPDATES": [
            ["About", "Reference", "Update", "Author", "When"],
            ["PROJECT", "", "Signed off by the CRO.", "carol", ""],
        ]})
        preview = upload(client, cast, own["id"], content).json()
        assert preview["summary"]["ok"], preview["issues"]
        client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                    headers=headers(cast["alice"]))
        feed = client.get(f"{PREFIX}/projects/{own['id']}/activity",
                          headers=headers(cast["alice"])).json()
        posted = [e for e in feed["items"]
                  if "Signed off by the CRO" in (e.get("narrative") or "")]
        assert posted, "the update was not posted"
        author = posted[0].get("author") or {}
        assert author.get("id") == cast["alice"], \
            f"recorded as {author} rather than the uploader"


class TestImportPermissions:
    def test_a_contributor_cannot_import(self, client, cast, own):
        content = book_of({"TASKS": [["Task Code", "Task"],
                                     ["", "Bob's import"]]})
        got = upload(client, cast, own["id"], content, who="bob")
        assert got.status_code == 403, got.text

    def test_an_outsider_gets_a_404(self, client, cast, own):
        content = book_of({"TASKS": [["Task Code", "Task"],
                                     ["", "Mallory's import"]]})
        got = upload(client, cast, own["id"], content, who="mallory")
        assert got.status_code == 404, got.text

    def test_a_workbook_for_another_project_is_refused(self, client, cast,
                                                       own):
        content = book_of({"PROJECT": [
            ["Project Code", "Project Name"],
            ["SOMEONE-ELSE", "Not this project"],
        ]})
        body = upload(client, cast, own["id"], content).json()
        assert body["issues"], "a mismatched project code was accepted"
        assert "SOMEONE-ELSE" in body["issues"][0]["message"]

    def test_an_outsider_cannot_export(self, client, cast, own):
        got = client.get(f"{PREFIX}/projects/{own['id']}/export",
                         headers=headers(cast["mallory"]))
        assert got.status_code == 404


class TestDependenciesSurviveTheTrip:
    """Dependencies are the part of a plan a spreadsheet usually loses.

    They also have four columns that have to line up with a model whose own
    fields are called something else, which is exactly the kind of mapping
    that is wrong until something exercises it end to end.
    """

    def test_a_link_exports_and_re_imports_as_itself(self, client, cast, own):
        detail = client.get(f"{PREFIX}/projects/{own['id']}",
                            headers=headers(cast["alice"])).json()
        tasks = sorted(detail["tasks"], key=lambda t: t["id"])[:2]
        assert len(tasks) == 2, "need two tasks to link"

        linked = client.post(
            f"{PREFIX}/projects/{own['id']}/dependencies",
            headers=headers(cast["alice"]),
            json={"predecessor_type": "TASK", "predecessor_id": tasks[0]["id"],
                  "successor_type": "TASK", "successor_id": tasks[1]["id"],
                  "dependency_type": "FS"})
        assert linked.status_code == 201, linked.text

        exported = client.get(f"{PREFIX}/projects/{own['id']}/export",
                              headers=headers(cast["alice"]))
        sheet = load_workbook(io.BytesIO(exported.content))["DEPENDENCIES"]
        rows = [[c.value for c in row] for row in sheet.iter_rows(min_row=2)]
        codes = {(r[1], r[3]) for r in rows if r[1]}
        assert (tasks[0]["code"], tasks[1]["code"]) in codes, codes

        preview = upload(client, cast, own["id"], exported.content).json()
        assert preview["summary"]["ok"], preview["issues"]
        deps = [c for c in preview["changes"] if c["entity"] == "dependency"]
        assert deps, "the exported link was not read back"
        assert all(c["action"] == "UNCHANGED" for c in deps), \
            [c["action"] for c in deps]

    def test_re_importing_does_not_duplicate_it(self, client, cast, own):
        exported = client.get(f"{PREFIX}/projects/{own['id']}/export",
                              headers=headers(cast["alice"]))
        before = len(client.get(f"{PREFIX}/projects/{own['id']}",
                                headers=headers(cast["alice"])
                                ).json()["dependencies"])
        preview = upload(client, cast, own["id"], exported.content).json()
        client.post(f"{PREFIX}/imports/{preview['import_id']}/commit",
                    headers=headers(cast["alice"]))
        after = len(client.get(f"{PREFIX}/projects/{own['id']}",
                               headers=headers(cast["alice"])
                               ).json()["dependencies"])
        assert after == before
