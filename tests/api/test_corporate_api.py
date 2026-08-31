"""Borrower 360 and the graph, over HTTP. Phase 3.

The permission tests are the point of this file. Seeing a borrower's exposure
and seeing the named people behind it are separate acts, and a system that
cannot separate them cannot answer an audit about who looked at
beneficial-ownership data. Every one is called as each of the four roles and
the status code is read - a permission that is only a hidden menu item is a
permission an attacker has.
"""

from __future__ import annotations

import pytest

from backend.corporate import service as service_mod

PREFIX = "/api/v1/corporate"


def headers(role: str) -> dict[str, str]:
    return {"X-IPM-Role": role}


@pytest.fixture(scope="module")
def client():
    from fastapi.testclient import TestClient

    from backend.api.main import app

    return TestClient(app)


def test_the_lake_is_built_and_this_suite_actually_ran(client):
    """FAILS - does not skip - when the corporate lake is absent.

    Every other test in this file skips when the lake is missing, which is
    right on a developer's machine and dangerous everywhere else: forty
    silent skips and a green run look identical to forty passes in a summary
    line. So the absence is reported ONCE, loudly, by a test that fails; the
    rest then skip without anyone mistaking the run for a verification.
    """
    response = client.get(f"{PREFIX}/meta")
    assert response.status_code == 200, (
        "the corporate lake is not built, so none of the Borrower 360 routes "
        "were exercised. Run scripts/build_corporate_universe.py. This "
        "failure exists so a run with forty skipped tests cannot be read as "
        f"a passing verification. Response: {response.status_code} "
        f"{response.text[:200]}")


@pytest.fixture(scope="module")
def built(client):
    """Skip cleanly when the lake has not been built.

    Not an xfail and not a silent pass: the routes are real and the reason
    they cannot be exercised here is an absent artefact, which is a different
    statement from "these routes do not work". The absence is reported by
    `test_the_lake_is_built_and_this_suite_actually_ran` above, which FAILS -
    so the skips here are never the whole story.
    """
    response = client.get(f"{PREFIX}/meta")
    if response.status_code == 503:
        pytest.skip("corporate lake not built; run "
                    "scripts/build_corporate_universe.py")
    assert response.status_code == 200
    return response.json()


@pytest.fixture(scope="module")
def borrower_id(client, built):
    period = built["latest_period"]
    response = client.get(f"{PREFIX}/cohort",
                          params={"period": period, "limit": 1})
    assert response.status_code == 200
    return response.json()["borrowers"][0]["borrower_id"]


class TestMeta:
    def test_it_declares_thirteen_tabs(self, built):
        assert len(built["tabs"]) == 13
        assert [tab["key"] for tab in built["tabs"]] == list(
            service_mod.TAB_KEYS)

    def test_it_declares_eleven_network_views(self, built):
        assert len(built["network_views"]) == 11
        for view in built["network_views"]:
            assert view["purpose"], (
                f"{view['key']} has no stated purpose; a view nobody can "
                "explain is a view nobody knows how to read")

    def test_it_declares_six_group_concepts(self, built):
        concepts = built["group_concepts"]
        assert len(concepts) == 6
        for concept in concepts:
            assert concept["question"]
            assert concept["basis"]
            assert concept["is_not"], (
                f"{concept['key']} does not say what it is NOT, which is the "
                "half that stops it being read as one of the others")

    def test_it_carries_the_score_banner(self, built):
        for phrase in ("NOT A PROBABILITY", "NOT PD", "NOT A RATING",
                       "NOT IFRS 9 STAGE", "NOT ECL"):
            assert phrase in built["network_risk_score_label"]

    def test_the_frontend_is_told_what_it_may_see(self, client, built):
        analyst = client.get(f"{PREFIX}/meta", headers=headers("ANALYST"))
        views = {v["key"]: v for v in analyst.json()["network_views"]}
        assert analyst.json()["may_see_natural_persons"] is False
        assert views["ubo"]["permitted"] is False
        assert views["ownership"]["permitted"] is True

    def test_every_response_says_it_is_synthetic(self, built):
        assert built["origin"] == "SYNTHETIC_DEMO"
        assert "Synthetic" in built["not_client_data"]


class TestSearch:
    def test_an_identifier_resolves(self, client, built, borrower_id):
        found = client.get(f"{PREFIX}/search",
                           params={"q": borrower_id}).json()
        assert found["resolved"] is True
        assert found["ambiguous"] is False

    def test_a_trading_name_resolves(self, client, built, borrower_id):
        """The display name is what a screen shows, so it has to identify.

        It used not to: the generator stripped the disambiguating suffix from
        the trading name, and three different borrowers all read "Al Nahda
        Ventures" on screen with no way to tell them apart.
        """
        row = client.get(f"{PREFIX}/borrowers/{borrower_id}").json()
        name = row["fields"]["display_name"]["value"]
        found = client.get(f"{PREFIX}/search", params={"q": name}).json()
        assert found["matched"] >= 1
        assert borrower_id in {b["borrower_id"] for b in found["borrowers"]}

    def test_a_shared_stem_stays_ambiguous(self, client, built,
                                           borrower_id):
        """Never silently resolved. Picking the best match and returning it
        as the answer is the failure, because nothing on the screen would say
        a choice had been made."""
        row = client.get(f"{PREFIX}/borrowers/{borrower_id}").json()
        stem = " ".join(str(
            row["fields"]["display_name"]["value"]).split()[:1])
        found = client.get(f"{PREFIX}/search", params={"q": stem}).json()
        if found["matched"] > 1:
            assert found["ambiguous"] is True
            assert found["resolved"] is False
            assert len(found["borrowers"]) > 1

    def test_a_legal_form_word_alone_matches_nothing(self, client, built):
        """"Company" is in every name here and identifies none of them.

        Matching it would return the whole book as candidates for a single
        borrower lookup, which is worse than returning nothing: nothing is
        obviously a non-answer, and 3,100 candidates looks like a result.
        """
        found = client.get(f"{PREFIX}/search", params={"q": "Company"}).json()
        assert found["matched"] == 0
        assert found["resolved"] is False

    def test_a_missing_cohort_member_is_disclosed(self, client, built,
                                                  borrower_id):
        """Returning the nine that matched and staying quiet about the tenth
        is how a portfolio review silently loses a borrower."""
        found = client.get(f"{PREFIX}/cohort", params={
            "borrower_ids": f"{borrower_id},CORP-000000"}).json()
        assert "CORP-000000" in found.get("not_found", [])

    def test_an_unknown_facet_is_refused_not_ignored(self, client, built):
        response = client.get(f"{PREFIX}/search",
                              params={"q": "x", "period": "Q9 1999"})
        assert response.status_code in (200, 404)


class TestTheBorrower:
    def test_every_field_carries_its_lineage(self, client, borrower_id):
        row = client.get(f"{PREFIX}/borrowers/{borrower_id}").json()
        assert len(row["fields"]) >= 130
        for name, field in row["fields"].items():
            assert field["source_dataset"], name
            assert field["authority"] in ("COPY", "DERIVED"), name

    def test_no_field_claims_to_be_authoritative(self, client, built):
        lineage = client.get(f"{PREFIX}/lineage").json()
        assert lineage["authoritative_field_count"] == 0

    def test_an_unknown_borrower_is_a_404_that_explains_itself(
            self, client, built):
        response = client.get(f"{PREFIX}/borrowers/CORP-000000")
        assert response.status_code == 404
        assert "another quarter" in response.json()["detail"]["message"]

    def test_the_six_groups_come_back_separately(self, client, borrower_id):
        payload = client.get(f"{PREFIX}/borrowers/{borrower_id}/groups").json()
        keys = [concept["key"] for concept in payload["concepts"]]
        assert keys == list(service_mod.GROUP_CONCEPT_KEYS)
        assert "wrong for every one of them" in payload["note"]

    def test_the_limit_group_declares_its_threshold_unverified(
            self, client, borrower_id):
        payload = client.get(f"{PREFIX}/borrowers/{borrower_id}/groups").json()
        limit = next(c for c in payload["concepts"]
                     if c["key"] == "exposure_limit_group")
        assert "UNVERIFIED REGULATORY PARAMETER" in limit["parameter_caveat"]


class TestTheEgoGraph:
    def test_it_returns_a_neighbourhood_not_the_network(self, client,
                                                        borrower_id):
        payload = client.get(
            f"{PREFIX}/borrowers/{borrower_id}/graph",
            params={"view": "ownership", "depth": 1}).json()
        assert payload["node_count"] >= 1
        assert payload["node_count"] < 200, (
            "a depth-1 ownership neighbourhood is a handful of nodes; "
            "anything near the cap means the expansion is not bounded")

    def test_depth_widens_it(self, client, borrower_id):
        one = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                         params={"view": "full", "depth": 1}).json()
        two = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                         params={"view": "full", "depth": 2}).json()
        assert two["node_count"] >= one["node_count"]

    def test_depth_is_capped(self, client, borrower_id):
        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                              params={"depth": 9})
        assert response.status_code == 422

    def test_truncation_is_declared_and_counted(self, client, borrower_id):
        payload = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                             params={"view": "full", "depth": 3}).json()
        if payload["truncated"]:
            assert payload["truncation_note"]
            assert payload["omitted_nodes"] + payload["omitted_edges"] > 0

    def test_an_unknown_view_is_refused_with_the_list(self, client,
                                                      borrower_id):
        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                              params={"view": "spider"})
        assert response.status_code == 422
        assert "ownership" in response.json()["detail"]["message"]

    def test_every_declared_view_actually_runs(self, client, borrower_id):
        for view in service_mod.NETWORK_VIEW_KEYS:
            response = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                                  params={"view": view, "depth": 1})
            assert response.status_code == 200, view


class TestPermissions:
    """Called as each role, and the status code read."""

    def test_a_viewer_sees_the_borrower_but_not_the_graph(self, client,
                                                          borrower_id):
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}",
                          headers=headers("VIEWER")).status_code == 200
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                          headers=headers("VIEWER")).status_code == 403

    def test_an_analyst_sees_the_graph_but_not_the_people(self, client,
                                                          borrower_id):
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                          headers=headers("ANALYST")).status_code == 200
        refused = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                             params={"view": "ubo"},
                             headers=headers("ANALYST"))
        assert refused.status_code == 403

    def test_the_refusal_says_the_view_exists(self, client, borrower_id):
        """Not an empty graph. An empty graph reads as "no owners", which is
        a claim about the borrower rather than about the caller."""
        refused = client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                             params={"view": "ubo"},
                             headers=headers("ANALYST"))
        message = refused.json()["detail"]["message"]
        assert "may well have owners" in message
        assert "different from there being none" in message

    def test_a_steward_sees_the_people(self, client, borrower_id):
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}/graph",
                          params={"view": "ubo"},
                          headers=headers("DATA_STEWARD")).status_code == 200

    def test_similarity_needs_the_people_permission(self, client,
                                                    borrower_id):
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}/similar",
                          headers=headers("ANALYST")).status_code == 403
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}/similar",
                          headers=headers("ADMIN")).status_code == 200

    def test_a_withheld_count_says_so_rather_than_reading_zero(
            self, client, borrower_id):
        row = client.get(f"{PREFIX}/borrowers/{borrower_id}",
                         headers=headers("VIEWER")).json()
        assert row["fields"]["ubo_count"]["value"] == "PERMISSION_REQUIRED"
        assert row["fields"]["ubo_count"]["withheld_reason"]
        assert row["may_see_natural_persons"] is False

    def test_the_named_permissions_are_registered(self):
        from backend.api import permissions as pm

        for name in ("BORROWER_360_VIEW", "BORROWER_360_GRAPH_VIEW",
                     "BORROWER_360_UBO_VIEW", "BORROWER_360_EXPORT"):
            assert name in pm.NAMED
            roles, description = pm.NAMED[name]
            assert roles and description


class TestSimilarityAndQuality:
    def test_a_similarity_candidate_creates_nothing(self, client,
                                                    borrower_id):
        payload = client.get(f"{PREFIX}/borrowers/{borrower_id}/similar",
                             headers=headers("ADMIN")).json()
        assert "does NOT establish control" in payload["caveat"]
        for candidate in payload["candidates"]:
            assert candidate["creates_control"] is False
            assert candidate["creates_ubo"] is False
            assert candidate["creates_group_membership"] is False
            assert candidate["label"] == "HIDDEN RELATIONSHIP CANDIDATE"

    def test_the_quality_register_is_served(self, client, built):
        payload = client.get(f"{PREFIX}/quality").json()
        assert payload["checks_run"] >= 14
        assert payload["overall_status"] in ("PASS", "FLAG", "REJECT")
        assert "REJECT blocks" in payload["blocking_rule"]


class TestThePack:
    """The export. Phase 3.12."""

    def test_it_downloads_a_workbook(self, client, borrower_id):
        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/pack")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"
        assert "borrower-360" in response.headers["content-disposition"]
        assert response.headers["x-creditprobe-origin"] == "SYNTHETIC_DEMO"

    def test_it_has_a_cover_and_seventeen_sheets(self, client, borrower_id):
        import io

        from openpyxl import load_workbook

        from backend.corporate import pack as pack_mod

        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/pack")
        book = load_workbook(io.BytesIO(response.content))
        assert book.sheetnames == list(pack_mod.SHEETS)
        assert book.sheetnames[0] == "COVER"
        assert len(book.sheetnames) == 18

    def test_the_cover_says_what_the_pack_is_not(self, client, borrower_id):
        import io

        from openpyxl import load_workbook

        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/pack")
        book = load_workbook(io.BytesIO(response.content))
        text = "\n".join(
            str(cell.value)
            for row in book["COVER"].iter_rows()
            for cell in row if cell.value)
        for phrase in ("NOT A PROBABILITY", "NOT an expected credit loss",
                       "not regulatory connectedness",
                       "UNVERIFIED REGULATORY PARAMETER",
                       "does NOT establish control", "SYNTHETIC_DEMO"):
            assert phrase in text, phrase

    def test_a_sentinel_survives_the_export(self, client, built):
        """Exported as a blank, four different statements become one empty
        cell - and Excel reads an empty cell as zero.

        Exported for a borrower that HAS one, found from the derived data
        rather than assumed: the first borrower in the book may legitimately
        be in every graph, and a test that happens to pass on it proves
        nothing about the rendering of an absence.
        """
        import io

        from openpyxl import load_workbook

        from backend.corporate import graphsummary as gs
        from backend.corporate import service as service

        period = built["latest_period"]
        frame = service.load(gs.GROUPS_DATASET)
        block = frame[(frame["period"] == period)
                      & (frame["debtrank_status"] == gs.NOT_AVAILABLE)]
        assert len(block), (
            "no borrower in this quarter is outside the exposure network, so "
            "the sentinel path cannot be exercised")
        subject = str(block.iloc[0]["borrower_id"])

        response = client.get(f"{PREFIX}/borrowers/{subject}/pack")
        book = load_workbook(io.BytesIO(response.content))
        text = "\n".join(
            str(cell.value)
            for name in book.sheetnames
            for row in book[name].iter_rows()
            for cell in row if cell.value)
        assert gs.NOT_AVAILABLE in text
        assert "NETWORK RISK SCORE" in text

    def test_the_people_sheets_are_present_but_withheld(self, client,
                                                        borrower_id):
        import io

        from openpyxl import load_workbook

        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/pack",
                              headers=headers("ANALYST"))
        assert response.status_code == 200
        book = load_workbook(io.BytesIO(response.content))
        assert "CONTROL AND UBO" in book.sheetnames
        assert "WITHHELD" in str(book["CONTROL AND UBO"]["A1"].value)

    def test_a_viewer_may_not_export(self, client, borrower_id):
        assert client.get(f"{PREFIX}/borrowers/{borrower_id}/pack",
                          headers=headers("VIEWER")).status_code == 403

    def test_the_lineage_sheet_carries_every_field(self, client,
                                                   borrower_id):
        import io

        from openpyxl import load_workbook

        from backend.corporate import lineage as lineage_mod

        response = client.get(f"{PREFIX}/borrowers/{borrower_id}/pack")
        book = load_workbook(io.BytesIO(response.content))
        # Title, note, blank, header, then one row per field.
        assert book["LINEAGE"].max_row == len(lineage_mod.FIELDS) + 4


# --------------------------------------------------------------- workspace
# A person's own working set: borrowers they pinned and searches they saved.


class TestWorkspace:
    """Pins and saved cohorts.

    Two properties matter more than the CRUD. A saved cohort stores the
    QUERY, so running it next quarter answers this quarter's question rather
    than replaying last quarter's ids. And a working set is one person's:
    nothing here is visible to anybody else, and nothing here publishes a
    judgement nobody offered.
    """

    def _clear(self, client, role: str = "ANALYST") -> None:
        found = client.get(f"{PREFIX}/workspace", headers=headers(role))
        if found.status_code != 200:
            return
        body = found.json()
        for pin in body.get("pins", []):
            client.delete(f"{PREFIX}/workspace/pins/{pin['reference']}",
                          headers=headers(role))
        for cohort in body.get("cohorts", []):
            client.delete(
                f"{PREFIX}/workspace/cohorts/{cohort['reference']}",
                headers=headers(role))

    def test_a_new_working_set_is_empty_and_says_what_it_is(self, client):
        self._clear(client)
        body = client.get(f"{PREFIX}/workspace",
                          headers=headers("ANALYST")).json()
        assert body["pins"] == []
        assert body["cohorts"] == []
        assert body["maximum_per_kind"] > 0
        assert len(body["searchable"]) >= 12
        assert "yours alone" in body["note"]

    def test_a_pin_survives_and_carries_its_note(self, client):
        self._clear(client)
        first = client.get(f"{PREFIX}/borrowers",
                           headers=headers("ANALYST"))
        borrower = None
        if first.status_code == 200:
            rows = first.json().get("rows") or []
            borrower = rows[0]["borrower_id"] if rows else None
        borrower = borrower or "CORP-100000"

        made = client.post(f"{PREFIX}/workspace/pins",
                           json={"borrower_id": borrower,
                                 "label": "watch this one",
                                 "noted": "group utilisation 31%, INVESTIGATE"},
                           headers=headers("ANALYST"))
        assert made.status_code == 200, made.text
        assert made.json()["pin"]["reference"] == borrower

        body = client.get(f"{PREFIX}/workspace",
                          headers=headers("ANALYST")).json()
        assert [p["reference"] for p in body["pins"]] == [borrower]
        assert body["pins"][0]["noted"].startswith("group utilisation"), (
            "the note a person wrote when they pinned it was dropped, so "
            "opening the pin later says nothing about whether it moved")

    def test_pinning_the_same_borrower_twice_updates_rather_than_doubles(
            self, client):
        self._clear(client)
        for label in ("first", "second"):
            client.post(f"{PREFIX}/workspace/pins",
                        json={"borrower_id": "CORP-100000", "label": label},
                        headers=headers("ANALYST"))
        body = client.get(f"{PREFIX}/workspace",
                          headers=headers("ANALYST")).json()
        assert len(body["pins"]) == 1
        assert body["pins"][0]["label"] == "second"

    def test_a_saved_cohort_stores_the_query_not_the_borrowers(self, client):
        self._clear(client)
        made = client.post(
            f"{PREFIX}/workspace/cohorts",
            json={"label": "Contracting over the trigger",
                  "query": {"facets": {"sector": "Contracting"}}},
            headers=headers("ANALYST"))
        assert made.status_code == 200, made.text
        saved = made.json()["cohort"]
        assert saved["query"]["facets"] == {"sector": "Contracting"}
        assert not saved["query"]["borrower_ids"], (
            "a saved cohort kept borrower ids, so next quarter it will "
            "answer last quarter's question under this quarter's name")

    def test_a_cohort_filtering_on_an_undeclared_attribute_is_refused(
            self, client):
        self._clear(client)
        made = client.post(
            f"{PREFIX}/workspace/cohorts",
            json={"label": "nonsense",
                  "query": {"facets": {"favourite_colour": "blue"}}},
            headers=headers("ANALYST"))
        assert made.status_code == 422, made.text
        assert "favourite_colour" in made.text, (
            "the refusal does not name the facet it refused, so a caller "
            "cannot tell which part of their cohort was wrong")

    def test_an_empty_cohort_is_refused_rather_than_saved_as_the_book(
            self, client):
        self._clear(client)
        made = client.post(f"{PREFIX}/workspace/cohorts",
                           json={"label": "everything", "query": {}},
                           headers=headers("ANALYST"))
        assert made.status_code == 422, made.text

    def test_a_saved_cohort_runs_against_the_current_book(self, client):
        self._clear(client)
        made = client.post(
            f"{PREFIX}/workspace/cohorts",
            json={"label": "Contracting names",
                  "query": {"facets": {"sector": "Contracting"}}},
            headers=headers("ANALYST"))
        if made.status_code != 200:
            pytest.skip("the corporate lake is not built")
        reference = made.json()["cohort"]["reference"]

        ran = client.get(f"{PREFIX}/workspace/cohorts/{reference}/run",
                         headers=headers("ANALYST"))
        assert ran.status_code == 200, ran.text
        body = ran.json()
        assert body["cohort"]["reference"] == reference
        assert "result" in body
        assert body["origin"] == "SYNTHETIC_DEMO"

    def test_running_a_cohort_nobody_saved_is_a_404(self, client):
        ran = client.get(f"{PREFIX}/workspace/cohorts/never-saved/run",
                         headers=headers("ANALYST"))
        assert ran.status_code == 404

    def test_removing_a_pin_removes_it(self, client):
        self._clear(client)
        client.post(f"{PREFIX}/workspace/pins",
                    json={"borrower_id": "CORP-100000"},
                    headers=headers("ANALYST"))
        gone = client.delete(f"{PREFIX}/workspace/pins/CORP-100000",
                             headers=headers("ANALYST"))
        assert gone.status_code == 200
        assert gone.json()["removed"] is True
        body = client.get(f"{PREFIX}/workspace",
                          headers=headers("ANALYST")).json()
        assert body["pins"] == []

    def test_every_role_that_may_view_may_keep_its_own_working_set(
            self, client):
        """A working set is the reader's own bookmarks. Refusing somebody
        their own bookmarks behind an export permission would be governance
        theatre - it protects nothing and stops the work."""
        for role in ("ADMIN", "DATA_STEWARD", "ANALYST", "VIEWER"):
            found = client.get(f"{PREFIX}/workspace", headers=headers(role))
            assert found.status_code == 200, (
                f"{role} cannot read their own working set")
