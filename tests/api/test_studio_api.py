"""
Analysis Studio over HTTP.

The refusals are the point. A description CreditProbe cannot read must come back
as a question rather than a guess; a method whose validation pack has not passed
must not be certifiable by anybody; and a fork must never inherit its parent's
tick. Everything else on this surface is reading.
"""

from __future__ import annotations

import io

import pytest

ADMIN = {"X-IPM-Role": "ADMIN"}
ANALYST = {"X-IPM-Role": "ANALYST"}
VIEWER = {"X-IPM-Role": "VIEWER"}

DESCRIPTION = ("the share of facilities performing at a reporting date that are "
               "90 or more days past due one year later")


@pytest.fixture(scope="module")
def client():
    from fastapi.testclient import TestClient

    from backend.api.main import app

    return TestClient(app)


@pytest.fixture
def built(client):
    """A freshly built method, saved into the live registry."""
    body = client.post("/api/v1/studio/build", headers=ANALYST, json={
        "name": "Test Forward Default Rate",
        "description": DESCRIPTION,
        "answers": {},
        "opening_period": "OPEN",
        "closing_period": "CLOSE",
        "save": True,
    })
    assert body.status_code == 200, body.text
    return body.json()


# ------------------------------------------------------------------- library


def test_the_library_is_large_and_mostly_not_certified(client):
    body = client.get("/api/v1/studio").json()
    stats = body["stats"]
    assert stats["total"] >= 300
    assert 40 <= stats["certified"] <= 60
    assert stats["certified"] < stats["total"] / 4, (
        "Most of a 300-method library is a definition, not a certified "
        "implementation, and the product should not pretend otherwise."
    )


def test_the_library_can_be_narrowed(client):
    body = client.get("/api/v1/studio",
                      params={"certified_only": True}).json()
    assert body["methods"]
    assert all(m["is_certified"] for m in body["methods"])


def test_search_matches_on_meaning_not_only_on_id(client):
    body = client.get("/api/v1/studio", params={"q": "concentration"}).json()
    assert body["total_matched"] >= 3


def test_the_certification_audit_is_public(client):
    body = client.get("/api/v1/studio/certification").json()
    assert body["certified_count"] >= 40
    # Every downgrade carries its reason. A claim refused without a reason is
    # indistinguishable from a bug.
    assert all(reason.strip() for reason in body["downgraded"].values())


def test_an_unknown_method_is_a_404(client):
    assert client.get("/api/v1/studio/no_such_method").status_code == 404


# ------------------------------------------------------------------ describing


def test_a_description_is_read_back_with_its_open_decisions(client):
    body = client.post("/api/v1/studio/describe", headers=ANALYST,
                       json={"description": DESCRIPTION}).json()
    reading = body["reading"]
    assert reading["understood"]
    assert reading["kind"] == "forward_rate"
    assert reading["horizon_periods"] == 4
    ids = {c["id"] for c in reading["clarifications"]}
    assert {"default_definition", "timing", "exits", "weighting"} <= ids
    assert all(c["because"] for c in reading["clarifications"])


def test_a_description_that_is_not_understood_says_so(client):
    body = client.post("/api/v1/studio/describe", headers=ANALYST,
                       json={"description": "just do the usual thing"}).json()
    assert not body["reading"]["understood"]
    assert body["reading"]["note"]


def test_building_from_an_unreadable_description_is_refused(client):
    response = client.post("/api/v1/studio/build", headers=ANALYST, json={
        "name": "Nonsense", "description": "just do the usual thing",
        "answers": {}, "opening_period": "OPEN", "closing_period": "CLOSE",
    })
    assert response.status_code == 422
    assert "could not tell" in response.json()["detail"]["message"]


def test_a_choice_the_runtime_cannot_honour_is_refused_not_approximated(client):
    response = client.post("/api/v1/studio/build", headers=ANALYST, json={
        "name": "Anytime Default Rate", "description": DESCRIPTION,
        "answers": {"timing": "anytime"},
        "opening_period": "OPEN", "closing_period": "CLOSE",
    })
    assert response.status_code == 422
    assert "every reporting period" in response.json()["detail"]["message"]


# -------------------------------------------------------------------- building


def test_a_built_method_arrives_validated_and_uncertified(built):
    method = built["method"]
    assert method["lifecycle"] == "validated"
    assert not method["is_certified"]
    assert method["source"] == "bank"
    assert built["validation"]["all_passed"]
    assert built["validation"]["failed"] == 0


def test_the_pack_shows_the_sql_the_number_came_from(built):
    validation = built["validation"]
    assert "SELECT" in validation["sql"].upper()
    # Every value bound, never concatenated. That separation is the safety
    # property, so the pack shows it separately too.
    assert validation["parameters"]
    assert validation["sql"].count("?") == len(validation["parameters"]), (
        "Every value is a placeholder in the statement and a parameter beside "
        "it. A value that appears in the statement text was concatenated."
    )


def test_the_pack_covers_the_contentious_cases(built):
    ids = {c["id"] for c in built["validation"]["cases"]}
    assert {"boundary_89", "boundary_90_closing", "cured_before_horizon",
            "no_forward_observation", "portfolio_total"} <= ids


def test_a_viewer_may_read_the_library_but_not_build(client):
    assert client.get("/api/v1/studio", headers=VIEWER).status_code == 200
    response = client.post("/api/v1/studio/describe", headers=VIEWER,
                           json={"description": DESCRIPTION})
    assert response.status_code == 403


# ---------------------------------------------------------------- certifying


def test_certification_needs_more_than_an_analyst(client, built):
    response = client.post(f"/api/v1/studio/{built['method']['id']}/certify",
                           headers=ANALYST, json={"certified_by": "Someone"})
    assert response.status_code == 403


def test_a_validated_method_can_be_certified_and_says_who_by(client, built):
    response = client.post(f"/api/v1/studio/{built['method']['id']}/certify",
                           headers=ADMIN,
                           json={"certified_by": "Model Validation"})
    assert response.status_code == 200, response.text
    method = response.json()["method"]
    assert method["is_certified"]
    assert method["certified_by"] == "Model Validation"
    assert method["certified_at"]


def test_a_method_with_no_evidence_cannot_be_certified(client):
    """A library definition nobody implemented stays a definition."""
    body = client.get("/api/v1/studio",
                      params={"lifecycle": "preconfigured", "limit": 5}).json()
    target = next(m for m in body["methods"] if not m["is_runnable"])
    response = client.post(f"/api/v1/studio/{target['id']}/certify",
                           headers=ADMIN, json={"certified_by": "Someone"})
    assert response.status_code == 422
    assert "missing" in response.json()["detail"]["message"]


# -------------------------------------------------------------------- forking


def test_a_fork_does_not_inherit_the_tick(client, built):
    source = built["method"]["id"]
    client.post(f"/api/v1/studio/{source}/certify", headers=ADMIN,
                json={"certified_by": "Model Validation"})
    response = client.post(f"/api/v1/studio/{source}/fork", headers=ANALYST,
                           json={"name": "Bank Variant Forward Default Rate"})
    assert response.status_code == 201, response.text
    fork = response.json()["method"]
    assert fork["lifecycle"] == "draft"
    assert not fork["is_certified"]
    assert fork["forked_from"] == source
    assert fork["version"] == "1.0.0"
    assert all(c["passed"] is None for c in fork["test_cases"])


def test_a_fork_computes_the_same_thing_until_it_is_changed(client, built):
    source = built["method"]["id"]
    fork = client.post(f"/api/v1/studio/{source}/fork", headers=ANALYST,
                       json={"name": "Identical Fork Of The Rate"}).json()["method"]
    assert fork["fingerprint"] == built["method"]["fingerprint"], (
        "A fork with no edit computes exactly what its source computes."
    )


def test_a_forked_method_can_be_revalidated(client, built):
    source = built["method"]["id"]
    fork = client.post(f"/api/v1/studio/{source}/fork", headers=ANALYST,
                       json={"name": "Revalidated Fork Of The Rate"}
                       ).json()["method"]
    body = client.post(f"/api/v1/studio/{fork['id']}/validate",
                       headers=ANALYST).json()
    assert body["validation"]["all_passed"]


# --------------------------------------------------------------------- editing


def test_editing_a_certified_method_keeps_the_signed_off_version(client, built):
    method_id = built["method"]["id"]
    client.post(f"/api/v1/studio/{method_id}/certify", headers=ADMIN,
                json={"certified_by": "Model Validation"})
    response = client.post(f"/api/v1/studio/{method_id}/edit", headers=ANALYST,
                           json={"changes": {"interpretation": "Read it twice."},
                                 "change_note": "Clearer wording."})
    assert response.status_code == 200, response.text
    method = response.json()["method"]
    assert method["lifecycle"] == "draft", "An edit drops the tick."
    assert not method["certified_at"]
    assert method["versions"], "The certified version stays in the history."
    assert method["versions"][0]["lifecycle"] == "certified"
    assert response.json()["changes"]


def test_the_plan_cannot_be_edited_as_prose(client, built):
    response = client.post(f"/api/v1/studio/{built['method']['id']}/edit",
                           headers=ANALYST,
                           json={"changes": {"plan": "whatever I like"}})
    assert response.status_code == 422


# ------------------------------------------------------------------- workbook


def test_the_validation_pack_downloads_as_a_workbook(client, built):
    import openpyxl

    from backend.studio.workbook import SHEETS

    response = client.get(
        f"/api/v1/studio/{built['method']['id']}/validation-pack.xlsx",
        headers=ANALYST)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats")
    book = openpyxl.load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == SHEETS

    reconciliation = list(book["Reconciliation"].iter_rows(values_only=True))
    verdicts = {row[-1] for row in reconciliation[1:]}
    assert verdicts == {"agrees"}

    sql = "\n".join(str(row[0] or "") for row in
                    book["SQL"].iter_rows(values_only=True))
    assert "SELECT" in sql.upper()


def test_a_method_with_no_plan_has_no_workbook(client):
    body = client.get("/api/v1/studio",
                      params={"lifecycle": "preconfigured", "limit": 5}).json()
    target = next(m for m in body["methods"] if not m["is_runnable"])
    response = client.get(f"/api/v1/studio/{target['id']}/validation-pack.xlsx",
                          headers=ANALYST)
    assert response.status_code == 422
