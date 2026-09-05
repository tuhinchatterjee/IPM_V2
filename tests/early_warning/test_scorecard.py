"""The borrower four-layer scorecard and its timeline. Sections 11C, 11D, 11I.

The screen showed which signals fired and how bad the worst one was. It did not
show, for any single condition, what the value is now, what it was last time,
which way it moved, what the line is, or whether it has been true before — so a
reader could not tell a name that has just crossed a line from one that has
been over it for a year, and those are different conversations.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from backend.early_warning import assessment as ea
from backend.early_warning import scorecard as sc
from backend.early_warning import signals as sg
from backend.early_warning import taxonomy as tx
from backend.early_warning import workbook as wb
from backend.product import methodology as me


@pytest.fixture(scope="module")
def standing() -> sg.Standing:
    book = sg._book("")
    ranked = book.get("_ranked") or []
    assert ranked, "no borrowers on the book"
    return ranked[0]


@pytest.fixture(scope="module")
def card(standing) -> dict:
    return sc.build(standing)


class TestEveryConditionIsShown:
    def test_the_layers_carry_every_governed_signal_between_them(self, card):
        # Not only the ones that fired. A layer showing three amber rows and
        # hiding the eleven green ones reads as an emergency whatever the
        # borrower is doing.
        keys = {c["signal"] for layer in card["layers"]
                for c in layer["components"]}
        assert keys == {s.key for s in tx.SIGNALS}, (
            "the scorecard drops signals the taxonomy governs")

    def test_conditions_within_threshold_are_present(self, card):
        within = [c for layer in card["layers"] for c in layer["components"]
                  if c["status"] == sc.WITHIN]
        assert within

    def test_untested_says_so_rather_than_reading_as_an_all_clear(self, card):
        for layer in card["layers"]:
            for component in layer["components"]:
                if not component["available"]:
                    assert component["status"] == sc.UNTESTED
                    assert component["unavailable"]

    def test_a_layer_counts_only_what_it_could_test(self, card):
        for layer in card["layers"]:
            assert layer["tested"] + layer["untested"] == len(
                layer["components"])
            assert layer["over"] <= layer["tested"]


class TestTheColumnsSection11cAsksFor:
    def test_every_column_is_named(self, card):
        assert card["columns"] == [
            "Current", "Previous", "Movement", "Threshold", "Status",
            "Severity", "Persistence", "Detection", "What it means"]

    def test_every_component_fills_every_column(self, card):
        for layer in card["layers"]:
            for component in layer["components"]:
                assert component["status"] in {sc.OVER, sc.WITHIN, sc.UNTESTED}
                assert component["severity"] in tx.SEVERITIES
                assert component["detection"] in tx.TAC_TYPES
                assert component["detection_letter"] in {"T", "A", "C"}
                assert component["state"] in ea.STATES
                assert component["means"], f"{component['signal']} says nothing"

    def test_a_value_carries_its_unit(self, card):
        for layer in card["layers"]:
            for component in layer["components"]:
                assert component["unit"], (
                    f"{component['signal']} publishes a bare number")

    def test_only_money_carries_a_currency(self, card):
        for layer in card["layers"]:
            for component in layer["components"]:
                if component["unit"] != tx.MONEY:
                    assert component["currency"] == "", (
                        f"{component['signal']} is {component['unit']} and "
                        "carries a currency, which invents money")


class TestTheLayers:
    def test_all_four_are_present_in_order(self, card):
        assert [layer["number"] for layer in card["layers"]] == [1, 2, 3, 4]

    def test_layer_four_is_not_zero(self, card):
        # §11F. It was, and said so. It is configured now.
        fourth = next(layer for layer in card["layers"]
                      if layer["number"] == 4)
        assert fourth["components"]
        assert fourth["tested"] >= 1

    def test_the_mapping_matches_the_published_methodology(self, card):
        for layer in card["layers"]:
            for component in layer["components"]:
                assert me.layer_of(component["family"]) == layer["layer"]

    def test_no_family_falls_outside_a_layer(self):
        assert me.unmapped_families() == ()

    def test_a_layer_says_what_it_found_in_a_sentence(self, card):
        for layer in card["layers"]:
            assert layer["sentence"].endswith(".")
            assert layer["matters"]


class TestTheAssessmentComesFirst:
    def test_the_level_and_its_evidence_are_on_the_payload(self, card):
        assert card["risk_level"] in ea.LEVELS
        assert card["assessment"]["level"] == card["risk_level"]
        assert card["assessment"]["primary_concern"]

    def test_there_is_no_score(self, card):
        for forbidden in ("score", "points", "weight", "weighted"):
            assert forbidden not in card["assessment"]


class TestTheDeepLink:
    """Section 11J."""

    def test_it_carries_the_borrower_and_the_reporting_date(self, card,
                                                            standing):
        link = card["borrower_360"]
        assert link["customer_id"] == standing.borrower_id
        assert link["reporting_period"] == standing.period
        assert f"{sc.BORROWER_PARAM}={standing.borrower_id}" in link["href"]
        assert f"{sc.PERIOD_PARAM}=" in link["href"]

    def test_the_period_travels_without_a_space(self, card):
        # A URL with a space in it is a URL people mangle.
        assert " " not in card["borrower_360"]["href"]

    def test_the_parameter_names_match_the_frontend(self):
        # Two surfaces agreeing on a URL by coincidence is how a deep link
        # quietly stops working. The frontend owns these names in
        # `lib/borrower-link.ts`; this asserts the backend spells them
        # the same way.
        from pathlib import Path

        source = Path("frontend/src/lib/borrower-link.ts").read_text()
        assert f'BORROWER_PARAM = "{sc.BORROWER_PARAM}"' in source
        assert f'PERIOD_PARAM = "{sc.PERIOD_PARAM}"' in source

    def test_a_borrower_with_no_period_still_gets_a_usable_link(self):
        quiet = sg.stand({"borrower_id": "CORP-9"}, period="")
        link = sc._deep_link(quiet)
        # `?period=` asks Borrower 360 for a quarter called nothing.
        assert "period=" not in link["href"]
        assert sc.BORROWER_PARAM in link["href"]


class TestTheTimeline:
    """Section 11I."""

    @pytest.fixture(scope="class")
    @classmethod
    def timeline(cls, standing):
        from backend.corporate import service as corporate

        snapshot = corporate._load(corporate.SNAPSHOT)
        periods = sorted((str(p) for p in snapshot["period"].unique()),
                         key=sg._period_key)
        return sc.timeline(standing.borrower_id, snapshot, periods, limit=6)

    def test_it_covers_more_than_one_reporting_date(self, timeline):
        assert len(timeline["entries"]) >= 2

    def test_each_period_is_evaluated_against_its_own_row(self, timeline):
        # A timeline that repeated the latest assessment would report the
        # same firing count at every date.
        counts = {e["fired"] for e in timeline["entries"] if e["on_book"]}
        assert len(counts) > 1

    def test_every_entry_says_what_it_found(self, timeline):
        for entry in timeline["entries"]:
            assert entry["sentence"]
            if entry["on_book"]:
                assert entry["risk_level"] in ea.LEVELS

    def test_the_first_period_is_the_oldest(self, timeline):
        keys = [sg._period_key(p) for p in timeline["periods"]]
        assert keys == sorted(keys)


class TestTheWorkbooks:
    """Section 11L."""

    def test_the_borrower_pack_opens_with_the_assessment(self, standing):
        book = openpyxl.load_workbook(io.BytesIO(wb.borrower(standing)))
        assert book.sheetnames[0] == "ASSESSMENT"
        assert "SOURCE" in book.sheetnames

    def test_it_carries_one_sheet_per_layer(self, standing):
        book = openpyxl.load_workbook(io.BytesIO(wb.borrower(standing)))
        for number in (1, 2, 3, 4):
            assert f"LAYER {number}" in book.sheetnames

    def test_a_layer_sheet_carries_every_column(self, standing, card):
        book = openpyxl.load_workbook(io.BytesIO(wb.borrower(standing)))
        sheet = book["LAYER 1"]
        header = [cell.value for cell in sheet[4]]
        for name in card["columns"]:
            assert name in header, f"the workbook drops the {name} column"

    def test_the_source_sheet_names_where_every_figure_came_from(self,
                                                                 standing):
        book = openpyxl.load_workbook(io.BytesIO(wb.borrower(standing)))
        sheet = book["SOURCE"]
        rows = list(sheet.iter_rows(min_row=5, values_only=True))
        assert len(rows) == len(standing.observations)
        for row in rows:
            assert row[1], "a condition with no dataset behind it"

    def test_the_disclosure_leaves_with_the_workbook(self, standing):
        book = openpyxl.load_workbook(io.BytesIO(wb.borrower(standing)))
        said = " ".join(str(c.value) for c in book["ASSESSMENT"]["A"][:4]
                        if c.value)
        assert wb.ORIGIN in said or wb.ORIGIN in str(
            book["ASSESSMENT"]["A2"].value)

    def test_the_watchlist_publishes_its_methodology(self):
        ranked = (sg._book("").get("_ranked") or [])[:20]
        book = openpyxl.load_workbook(
            io.BytesIO(wb.watchlist(ranked, period="Q2 2026", limit=20)))
        assert book.sheetnames == ["WATCHLIST", "METHODOLOGY"]
        said = " ".join(str(c.value) for row in book["METHODOLOGY"].iter_rows()
                        for c in row if c.value)
        assert "Signal count" in said, (
            "the workbook does not say what the level deliberately ignores")

    def test_the_watchlist_row_count_matches_what_was_asked_for(self):
        ranked = (sg._book("").get("_ranked") or [])[:40]
        book = openpyxl.load_workbook(
            io.BytesIO(wb.watchlist(ranked, period="Q2 2026", limit=10)))
        # Title rows plus a header plus ten borrowers.
        assert book["WATCHLIST"].max_row == 14
