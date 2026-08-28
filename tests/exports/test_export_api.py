"""
The download endpoints: headers, roles, and every way an export can fail.

§39 says authorisation is enforced in the backend and not by hiding a button,
so these tests call the endpoints directly as each role. §44 says no failure
returns an unexplained 500, so every refusal here is asserted to carry a status
that says which refusal it was and a message a person could act on.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

RESULTS_URL = "/api/v1/analysis-runs/{run}/export/results.xlsx"
PACK_URL = "/api/v1/trace/{run}/export/calculation-pack.xlsx"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def as_role(role: str) -> dict[str, str]:
    return {"X-IPM-Role": role}


# ================================================================ §38 the download


class TestDownloadResponse:
    def test_the_results_workbook_downloads_as_a_file(self, client, rating_run):
        response = client.get(RESULTS_URL.format(run=rating_run),
                              headers=as_role("ADMIN"))
        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX
        assert response.headers["content-disposition"].startswith("attachment;")
        assert ".xlsx" in response.headers["content-disposition"]
        assert int(response.headers["content-length"]) == len(response.content)
        assert response.content[:2] == b"PK"

    def test_the_calculation_pack_downloads_as_a_file(self, client, rating_run):
        response = client.get(PACK_URL.format(run=rating_run),
                              headers=as_role("ADMIN"))
        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX
        assert load_workbook(io.BytesIO(response.content)).sheetnames[0] == "COVER"

    def test_a_workbook_is_never_cached(self, client, rating_run):
        """A cached file would be served after the analysis was re-run."""
        response = client.get(RESULTS_URL.format(run=rating_run),
                              headers=as_role("ADMIN"))
        assert "no-store" in response.headers["cache-control"]

    def test_the_filename_is_sanitised(self, client, rating_run):
        disposition = client.get(RESULTS_URL.format(run=rating_run),
                                 headers=as_role("ADMIN")
                                 ).headers["content-disposition"]
        assert not set(disposition.split('"')[1]) & set('<>:/\\|?*')

    def test_the_response_names_the_run_it_came_from(self, client, rating_run):
        response = client.get(PACK_URL.format(run=rating_run),
                              headers=as_role("ADMIN"))
        assert response.headers["x-creditprobe-run"] == str(rating_run)
        assert response.headers["x-creditprobe-trace-version"]


# ============================================================ §39 authorisation


class TestAuthorization:
    @pytest.mark.parametrize("role", ["ADMIN", "DATA_STEWARD", "ANALYST"])
    def test_every_working_role_may_download_results(self, client, rating_run, role):
        assert client.get(RESULTS_URL.format(run=rating_run),
                          headers=as_role(role)).status_code == 200

    @pytest.mark.parametrize("role", ["ADMIN", "DATA_STEWARD"])
    def test_governance_roles_may_download_the_full_pack(self, client, rating_run,
                                                         role):
        assert client.get(PACK_URL.format(run=rating_run),
                          headers=as_role(role)).status_code == 200

    def test_a_viewer_is_refused_the_full_pack(self, client, rating_run):
        response = client.get(PACK_URL.format(run=rating_run),
                              headers=as_role("VIEWER"))
        assert response.status_code == 403
        detail = response.json()["detail"]
        assert detail["error"] == "forbidden"
        assert "row-level" in detail["message"]

    def test_the_refusal_explains_what_is_available_instead(self, client,
                                                            rating_run):
        message = client.get(PACK_URL.format(run=rating_run),
                             headers=as_role("VIEWER")).json()["detail"]["message"]
        assert "results workbook" in message.lower()

    def test_hiding_the_button_is_not_the_control(self, client, rating_run):
        """A Viewer who constructs the URL by hand is still refused."""
        assert client.get(PACK_URL.format(run=rating_run),
                          headers=as_role("VIEWER")).status_code == 403

    def test_availability_tells_the_interface_what_to_offer(self, client,
                                                            rating_run):
        body = client.get(
            f"/api/v1/analysis-runs/{rating_run}/export/availability",
            headers=as_role("VIEWER")).json()
        assert body["results"]["label"] == "DOWNLOAD RESULTS"
        assert body["calculation_pack"]["label"] == "DOWNLOAD FULL CALCULATION"
        assert body["calculation_pack"]["allowed"] is False
        assert body["calculation_pack"]["reason"]

    def test_availability_matches_what_the_endpoint_does(self, client, rating_run):
        for role in ("ADMIN", "DATA_STEWARD", "ANALYST", "VIEWER"):
            body = client.get(
                f"/api/v1/analysis-runs/{rating_run}/export/availability",
                headers=as_role(role)).json()
            served = client.get(PACK_URL.format(run=rating_run),
                                headers=as_role(role)).status_code
            assert body["calculation_pack"]["allowed"] is (served == 200), role


# ============================================================== §44 failure UX


class TestFailures:
    def test_a_run_that_does_not_exist_is_a_404(self, client):
        response = client.get(RESULTS_URL.format(run=99999999),
                              headers=as_role("ADMIN"))
        assert response.status_code == 404
        assert response.json()["detail"]["error"] == "run_not_found"

    def test_a_clarification_is_not_an_error(self, client):
        """A question CreditProbe stopped to ask about has no result to export."""
        response = client.post("/api/v1/ask",
                               json={"question": "Show me exposure."},
                               headers=as_role("ADMIN"))
        body = response.json()
        if body.get("status") != "needs_clarification" or not body.get("analysis_run_id"):
            pytest.skip("this question did not produce a persisted clarification")
        refused = client.get(RESULTS_URL.format(run=body["analysis_run_id"]),
                             headers=as_role("ADMIN"))
        assert refused.status_code == 409
        assert refused.json()["detail"]["error"] == "not_exportable"
        assert "clarification" in refused.json()["detail"]["message"]

    def test_no_failure_is_an_unexplained_500(self, client, rating_run):
        for url, role in (
            (RESULTS_URL.format(run=99999999), "ADMIN"),
            (PACK_URL.format(run=99999999), "ADMIN"),
            (PACK_URL.format(run=rating_run), "VIEWER"),
        ):
            response = client.get(url, headers=as_role(role))
            assert response.status_code != 500
            detail = response.json()["detail"]
            assert detail["error"] and detail["message"]


# =============================================================== §41 audit log


class TestAuditLog:
    def test_a_download_is_recorded(self, client, rating_run):
        before = client.get(f"/api/v1/analysis-runs/{rating_run}/export/history",
                            headers=as_role("ADMIN")).json()["exports"]
        client.get(RESULTS_URL.format(run=rating_run), headers=as_role("ADMIN"))
        after = client.get(f"/api/v1/analysis-runs/{rating_run}/export/history",
                           headers=as_role("ADMIN")).json()["exports"]
        assert len(after) == len(before) + 1
        latest = after[0]
        assert latest["kind"] == "results"
        assert latest["status"] == "allowed"
        assert latest["content_hash"]
        assert latest["filename"].endswith(".xlsx")
        assert latest["run_id"] == rating_run

    def test_a_refusal_is_recorded_too(self, client, rating_run):
        """A log that only records successes cannot answer 'who tried'."""
        client.get(PACK_URL.format(run=rating_run), headers=as_role("VIEWER"))
        history = client.get(
            f"/api/v1/analysis-runs/{rating_run}/export/history",
            headers=as_role("ADMIN")).json()["exports"]
        denied = [e for e in history if e["status"] == "denied"]
        assert denied, "a refused download left no trace"
        assert denied[0]["authorization"]
        assert denied[0]["reason"]

    def test_the_record_says_which_data_was_in_the_file(self, client, rating_run,
                                                        rating_pack):
        client.get(PACK_URL.format(run=rating_run), headers=as_role("ADMIN"))
        history = client.get(
            f"/api/v1/analysis-runs/{rating_run}/export/history",
            headers=as_role("ADMIN")).json()["exports"]
        pack = next(e for e in history
                    if e["kind"] == "calculation_pack" and e["status"] == "allowed")
        assert set(pack["datasets"]) == {s.dataset for s in rating_pack.sources}
        assert pack["row_count"] == len(rating_pack.rows)
        assert pack["trace_version"] == rating_pack.version

    def test_the_history_is_not_public(self, client, rating_run):
        assert client.get(
            f"/api/v1/analysis-runs/{rating_run}/export/history",
            headers=as_role("VIEWER")).status_code == 403


# ======================================================== §42 reproducibility


class TestReproducibility:
    def test_the_same_run_exports_the_same_analytical_content(self, client,
                                                              rating_run):
        """Two downloads differ in their timestamp and downloader, not their figures."""
        first = client.get(RESULTS_URL.format(run=rating_run),
                           headers=as_role("ADMIN")).content
        second = client.get(RESULTS_URL.format(run=rating_run),
                            headers=as_role("ADMIN")).content

        def table(content: bytes) -> list:
            ws = load_workbook(io.BytesIO(content))["RESULTS"]
            return [[ws.cell(row=r, column=c).value for c in range(1, 8)]
                    for r in range(1, 40)]

        assert table(first) == table(second)

    def test_a_named_trace_version_is_honoured(self, client, rating_run,
                                               rating_pack):
        response = client.get(
            RESULTS_URL.format(run=rating_run) + f"?version={rating_pack.version}",
            headers=as_role("ADMIN"))
        assert response.status_code == 200
        assert response.headers["x-creditprobe-trace-version"] == str(
            rating_pack.version)
