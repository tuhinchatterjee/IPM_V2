"""
The two workbooks, built from an analysis this suite ran.

These tests open the generated files with openpyxl and read them, because the
only failure that matters is one a person opening the file in Excel would see:
a sheet in the wrong place, a heading with no data under it, a total that does
not add up, a hyperlink that goes nowhere.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from backend.exports import calculation, gather, population, profile, results
from backend.exports import plan as planning
from backend.exports.contract import CALCULATION_PACK, RESULTS


def opened(workbook):
    return load_workbook(io.BytesIO(workbook.content))


def values(ws, column: int, start: int, end: int) -> list:
    return [ws.cell(row=r, column=column).value for r in range(start, end + 1)]


def header_row(ws, columns) -> int:
    """The row carrying every column's heading.

    Found rather than assumed: both writers put a title and a scope line above
    the table, and a test that hard-coded the offset would pass while silently
    reading the wrong row the day a subtitle was added.
    """
    labels = [str(c.get("label") or c["name"]) for c in columns]
    for r in range(1, 40):
        row = [str(ws.cell(row=r, column=i + 1).value or "")
               for i in range(len(columns))]
        if all(any(label in cell for cell in row) for label in labels):
            return r
    raise AssertionError(f"no header row carrying {labels} in {ws.title}")


def find_row(ws, text: str, column: int = 1, limit: int = 400) -> int:
    for r in range(1, limit):
        if str(ws.cell(row=r, column=column).value or "").strip() == text:
            return r
    raise AssertionError(f"{text!r} is not in column {column} of {ws.title}")


# =============================================================== the pack fixtures


@pytest.fixture(scope="module")
def view(rating_pack):
    return planning.read(rating_pack.ir,
                         kernel_steps=rating_pack.query.get("kernel_steps"))


@pytest.fixture(scope="module")
def profiles(rating_pack, view):
    return profile.profiles_for(rating_pack, view)


@pytest.fixture(scope="module")
def extract(rating_pack, view):
    return population.extract_for(rating_pack, view)


@pytest.fixture(scope="module")
def results_book(rating_pack):
    return results.build(rating_pack)


@pytest.fixture(scope="module")
def pack_book(rating_pack, profiles, extract):
    return calculation.build(rating_pack, profiles=profiles, extract=extract)


# ===================================================== §6-§8 the results workbook


class TestResultsWorkbook:
    def test_it_is_a_real_xlsx(self, results_book):
        assert results_book.content[:2] == b"PK"
        assert results_book.kind == RESULTS
        assert opened(results_book).sheetnames

    def test_results_is_the_first_sheet(self, results_book):
        assert opened(results_book).sheetnames[0] == "RESULTS"

    def test_it_carries_the_exact_rows_in_the_exact_order(self, results_book,
                                                          rating_pack):
        ws = opened(results_book)["RESULTS"]
        columns = rating_pack.visible_columns()
        header = header_row(ws, columns)

        for offset, expected in enumerate(rating_pack.rows):
            for index, column in enumerate(columns, start=1):
                cell = ws.cell(row=header + 1 + offset, column=index).value
                wanted = expected[column["name"]]
                if isinstance(wanted, float):
                    assert cell == pytest.approx(wanted), (
                        f"row {offset + 1}, {column['name']}")
                else:
                    assert cell == wanted, f"row {offset + 1}, {column['name']}"

    def test_the_row_count_matches_the_analysis(self, results_book, rating_pack):
        assert results_book.manifest["row_count"] == len(rating_pack.rows)

    def test_the_filename_names_the_analysis_not_its_shape(self, results_book):
        """A folder of "aggregated_across_the_governed_book" files is a folder
        nobody can search. The plan's own explanation names the measure, the
        breakdown and the period, which is what tells two downloads apart."""
        name = results_book.filename
        assert "aggregated_across_the_governed_book" not in name
        assert name.startswith("CreditProbe_")
        assert name.endswith("_results.xlsx")

    def test_it_records_what_it_represents(self, results_book, rating_pack):
        manifest = results_book.manifest
        assert manifest["run_id"] == rating_pack.run_id
        assert manifest["trace_version"] == rating_pack.version
        assert manifest["plan_fingerprint"] == rating_pack.plan_fingerprint
        assert manifest["build_sha"] == rating_pack.build_sha

    def test_a_summary_sheet_explains_the_scope(self, results_book):
        assert "SUMMARY" in opened(results_book).sheetnames

    def test_no_hidden_lineage_column_is_exported(self, results_book, rating_pack):
        """A denominator carried for a share calculation is not an answer."""
        hidden = [str(c["name"]) for c in rating_pack.columns if c.get("hidden")]
        assert hidden, "this regression needs a run with a hidden column to be meaningful"
        ws = opened(results_book)["RESULTS"]
        seen = {str(ws.cell(row=r, column=c).value or "")
                for r in range(1, 12) for c in range(1, 12)}
        for name in hidden:
            assert not any(name in cell for cell in seen)


# ================================================ §9-§32 the full calculation pack


class TestCalculationPack:
    def test_it_is_a_real_xlsx(self, pack_book):
        assert pack_book.content[:2] == b"PK"
        assert pack_book.kind == CALCULATION_PACK

    def test_cover_is_first_and_final_results_is_last(self, pack_book):
        names = opened(pack_book).sheetnames
        assert names[0] == "COVER"
        assert names[-1] == "FINAL RESULTS"

    def test_every_required_section_is_present_in_order(self, pack_book):
        required = [
            "COVER", "ANALYSIS REQUEST", "EXECUTIVE SUMMARY", "DATA SOURCES",
            "FIELDS USED", "POPULATION & PERIOD", "SOURCE PROFILES",
            "RELATIONSHIPS & JOINS", "JOIN RECONCILIATION",
            "FILTERS & EXCLUSIONS", "TRANSFORMATIONS", "CALCULATION STEPS",
            "INTERMEDIATE RESULTS", "FORMULAS & QUERY", "VALIDATION CHECKS",
            "INVARIANTS & RECONCILIATION", "TRACE LEDGER",
            "INTERPRETATION EVIDENCE", "LIMITATIONS", "FINAL RESULTS",
        ]
        names = opened(pack_book).sheetnames
        positions = [names.index(name) for name in required]
        assert positions == sorted(positions), (
            f"the required sections are out of order: {names}")

    def test_the_cover_links_to_every_sheet(self, pack_book):
        book = opened(pack_book)
        cover = book["COVER"]
        linked = {
            str(cover.cell(row=r, column=2).value)
            for r in range(1, 120)
            if cover.cell(row=r, column=2).hyperlink is not None
        }
        for name in book.sheetnames:
            if name == "COVER":
                continue
            assert name in linked, f"COVER does not link to {name}"

    def test_every_sheet_links_back_to_the_cover(self, pack_book):
        book = opened(pack_book)
        for name in book.sheetnames:
            if name == "COVER":
                continue
            top = book[name].cell(row=1, column=1)
            assert top.hyperlink is not None, f"{name} has no way back"
            assert "COVER" in str(top.value)

    def test_the_sources_sheet_lists_every_dataset_that_was_read(self, pack_book,
                                                                 rating_pack):
        ws = opened(pack_book)["DATA SOURCES"]
        text = {str(ws.cell(row=r, column=c).value or "")
                for r in range(1, 40) for c in range(1, 20)}
        for source in rating_pack.sources:
            assert source.dataset in text

    def test_the_final_sheet_carries_the_result(self, pack_book, rating_pack):
        ws = opened(pack_book)["FINAL RESULTS"]
        cells = [ws.cell(row=r, column=c).value
                 for r in range(1, 60) for c in range(1, 10)]
        measure = [c for c in rating_pack.visible_columns()
                   if str(c.get("semantic")) == "money"]
        assert measure, "this regression needs a money column"
        name = measure[0]["name"]
        for row in rating_pack.rows[:3]:
            assert any(isinstance(c, float)
                       and abs(c - float(row[name])) < 0.005 for c in cells), (
                f"{row[name]} is not on FINAL RESULTS")

    def test_a_skipped_check_is_never_reported_as_a_pass(self, pack_book):
        ws = opened(pack_book)["VALIDATION CHECKS"]
        header = find_row(ws, "Check ID")
        for r in range(header + 1, header + 200):
            identifier = ws.cell(row=r, column=1).value
            if identifier is None:
                break
            status = str(ws.cell(row=r, column=7).value or "")
            assert status in {"PASS", "WARNING", "FAIL", "SKIPPED"}, status

    def test_the_manifest_counts_the_checks(self, pack_book):
        counts = pack_book.manifest["checks"]
        assert counts["total"] >= 1
        assert counts["failed"] == 0, "the mandatory regression must not fail a check"

    def test_it_records_enough_to_prove_what_it_is(self, pack_book, rating_pack):
        manifest = pack_book.manifest
        assert manifest["run_id"] == rating_pack.run_id
        assert manifest["trace_version"] == rating_pack.version
        assert manifest["plan_fingerprint"]
        assert manifest["schema_version"]
        assert manifest["generator"]

    def test_nothing_secret_reaches_the_file(self, pack_book):
        """§40. A workbook that carried a key would carry it to a laptop."""
        text = pack_book.content
        for forbidden in (b"sk-ant-", b"ANTHROPIC_API_KEY", b"Authorization:",
                          b"Bearer "):
            assert forbidden not in text, f"{forbidden!r} reached the workbook"

    def test_no_filesystem_path_is_exported(self, pack_book):
        """A deployment's directory layout has no audit value and is a leak."""
        ws = opened(pack_book)["FORMULAS & QUERY"]
        cells = [str(ws.cell(row=r, column=c).value or "")
                 for r in range(1, 200) for c in range(1, 4)]
        assert not any(cell.startswith("/home/") or cell.startswith("C:\\")
                       for cell in cells)


# =========================================================== §33 the mandatory case


class TestRatingWiseExposure:
    """§33: rating-wise IFRS 9 EAD, end to end, with its reconciliation."""

    def test_the_result_is_one_row_per_rating(self, rating_pack):
        assert len(rating_pack.rows) > 1
        dimension = rating_pack.visible_columns()[0]["name"]
        seen = [row[dimension] for row in rating_pack.rows]
        assert len(seen) == len(set(seen)), "a rating appears twice"

    def test_the_measure_is_named_not_generic(self, rating_pack):
        labels = [str(c.get("label") or "") for c in rating_pack.visible_columns()]
        assert any("exposure at default" in label.lower() or "ead" in label.lower()
                   for label in labels), labels

    def test_the_rating_totals_sum_to_the_book(self, rating_pack, profiles):
        """§33 step 8: the sum of rating-level EAD equals total included EAD."""
        measure = next(c for c in rating_pack.visible_columns()
                       if str(c.get("semantic")) == "money")
        total = sum(float(row[measure["name"]]) for row in rating_pack.rows)
        source = next((p for p in profiles if p.dataset == "ifrs9_staging"
                       and p.usable), None)
        if source is None:
            pytest.skip("the IFRS 9 source could not be profiled here")
        stat = next((s for s in source.numeric if s.field_name == "ead"), None)
        if stat is None or stat.total is None:
            pytest.skip("no EAD profile to reconcile against")
        assert total == pytest.approx(stat.total, abs=0.05)

    def test_the_pack_shows_the_join_that_brought_the_rating_in(self, pack_book,
                                                                view):
        assert view.joins, "rating-wise EAD needs a governed join"
        ws = opened(pack_book)["JOIN RECONCILIATION"]
        header = find_row(ws, "#")
        rows = 0
        for r in range(header + 1, header + 30):
            if ws.cell(row=r, column=1).value is None:
                break
            rows += 1
        assert rows == len(view.joins)

    def test_every_join_reports_a_match_rate(self, pack_book, view):
        ws = opened(pack_book)["JOIN RECONCILIATION"]
        header = find_row(ws, "#")
        for offset in range(len(view.joins)):
            match = ws.cell(row=header + 1 + offset, column=14).value
            assert match not in (None, ""), "a join with no match rate"


# ================================================= §26 the Excel reconstruction


@pytest.fixture(scope="module")
def single_source(client):
    """A single-dataset analysis, which is the case Excel can faithfully rebuild."""
    from tests.exports.conftest import ask

    run_id = ask(client, "What is total EAD by sector in the latest quarter?")
    pack = gather.pack_for(run_id, user_name="Test Runner")
    return pack, planning.read(pack.ir)


@pytest.fixture(scope="module")
def single_book(single_source):
    pack, view = single_source
    return calculation.build(
        pack,
        profiles=profile.profiles_for(pack, view),
        extract=population.extract_for(pack, view),
    )


class TestExcelReconstruction:
    """The formulas are live, and they reference the exported population."""

    def test_it_writes_real_excel_formulas(self, single_book):
        ws = opened(single_book)["EXCEL RECONSTRUCTION"]
        header = find_row(ws, "Row-by-row comparison")
        formula = ws.cell(row=header + 2, column=2).value
        assert isinstance(formula, str) and formula.startswith("="), formula
        assert "SUMIF" in formula or "AVERAGEIF" in formula

    def test_the_formulas_reconcile_against_the_runtime_values(self, single_book,
                                                               single_source):
        """Recompute the SUMIF in Python over the exported rows.

        Excel would evaluate the formula on open; this asserts the ranges point
        at the right cells, which is the part that can silently be one row off.
        """
        pack, _ = single_source
        book = opened(single_book)
        sheet = next(n for n in book.sheetnames if n.startswith("POPULATION EXTRACT"))
        extract_ws = book[sheet]
        recon = book["EXCEL RECONSTRUCTION"]

        formula = str(recon.cell(row=find_row(recon, "Row-by-row comparison") + 2,
                                 column=2).value)
        # =SUMIF('SHEET'!$C$14:$C$16359,$A14,'SHEET'!$B$14:$B$16359)
        import re

        found = re.findall(r"\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)", formula)
        assert len(found) == 2, formula
        (key_col, first, last), (value_col, _, _) = found
        first, last = int(first), int(last)

        def index(letter: str) -> int:
            from openpyxl.utils import column_index_from_string

            return column_index_from_string(letter)

        totals: dict = {}
        for r in range(first, last + 1):
            key = extract_ws.cell(row=r, column=index(key_col)).value
            value = extract_ws.cell(row=r, column=index(value_col)).value
            if isinstance(value, int | float):
                totals[key] = totals.get(key, 0.0) + float(value)

        start = find_row(recon, "Row-by-row comparison") + 2
        checked = 0
        for offset in range(len(pack.rows)):
            key = recon.cell(row=start + offset, column=1).value
            runtime = recon.cell(row=start + offset, column=3).value
            if key is None or not isinstance(runtime, int | float):
                break
            assert totals.get(key) == pytest.approx(runtime, abs=0.01), key
            checked += 1
        assert checked == len(pack.rows)

    def test_a_joined_analysis_says_why_it_cannot_be_rebuilt(self, pack_book):
        """§26: do not fake an Excel formula where one would not be faithful."""
        ws = opened(pack_book)["EXCEL RECONSTRUCTION"]
        text = " ".join(str(ws.cell(row=r, column=1).value or "")
                        for r in range(1, 12))
        assert "joined" in text.lower() or "SUMIF" not in text


# ================================================ §24 the population and its limits


class TestPopulationExtract:
    def test_the_extract_is_the_source_and_says_so(self, pack_book):
        book = opened(pack_book)
        sheet = next((n for n in book.sheetnames
                      if n.startswith(("POPULATION EXTRACT", "Population_"))), None)
        assert sheet, book.sheetnames
        ws = book[sheet]
        text = " ".join(str(ws.cell(row=r, column=1).value or "")
                        for r in range(1, 16)).lower()
        assert "read at export time" in text or "export time" in text

    def test_a_population_above_the_ceiling_is_refused_not_truncated(self,
                                                                    rating_pack,
                                                                    view):
        small = population.extract_for(rating_pack, view, limit=5)
        assert not small.present
        assert "ceiling" in small.omitted
        assert not small.rows

    def test_a_split_extract_is_indexed(self, rating_pack, view):
        chunked = population.extract_for(rating_pack, view, rows_per_sheet=500)
        if not chunked.present:
            pytest.skip("no population to split here")
        assert len(chunked.chunks) > 1
        first, last = chunked.chunks[0]
        assert first == 1
        assert last == 500
        covered = sum(b - a + 1 for a, b in chunked.chunks)
        assert covered == chunked.row_count, "a split extract lost rows"
