"""
The result shapes an export has to survive.

§56 lists them, and each one has broken something in a workbook writer
somewhere: an empty result that produced a sheet with a heading and nothing
under it; a certified method with no composed plan whose pack came out as
twenty blank sheets; a multi-dataset analysis whose "one dataset" assumption
put the wrong name on the cover.

Every case runs its own analysis rather than asserting against records that
happen to exist (§58).
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from backend.exports import calculation, gather, population, results
from backend.exports import plan as planning
from backend.exports import profile as profiling
from backend.exports.contract import NotExportable, RunNotFound
from tests.exports.conftest import ask


def built(pack, *, profiles=None, extract=None):
    return calculation.build(pack, profiles=profiles, extract=extract)


def opened(workbook):
    return load_workbook(io.BytesIO(workbook.content))


def packed(client, question: str):
    run_id = ask(client, question)
    pack = gather.pack_for(run_id, user_name="Test Runner")
    return pack, planning.read(pack.ir)


# ===================================================================== shapes


class TestOneDataset:
    def test_a_single_source_analysis_names_one_dataset(self, client):
        pack, view = packed(client, "What is total EAD by sector in the latest quarter?")
        assert len(view.scans) == 1
        workbook = built(pack, profiles=profiling.profiles_for(pack, view))
        assert workbook.manifest["datasets"] == [view.scans[0].dataset]

    def test_it_says_plainly_that_nothing_was_joined(self, client):
        pack, view = packed(client, "What is total EAD by sector in the latest quarter?")
        ws = opened(built(pack))["RELATIONSHIPS & JOINS"]
        text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 12))
        assert "made no joins" in text


class TestManyDatasets:
    def test_every_source_reaches_the_pack(self, client):
        pack, view = packed(
            client,
            "For each rating grade, show average ECL coverage and average DSCR.",
        )
        assert len(view.scans) > 1, "this regression needs a joined analysis"
        workbook = built(pack, profiles=profiling.profiles_for(pack, view))
        listed = set(workbook.manifest["datasets"])
        assert {s.dataset for s in view.scans} <= listed

    def test_the_population_extract_says_it_is_one_source_only(self, client):
        pack, view = packed(
            client,
            "For each rating grade, show average ECL coverage and average DSCR.",
        )
        extract = population.extract_for(pack, view)
        assert extract.stands_for_calculation is False
        assert any("more than one source" in note for note in extract.notes)


class TestCertifiedMethod:
    """A registered analysis has no composed IR, and the pack must say so."""

    def test_a_plan_with_no_operations_is_an_empty_view(self):
        view = planning.read({})
        assert view.empty
        assert view.scans == []

    def test_the_pack_explains_a_registered_method_rather_than_blanking(self):
        pack = gather.Pack(run_id=0, title="Certified", certification="certified",
                           analysis_id="portfolio_summary", analysis_version="1.2.0")
        workbook = built(pack)
        ws = opened(workbook)["CALCULATION STEPS"]
        text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 14))
        assert "certified method" in text
        assert "no step-level ledger" in text


class TestEmptyResult:
    def test_an_empty_result_is_a_conclusion_not_a_blank_sheet(self):
        pack = gather.Pack(
            run_id=0,
            title="Nothing matched",
            narrative={"direct_answer": "No facility matched those conditions."},
            columns=[{"name": "sector", "label": "Sector"}],
            rows=[],
        )
        ws = opened(results.build(pack))["RESULTS"]
        text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 8))
        assert "No facility matched" in text

    def test_the_pack_still_builds_and_still_ends_on_final_results(self):
        pack = gather.Pack(run_id=0, title="Nothing matched", rows=[],
                           columns=[{"name": "sector"}])
        assert opened(built(pack)).sheetnames[-1] == "FINAL RESULTS"


# ================================================================== failures


class TestFailureHandling:
    def test_a_run_that_does_not_exist_is_a_404_not_a_500(self):
        with pytest.raises(RunNotFound) as raised:
            gather.pack_for(99999999)
        assert raised.value.status == 404

    def test_a_clarification_is_explained_rather_than_exported(self):
        from backend.orchestration import store

        original = store.load_version
        try:
            store.load_version = lambda *a, **k: {"status": "needs_clarification"}
            with pytest.raises(NotExportable) as raised:
                gather.pack_for(1)
            assert raised.value.status == 409
            assert "clarification" in raised.value.message
        finally:
            store.load_version = original

    def test_a_refusal_is_explained_rather_than_exported(self):
        from backend.orchestration import store

        original = store.load_version
        try:
            store.load_version = lambda *a, **k: {"status": "rejected"}
            with pytest.raises(NotExportable) as raised:
                gather.pack_for(1)
            assert "refused" in raised.value.message
        finally:
            store.load_version = original

    def test_a_source_that_cannot_be_profiled_does_not_lose_the_export(self):
        found = profiling.profile_one("no_such_dataset_at_all", period="Q2 2026")
        assert not found.usable
        assert found.error
        # And the pack still builds around it, saying why the profile is missing.
        pack = gather.Pack(run_id=0, title="Test", rows=[{"a": 1}],
                           columns=[{"name": "a"}])
        ws = opened(built(pack, profiles=[found]))["SOURCE PROFILES"]
        text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, 20))
        assert "could not be" in text


# ============================================================ §55 the details


@pytest.fixture(scope="module")
def sector(client):
    """One simple, ordered, single-source result — the case §55 describes."""
    return packed(client, "What is total EAD by sector in the latest quarter?")


class TestResultsFidelity:
    def test_the_sort_order_is_the_analysis_own(self, sector):
        pack, _ = sector
        measure = next(c for c in pack.visible_columns()
                       if str(c.get("semantic")) == "money")
        ws = opened(results.build(pack))["RESULTS"]
        header = _header_row(ws, pack.visible_columns())
        column = [c["name"] for c in pack.visible_columns()].index(measure["name"]) + 1
        written = [
            ws.cell(row=header + 1 + i, column=column).value
            for i in range(len(pack.rows))
        ]
        expected = [row[measure["name"]] for row in pack.rows]
        assert written == pytest.approx(expected)
        assert written == sorted(written, reverse=True), "the exported order moved"

    def test_units_are_in_the_headers_and_not_in_every_cell(self, sector):
        pack, _ = sector
        ws = opened(results.build(pack))["RESULTS"]
        header = _header_row(ws, pack.visible_columns())
        headers = [
            str(ws.cell(row=header, column=i + 1).value or "")
            for i in range(len(pack.visible_columns()))
        ]
        assert any("SAR mn" in h for h in headers)
        first = ws.cell(row=header + 1, column=2).value
        assert isinstance(first, int | float), "a figure was written as text"

    def test_money_carries_a_number_format(self, sector):
        pack, _ = sector
        ws = opened(results.build(pack))["RESULTS"]
        header = _header_row(ws, pack.visible_columns())
        assert ws.cell(row=header + 1, column=2).number_format != "General"

    def test_the_summary_records_the_scope_and_the_provenance(self, sector):
        pack, _ = sector
        ws = opened(results.build(pack))["SUMMARY"]
        text = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, 40) for c in (1, 2)
        )
        assert pack.question in text
        assert pack.plan_fingerprint in text

    def test_a_chart_sheet_references_the_result_and_invents_nothing(self, sector):
        pack, _ = sector
        workbook = results.build(pack)
        described = workbook.manifest.get("chart")
        if not described:
            pytest.skip("this result was not charted")
        assert "referencing RESULTS" in described


def _header_row(ws, columns) -> int:
    labels = [str(c.get("label") or c["name"]) for c in columns]
    for r in range(1, 30):
        row = [str(ws.cell(row=r, column=i + 1).value or "")
               for i in range(len(columns))]
        if all(any(label in cell for cell in row) for label in labels):
            return r
    raise AssertionError(f"no header row carrying {labels}")
