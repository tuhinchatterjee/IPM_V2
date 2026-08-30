"""Continuous Learning. §56-§93.

Almost every test here is about a claim the system is not allowed to make.
That is the shape of the subsystem: measuring improvement is easy, and
measuring it honestly means refusing four or five very natural ways to
report it well.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta

import pytest

from backend.continuous import measurement as ms
from backend.continuous import partitions as pt
from backend.continuous import snapshots as sn

# ==================================================== §58 the three sets


def test_the_three_partitions_each_say_what_they_are_for():
    assert len(pt.PARTITIONS) == 3
    for name in pt.PARTITIONS:
        assert len(pt.MEANS[name]) > 40, name
        assert pt.USED_FOR[name]


def test_only_development_may_be_tuned_against():
    """§58: the validation set is not used to optimise individual fixes.
    Tuning against it destroys the only thing it can tell you."""
    assert pt.tuning_allowed(pt.DEVELOPMENT)[0] is True

    allowed, why = pt.tuning_allowed(pt.VALIDATION)
    assert allowed is False
    assert "generalised" in why

    allowed, why = pt.tuning_allowed(pt.SEALED_HOLDOUT)
    assert allowed is False
    assert "certifies its own tuning" in why


@pytest.mark.parametrize("audience", [a for a, _ in pt.NEVER_EXPOSE_TO])
def test_sealed_holdout_content_reaches_none_of_section_58s_six(audience):
    allowed, why = pt.may_expose(pt.SEALED_HOLDOUT, audience)

    assert allowed is False
    assert audience in why


def test_certification_does_not_open_the_holdout():
    """Certification establishes that a score is meaningful. It does not
    establish that the exam may be circulated."""
    assert pt.may_expose(pt.SEALED_HOLDOUT, "continuous_learning_ui",
                         certified=True)[0] is False


def test_an_unknown_audience_is_refused_rather_than_allowed():
    """A typo in a caller should not open the holdout."""
    allowed, why = pt.may_expose(pt.SEALED_HOLDOUT, "some_new_screen")

    assert allowed is False
    assert "unknown audience is refused" in why


def test_only_aggregate_fields_survive_the_filter():
    """An allowlist, not a blocklist. The field that leaks the questions
    will be called something nobody thought of."""
    filtered = pt.aggregate({
        "score": 0.91, "case_count": 300,
        "question": "what is total ECL?", "gold_answer": "4.2bn",
        "case_ids": ["h-1", "h-2"],
    })

    assert filtered == {"score": 0.91, "case_count": 300}


# ===================================================== §72 hygiene


def _uses(partition: str, count: int, *, when=None) -> list[pt.Use]:
    at = when or datetime.now(UTC)
    return [pt.Use(partition=partition, at=at) for _ in range(count)]


def test_validation_run_too_often_is_flagged():
    report = pt.hygiene(_uses(pt.VALIDATION, 12)
                        + _uses(pt.DEVELOPMENT, 40))

    assert report.healthy is False
    assert any("second development set" in f for f in report.findings)


def test_validation_run_as_often_as_development_is_one_set_doing_both_jobs():
    report = pt.hygiene(_uses(pt.VALIDATION, 6) + _uses(pt.DEVELOPMENT, 8))

    assert any("one set being used for both" in f for f in report.findings)


def test_running_the_sealed_holdout_repeatedly_is_flagged():
    """§58: do not run the sealed holdout every hour. Each run spends some
    of what makes it meaningful."""
    report = pt.hygiene(_uses(pt.SEALED_HOLDOUT, 5))

    assert any("spends some of what makes it meaningful" in f
               for f in report.findings)


def test_development_measured_with_no_validation_check_is_flagged():
    report = pt.hygiene(_uses(pt.DEVELOPMENT, 10))

    assert any("not evidence that anything generalised" in f
               for f in report.findings)


def test_old_uses_fall_outside_the_window():
    old = datetime.now(UTC) - timedelta(days=90)
    report = pt.hygiene(_uses(pt.VALIDATION, 40, when=old))

    assert report.validation_runs == 0


# =============================================== §61 the three forms


def test_section_61s_worked_example_reproduces_exactly():
    """82.0% → 88.5%."""
    change = ms.Change("Understanding & Context", 0.82, 0.885, cases=200)

    assert change.points == 6.5
    assert change.relative == 7.93
    assert change.error_reduction == 36.11


def test_a_small_base_makes_the_relative_figure_flattering():
    """A 2 pp move reported as a 40% improvement is the oldest trick, and
    the reason all three forms are shown together."""
    change = ms.Change("x", 0.05, 0.07, cases=200)

    assert change.points == 2.0
    assert change.relative == 40.0
    assert "percentage points" in change.sentence()


def test_too_few_cases_refuses_to_report_a_percentage():
    change = ms.Change("x", 0.5, 0.9, cases=6)

    assert change.evidence == ms.NO_EVIDENCE
    assert change.verdict == ms.INSUFFICIENT_EVIDENCE
    assert "not a small improvement" in change.sentence()


def test_every_sentence_carries_its_sample():
    """§77: do not claim 'improved 12%' without showing sample context."""
    change = ms.Change("x", 0.8, 0.92, cases=150)

    assert "150 case(s)" in change.sentence()
    assert change.evidence in ms.EVIDENCE_LEVELS


def test_a_critical_regression_makes_the_verdict_regressed():
    change = ms.Change("x", 0.6, 0.95, cases=400, critical_introduced=2)

    assert change.points == 35.0
    assert change.verdict == ms.REGRESSED


# ============================================ §62 the six dimensions


def _dimension(dev_after: float, val_after: float, *, cases: int = 200,
               critical: int = 0, days: int = 0) -> ms.DimensionResult:
    return ms.DimensionResult(
        dimension="Computation & Evidence",
        development=ms.Change("dev", 0.80, dev_after, cases=cases),
        validation=ms.Change("val", 0.80, val_after, cases=cases,
                             critical_introduced=critical),
        days_since_run=days)


def test_development_improvement_validation_does_not_confirm_is_mixed():
    """Development is the set that was tuned against. A screen taking its
    verdict would report every round of tuning as a win."""
    assert _dimension(0.88, 0.801).verdict == ms.MIXED


def test_both_partitions_agreeing_is_an_improvement():
    assert _dimension(0.88, 0.87).verdict == ms.IMPROVED


def test_a_critical_validation_regression_outranks_everything():
    assert _dimension(0.95, 0.94, critical=1).verdict == ms.REGRESSED


def test_an_old_measurement_is_stale_rather_than_current():
    assert _dimension(0.88, 0.87, days=60).verdict == ms.STALE


def test_the_mixed_sentence_says_which_number_to_believe():
    reading = _dimension(0.88, 0.801).to_dict()["reads_as"]

    assert "validation figure is the one to believe" in reading


# ============================================== §63 quantity vs quality


def test_capture_with_no_measured_lift_says_so_in_section_63s_words():
    verdict = ms.quality_verdict(
        quantity={"new_observations": 400, "new_approved_cases": 60},
        dimensions=[])

    assert verdict["headline"] == (
        "MORE KNOWLEDGE CAPTURED — NO MEASURED PERFORMANCE IMPROVEMENT YET")


def test_quantity_and_quality_are_never_one_number():
    verdict = ms.quality_verdict(quantity={"new_observations": 400},
                                 dimensions=[_dimension(0.88, 0.87)])

    assert "learning_quantity" in verdict
    assert "learning_quality" in verdict
    assert "Adding cases is not improving" in verdict["why_they_are_separate"]


def test_a_regression_is_the_finding_whatever_the_others_did():
    verdict = ms.quality_verdict(
        quantity={},
        dimensions=[_dimension(0.9, 0.9), _dimension(0.9, 0.7, critical=1)])

    assert "regressed on validation" in verdict["headline"]
    assert "whatever the others did" in verdict["headline"]


# ================================================== §76 overfitting


def test_section_76s_worked_example_is_possible_overfitting():
    """Development +8 pp, validation +0.5 pp."""
    drift = ms.overfitting([_dimension(0.88, 0.805)])

    assert drift.development_points == 8.0
    assert drift.validation_points == 0.5
    assert drift.suspected is True
    assert "has not been shown to generalise" in drift.to_dict()[
        "recommended_review"]


def test_validation_regressing_while_development_improves_is_the_clearest_case():
    drift = ms.overfitting([_dimension(0.88, 0.78)])

    assert drift.suspected is True
    assert "specific to the cases they were written against" in \
        drift.to_dict()["recommended_review"]


def test_both_moving_together_is_generalisation_not_overfitting():
    drift = ms.overfitting([_dimension(0.88, 0.87)])

    assert drift.suspected is False
    assert "what generalisation looks like" in drift.to_dict()[
        "recommended_review"]


def test_a_critical_validation_regression_blocks_activation():
    """§76's closing line. Not weighed against the improvements."""
    ok, why = ms.may_activate([_dimension(0.99, 0.99, critical=1)])

    assert ok is False
    assert "wrong answer the bank would have shown a client" in why


def test_activation_is_allowed_when_validation_is_clean():
    assert ms.may_activate([_dimension(0.88, 0.87)])[0] is True


# ================================================== §78 the waterfall


def test_only_isolated_contributions_are_attributed():
    """§78: only use additive attribution where isolated evaluations
    support it."""
    result = ms.waterfall(
        starting=0.80, ending=0.88,
        contributions=[
            ms.Contribution("Teaching Cases", 5.0, isolated=True),
            ms.Contribution("Brain imports", 9.0),
        ])

    bars = {b["source"]: b["points"] for b in result["bars"]}
    assert bars["Teaching Cases"] == 5.0
    assert "Brain imports" not in bars
    assert "Brain imports" in result["sources_not_isolated"]


def test_the_residual_is_named_rather_than_distributed():
    """A waterfall that always balances is one somebody made balance."""
    result = ms.waterfall(starting=0.80, ending=0.88, contributions=[
        ms.Contribution("Teaching Cases", 3.0, isolated=True)])

    residual = next(b for b in result["bars"]
                    if b["source"] == "UNATTRIBUTED / INTERACTION")
    assert residual["points"] == 5.0
    assert "made balance" in result["why_there_is_a_residual"]


def test_a_regression_bar_is_negative():
    result = ms.waterfall(
        starting=0.80, ending=0.82,
        contributions=[ms.Contribution("Teaching Cases", 4.0,
                                       isolated=True)],
        regressions=[ms.Contribution("Routing/model changes", 2.0,
                                     isolated=True)])

    bars = {b["source"]: b["points"] for b in result["bars"]}
    assert bars["Routing/model changes"] == -2.0


def test_the_eight_attribution_sources_are_section_78s():
    assert "Teaching Cases" in ms.SOURCES
    assert "Brain imports" in ms.SOURCES
    assert "Regulatory learning" in ms.SOURCES
    assert len(ms.SOURCES) == 7


# =============================================== §57/§59/§60 records


def _baseline(**over) -> sn.Baseline:
    fields = {"instance_id": "inst-1", "build_sha": "abc123",
              "development_set_version": "dev-1",
              "validation_set_version": "val-1",
              "six_dimension_scores": {"Computation & Evidence": 0.82}}
    fields.update(over)
    return sn.Baseline(**fields)


def test_a_baseline_says_what_a_comparison_against_it_establishes():
    """"Compared to what?" is answered by an ontology version and a
    case-set version, not by a date."""
    baseline = _baseline(ontology_version="2.0.0",
                         teaching_release_id="tr-9")

    assert "ontology 2.0.0" in baseline.comparable_to
    assert "development set dev-1" in baseline.comparable_to


def test_a_baseline_with_no_case_set_version_is_refused():
    problems = sn.validate_baseline(_baseline(development_set_version=""))

    assert any("oldest way to report one" in p for p in problems)


def test_a_baseline_with_no_build_cannot_separate_improving_from_deploying():
    problems = sn.validate_baseline(_baseline(build_sha=""))

    assert any("'we improved' from 'we deployed'" in p for p in problems)


def test_a_baseline_never_carries_holdout_content():
    body = _baseline(sealed_holdout_version="hold-3").to_dict()

    assert body["sealed_holdout_version"] == "hold-3"
    assert body["sealed_holdout_content_included"] is False


def test_a_snapshot_with_no_baseline_is_refused():
    """A number with no reference point gets compared to whichever earlier
    number flatters it."""
    problems = sn.validate_snapshot(sn.Snapshot())

    assert any("flatters it" in p for p in problems)


def test_validation_scores_without_a_set_version_are_refused():
    problems = sn.validate_snapshot(sn.Snapshot(
        comparison_baseline_id="b-1",
        validation_scores={"x": 0.9}))

    assert any("not comparable" in p for p in problems)


def test_the_thirteen_windows_are_section_60s():
    assert len(sn.WINDOWS) == 13
    assert sn.SINCE_INSTALLATION in sn.ANCHORED
    assert sn.LAST_7_DAYS not in sn.ANCHORED


def test_an_anchored_window_without_its_anchor_is_refused():
    """Answering "since the current Brain" with "the last thirty days" is
    the kind of wrong nobody catches."""
    with pytest.raises(sn.SnapshotError) as caught:
        sn.window_bounds(sn.SINCE_CURRENT_BRAIN)

    assert "different question" in str(caught.value)


def test_all_time_has_no_start():
    start, end = sn.window_bounds(sn.ALL_TIME)

    assert start is None
    assert end is not None


def test_captured_and_changed_are_two_blocks_that_are_never_added():
    baseline = _baseline()
    snapshot = sn.Snapshot(
        comparison_baseline_id=baseline.baseline_id,
        development_set_version="dev-1", validation_set_version="val-1",
        six_dimension_scores_dev={"Computation & Evidence": 0.885},
        new_learning_captured=400)

    body = sn.compare(baseline, snapshot)

    assert body["learning_captured_during_window"]["new_observations"] == 400
    assert body["performance_change_during_window"]["development"][0][
        "points"] == 6.5
    assert "never added" in body["these_are_not_the_same_thing"]


def test_a_comparison_across_different_case_sets_says_so():
    baseline = _baseline()
    snapshot = sn.Snapshot(comparison_baseline_id=baseline.baseline_id,
                           development_set_version="dev-2",
                           validation_set_version="val-1")

    assert sn.compare(baseline, snapshot)["case_sets_match"] is False


def test_an_unmoved_dimension_is_reported_rather_than_omitted():
    """A dimension omitted because it did not move reads as a dimension
    that was not measured, and the two mean opposite things."""
    baseline = _baseline(six_dimension_scores={"A": 0.8, "B": 0.9})
    snapshot = sn.Snapshot(comparison_baseline_id=baseline.baseline_id,
                           development_set_version="dev-1",
                           six_dimension_scores_dev={"A": 0.85})

    rows = sn.compare(baseline, snapshot)[
        "performance_change_during_window"]["development"]
    by_name = {r["dimension"]: r for r in rows}

    assert by_name["B"]["measured"] is False
    assert by_name["A"]["measured"] is True


def test_the_eleven_triggers_separate_change_from_the_clock():
    """A snapshot taken because something changed is worth comparing
    against; a daily one taken while nothing happened is noise."""
    assert len(sn.TRIGGERS) == 11
    assert sn.BRAIN_IMPORT in sn.CHANGE_TRIGGERS
    assert sn.DAILY not in sn.CHANGE_TRIGGERS


# ==================================================== §66 velocity


def test_captured_and_activated_are_separate_rates():
    """Forty observations a week with nothing activated is a backlog, and
    the capture rate alone reads as a healthy learning system."""
    snapshots = [sn.Snapshot(comparison_baseline_id="b",
                             new_learning_captured=300,
                             new_learning_activated=0)]

    rate = ms.velocity(snapshots, days=30)

    assert rate["captured_per_day"] == 10.0
    assert rate["activated_per_day"] == 0.0
    assert rate["conversion"] == 0.0
    assert "backlog" in rate["note"]


# ==================================================== §67 attribution


def test_attribution_is_named_as_belief_rather_than_cause():
    """A list of what happened at the same time is a list of suspects."""
    body = ms.attribution(_dimension(0.88, 0.87))

    assert body["established"] is False
    assert "list of suspects" in body["note"]


# ==================================================== §87 staleness


def test_the_nine_staleness_axes_each_say_what_they_affect():
    """"Stale" tells a reader to re-run something. "Stale, the ontology
    changed" tells them what to re-run and roughly how much to expect."""
    from backend.continuous import staleness as st

    assert len(st.AXES) == 9
    for axis, what, affects in st.AXES:
        assert what, axis
        assert len(affects) > 30, axis


def test_a_changed_ontology_makes_the_display_stale():
    from backend.continuous import staleness as st

    report = st.assess({"ontology_version": "2.0.0"},
                       {"ontology_version": "2.1.0"})

    assert report.stale is True
    assert report.label == st.STALE
    assert report.to_dict()["changed_axes"][0]["axis"] == "ontology_version"


def test_a_changed_evaluation_set_is_incomparable_rather_than_merely_stale():
    """Two scores over two different case sets are not comparable, and the
    difference between them is not improvement."""
    from backend.continuous import staleness as st

    report = st.assess({"development_set_version": "d1"},
                       {"development_set_version": "d2"})

    assert report.label == st.INCOMPARABLE_LABEL
    assert any("not improvement" in f for f in report.findings)


def test_an_unknown_version_counts_as_changed_rather_than_matching():
    """"We do not know what it was" and "it is the same" are different, and
    defaulting towards CURRENT leaves a stale number on screen unlabelled."""
    from backend.continuous import staleness as st

    assert st.assess({}, {"brain_version": "2.0"}).stale is True


def test_a_mapping_version_is_compared_by_content_not_identity():
    """Comparing prompt_versions by identity would make every reload look
    like a change."""
    from backend.continuous import staleness as st

    assert st.assess({"prompt_versions": {"planner": "3", "critic": "2"}},
                     {"prompt_versions": {"critic": "2", "planner": "3"}}
                     ).stale is False


def test_staleness_never_rewrites_the_snapshot():
    """§87: the historical snapshot remains immutable. Staleness is a label
    on what is displayed now."""
    from backend.continuous import staleness as st

    body = st.assess({"build_sha": "a"}, {"build_sha": "b"}).to_dict()

    assert body["historical_snapshot_unchanged"] is True
    assert "still true about the moment it was taken" in body["note"]


# ==================================================== §83 the report


def _payload(**over):
    body = {
        "window": "LAST_30_DAYS",
        "headline": "MORE KNOWLEDGE CAPTURED — NO MEASURED PERFORMANCE "
                    "IMPROVEMENT YET",
        "baseline": {"baseline_id": "base-1",
                     "comparable_to": "ontology 2.0.0",
                     "created_at": "2026-01-01T00:00:00+00:00"},
        "learning_captured": {"new_observations": 400},
        "measured_change": {},
        "dimensions": [],
        "windows_available": ["LAST_30_DAYS"],
        "these_are_not_the_same_thing": "Captured is what went in.",
    }
    body.update(over)
    return body


def test_the_report_has_all_twenty_sheets_even_when_empty():
    """A report whose shape changes with its content cannot be compared to
    last quarter's."""
    from backend.continuous import report as rp

    book = rp.build(_payload())

    assert len(rp.SHEETS) == 20
    assert len(book.manifest["sheets"]) == 20
    assert book.manifest["sheets"][0] == "Overview"
    assert book.manifest["sheets"][-1] == "Methodology"


def test_an_empty_sheet_says_what_would_have_been_there():
    import io

    from openpyxl import load_workbook

    from backend.continuous import report as rp

    book = rp.build(_payload())
    loaded = load_workbook(io.BytesIO(book.content))
    text = "\n".join(
        str(cell.value) for row in loaded["Validation Performance"].iter_rows()
        for cell in row if cell.value)

    assert "not evidence that anything generalised" in text


def test_the_report_refuses_to_carry_a_secret():
    """The scan runs over the assembled cells, because the leak that matters
    is the one that reached a cell."""
    from backend.continuous import report as rp

    with pytest.raises(rp.ReportError) as caught:
        rp.build(_payload(headline="key is sk-ant-api03-" + "x" * 60))

    assert "may not be written" in str(caught.value)


def test_the_report_declares_what_it_does_not_contain():
    from backend.continuous import report as rp

    book = rp.build(_payload())

    assert book.manifest["contains_sealed_holdout_content"] is False
    assert book.manifest["contains_client_rows"] is False


def test_the_methodology_sheet_states_the_thresholds():
    import io

    from openpyxl import load_workbook

    from backend.continuous import report as rp

    book = rp.build(_payload())
    loaded = load_workbook(io.BytesIO(book.content))
    text = "\n".join(
        str(cell.value) for row in loaded["Methodology"].iter_rows()
        for cell in row if cell.value)

    assert "40% improvement" in text
    assert "tuned against" in text
    assert "no client row" in text
